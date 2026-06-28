# -*- coding: utf-8 -*-
"""
محلّل تقارير أبشر / تم لقوائم المركبات (sheet: vehicle_list).

يستخرج لكل مركبة:
  - رقم اللوحة           (العمود: رقم اللوحة)
  - نوع المركبة          (الماركة + الطراز + سنة الصنع)
  - رقم الإقامة/الهوية    (العمود: رقم هوية المستخدم الفعلي)
  - اسم الموظف           (العمود: اسم المستخدم الفعلي)
  - نوع التسجيل + تاريخ انتهاء رخصة السير + تاريخ انتهاء الفحص (إضافية مفيدة)

ملاحظة: تقرير أبشر لا يحتوي رقم الجوال — يُؤخذ الجوال من قاعدة البيانات أو من تصدير «تم».

الأعمدة تُكتشف بالاسم وليس بالترتيب، فلو تغيّر ترتيب الأعمدة في تصدير لاحق يظل يعمل.
"""
import re

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl مطلوب:  pip install openpyxl")

# مرادفات أسماء الأعمدة كما تظهر في صف العناوين بتقرير أبشر
COL_ALIASES = {
    "plate":    ["رقم اللوحة"],
    "reg":      ["نوع التسجيل"],
    "brand":    ["الماركة"],
    "model":    ["الطراز"],
    "year":     ["سنة الصنع"],
    "iqama":    ["رقم هوية المستخدم الفعلي", "رقم هوية المستخدم", "رقم الهوية"],
    "name":     ["اسم المستخدم الفعلي", "اسم المستخدم"],
    "lic_exp":  ["تاريخ انتهاء رخصة السير"],
    "insp_exp": ["تاريخ انتهاء الفحص"],
}

_AR2EN = {ord(a): e for a, e in zip("٠١٢٣٤٥٦٧٨٩", "0123456789")}


def to_en_digits(s):
    return ("" if s is None else str(s)).translate(_AR2EN)


def norm_plate(p):
    """توحيد رقم اللوحة للمطابقة: أرقام إنجليزية + إزالة المسافات."""
    return re.sub(r"\s+", "", to_en_digits(p)) if p else ""


def norm_id(v):
    """رقم الإقامة/الهوية: أرقام فقط؛ None لو '-' أو فارغ."""
    if v is None:
        return None
    s = re.sub(r"[^\d]", "", to_en_digits(v)).strip()
    return s or None


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("-", "—") else s


def _find_header(rows):
    for idx, row in enumerate(rows):
        vals = [clean(c) for c in row]
        if "رقم اللوحة" in vals and ("اسم المستخدم الفعلي" in vals or "رقم هوية المستخدم الفعلي" in vals):
            return idx, vals
    return None, None


def _map_columns(header):
    idx = {}
    for key, aliases in COL_ALIASES.items():
        for a in aliases:
            if a in header:
                idx[key] = header.index(a)
                break
    return idx


def parse_absher(path):
    """يرجع قائمة سجلات (مركبة واحدة لكل عنصر)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["vehicle_list"] if "vehicle_list" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    h_idx, header = _find_header(rows)
    if h_idx is None:
        raise ValueError("لم يُعثر على صف العناوين (رقم اللوحة) في: %s" % path)
    cmap = _map_columns(header)

    out = []
    for row in rows[h_idx + 1:]:
        if not row:
            continue

        def get(key):
            i = cmap.get(key)
            return row[i] if (i is not None and i < len(row)) else None

        plate = clean(get("plate"))
        if not plate:
            continue
        brand, model, year = clean(get("brand")), clean(get("model")), clean(get("year"))
        vehicle = " ".join(x for x in (brand, model, year) if x).strip()
        out.append({
            "plate": plate,
            "plate_norm": norm_plate(plate),
            "reg": clean(get("reg")),
            "vehicle": vehicle,
            "brand": brand, "model": model, "year": year,
            "iqama": norm_id(get("iqama")),
            "name": clean(get("name")),
            "lic_exp": clean(get("lic_exp")),
            "insp_exp": clean(get("insp_exp")),
            "source": path,
        })
    return out


def parse_many(paths):
    recs = []
    for p in paths:
        recs += parse_absher(p)
    return recs


if __name__ == "__main__":
    import sys
    for p in sys.argv[1:]:
        recs = parse_absher(p)
        auth = [r for r in recs if r["iqama"]]
        print("%-60s vehicles=%-4d authorized_user=%-4d" % (p[:60], len(recs), len(auth)))
