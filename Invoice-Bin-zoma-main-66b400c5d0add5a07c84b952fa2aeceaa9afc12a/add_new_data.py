import openpyxl, sqlite3, os, re
import sys
sys.stdout.reconfigure(encoding='utf-8')

f1 = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\مركبات لها مستخدم فعلي 2026-05-03.xlsx'
f2 = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\مركبات لها مستخدم فعلي 2026-05-03 2.xlsx'
db_path = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\InvoiceApp\database.sqlite'

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

conn = sqlite3.connect(db_path)
c = conn.cursor()

def process_file(file_path):
    print(f"Reading {file_path} ...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    count = 0
    for r in range(3, ws.max_row + 1):
        plate = str(ws.cell(row=r, column=1).value or '').strip()
        make = str(ws.cell(row=r, column=4).value or '').strip()
        model = str(ws.cell(row=r, column=5).value or '').strip()
        year = str(ws.cell(row=r, column=6).value or '').strip()
        name = str(ws.cell(row=r, column=15).value or '').strip()
        
        if plate and name and name != 'None':
            car = f"{make} {model} {year}".strip()
            # check if exists
            c.execute('SELECT id FROM drivers WHERE name = ? OR plate = ?', (name, plate))
            if not c.fetchone():
                c.execute('INSERT INTO drivers (name, plate, car) VALUES (?, ?, ?)', (name, plate, car))
                count += 1
    return count

added = 0
added += process_file(f1)
added += process_file(f2)

conn.commit()
conn.close()
print(f"Added {added} new drivers/vehicles to the database successfully.")

