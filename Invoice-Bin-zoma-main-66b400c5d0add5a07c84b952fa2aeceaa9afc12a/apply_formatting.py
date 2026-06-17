"""
apply_formatting.py
-------------------
يطبق تنسيقاً احترافياً على جميع صفوف البيانات في ملف الكشف
عن طريق نسخ تنسيق الصف الأول كقالب.
"""
import os
import sys
# Fix Windows console encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from copy import copy
from openpyxl import load_workbook

# ── مسارات الملفات ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_TARGET = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026_محدث_مطابق.xlsx")
FILE_OUTPUT = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026_محدث_نهائي.xlsx")

# ── التحقق من وجود الملف ─────────────────────────────────────────────────────
if not os.path.exists(FILE_TARGET):
    print(f"[ERROR] الملف غير موجود: {FILE_TARGET}")
    print("شغّل sync_gps_rows.py أولاً لإنشاء هذا الملف.")
    sys.exit(1)

print("تطبيق التنسيق الاحترافي...")
wb = load_workbook(FILE_TARGET)
ws = wb.active

TEMPLATE_ROW = 6        # الصف الأول كقالب للتنسيق
DATA_ROWS    = 77       # عدد صفوف البيانات
DATA_END_ROW = TEMPLATE_ROW + DATA_ROWS - 1
MAX_COL      = 10       # A:J

# نسخ التنسيق من الصف القالب إلى جميع صفوف البيانات
for r in range(TEMPLATE_ROW + 1, DATA_END_ROW + 1):
    for c in range(1, MAX_COL + 1):
        src = ws.cell(row=TEMPLATE_ROW, column=c)
        tgt = ws.cell(row=r, column=c)
        if src.has_style:
            tgt.font          = copy(src.font)
            tgt.border        = copy(src.border)
            tgt.fill          = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.protection    = copy(src.protection)
            tgt.alignment     = copy(src.alignment)

# ضبط ارتفاع الصفوف
template_h = ws.row_dimensions[TEMPLATE_ROW].height
if template_h:
    for r in range(TEMPLATE_ROW, DATA_END_ROW + 1):
        ws.row_dimensions[r].height = template_h

wb.save(FILE_OUTPUT)
print(f"تم تطبيق التنسيق وحفظ: {FILE_OUTPUT}")

