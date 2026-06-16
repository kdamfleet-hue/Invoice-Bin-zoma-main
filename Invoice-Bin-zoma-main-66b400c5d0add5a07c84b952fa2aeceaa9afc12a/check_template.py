import openpyxl

wb = openpyxl.load_workbook("washing_template.xlsx")
ws = wb.active

f = open("template_detail.txt", "w", encoding="utf-8")
f.write("Merged cells:\n")
for mc in ws.merged_cells.ranges:
    f.write("  %s\n" % str(mc))

f.write("\nRow 1-5 values:\n")
for r in range(1, 6):
    vals = []
    for c in range(1, 19):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            vals.append("(%d,%d)=%s" % (r, c, repr(val)[:80]))
    if vals:
        f.write("  Row %d: %s\n" % (r, " | ".join(vals)))

f.write("\nRow 34-37 values:\n")
for r in range(34, 38):
    vals = []
    for c in range(1, 19):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            vals.append("(%d,%d)=%s" % (r, c, repr(val)[:80]))
    if vals:
        f.write("  Row %d: %s\n" % (r, " | ".join(vals)))

f.write("\nTotal rows: %d\n" % ws.max_row)
f.write("Total cols: %d\n" % ws.max_column)

# Check if rows 5+ are merged
f.write("\nChecking if data rows (5-35) have merged cells:\n")
for mc in ws.merged_cells.ranges:
    mc_str = str(mc)
    f.write("  Merged range: %s\n" % mc_str)

f.close()
print("Saved template_detail.txt")
