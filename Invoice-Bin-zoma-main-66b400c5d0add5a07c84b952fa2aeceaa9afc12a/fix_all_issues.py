# -*- coding: utf-8 -*-
"""Fix all remaining issues: reversed plates, duplicates, missing data."""
import openpyxl, sqlite3, sys, io, os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
BASE = os.path.dirname(os.path.abspath(__file__))

# Load the most reliable source: ربط_الأسماء_باللوحات
name_to_plate = {}
plate_file = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\ربط_الأسماء_باللوحات.xlsx'
wb = openpyxl.load_workbook(plate_file, data_only=True)
ws = wb.active
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(row=r, column=1).value or '').strip()
    plate = str(ws.cell(row=r, column=2).value or '').strip()
    if name and plate:
        name_to_plate[name] = plate
wb.close()

# Load vehicle details from قائمة_المركبات
iqama_to_vehicle = {}
veh_file = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\قائمة_المركبات_بعد_ربط_الأسماء_باللوحة.xlsx'
wb2 = openpyxl.load_workbook(veh_file, data_only=True)
ws2 = wb2.active
for r in range(2, ws2.max_row + 1):
    plate = str(ws2.cell(row=r, column=1).value or '').strip()
    brand = str(ws2.cell(row=r, column=4).value or '').strip()
    model_name = str(ws2.cell(row=r, column=5).value or '').strip()
    year = str(ws2.cell(row=r, column=6).value or '').strip()
    iqama = str(ws2.cell(row=r, column=14).value or '').strip()
    name = str(ws2.cell(row=r, column=15).value or '').strip()
    if iqama and len(iqama) >= 8:
        car_desc = f"{year} {brand} {model_name}".strip()
        iqama_to_vehicle[iqama] = {'plate': plate, 'car': car_desc, 'name_src': name}
    # Also map by plate
    if plate:
        iqama_to_vehicle[f'plate:{plate}'] = {'iqama': iqama, 'car': car_desc, 'name_src': name}
wb2.close()

db = sqlite3.connect(os.path.join(BASE, 'database.sqlite'))
db.row_factory = sqlite3.Row

# Step 1: Fix plates from ربط_الأسماء (most reliable name->plate mapping)
print("=== Step 1: Fix plates from ربط_الأسماء ===")
drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()
for d in drivers:
    name = d['name']
    if name in name_to_plate:
        correct_plate = name_to_plate[name]
        if d['plate'] != correct_plate:
            db.execute("UPDATE drivers SET plate = ? WHERE id = ?", (correct_plate, d['id']))
            print(f"  PLATE FIX: {name}: '{d['plate']}' -> '{correct_plate}'")

# Step 2: Fix car types from قائمة_المركبات (by iqama)
print("\n=== Step 2: Fix car types ===")
drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()
for d in drivers:
    iqama = d['iqama'] or ''
    if iqama in iqama_to_vehicle:
        src = iqama_to_vehicle[iqama]
        # Fix car if missing or from reversed PDF text
        car = d['car'] or ''
        if not car or 'وزوسيا' in car or 'يشيبوستم' in car or 'يرول' in car or 'انيد' in car:
            if src.get('car'):
                db.execute("UPDATE drivers SET car = ? WHERE id = ?", (src['car'], d['id']))
                print(f"  CAR FIX: {d['name']}: '{car}' -> '{src['car']}'")

# Step 3: Fix iqama for drivers who have plate but no iqama
print("\n=== Step 3: Fix missing iqamas ===")
drivers = db.execute("SELECT * FROM drivers WHERE iqama IS NULL OR iqama = '' ORDER BY name").fetchall()
for d in drivers:
    plate = d['plate'] or ''
    if plate:
        key = f'plate:{plate}'
        if key in iqama_to_vehicle:
            src = iqama_to_vehicle[key]
            if src.get('iqama'):
                db.execute("UPDATE drivers SET iqama = ? WHERE id = ?", (src['iqama'], d['id']))
                print(f"  IQAMA FIX: {d['name']}: added iqama={src['iqama']}")

# Step 4: Merge remaining duplicates by iqama
print("\n=== Step 4: Merge duplicates ===")
db.commit()
drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()
iqama_groups = {}
for d in drivers:
    iq = (d['iqama'] or '').strip()
    if iq and len(iq) >= 8:
        if iq not in iqama_groups:
            iqama_groups[iq] = []
        iqama_groups[iq].append(dict(d))

for iq, group in iqama_groups.items():
    if len(group) <= 1:
        continue
    
    def score(d):
        s = len(d.get('name',''))
        if d.get('plate') and 'None' not in str(d.get('plate','')): s += 10
        if d.get('car') and 'None' not in str(d.get('car','')): s += 10
        if d.get('phone') and 'None' not in str(d.get('phone','')): s += 5
        return s
    
    group.sort(key=score, reverse=True)
    best = group[0]
    
    for other in group[1:]:
        for field in ['plate', 'car', 'phone', 'empid']:
            ov = other.get(field, '')
            bv = best.get(field, '')
            if ov and ov != 'None' and (not bv or bv == 'None'):
                best[field] = ov
        
        db.execute("DELETE FROM drivers WHERE id = ?", (other['id'],))
        print(f"  MERGED: '{other['name']}' (id={other['id']}) -> '{best['name']}' (id={best['id']})")
    
    db.execute("UPDATE drivers SET plate=?, car=?, phone=?, empid=? WHERE id=?",
        (best.get('plate'), best.get('car'), best.get('phone'), best.get('empid'), best['id']))

db.commit()

# Final report
print(f"\n{'='*60}")
total = db.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
no_plate = db.execute("SELECT COUNT(*) FROM drivers WHERE plate IS NULL OR plate = '' OR plate = 'None'").fetchone()[0]
no_car = db.execute("SELECT COUNT(*) FROM drivers WHERE car IS NULL OR car = '' OR car = 'None'").fetchone()[0]
no_iqama = db.execute("SELECT COUNT(*) FROM drivers WHERE iqama IS NULL OR iqama = ''").fetchone()[0]
no_phone = db.execute("SELECT COUNT(*) FROM drivers WHERE phone IS NULL OR phone = '' OR phone = 'None'").fetchone()[0]
print(f"FINAL STATS:")
print(f"  Total drivers: {total}")
print(f"  Missing plate: {no_plate}")
print(f"  Missing car:   {no_car}")
print(f"  Missing iqama: {no_iqama}")
print(f"  Missing phone: {no_phone}")

print(f"\n=== ALL DRIVERS ===")
drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()
for i, d in enumerate(drivers, 1):
    ok = '✅' if d['plate'] and d['car'] and d['iqama'] else '⚠️'
    print(f"  {ok} [{i}] {d['name']} | 🪪{d['iqama'] or '-'} | 🚗{d['plate'] or '-'} | 🚛{d['car'] or '-'} | 📱{d['phone'] or '-'}")

db.close()

