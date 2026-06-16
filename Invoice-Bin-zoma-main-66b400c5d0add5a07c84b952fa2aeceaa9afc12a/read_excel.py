# -*- coding: utf-8 -*-
import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]

for fname in files:
    fpath = os.path.join(folder, fname)
    try:
        wb = openpyxl.load_workbook(fpath)
    except:
        continue
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Print first 8 rows only
        for r in range(1, 9):
            row_data = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_data.append(f'C{c}={val}')
            sep = ' | '
            if row_data:
                print(f'Row {r}: {sep.join(row_data)}')
            else:
                print(f'Row {r}: (empty)')

# Check for logos
archive = os.path.join(folder, 'أرشيف')
logos = [f for f in os.listdir(archive) if f.endswith('.png')]
print(f'\nLogos in archive: {logos}')

