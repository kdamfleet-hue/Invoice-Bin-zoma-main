import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

for f in os.listdir('.'):
    if f.endswith('.xlsx') and 'تحديث' in f:
        wb = openpyxl.load_workbook(f)
        
        # Sheet 1 - main data
        ws = wb['جدول تحديث المركبات']
        print("=== MAIN SHEET: جدول تحديث المركبات ===")
        print(f"Total data rows (after row 8 headers): counting...")
        count = 0
        for r in range(9, ws.max_row+1):
            if ws.cell(row=r, column=1).value is not None:
                count += 1
        print(f"  Data rows: {count}")
        
        # Check summary/footer area
        print("\n=== Summary rows (38+) ===")
        for r in range(38, min(62, ws.max_row+1)):
            row_data = []
            for c in range(1, 18):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    row_data.append(f'C{c}={cell.value}')
            if row_data:
                print(f'  Row {r}: {" | ".join(row_data)}')
        
        # Sheet 2 - spare/broken vehicles
        ws2 = wb['الورقة 1']
        print("\n=== SHEET 2: الورقة 1 (Spare/Broken) ===")
        for r in range(7, min(30, ws2.max_row+1)):
            row_data = []
            for c in range(1, 20):
                cell = ws2.cell(row=r, column=c)
                if cell.value is not None:
                    row_data.append(f'C{c}={cell.value}')
            if row_data:
                print(f'  Row {r}: {" | ".join(row_data)}')
        break

