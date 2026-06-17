# -*- coding: utf-8 -*-
"""
Style vehicles Excel with driver card numbers from بيانات السائقين file.
Matches by plate number. Adds رقم بطاقة السائق + تاريخ انتهاء البطاقة columns.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os, re

script_dir = os.path.dirname(os.path.abspath(__file__))
archive_dir = os.path.join(script_dir, 'أرشيف')

# ===== FIND FILES =====
source_file = kart_file = driver_file = None
for f in os.listdir(script_dir):
    if not f.endswith('.xlsx') or 'احترافي' in f:
        continue
    if 'تحديث' in f:
        source_file = os.path.join(script_dir, f)
    elif 'دينا' in f and 'لوري' in f:
        kart_file = os.path.join(script_dir, f)
    elif 'بيانات' in f:
        driver_file = os.path.join(script_dir, f)

print(f"Main: {source_file}\nKart: {kart_file}\nDriver: {driver_file}")

# Logos
logo_ar = os.path.join(archive_dir, 'source_img_1.png')
logo_en = os.path.join(archive_dir, 'source_img_0.png')
logo_center = os.path.join(archive_dir, 'output-onlinepngtools.png')

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

def clean_val(val):
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    return str(val).strip().replace('\n','').replace('\r','')

# ===== READ كرت التشغيل =====
kart_map = {}
if kart_file:
    ks = openpyxl.load_workbook(kart_file).active
    for r in list(range(6,22)) + list(range(28,50)):
        plate = ks.cell(row=r, column=11).value
        kart = ks.cell(row=r, column=5).value
        if plate and kart:
            kart_map[normalize_plate(plate)] = clean_val(kart)
print(f"Kart entries: {len(kart_map)}")

# ===== READ DRIVER CARD DATA =====
driver_card_map = {}  # plate -> (card_number, expiry_date)
if driver_file:
    ds = openpyxl.load_workbook(driver_file).active
    for r in range(9, ds.max_row + 1):
        plate = ds.cell(row=r, column=7).value
        card_num = ds.cell(row=r, column=8).value
        card_exp = ds.cell(row=r, column=9).value
        name = ds.cell(row=r, column=3).value
        if plate and card_num:
            np = normalize_plate(plate)
            driver_card_map[np] = (clean_val(card_num), clean_val(card_exp), clean_val(name))
            print(f"  Driver: {plate} -> card={card_num}, exp={clean_val(card_exp)}, name={name}")
print(f"Driver card entries: {len(driver_card_map)}")

# ===== READ MAIN SOURCE =====
src_ws = openpyxl.load_workbook(source_file).active

def read_section(ws, hdr_row, start, end):
    rows = []
    for r in range(start, end + 1):
        rd = []
        has = False
        for c in range(1, 17):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip(): has = True
            rd.append(v)
        if has:
            plate = rd[11]  # C12
            np = normalize_plate(plate)
            if np in kart_map: rd[4] = kart_map[np]
            card_info = driver_card_map.get(np, ('', '', ''))
            rd.append(card_info[0])  # رقم بطاقة السائق (index 16)
            rd.append(card_info[1])  # تاريخ انتهاء البطاقة (index 17)
            if card_info[0]:
                print(f"  Matched {plate}: card={card_info[0]}, name_src={card_info[2]}")
            rows.append(rd)
    return rows

print("\n=== Dina ===")
dina_data = read_section(src_ws, 5, 6, 22)
print(f"\n=== Lorry ===")
lorry_data = read_section(src_ws, 28, 29, 48)

# ================================================================
# CREATE STYLED WORKBOOK
# ================================================================
wb = Workbook()
ws = wb.active
ws.title = "تحديث تواريخ السيارات"
ws.sheet_view.rightToLeft = True
ws.sheet_view.showGridLines = False

# Theme
P = "1A3A5C"; HB = "2C4A6E"; AG = "C8A45A"; W = "FFFFFF"
RA1 = "EDF2F7"; RA2 = "FFFFFF"; TD = "1A1A2E"; TS = "5A6A7A"
GA = "27AE60"; RI = "C0392B"; AN = "B8860B"; LB = "F8F9FA"
S2 = "34495E"; SB = "F0F2F5"

ts = Side(style='thin', color='D0D5DD')
mn = Side(style='medium', color=P)
ag = Side(style='medium', color=AG)
tb = Border(left=ts, right=ts, top=ts, bottom=ts)
hb = Border(left=Side(style='thin',color=HB), right=Side(style='thin',color=HB), top=mn, bottom=mn)
sb = Border(left=Side(style='thin',color=P), right=Side(style='thin',color=P), top=mn, bottom=Side(style='medium',color=AG))
def fl(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')

# 18 columns now (added 2: بطاقة السائق + انتهاء البطاقة)
headers = ['م','انتهاء التأمين م','انتهاء الفحص هـ','انتهاء الاستمارة',
    'انتهاء كرت التشغيل','انتهاء رخصة لوحه إعلانية','اسم المستخدم',
    'الرقم الوظيفي','شريحة الديزل','GPS','الملاحظات',
    'رقم اللوحة','الموديل','نوع السيارة','الحمولة','الرقم التسلسلي',
    'رقم بطاقة السائق','تاريخ انتهاء البطاقة']
NCOLS = 18
last_col = get_column_letter(NCOLS)

widths = [5,16,16,16,16,20,18,12,12,8,25,16,10,18,10,16,18,16]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w

row = 1
# Gold top stripe
for c in range(1,NCOLS+1):
    ws.cell(row=row,column=c).fill = fl(AG)
ws.row_dimensions[row].height = 4
row = 2

# Logos
ws.merge_cells(f'A{row}:{last_col}{row+2}')
for r in range(row,row+3): ws.row_dimensions[r].height = 28
try:
    i1 = XlImage(logo_ar); i1.width=200; i1.height=70; ws.add_image(i1,'A2')
except: pass
try:
    i2 = XlImage(logo_en); i2.width=200; i2.height=70; ws.add_image(i2,'P2')
except: pass
try:
    i3 = XlImage(logo_center); i3.width=280; i3.height=70; ws.add_image(i3,'G2')
except: pass
row = 5

# Date
ws.merge_cells(f'A{row}:C{row}')
ws[f'A{row}'] = "2026/04/30 م"
ws[f'A{row}'].font = Font(name='Cairo',size=10,color=TS)
ws[f'A{row}'].alignment = Alignment(horizontal='left',vertical='center')
ws.row_dimensions[row].height = 22
row = 6

# Title
ws.merge_cells(f'A{row}:{last_col}{row}')
ws[f'A{row}'] = "تحديث تواريخ الانتهاء للسيارات لوري ودينات — فرع الدمام 2026"
ws[f'A{row}'].font = Font(name='Cairo',size=14,bold=True,color=P)
ws[f'A{row}'].alignment = Alignment(horizontal='center',vertical='center')
ws.row_dimensions[row].height = 40
row = 7
ws.row_dimensions[row].height = 8
row = 8

def write_section(ws, row, title, hdrs, data, sc):
    # Section title bar
    ws.merge_cells(f'A{row}:{last_col}{row}')
    c = ws[f'A{row}']
    c.value = title; c.font = Font(name='Cairo',size=13,bold=True,color=W)
    c.fill = fl(sc); c.alignment = Alignment(horizontal='center',vertical='center')
    c.border = sb
    for cc in range(2,NCOLS+1):
        ws.cell(row=row,column=cc).fill = fl(sc)
        ws.cell(row=row,column=cc).border = sb
    ws.row_dimensions[row].height = 36
    row += 1

    # Column headers
    ws.row_dimensions[row].height = 36
    for i,h in enumerate(hdrs):
        c = ws.cell(row=row,column=i+1,value=h)
        c.font = Font(name='Cairo',size=9,bold=True,color=W)
        c.fill = fl(HB); c.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border = hb
    row += 1

    # Data rows
    for ri,dr in enumerate(data):
        ws.row_dimensions[row].height = 27
        bg = fl(RA1) if ri%2==0 else fl(RA2)
        for ci in range(NCOLS):
            col = ci+1
            val = dr[ci] if ci < len(dr) else None
            cell = ws.cell(row=row,column=col)
            cell.border = tb; cell.fill = bg
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            if col == 1:
                cell.value = ri+1
                cell.font = Font(name='Cairo',size=10,bold=True,color=P)
            else:
                dv = clean_val(val)
                cell.value = dv
                if col in (2,3,4,5,6):
                    cell.font = Font(name='Cairo',size=10,color=TD)
                    if col == 5 and dv: cell.font = Font(name='Cairo',size=10,bold=True,color=TD)
                elif col == 7:
                    cell.font = Font(name='Cairo',size=10,bold=True,color=TD)
                elif col == 8:
                    cell.font = Font(name='Cairo',size=10,color=TS)
                elif col == 9:
                    if dv in ('مفعل','مفعله'): cell.font = Font(name='Cairo',size=10,bold=True,color=GA)
                    elif 'غير' in dv or 'لا' in dv: cell.font = Font(name='Cairo',size=10,bold=True,color=RI)
                    else: cell.font = Font(name='Cairo',size=10,color=TD)
                elif col == 11:
                    cell.font = Font(name='Cairo',size=9,italic=True,color=AN) if dv else Font(name='Cairo',size=9,color=TS)
                    cell.alignment = Alignment(horizontal='right',vertical='center',wrap_text=True)
                elif col == 12:
                    cell.font = Font(name='Cairo',size=10,bold=True,color=P)
                elif col == 14:
                    cell.font = Font(name='Cairo',size=10,bold=True,color=TD)
                elif col == 16:
                    cell.font = Font(name='Consolas',size=9,color=TS)
                elif col == 17:  # رقم بطاقة السائق
                    cell.font = Font(name='Consolas',size=9,bold=True,color=P)
                elif col == 18:  # تاريخ انتهاء البطاقة
                    cell.font = Font(name='Cairo',size=10,color=TD)
                else:
                    cell.font = Font(name='Cairo',size=10,color=TD)
        row += 1

    # Summary
    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws[f'A{row}'] = f"إجمالي عدد المركبات: {len(data)}"
    ws[f'A{row}'].font = Font(name='Cairo',size=11,bold=True,color=P)
    ws[f'A{row}'].fill = fl(LB)
    ws[f'A{row}'].alignment = Alignment(horizontal='center',vertical='center')
    ws[f'A{row}'].border = Border(top=Side(style='thin',color=AG),bottom=ag,left=ts,right=ts)
    for cc in range(2,NCOLS+1):
        ws.cell(row=row,column=cc).fill = fl(LB)
        ws.cell(row=row,column=cc).border = Border(top=Side(style='thin',color=AG),bottom=ag,left=ts,right=ts)
    ws.row_dimensions[row].height = 30
    row += 1
    return row

row = write_section(ws,row,"جدول سيارات الدينا (ايسوزو - متسوبيشي - هينو)",headers,dina_data,P)
ws.row_dimensions[row].height = 18; row += 1
row = write_section(ws,row,"جدول سيارات اللوري (ايسوزو - متسوبيشي)",headers,lorry_data,S2)
ws.row_dimensions[row].height = 12; row += 1

# Closing
ws.merge_cells(f'A{row}:{last_col}{row}')
ws[f'A{row}'] = "شركة بن زومة للتجارة الدولية والإنماء المحدودة  —  فرع الدمام  —  قسم الحركة"
ws[f'A{row}'].font = Font(name='Cairo',size=12,bold=True,color=P)
ws[f'A{row}'].fill = fl(LB)
ws[f'A{row}'].alignment = Alignment(horizontal='center',vertical='center')
ws[f'A{row}'].border = Border(top=ag,bottom=ag)
for cc in range(2,NCOLS+1):
    ws.cell(row=row,column=cc).fill = fl(LB)
    ws.cell(row=row,column=cc).border = Border(top=ag,bottom=ag)
ws.row_dimensions[row].height = 38; row += 1
ws.row_dimensions[row].height = 25; row += 1

# Signatures (9 blocks across 18 cols = 2 cols each)
sigs = ['السائق','الميكانيكي','مسئول الورشة','قسم الحركة','مسئول الشراء','الحسابات','مدير الفرع','اعتماد الإدارة','']
ws.row_dimensions[row].height = 28
for i,label in enumerate(sigs):
    if not label: continue
    c = i*2+1
    if c+1 <= NCOLS:
        ws.merge_cells(start_row=row,start_column=c,end_row=row,end_column=c+1)
        cell = ws.cell(row=row,column=c,value=label)
        cell.font = Font(name='Cairo',size=9,bold=True,color=P)
        cell.fill = fl(SB); cell.alignment = Alignment(horizontal='center',vertical='center')
        cell.border = tb; ws.cell(row=row,column=c+1).fill = fl(SB); ws.cell(row=row,column=c+1).border = tb
row += 1

ws.row_dimensions[row].height = 45
for i in range(8):
    c = i*2+1
    if c+1 <= NCOLS:
        ws.merge_cells(start_row=row,start_column=c,end_row=row,end_column=c+1)
        cell = ws.cell(row=row,column=c)
        cell.border = Border(left=ts,right=ts,bottom=Side(style='thin',color='CCCCCC'))
        ws.cell(row=row,column=c+1).border = Border(left=ts,right=ts,bottom=Side(style='thin',color='CCCCCC'))
        if i == 3:
            cell.value = "خالد الغامدي"
            cell.font = Font(name='Cairo',size=9,color=TS)
            cell.alignment = Alignment(horizontal='center',vertical='top')
row += 1; ws.row_dimensions[row].height = 15; row += 1

# Notes
ws.merge_cells(f'A{row}:{last_col}{row}')
ws[f'A{row}'] = "ملاحظات هامة"
ws[f'A{row}'].font = Font(name='Cairo',size=12,bold=True,color=W)
ws[f'A{row}'].fill = fl(P); ws[f'A{row}'].alignment = Alignment(horizontal='center',vertical='center')
for cc in range(2,NCOLS+1): ws.cell(row=row,column=cc).fill = fl(P)
ws.row_dimensions[row].height = 30; row += 1

for note in [
    "1- يعبأ الجدول من قبل مسئول الحركة بعد مراجعة تواريخ انتهاء كل مركبة",
    "2- يتم تحديث الجدول بشكل دوري عند تجديد أي من الوثائق (تأمين، فحص، استمارة، كرت تشغيل)",
    "3- عند وجود ملاحظة لأي من المركبات يمكن تدوينها في حقل الملاحظات",
    "4- ترفق صورة للإدارة المالية وصورة لقسم الحركة بالإدارة، ويتم حفظ صورة من الجدول لدى الفرع"
]:
    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws[f'A{row}'] = note
    ws[f'A{row}'].font = Font(name='Cairo',size=10,color=TD)
    ws[f'A{row}'].alignment = Alignment(horizontal='right',vertical='center',wrap_text=True)
    ws[f'A{row}'].fill = fl(LB)
    for cc in range(2,NCOLS+1): ws.cell(row=row,column=cc).fill = fl(LB)
    ws.row_dimensions[row].height = 22; row += 1

# Print setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25; ws.page_margins.right = 0.25
ws.page_margins.top = 0.3; ws.page_margins.bottom = 0.3

out = os.path.join(script_dir, 'تحديث_تواريخ_السيارات_احترافي_2026.xlsx')
wb.save(out)
print(f"\n✅ Saved: {out}")

