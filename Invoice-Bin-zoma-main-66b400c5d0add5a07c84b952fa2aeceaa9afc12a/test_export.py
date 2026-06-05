import openpyxl, json, io
from copy import copy as shallow_copy

# Simulate the export with test data
vehicles = [
    {"id": i, "plate": "ب ط س %d" % (3260+i), "type": "ايسوزو 2023", "driver": "سائق %d" % i, "m": [1,0,1,0,1,0,0,0,0,0,0,0]}
    for i in range(1, 85)  # 84 vehicles to test beyond template limit
]

template_path = "washing_template.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Unmerge ALL merged cells
merged_ranges = list(ws.merged_cells.ranges)
for mr in merged_ranges:
    ws.unmerge_cells(str(mr))

# Copy styles
data_fills = {}
data_fonts = {}
data_aligns = {}
data_borders = {}
for c in range(1, 19):
    cell = ws.cell(row=5, column=c)
    data_fills[c] = shallow_copy(cell.fill)
    data_fonts[c] = shallow_copy(cell.font)
    data_aligns[c] = shallow_copy(cell.alignment)
    data_borders[c] = shallow_copy(cell.border)

# Write data
for idx, v in enumerate(vehicles):
    r = 5 + idx
    ws.cell(row=r, column=1, value=v.get("id", idx + 1))
    ws.cell(row=r, column=2, value=v.get("plate", ""))
    ws.cell(row=r, column=3, value=v.get("type", ""))
    ws.cell(row=r, column=4, value=v.get("driver", ""))
    months = v.get("m", [])
    total = sum(months)
    for m_idx in range(12):
        val = "استلم" if m_idx < len(months) and months[m_idx] == 1 else None
        ws.cell(row=r, column=5 + m_idx, value=val)
    ws.cell(row=r, column=17, value=total)
    for c in range(1, 19):
        cell = ws.cell(row=r, column=c)
        cell.fill = shallow_copy(data_fills.get(c, data_fills[1]))
        cell.font = shallow_copy(data_fonts.get(c, data_fonts[1]))
        cell.alignment = shallow_copy(data_aligns.get(c, data_aligns[1]))
        cell.border = shallow_copy(data_borders.get(c, data_borders[1]))

# Summary row
summary_row = 5 + len(vehicles)
ws.cell(row=summary_row, column=1, value="إجمالي الغسيل الشهري")
ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
for m_idx in range(12):
    col = 5 + m_idx
    start_cell = ws.cell(row=5, column=col).coordinate.replace("5", "")
    formula = '=COUNTIF(%s5:%s%d,"استلم")' % (start_cell, start_cell, summary_row - 1)
    ws.cell(row=summary_row, column=col, value=formula)
ws.cell(row=summary_row, column=17, value="=SUM(Q5:Q%d)" % (summary_row - 1))

# Footer
footer_row = summary_row + 2
ws.cell(row=footer_row, column=1, value="تم إعداد هذا الجدول بواسطة قسم الحركة - فرع الدمام")
ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=18)

# Re-merge headers
ws.merge_cells("A1:R1")
ws.merge_cells("A2:D2")
ws.merge_cells("A3:R3")

total_washes = sum(sum(v.get("m", [])) for v in vehicles)
ws.cell(row=2, column=12, value=total_washes)

output = io.BytesIO()
wb.save(output)
print("SUCCESS! File exported with %d vehicles, size: %d bytes" % (len(vehicles), output.tell()))
