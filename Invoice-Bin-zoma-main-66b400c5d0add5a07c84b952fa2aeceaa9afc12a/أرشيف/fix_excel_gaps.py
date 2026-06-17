import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_محسن تم التعديل عليه 3.xlsx')
output_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_بدون_فراغات.xlsx')

wb = openpyxl.load_workbook(input_file)
ws = wb.active

# ===== Delete empty rows =====
# We scan from bottom to top to safely delete rows
for row_idx in range(ws.max_row, 0, -1):
    row_is_empty = True
    for cell in ws[row_idx]:
        if cell.value is not None and str(cell.value).strip() != "":
            row_is_empty = False
            break
    if row_is_empty:
        ws.delete_rows(row_idx)

# ===== Clean and Merge Columns in Main Table =====
# The user complains about "فراغات" (gaps). Let's see if we should merge the columns in the main table
# so there are no empty gaps between data.
# First, let's find the main table.
main_header_row = None
for row_idx in range(1, ws.max_row + 1):
    val = str(ws.cell(row=row_idx, column=1).value).strip()
    if val == "م":
        main_header_row = row_idx
        break

if main_header_row:
    # Now loop through data of main table
    # Columns are likely A, B, D, E, F, G, I based on dump.
    # C, H, J are empty. We should merge B:C, G:H, I:J for the main table so the data spans beautifully.
    for r in range(main_header_row, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip() == "📋  ملاحظات وتفاصيل الفلاتر":
            break
        # Merge B and C
        if not type(ws.cell(row=r, column=2)).__name__ == 'MergedCell':
            try: ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            except: pass
        # Merge G and H
        if not type(ws.cell(row=r, column=7)).__name__ == 'MergedCell':
            try: ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
            except: pass
        # Merge I and J
        if not type(ws.cell(row=r, column=9)).__name__ == 'MergedCell':
            try: ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
            except: pass

# Also center everything
for row in ws.iter_rows():
    for cell in row:
        if cell.alignment.horizontal != 'center':
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb.save(output_file)
print("Saved cleaned file to", output_file)

