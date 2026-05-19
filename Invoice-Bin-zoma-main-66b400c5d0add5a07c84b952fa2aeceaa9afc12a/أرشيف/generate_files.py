# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import base64
from data import *

script_dir = os.path.dirname(os.path.abspath(__file__))

# Read logo as base64
logo_path = os.path.join(script_dir, 'output-onlinepngtools.png')
with open(logo_path, 'rb') as f:
    logo_b64 = base64.b64encode(f.read()).decode('utf-8')

# ============ GENERATE HTML ============
html = f'''<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>بيان استهلاك الزيوت والفلاتر - فرع الدمام</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;800&display=swap');

:root {{
  --primary: #1a3a5c;
  --primary-light: #2a5a8c;
  --accent: #c8a45a;
  --accent-light: #e8c97a;
  --bg: #f8f6f1;
  --card-bg: #ffffff;
  --text: #2c3e50;
  --text-light: #6b7b8d;
  --border: #e0ddd5;
  --header-gradient: linear-gradient(135deg, #1a3a5c 0%, #2a5a8c 50%, #1a3a5c 100%);
  --row-even: #f9f7f2;
  --row-hover: #eef5ff;
  --success: #27ae60;
  --shadow: 0 4px 24px rgba(26, 58, 92, 0.08);
  --shadow-lg: 0 8px 40px rgba(26, 58, 92, 0.12);
}}

* {{ margin: 0; padding: 0; box-sizing: border-box; }}

body {{
  font-family: 'Cairo', sans-serif;
  background: var(--bg);
  color: var(--text);
  line-height: 1.7;
  min-height: 100vh;
}}

.page-container {{
  max-width: 1100px;
  margin: 0 auto;
  padding: 30px 20px;
}}

/* ===== HEADER / LOGO ===== */
.header {{
  text-align: center;
  margin-bottom: 30px;
  padding: 30px 20px 20px;
  background: var(--card-bg);
  border-radius: 16px;
  box-shadow: var(--shadow);
  border-bottom: 3px solid var(--accent);
  position: relative;
  overflow: hidden;
}}
.header::before {{
  content: '';
  position: absolute;
  top: 0; left: 0; right: 0;
  height: 4px;
  background: var(--header-gradient);
}}
.header img {{
  max-width: 420px;
  width: 80%;
  height: auto;
  margin-bottom: 18px;
}}
.header h1 {{
  font-size: 1.3rem;
  font-weight: 700;
  color: var(--primary);
  margin-bottom: 4px;
  letter-spacing: 0.3px;
}}
.header .subtitle {{
  font-size: 0.95rem;
  color: var(--text-light);
  font-weight: 400;
}}

/* ===== TABLE SECTION ===== */
.table-section {{
  background: var(--card-bg);
  border-radius: 16px;
  box-shadow: var(--shadow);
  overflow: hidden;
  margin-bottom: 30px;
}}
.table-section-header {{
  background: var(--header-gradient);
  color: #fff;
  padding: 16px 24px;
  font-size: 1.1rem;
  font-weight: 700;
  display: flex;
  align-items: center;
  gap: 10px;
}}
.table-section-header .icon {{
  width: 32px; height: 32px;
  background: rgba(255,255,255,0.15);
  border-radius: 8px;
  display: flex; align-items: center; justify-content: center;
  font-size: 1.1rem;
}}

.table-wrapper {{
  overflow-x: auto;
  padding: 0;
}}

table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 0.92rem;
}}
thead th {{
  background: var(--primary);
  color: #fff;
  padding: 14px 12px;
  font-weight: 700;
  text-align: center;
  white-space: nowrap;
  border-left: 1px solid rgba(255,255,255,0.15);
  font-size: 0.88rem;
  letter-spacing: 0.3px;
}}
thead th:last-child {{ border-left: none; }}

tbody td {{
  padding: 11px 12px;
  text-align: center;
  border-bottom: 1px solid var(--border);
  border-left: 1px solid var(--border);
  transition: background 0.2s;
  font-size: 0.88rem;
}}
tbody td:last-child {{ border-left: none; }}

tbody tr:nth-child(even) {{ background: var(--row-even); }}
tbody tr:hover {{ background: var(--row-hover); }}

tbody tr td:first-child {{
  font-weight: 700;
  color: var(--primary);
  width: 40px;
}}
.badge {{
  display: inline-block;
  padding: 2px 10px;
  border-radius: 20px;
  font-size: 0.8rem;
  font-weight: 600;
}}
.badge-zero {{
  background: #f0f0f0;
  color: #999;
}}
.badge-filter {{
  background: #eef7ee;
  color: #27ae60;
  border: 1px solid #c8e6c8;
}}

/* ===== NOTES SECTION ===== */
.notes-section {{
  background: var(--card-bg);
  border-radius: 16px;
  box-shadow: var(--shadow);
  overflow: hidden;
  margin-bottom: 30px;
}}
.notes-header {{
  background: linear-gradient(135deg, #c8a45a 0%, #a8843a 100%);
  color: #fff;
  padding: 16px 24px;
  font-size: 1.1rem;
  font-weight: 700;
  display: flex; align-items: center; gap: 10px;
}}
.notes-body {{
  padding: 24px;
}}
.notes-body h3 {{
  color: var(--primary);
  font-size: 1.05rem;
  margin-bottom: 8px;
  padding-bottom: 8px;
  border-bottom: 2px solid var(--accent);
  display: inline-block;
}}
.notes-body .info-box {{
  background: #f0f4f8;
  border-right: 4px solid var(--primary);
  padding: 14px 18px;
  border-radius: 0 10px 10px 0;
  margin: 14px 0 20px;
  font-size: 0.95rem;
  line-height: 2;
}}
.filters-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
  gap: 8px;
  margin-top: 14px;
}}
.filter-item {{
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 8px 14px;
  background: #fafaf8;
  border-radius: 8px;
  border: 1px solid var(--border);
  font-size: 0.88rem;
  transition: all 0.2s;
}}
.filter-item:hover {{
  background: #eef5ff;
  border-color: var(--primary-light);
  transform: translateX(-2px);
}}
.filter-num {{
  min-width: 26px; height: 26px;
  background: var(--primary);
  color: #fff;
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 0.75rem;
  font-weight: 700;
  flex-shrink: 0;
}}

/* ===== DIESEL SECTION ===== */
.diesel-section {{
  margin-top: 20px;
  padding-top: 18px;
  border-top: 2px dashed var(--border);
}}
.diesel-item {{
  padding: 6px 14px;
  font-size: 0.88rem;
  color: var(--text);
}}
.diesel-item.sub {{
  color: var(--text-light);
  font-size: 0.82rem;
  padding-right: 34px;
}}

/* ===== SIGNATURE ===== */
.signature-section {{
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  padding: 24px;
  border-top: 2px solid var(--border);
  margin-top: 10px;
}}
.sig-block {{
  text-align: center;
}}
.sig-block .label {{
  font-weight: 700;
  color: var(--primary);
  font-size: 0.95rem;
}}
.sig-block .name {{
  color: var(--text-light);
  font-size: 0.88rem;
  margin-top: 4px;
}}
.sig-line {{
  width: 120px;
  height: 2px;
  background: var(--border);
  margin: 30px auto 0;
}}

.closing-note {{
  text-align: center;
  padding: 18px;
  background: #f0f4f8;
  border-radius: 10px;
  margin: 20px 24px 24px;
  font-weight: 600;
  color: var(--primary);
}}

/* ===== FOOTER ===== */
.footer {{
  text-align: center;
  padding: 18px;
  color: var(--text-light);
  font-size: 0.8rem;
}}

/* ===== PRINT ===== */
@media print {{
  body {{ background: #fff; }}
  .page-container {{ max-width: 100%; padding: 10px; }}
  .header, .table-section, .notes-section {{ box-shadow: none; border-radius: 0; }}
  tbody tr:hover {{ background: inherit; }}
}}
</style>
</head>
<body>
<div class="page-container">

  <!-- HEADER -->
  <div class="header">
    <img src="data:image/png;base64,{logo_b64}" alt="شركة بن زومة للتجارة الدولية والإنماء المحدودة">
    <h1>{TITLE}</h1>
  </div>

  <!-- MAIN TABLE -->
  <div class="table-section">
    <div class="table-section-header">
      <div class="icon">📋</div>
      <span>جدول بيان الاستهلاك</span>
    </div>
    <div class="table-wrapper">
      <table>
        <thead>
          <tr>
'''

for h in HEADERS:
    html += f'            <th>{h}</th>\n'

html += '''          </tr>
        </thead>
        <tbody>
'''

for row in TABLE_DATA:
    html += '          <tr>\n'
    for idx, cell in enumerate(row):
        if idx == 6:  # filter column
            if str(cell).strip() == "0":
                html += f'            <td><span class="badge badge-zero">لا يوجد</span></td>\n'
            else:
                html += f'            <td><span class="badge badge-filter">{cell}</span></td>\n'
        else:
            html += f'            <td>{cell}</td>\n'
    html += '          </tr>\n'

html += '''        </tbody>
      </table>
    </div>
  </div>

  <!-- NOTES SECTION -->
  <div class="notes-section">
    <div class="notes-header">
      <div class="icon" style="background:rgba(255,255,255,0.2);border-radius:8px;width:32px;height:32px;display:flex;align-items:center;justify-content:center;">🔧</div>
      <span>ملاحظات وتفاصيل الفلاتر</span>
    </div>
    <div class="notes-body">
      <h3>''' + NOTES_TITLE + '''</h3>
      <div class="info-box">''' + NOTES_SUBTITLE.replace('\n', '<br>') + '''</div>

      <h3>''' + FILTERS_TITLE + '''</h3>
      <div class="filters-grid">
'''

for i, f_item in enumerate(FILTERS_LIST):
    html += f'        <div class="filter-item"><div class="filter-num">{i+1}</div><span>{f_item}</span></div>\n'

html += '''      </div>

      <div class="diesel-section">
        <h3>فلاتر الديزل</h3>
        <div style="margin-top:10px;">
'''

for d_item in DIESEL_FILTERS:
    if d_item.startswith("تركب"):
        html += f'          <div class="diesel-item sub">↳ {d_item}</div>\n'
    else:
        html += f'          <div class="diesel-item">• {d_item}</div>\n'

html += '''        </div>
      </div>

      <div class="closing-note">''' + CLOSING_NOTE.replace('\n', '<br>') + '''</div>

      <div class="signature-section">
        <div class="sig-block">
          <div class="label">قسم الحركة</div>
          <div class="name">خالد الغامدي</div>
          <div class="sig-line"></div>
        </div>
        <div class="sig-block">
          <div class="label">التعميد</div>
          <div class="sig-line"></div>
        </div>
      </div>
    </div>
  </div>

  <div class="footer">شركة بن زومة للتجارة الدولية والإنماء المحدودة &copy; 2026</div>
</div>
</body>
</html>'''

html_path = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_محسن.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"HTML saved: {html_path}")

# ============ GENERATE EXCEL ============
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "بيان الاستهلاك"
ws.sheet_view.rightToLeft = True

# Colors
primary_color = "1A3A5C"
accent_color = "C8A45A"
header_font_color = "FFFFFF"
even_row_color = "F5F3EE"
border_color = "D5D0C8"

# Borders
thin_border = Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
    top=Side(style='thin', color=border_color),
    bottom=Side(style='thin', color=border_color)
)
header_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color=primary_color),
    bottom=Side(style='medium', color=primary_color)
)

# Column widths
col_widths = [6, 16, 18, 18, 16, 14, 26]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row 1-3: Logo space
ws.merge_cells('A1:G3')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Insert logo
try:
    img = XlImage(logo_path)
    img.width = 380
    img.height = 75
    ws.add_image(img, 'C1')
except:
    ws['A1'] = "شركة بن زومة للتجارة الدولية والإنماء المحدودة"
    ws['A1'].font = Font(name='Cairo', size=16, bold=True, color=primary_color)

# Row 4: Spacer
ws.row_dimensions[4].height = 8

# Row 5: Title
ws.merge_cells('A5:G5')
ws['A5'] = TITLE
ws['A5'].font = Font(name='Cairo', size=12, bold=True, color=primary_color)
ws['A5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[5].height = 40

# Row 6: Spacer
ws.row_dimensions[6].height = 8

# Row 7: Headers
header_row = 7
ws.row_dimensions[header_row].height = 32
for i, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=header_row, column=i, value=h)
    cell.font = Font(name='Cairo', size=10, bold=True, color=header_font_color)
    cell.fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = header_border

# Data rows
start_row = 8
for r_idx, row in enumerate(TABLE_DATA):
    excel_row = start_row + r_idx
    ws.row_dimensions[excel_row].height = 26
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=excel_row, column=c_idx+1, value=val)
        cell.font = Font(name='Cairo', size=10, color='2C3E50')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if r_idx % 2 == 0:
            cell.fill = PatternFill(start_color=even_row_color, end_color=even_row_color, fill_type='solid')
        # Bold row number
        if c_idx == 0:
            cell.font = Font(name='Cairo', size=10, bold=True, color=primary_color)

# ===== Sheet 2: Notes =====
ws2 = wb.create_sheet("ملاحظات الفلاتر")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 50

# Title
ws2.merge_cells('A1:B1')
ws2['A1'] = NOTES_TITLE
ws2['A1'].font = Font(name='Cairo', size=13, bold=True, color=primary_color)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[1].height = 40

ws2.merge_cells('A2:B2')
ws2['A2'] = NOTES_SUBTITLE.replace('\n', ' | ')
ws2['A2'].font = Font(name='Cairo', size=10, color='6B7B8D')
ws2['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[2].height = 35

# Filters header
ws2.merge_cells('A4:B4')
ws2['A4'] = FILTERS_TITLE
ws2['A4'].font = Font(name='Cairo', size=11, bold=True, color=header_font_color)
ws2['A4'].fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type='solid')
ws2['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[4].height = 35

# Filters list
for i, f_item in enumerate(FILTERS_LIST):
    row_n = 5 + i
    ws2.cell(row=row_n, column=1, value=i+1).font = Font(name='Cairo', size=10, bold=True, color=primary_color)
    ws2.cell(row=row_n, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=row_n, column=1).border = thin_border
    ws2.cell(row=row_n, column=2, value=f_item).font = Font(name='Cairo', size=10)
    ws2.cell(row=row_n, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws2.cell(row=row_n, column=2).border = thin_border
    if i % 2 == 0:
        ws2.cell(row=row_n, column=1).fill = PatternFill(start_color=even_row_color, end_color=even_row_color, fill_type='solid')
        ws2.cell(row=row_n, column=2).fill = PatternFill(start_color=even_row_color, end_color=even_row_color, fill_type='solid')

# Diesel section
d_start = 5 + len(FILTERS_LIST) + 1
ws2.merge_cells(f'A{d_start}:B{d_start}')
ws2[f'A{d_start}'] = "فلاتر الديزل"
ws2[f'A{d_start}'].font = Font(name='Cairo', size=11, bold=True, color=header_font_color)
ws2[f'A{d_start}'].fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type='solid')
ws2[f'A{d_start}'].alignment = Alignment(horizontal='center', vertical='center')

for i, d_item in enumerate(DIESEL_FILTERS):
    row_n = d_start + 1 + i
    ws2.merge_cells(f'A{row_n}:B{row_n}')
    ws2[f'A{row_n}'] = d_item
    ws2[f'A{row_n}'].font = Font(name='Cairo', size=10, color='2C3E50' if not d_item.startswith("تركب") else '6B7B8D')
    ws2[f'A{row_n}'].alignment = Alignment(horizontal='right', vertical='center')
    ws2[f'A{row_n}'].border = thin_border

# Closing
c_row = d_start + 1 + len(DIESEL_FILTERS) + 1
ws2.merge_cells(f'A{c_row}:B{c_row}')
ws2[f'A{c_row}'] = CLOSING_NOTE.replace('\n', ' - ')
ws2[f'A{c_row}'].font = Font(name='Cairo', size=11, bold=True, color=primary_color)
ws2[f'A{c_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Signatures
s_row = c_row + 2
ws2[f'A{s_row}'] = "التعميد"
ws2[f'A{s_row}'].font = Font(name='Cairo', size=10, bold=True, color=primary_color)
ws2[f'A{s_row}'].alignment = Alignment(horizontal='center')
ws2[f'B{s_row}'] = "قسم الحركة - خالد الغامدي"
ws2[f'B{s_row}'].font = Font(name='Cairo', size=10, bold=True, color=primary_color)
ws2[f'B{s_row}'].alignment = Alignment(horizontal='center')

excel_path = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_محسن.xlsx')
wb.save(excel_path)
print(f"Excel saved: {excel_path}")
print("Done!")

