import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, 'جدول غسيل 2026-4-25.xlsx')
output_file = os.path.join(script_dir, 'جدول_غسيل_احترافي_2026.xlsx')
img_0_path = os.path.join(script_dir, 'source_img_0.png')
img_1_path = os.path.join(script_dir, 'source_img_1.png')

# 1. Read Data from Old File
old_wb = openpyxl.load_workbook(input_file, data_only=True)
old_ws = old_wb.active

data = []
header_row_idx = None

# Find header
for r in range(1, 15):
    val = str(old_ws.cell(row=r, column=1).value).strip()
    if val == 'م' or val == '1':
        if val == 'م':
            header_row_idx = r
        break

if not header_row_idx:
    header_row_idx = 3 # fallback

# Extract cars
for r in range(header_row_idx + 1, old_ws.max_row + 1):
    row_data = []
    # Only read if there is an ID or Name
    if old_ws.cell(row=r, column=1).value is None and old_ws.cell(row=r, column=2).value is None:
        continue
    for c in range(1, 17):
        val = old_ws.cell(row=r, column=c).value
        row_data.append(val if val is not None else "")
    
    # ensure it's not a legend row like "المغسلة الجديدة"
    if "المغسلة" in str(row_data[1]) or "المغسلة" in str(row_data[2]):
        continue
    
    data.append(row_data)

# 2. Create Brand New Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "جدول غسيل"
ws.sheet_view.rightToLeft = True
ws.sheet_view.showGridLines = False

# ===== COLORS & STYLES =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
ODD_ROW = "FFFFFF"
TEXT_DARK = "2C3E50"
GREEN_TEXT = "27AE60"
RED_TEXT = "C0392B"

thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
accent_side = Side(style='medium', color=ACCENT)

thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'), top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# Base font
for r in range(1, len(data) + 15):
    for c in range(1, 17):
        ws.cell(row=r, column=c).font = Font(name='Cairo', size=11, color=TEXT_DARK)
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')

row = 1
# Try to add logos
try:
    img_ar = XlImage(img_1_path)
    img_ar.width = 200; img_ar.height = 70
    ws.add_image(img_ar, 'A1')
    
    img_en = XlImage(img_0_path)
    img_en.width = 200; img_en.height = 70
    ws.add_image(img_en, 'N1')
except Exception:
    pass

ws.row_dimensions[row].height = 40; row += 1
ws.row_dimensions[row].height = 40; row += 1

# Title
ws.merge_cells(f'A{row}:P{row}')
title_cell = ws[f'A{row}']
title_cell.value = "جدول غسيل السيارات لفرع الدمام لعام 2026"
title_cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
for c in range(1, 17):
    ws.cell(row=row, column=c).border = Border(bottom=accent_side)
ws.row_dimensions[row].height = 40
row += 2

# Headers
headers = ["م", "اسم السائق", "نوع السيارة", "رقم اللوحة", "شهر 1", "شهر 2", "شهر 3", "شهر 4", "شهر 5", "شهر 6", "شهر 7", "شهر 8", "شهر 9", "شهر 10", "شهر 11", "شهر 12"]
for c, text in enumerate(headers, 1):
    cell = ws.cell(row=row, column=c, value=text)
    cell.font = Font(name='Cairo', size=12, bold=True, color=WHITE)
    cell.fill = fl(PRIMARY)
    cell.border = header_border
ws.row_dimensions[row].height = 35
row += 1

# Data
for i, row_data in enumerate(data):
    ws.row_dimensions[row].height = 25
    bg = fl(EVEN_ROW) if i % 2 == 0 else fl(ODD_ROW)
    
    for c in range(1, 17):
        cell = ws.cell(row=row, column=c)
        val = str(row_data[c-1]).strip() if c-1 < len(row_data) else ""
        if val == "None": val = ""
        
        cell.value = val
        cell.fill = bg
        cell.border = thin_border
        
        if c == 1 and val:
            cell.font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
        elif c in (2, 3, 4) and val:
            cell.font = Font(name='Cairo', size=11, bold=True, color=TEXT_DARK)
        elif c >= 5 and val:
            if "استلم" in val:
                cell.font = Font(name='Cairo', size=11, bold=True, color=GREEN_TEXT)
            elif "لا" in val or "لم" in val:
                cell.font = Font(name='Cairo', size=11, bold=True, color=RED_TEXT)
            else:
                cell.font = Font(name='Cairo', size=11, color=TEXT_DARK)
    row += 1

# Footer / Legend
row += 1
ws.merge_cells(f'F{row}:H{row}')
ws[f'F{row}'] = "المغسلة الجديدة"
ws[f'F{row}'].font = Font(name='Cairo', size=11, bold=True, color=TEXT_DARK)
ws[f'F{row}'].fill = fl("BDD7EE") # Light blue
for c in range(6, 9): ws.cell(row=row, column=c).border = thin_border

ws.merge_cells(f'I{row}:K{row}')
ws[f'I{row}'] = "المغسلة القديمة"
ws[f'I{row}'].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
ws[f'I{row}'].fill = fl("000000") # Black
for c in range(9, 12): ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 30

# Set Column Widths
widths = [5, 20, 20, 15] + [10]*12
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25

wb.save(output_file)
print("Saved clean file to", output_file)

