import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_محسن تم التعديل عليه 3.xlsx')
output_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_نهائي_مميز.xlsx')
logo_path = os.path.join(script_dir, 'output-onlinepngtools.png')

wb = openpyxl.load_workbook(input_file)
ws = wb.active
ws.sheet_view.rightToLeft = True

# ===== COLORS & STYLES =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
ODD_ROW = "FFFFFF"
TEXT_DARK = "2C3E50"
TEXT_LIGHT = "6B7B8D"
LIGHT_BG = "F5F3EE"
GREEN_TEXT = "27AE60"
RED_TEXT = "C0392B"

thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'), top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# Base font
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Header image
try:
    img = XlImage(logo_path)
    img.width = 400
    img.height = 80
    ws.add_image(img, 'D1')
except:
    pass

for r in range(1, 5):
    ws.row_dimensions[r].height = 22

# Find sections dynamically
main_header_row = None
filter_header_row = None
diesel_row = None
closing_row = None

for row in ws.iter_rows(min_row=1, max_row=150):
    val = str(row[0].value).strip() if row[0].value else ""
    
    if "بيان استهلاك الزيوت" in val:
        ws.row_dimensions[row[0].row].height = 45
        row[0].font = Font(name='Cairo', size=14, bold=True, color=PRIMARY)
        # Accent line above it
        for c in range(1, 11):
            ws.cell(row=row[0].row - 1, column=c).border = Border(bottom=Side(style='medium', color=ACCENT))
            
    elif val == 'م' and not main_header_row:
        main_header_row = row[0].row
        ws.row_dimensions[main_header_row].height = 35
        for cell in row:
            if cell.value:
                cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
                cell.fill = fl(PRIMARY)
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
    elif "ملاحظات وتفاصيل الفلاتر" in val:
        ws.row_dimensions[row[0].row].height = 38
        row[0].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
        row[0].fill = fl(ACCENT)
        
    elif "شراء عدد 3 براميل" in val:
        ws.row_dimensions[row[0].row].height = 35
        row[0].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
        
    elif "الكمية= 625" in val:
        ws.row_dimensions[row[0].row].height = 32
        row[0].font = Font(name='Cairo', size=10, color=TEXT_LIGHT)
        row[0].fill = fl(LIGHT_BG)
        
    elif "شراء فلاتر زيت وديزل" in val:
        ws.row_dimensions[row[0].row].height = 38
        row[0].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        row[0].fill = fl(PRIMARY)
        
    elif val == 'م' and main_header_row and row[0].row > main_header_row:
        filter_header_row = row[0].row
        ws.row_dimensions[filter_header_row].height = 30
        for cell in row:
            if cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=WHITE)
                cell.fill = fl(PRIMARY)
                cell.border = header_border
                
    elif "فلاتر الديزل" in val:
        diesel_row = row[0].row
        ws.row_dimensions[diesel_row].height = 32
        row[0].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        row[0].fill = fl("34495E")
        
    elif "وبناءا عليه تم طلب" in val:
        closing_row = row[0].row
        ws.row_dimensions[closing_row].height = 38
        row[0].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
        row[0].fill = fl(LIGHT_BG)
        row[0].border = Border(top=Side(style='medium', color=ACCENT), bottom=Side(style='medium', color=ACCENT))
        
    elif "التعميد" in val:
        row[0].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
        row[4].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY) # قسم الحركة
        
    elif "خالد الغامدي" in val:
        pass # row[4] usually
        
# Style Main Data
if main_header_row:
    for r in range(main_header_row + 1, 100):
        if ws.cell(row=r, column=1).value == "📋  ملاحظات وتفاصيل الفلاتر" or str(ws.cell(row=r, column=1).value).strip() == "":
            break
        ws.row_dimensions[r].height = 27
        bg = fl(EVEN_ROW) if r % 2 == 0 else fl(ODD_ROW)
        
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            # Find the merged cell boundary or just apply to all
            cell.border = thin_border
            cell.fill = bg
            if c == 1 and cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
            elif c == 2 and cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
            elif c in (9, 10) and cell.value:
                if "لا يوجد" in str(cell.value):
                    cell.font = Font(name='Cairo', size=9, color=TEXT_LIGHT)
                else:
                    cell.font = Font(name='Cairo', size=10, bold=True, color=GREEN_TEXT)

# Style Filter Data
if filter_header_row:
    for r in range(filter_header_row + 1, 150):
        if ws.cell(row=r, column=1).value == "فلاتر الديزل" or str(ws.cell(row=r, column=1).value).strip() == "":
            break
        ws.row_dimensions[r].height = 25
        bg = fl(EVEN_ROW) if r % 2 == 0 else fl(ODD_ROW)
        
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if c in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                cell.fill = bg
                cell.border = thin_border
            
            val = str(cell.value) if cell.value is not None else ""
            if c in (1, 6) and val: # م
                cell.font = Font(name='Cairo', size=9, bold=True, color=PRIMARY)
            elif c in (2, 7) and val: # نوع الفلتر
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.font = Font(name='Cairo', size=9, color=TEXT_DARK)
            elif c in (3, 8) and val: # السابق
                cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
            elif c in (4, 9) and val: # المستهلك
                cell.font = Font(name='Cairo', size=10, bold=True, color=RED_TEXT)
            elif c in (5, 10) and val: # المتبقي
                cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)

# Style Diesel items
if diesel_row:
    for r in range(diesel_row + 1, diesel_row + 10):
        val = str(ws.cell(row=r, column=1).value or "")
        if "وبناءا عليه" in val or not val:
            break
        ws.row_dimensions[r].height = 25
        cell = ws.cell(row=r, column=1)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)

# Set widths
# A:م(5), B:نوع(17), C:السابق(8), D:المستهلك(8), E:المتبقي(9) 
# F:م(5), G:نوع(17), H:السابق(8), I:المستهلك(8), J:المتبقي(9)
# Total = 94.
# For main table: B+C(25)=رقم لوحة, D(8)=المستخدم(wait, 8 is too small). 
# Let's adjust for a better fit for both!
# A:5, B:14, C:8, D:14, E:12, F:8, G:14, H:8, I:10, J:10
widths = [5, 15, 8, 12, 11, 6, 17, 8, 10, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Print setup
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3

wb.save(output_file)
print("Saved to", output_file)

