# -*- coding: utf-8 -*-
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# Colors & Styles (Premium theme)
PRIMARY = "1A3A5C"
HEADER_BG = "2C4A6E"
ACCENT_GOLD = "C8A45A"
WHITE = "FFFFFF"
ROW_ALT_1 = "EDF2F7"
ROW_ALT_2 = "FFFFFF"
TEXT_DARK = "1A1A2E"
TEXT_SEC = "5A6A7A"
AMBER_NOTE = "B8860B"
LIGHT_BG = "F8F9FA"

ts = Side(style='thin', color='D0D5DD')
mn = Side(style='medium', color=PRIMARY)
ag = Side(style='medium', color=ACCENT_GOLD)
thin_border = Border(left=ts, right=ts, top=ts, bottom=ts)
header_border = Border(left=Side(style='thin', color=HEADER_BG), right=Side(style='thin', color=HEADER_BG), top=mn, bottom=mn)
section_border = Border(left=Side(style='thin', color=PRIMARY), right=Side(style='thin', color=PRIMARY), top=mn, bottom=ag)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

def clean_val(val):
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    return str(val).strip().replace('\n',' ').replace('\r','')

def parse_date(val):
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if '00:00:00' in s: s = s.split(' ')[0]
    return s

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New'

# Find files
target_file = [f for f in os.listdir(folder) if 'تحديث' in f and 'الدمام' in f][0]
driver_file = [f for f in os.listdir(folder) if 'بيانات السائ' in f.replace('ئ', 'ئ') or 'بيانات' in f][0]
vehicle_file = [f for f in os.listdir(folder) if 'دينا' in f and 'لوري' in f][0]

print(f"Target: {target_file}\nDriver: {driver_file}\nVehicle: {vehicle_file}")

target_path = os.path.join(folder, target_file)
driver_path = os.path.join(folder, driver_file)
vehicle_path = os.path.join(folder, vehicle_file)

# 1. Load Driver Info
# Mapping by Emp ID and Plate Number
emp_map = {} # Emp_ID -> {iqama, driver_card_exp}
plate_to_emp = {} # plate -> Emp_ID
name_to_emp = {} # name -> Emp_ID

wb_drv = openpyxl.load_workbook(driver_path, data_only=True)
ws_drv = wb_drv.active
for r in range(9, ws_drv.max_row + 1):
    emp_id = str(ws_drv.cell(row=r, column=2).value).strip()
    name = str(ws_drv.cell(row=r, column=3).value).strip()
    iqama = str(ws_drv.cell(row=r, column=4).value).strip()
    plate = ws_drv.cell(row=r, column=7).value
    card_exp = parse_date(ws_drv.cell(row=r, column=9).value)
    
    if emp_id and emp_id != 'None':
        emp_map[emp_id] = {'iqama': iqama if iqama != 'None' else '', 'card_exp': card_exp}
        name_to_emp[name] = emp_id
    if plate:
        plate_to_emp[normalize_plate(plate)] = emp_id

# 2. Load Vehicle Info
# Mapping by Plate Number
veh_map = {} # plate -> {serial, istimara, fahs, tasgheel}
wb_veh = openpyxl.load_workbook(vehicle_path, data_only=True)
ws_veh = wb_veh.active
# Dina and Lorry sections are in the same sheet
for r in range(6, ws_veh.max_row + 1):
    plate = ws_veh.cell(row=r, column=11).value
    if plate and str(plate).strip() and str(plate) != 'رقم اللوحة':
        np = normalize_plate(plate)
        istimara = parse_date(ws_veh.cell(row=r, column=4).value)
        tasgheel = parse_date(ws_veh.cell(row=r, column=5).value)
        serial = str(ws_veh.cell(row=r, column=15).value).strip()
        serial = '' if serial == 'None' else serial
        # Fahs: try C2 then C3
        fahs_m = parse_date(ws_veh.cell(row=r, column=2).value)
        fahs_h = parse_date(ws_veh.cell(row=r, column=3).value)
        fahs = fahs_m if fahs_m and 'لا' not in fahs_m and 'سيارة' not in fahs_m else fahs_h
        if not fahs or fahs == 'None': fahs = ''
        
        veh_map[np] = {
            'serial': serial, 'istimara': istimara, 'fahs': fahs, 'tasgheel': tasgheel
        }

print(f"Loaded {len(emp_map)} drivers, {len(veh_map)} vehicles.")

# 3. Read Target Data
wb_target = openpyxl.load_workbook(target_path, data_only=True)
ws_target = wb_target.active

main_data = []
vacation_data = []

in_main = False
in_vac = False

for r in range(1, ws_target.max_row + 1):
    c1 = str(ws_target.cell(row=r, column=1).value).strip()
    if c1 == 'م' and str(ws_target.cell(row=r, column=2).value).strip() == 'الرقم الوظيفي':
        in_main = True
        continue
    if c1 == 'م' and 'اجازة' in str(ws_target.cell(row=r, column=2).value):
        in_main = False
        in_vac = True
        continue
    if c1 == 'ملخص الأعداد':
        break
    
    if in_main:
        if c1 and c1 != 'None' and c1.isdigit():
            row_data = [ws_target.cell(row=r, column=c).value for c in range(1, 12)]
            main_data.append(row_data)
    elif in_vac:
        if c1 and c1 != 'None' and c1.isdigit():
            row_data = [ws_target.cell(row=r, column=c).value for c in range(1, 8)]
            vacation_data.append(row_data)

# Process Main Data
processed_main = []
for rd in main_data:
    emp_id = str(rd[1]).strip()
    name = str(rd[2]).strip()
    plate = str(rd[5]).strip()
    np = normalize_plate(plate)
    
    # Defaults
    iqama = rd[3]
    card_exp = ''
    serial = ''
    fahs = ''
    istimara = ''
    tasgheel = ''
    
    if emp_id in emp_map:
        if emp_map[emp_id]['iqama']: iqama = emp_map[emp_id]['iqama']
        card_exp = emp_map[emp_id]['card_exp']
    elif np in plate_to_emp and plate_to_emp[np] in emp_map:
        e_id = plate_to_emp[np]
        if emp_map[e_id]['iqama']: iqama = emp_map[e_id]['iqama']
        card_exp = emp_map[e_id]['card_exp']
        
    if np in veh_map:
        serial = veh_map[np]['serial']
        fahs = veh_map[np]['fahs']
        istimara = veh_map[np]['istimara']
        tasgheel = veh_map[np]['tasgheel']
        
    new_rd = [
        rd[0], # م
        rd[1], # الرقم الوظيفي
        rd[2], # اسم السائق
        iqama, # رقم الإقامة
        card_exp, # تاريخ انتهاء بطاقة السائق
        rd[4], # الوظيفة
        rd[5], # رقم اللوحة
        rd[6], # موديل المركبة
        rd[7], # نوع المركبة
        rd[8], # عدد الطبالي
        rd[9], # حمولة المركبة
        serial, # الرقم التسلسلي
        fahs, # تاريخ الفحص
        istimara, # تاريخ الاستمارة
        tasgheel, # كرت التشغيل
        rd[10] # الملاحظات
    ]
    processed_main.append(new_rd)

# Process Vacation Data
processed_vac = []
for rd in vacation_data:
    name = str(rd[1]).strip()
    plate = str(rd[2]).strip()
    np = normalize_plate(plate)
    
    emp_id = ''
    iqama = ''
    
    # Try to find driver info
    if np in plate_to_emp:
        emp_id = plate_to_emp[np]
    else:
        # try matching name
        for n, eid in name_to_emp.items():
            if name in n or n in name:
                emp_id = eid
                break
                
    if emp_id in emp_map:
        iqama = emp_map[emp_id]['iqama']
        
    new_rd = [
        rd[0], # م
        emp_id, # الرقم الوظيفي (new)
        rd[1], # اسم السائق
        iqama, # رقم الإقامة (new)
        rd[2], # رقم اللوحة
        rd[3], # نوع المركبة والموديل
        rd[4], # عدد الطبالي
        rd[5], # حمولة المركبة
        rd[6]  # الملاحظات
    ]
    processed_vac.append(new_rd)

# 4. Generate New Styled Excel
wb = Workbook()
ws = wb.active
ws.title = "تحديث المركبات"
ws.sheet_view.rightToLeft = True
ws.sheet_view.showGridLines = False

main_headers = [
    'م', 'الرقم الوظيفي', 'اسم السائق', 'رقم الإقامة', 'انتهاء بطاقة السائق',
    'الوظيفة', 'رقم اللوحة', 'الموديل', 'النوع', 'عدد الطبالي', 'الحمولة',
    'الرقم التسلسلي', 'انتهاء الفحص الدوري', 'انتهاء الاستمارة', 'انتهاء كرت التشغيل', 'الملاحظات'
]
col_widths = [5, 12, 18, 14, 15, 12, 14, 10, 16, 10, 10, 16, 15, 15, 15, 20]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1
# Gold top stripe
for c in range(1, len(main_headers)+1):
    ws.cell(row=row,column=c).fill = fl(ACCENT_GOLD)
ws.row_dimensions[row].height = 4
row += 1

# Logos (placeholder if images missing)
logo_ar = os.path.join(folder, '..', 'أرشيف', 'source_img_1.png')
logo_en = os.path.join(folder, '..', 'أرشيف', 'source_img_0.png')
ws.merge_cells(f'A{row}:{get_column_letter(len(main_headers))}{row+2}')
for r in range(row, row+3): ws.row_dimensions[r].height = 25
try:
    if os.path.exists(logo_ar):
        i1 = XlImage(logo_ar); i1.width=200; i1.height=70; ws.add_image(i1,'A2')
    if os.path.exists(logo_en):
        i2 = XlImage(logo_en); i2.width=200; i2.height=70; ws.add_image(i2, f'{get_column_letter(len(main_headers)-2)}2')
except Exception as e: print("Logos skip:", e)
row += 3

ws.merge_cells(f'A{row}:C{row}')
ws[f'A{row}'] = "تاريخ التحديث: 2026/04/30 م"
ws[f'A{row}'].font = Font(name='Cairo', size=10, color=TEXT_SEC)
row += 1

ws.merge_cells(f'A{row}:{get_column_letter(len(main_headers))}{row}')
ws[f'A{row}'] = "جدول تحديث مركبات وموصلين و موزعين فرع الدمام — 2026"
ws[f'A{row}'].font = Font(name='Cairo', size=15, bold=True, color=PRIMARY)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 40
row += 2

# Main Section
ws.merge_cells(f'A{row}:{get_column_letter(len(main_headers))}{row}')
c = ws[f'A{row}']
c.value = "بيانات المتابعة الدورية للمركبات والسائقين"
c.font = Font(name='Cairo', size=13, bold=True, color=WHITE)
c.fill = fl(PRIMARY)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = section_border
for cc in range(2, len(main_headers)+1):
    ws.cell(row=row,column=cc).fill = fl(PRIMARY)
    ws.cell(row=row,column=cc).border = section_border
ws.row_dimensions[row].height = 36
row += 1

ws.row_dimensions[row].height = 45
for i, h in enumerate(main_headers):
    c = ws.cell(row=row, column=i+1, value=h)
    c.font = Font(name='Cairo', size=9, bold=True, color=WHITE)
    c.fill = fl(HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = header_border
row += 1

for ri, rd in enumerate(processed_main):
    ws.row_dimensions[row].height = 28
    bg = fl(ROW_ALT_1) if ri % 2 == 0 else fl(ROW_ALT_2)
    for ci in range(len(main_headers)):
        v = clean_val(rd[ci])
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)
        if ci == 0: cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
        elif ci == 4 and v: cell.font = Font(name='Cairo', size=9, bold=True, color='27AE60') # Card Exp
        elif ci == 6: cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY) # Plate
        elif ci == 11: cell.font = Font(name='Consolas', size=9, color=TEXT_SEC) # Serial
        elif ci >= 12 and ci <= 14 and v: cell.font = Font(name='Cairo', size=9, bold=True, color='27AE60') # Dates
        elif ci == 15 and v: cell.font = Font(name='Cairo', size=9, italic=True, color=AMBER_NOTE)
    row += 1

row += 2

# Vacation Section
vac_headers = ['م', 'الرقم الوظيفي', 'اسم السائق', 'رقم الإقامة', 'رقم اللوحة', 'نوع المركبة والموديل', 'عدد الطبالي', 'الحمولة', 'الملاحظات']
ws.merge_cells(f'A{row}:{get_column_letter(len(vac_headers))}{row}')
c = ws[f'A{row}']
c.value = "أسماء السائقين في اجازة والمركبات المتوقفة"
c.font = Font(name='Cairo', size=13, bold=True, color=WHITE)
c.fill = fl("34495E")
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = section_border
for cc in range(2, len(vac_headers)+1):
    ws.cell(row=row,column=cc).fill = fl("34495E")
    ws.cell(row=row,column=cc).border = section_border
ws.row_dimensions[row].height = 36
row += 1

ws.row_dimensions[row].height = 36
for i, h in enumerate(vac_headers):
    c = ws.cell(row=row, column=i+1, value=h)
    c.font = Font(name='Cairo', size=9, bold=True, color=WHITE)
    c.fill = fl(HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = header_border
row += 1

for ri, rd in enumerate(processed_vac):
    ws.row_dimensions[row].height = 28
    bg = fl(ROW_ALT_1) if ri % 2 == 0 else fl(ROW_ALT_2)
    for ci in range(len(vac_headers)):
        v = clean_val(rd[ci])
        cell = ws.cell(row=row, column=ci+1, value=v)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)
        if ci == 0: cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
        elif ci == 4: cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
    row += 1

row += 2
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "تم إعداد هذا الجدول بواسطة قسم الحركة بفرع الدمام"
ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)

# Print setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25; ws.page_margins.right = 0.25
ws.page_margins.top = 0.3; ws.page_margins.bottom = 0.3

out_path = os.path.join(folder, "تحديث_المركبات_والسائقين_الدمام_2026_محدث_نهائي.xlsx")
wb.save(out_path)
print(f"\nSaved successfully to:\n{out_path}")


