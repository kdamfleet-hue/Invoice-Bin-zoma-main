import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\نسخة من تحديث المركبات والسائقين-الدمام2026 -5-3 تم اضافه ارقام الاقامه-1.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print(f"File loaded. Max rows: {ws.max_row}, Max cols: {ws.max_column}")
for c in range(1, ws.max_column + 1):
    header = ws.cell(row=6, column=c).value
    # let's look at the first few rows of data to spot Hijri years (144x)
    samples = []
    for r in range(7, 15):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            samples.append(str(val))
    print(f"Col {c} ({header}): {samples[:3]}")
