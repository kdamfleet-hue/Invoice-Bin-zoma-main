import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\نسخة من تحديث المركبات والسائقين-الدمام2026 -5-3 تم اضافه ارقام الاقامه-1.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name} | Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    for r in range(1, min(10, ws.max_row + 1)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            row_data.append(str(ws.cell(row=r, column=c).value))
        print("  |  ".join(row_data))



