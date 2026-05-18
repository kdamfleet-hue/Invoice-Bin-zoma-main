# -*- coding: utf-8 -*-
import openpyxl, os, sys, re
from hijridate import Hijri
sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\نسخة من تحديث المركبات والسائقين-الدمام2026 -5-3 تم اضافه ارقام الاقامه-1.xlsx'

def convert_hijri_to_gregorian(date_str):
    if not isinstance(date_str, str):
        return date_str
    
    # Match YYYY-MM-DD or YYYY/MM/DD
    match_ymd = re.search(r'(14[0-9]{2})[-/]([0-9]{1,2})[-/]([0-9]{1,2})', date_str)
    if match_ymd:
        y, m, d = int(match_ymd.group(1)), int(match_ymd.group(2)), int(match_ymd.group(3))
        try:
            greg = Hijri(y, m, d).to_gregorian()
            return date_str.replace(match_ymd.group(0), greg.strftime('%Y-%m-%d'))
        except Exception as e:
            print(f"Error converting {date_str}: {e}")
            return date_str

    # Match DD-MM-YYYY or DD/MM/YYYY
    match_dmy = re.search(r'([0-9]{1,2})[-/]([0-9]{1,2})[-/](14[0-9]{2})', date_str)
    if match_dmy:
        d, m, y = int(match_dmy.group(1)), int(match_dmy.group(2)), int(match_dmy.group(3))
        try:
            greg = Hijri(y, m, d).to_gregorian()
            return date_str.replace(match_dmy.group(0), greg.strftime('%Y-%m-%d'))
        except Exception as e:
            print(f"Error converting {date_str}: {e}")
            return date_str

    return date_str

wb = openpyxl.load_workbook(file_path)
count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val and isinstance(val, str):
                new_val = convert_hijri_to_gregorian(val)
                if new_val != val:
                    print(f"Converted: {val} -> {new_val}")
                    cell.value = new_val
                    count += 1

output_path = file_path.replace('.xlsx', '_ميلادي.xlsx')
wb.save(output_path)
print(f"Total converted: {count}")
print(f"Saved as: {output_path}")


