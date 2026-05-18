import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

f1 = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\مركبات لها مستخدم فعلي 2026-05-03.xlsx'
f2 = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New\مركبات لها مستخدم فعلي 2026-05-03 2.xlsx'

def inspect(file_path):
    print(f"\nInspecting: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Check rows 1 to 5 for headers
    for r in range(1, min(6, ws.max_row + 1)):
        row_data = [str(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        print(f"R{r}: " + " | ".join(row_data))

inspect(f1)
inspect(f2)



