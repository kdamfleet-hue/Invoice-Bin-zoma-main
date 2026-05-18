# -*- coding: utf-8 -*-
import sys, os, re, copy
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New'

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

def parse_date(val):
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if '00:00:00' in s: s = s.split(' ')[0]
    return s

target_file = [f for f in os.listdir(folder) if 'تحديث' in f and 'الدمام' in f][0]
driver_file = [f for f in os.listdir(folder) if 'بيانات السائ' in f.replace('ئ', 'ئ') or 'بيانات' in f][0]
vehicle_file = [f for f in os.listdir(folder) if 'دينا' in f and 'لوري' in f][0]

target_path = os.path.join(folder, target_file)
driver_path = os.path.join(folder, driver_file)
vehicle_path = os.path.join(folder, vehicle_file)

# 1. Load Driver Info
emp_map = {} 
plate_to_emp = {} 
name_to_emp = {} 

wb_drv = openpyxl.load_workbook(driver_path, data_only=True)
ws_drv = wb_drv.active
for r in range(9, ws_drv.max_row + 1):
    emp_id = str(ws_drv.cell(row=r, column=2).value).strip()
    name = str(ws_drv.cell(row=r, column=3).value).strip()
    iqama = str(ws_drv.cell(row=r, column=4).value).strip()
    plate = ws_drv.cell(row=r, column=7).value
    card_exp = parse_date(ws_drv.cell(row=r, column=9).value)
    
    if emp_id and emp_id != 'None':
        emp_map[emp_id] = {'iqama': iqama if iqama != 'None' else '', 'card_exp': card_exp}
        name_to_emp[name] = emp_id
    if plate:
        plate_to_emp[normalize_plate(plate)] = emp_id

# 2. Load Vehicle Info
veh_map = {} 
wb_veh = openpyxl.load_workbook(vehicle_path, data_only=True)
ws_veh = wb_veh.active
for r in range(6, ws_veh.max_row + 1):
    plate = ws_veh.cell(row=r, column=11).value
    if plate and str(plate).strip() and str(plate) != 'رقم اللوحة':
        np = normalize_plate(plate)
        istimara = parse_date(ws_veh.cell(row=r, column=4).value)
        tasgheel = parse_date(ws_veh.cell(row=r, column=5).value)
        serial = str(ws_veh.cell(row=r, column=15).value).strip()
        serial = '' if serial == 'None' else serial
        fahs_m = parse_date(ws_veh.cell(row=r, column=2).value)
        fahs_h = parse_date(ws_veh.cell(row=r, column=3).value)
        fahs = fahs_m if fahs_m and 'لا' not in fahs_m and 'سيارة' not in fahs_m else fahs_h
        if not fahs or fahs == 'None': fahs = ''
        veh_map[np] = {
            'serial': serial, 'istimara': istimara, 'fahs': fahs, 'tasgheel': tasgheel
        }

# 3. Modify Original Target File
wb_target = openpyxl.load_workbook(target_path)
ws_target = wb_target.active

# Unmerge cells from column 8 onwards to avoid MergedCell error
merged_ranges = list(ws_target.merged_cells.ranges)
for merged_range in merged_ranges:
    if merged_range.max_col >= 8:
        ws_target.unmerge_cells(str(merged_range))

in_main = False
in_vac = False

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = copy.copy(src_cell.number_format)
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)

main_hdr_row = 6
new_cols = [
    "الرقم التسلسلي", 
    "تاريخ انتهاء الفحص الدوري", 
    "تاريخ انتهاء رخصة السير (الاستمارة)", 
    "تاريخ انتهاء بطاقة السائق", 
    "تاريخ انتهاء بطاقة التشغيل"
]

for i, hdr in enumerate(new_cols):
    col = 12 + i
    cell = ws_target.cell(row=main_hdr_row, column=col, value=hdr)
    copy_style(ws_target.cell(row=main_hdr_row, column=11), cell)
    ws_target.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

for r in range(7, ws_target.max_row + 1):
    c1 = str(ws_target.cell(row=r, column=1).value).strip()
    c2 = str(ws_target.cell(row=r, column=2).value).strip()
    
    if 'اجازة' in c2:
        in_main = False
        in_vac = True
        
        ws_target.cell(row=r, column=8, value="الرقم الوظيفي")
        copy_style(ws_target.cell(row=r, column=7), ws_target.cell(row=r, column=8))
        ws_target.column_dimensions['H'].width = 12
        
        ws_target.cell(row=r, column=9, value="رقم الإقامة")
        copy_style(ws_target.cell(row=r, column=7), ws_target.cell(row=r, column=9))
        ws_target.column_dimensions['I'].width = 14
        continue

    if c1 == 'ملخص الأعداد':
        break

    if c1 and c1.isdigit():
        if not in_vac:
            emp_id = str(ws_target.cell(row=r, column=2).value).strip()
            name = str(ws_target.cell(row=r, column=3).value).strip()
            plate = str(ws_target.cell(row=r, column=6).value).strip()
            np = normalize_plate(plate)
            
            card_exp = ''
            serial = ''
            fahs = ''
            istimara = ''
            tasgheel = ''
            
            if emp_id in emp_map:
                card_exp = emp_map[emp_id]['card_exp']
            elif np in plate_to_emp and plate_to_emp[np] in emp_map:
                card_exp = emp_map[plate_to_emp[np]]['card_exp']
                
            if np in veh_map:
                serial = veh_map[np]['serial']
                fahs = veh_map[np]['fahs']
                istimara = veh_map[np]['istimara']
                tasgheel = veh_map[np]['tasgheel']
            
            vals = [serial, fahs, istimara, card_exp, tasgheel]
            for i, v in enumerate(vals):
                col = 12 + i
                cell = ws_target.cell(row=r, column=col, value=v)
                copy_style(ws_target.cell(row=r, column=11), cell)
        else:
            name = str(ws_target.cell(row=r, column=2).value).strip()
            plate = str(ws_target.cell(row=r, column=3).value).strip()
            np = normalize_plate(plate)
            
            emp_id = ''
            iqama = ''
            
            if np in plate_to_emp:
                emp_id = plate_to_emp[np]
            else:
                for n, eid in name_to_emp.items():
                    if name in n or n in name:
                        emp_id = eid
                        break
                        
            if emp_id in emp_map:
                iqama = emp_map[emp_id]['iqama']
                
            cell_emp = ws_target.cell(row=r, column=8, value=emp_id)
            copy_style(ws_target.cell(row=r, column=7), cell_emp)
            
            cell_iqama = ws_target.cell(row=r, column=9, value=iqama)
            copy_style(ws_target.cell(row=r, column=7), cell_iqama)

out_path = os.path.join(folder, "تحديث المركبات والسائقين-الدمام2026.xlsx")
wb_target.save(out_path)
print(f"Updated successfully in place: {out_path}")
