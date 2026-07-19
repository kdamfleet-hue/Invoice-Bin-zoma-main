from gevent import monkey
monkey.patch_all()

import os
import io
import psutil
import logging
import shutil
import secrets
import hmac
import openpyxl
from copy import copy
from contextlib import contextmanager
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell as MC
import sqlite3
import base64
import requests
import time
import threading
import uuid
from collections import deque
import json
import re
import html
import absher_sync          # محرّك مزامنة أبشر (قارئ xlsx بالمكتبة القياسية + الفروقات)
from datetime import datetime
import pandas as pd
import numpy as np
from functools import wraps

try:
    import psycopg2
    import psycopg2.extras
except ImportError:  # psycopg2 only required when DATABASE_URL (PostgreSQL) is used
    psycopg2 = None
from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_cors import CORS
from dotenv import load_dotenv
from flask_mail import Mail, Message
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

from flask_caching import Cache
from flask_talisman import Talisman

from models.database import init_db, db_connection, get_db, DB_PATH, DATABASE_URL, USE_POSTGRES
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("InvoiceApp")

# Directories
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "user_templates")
try:
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
except PermissionError:
    # Fallback to /tmp for read-only deployment environments (like Render/Heroku)
    TEMPLATE_DIR = "/tmp/user_templates"
    os.makedirs(TEMPLATE_DIR, exist_ok=True)

LOGO_PATH = os.path.join(os.path.dirname(__file__), "static", "excel_logo.png")

# Valid tab names for template upload
VALID_TABS = {"invoice", "oils", "purchase", "schedule", "workshop"}
# Default template filenames (fallback)
DEFAULT_TEMPLATES = {
    "oils": "oils_template.xlsx",
    "purchase": "po_template.xlsx",
    "schedule": "schedule_base.xlsx",
    "workshop": "تقرير الورشة.xlsx",
}

# Logo for email headers (120x80px PNG, white bg)
EMAIL_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAHgAAABQCAIAAABd+SbeAAAGl0lEQVR42u3abWybVxUH8HPu47eSxHFiu3GTKW5e3KR5aZhDOkqmFhoxadpg0wqCIugEqwZsIAqIicKmdqNiCLURQhvdBmXVVrUrRJ3WoWpsq1iaboNuTZs0IU7avGyu06RJ7Dh+t597Dx9cYOpgdBMTj9H5f/AXPx8e/XR17z3nPEhEwPnwI5iAoRmaw9AMzdAchmZoDkMzNENzGJqhOQzN0AzNYWiG5jA0QzM0h6HfHQIoxMG9KDhlBEAsPOtCgiYCBFBSZbKy4KxFgRBfSSSWvXv30Bd2nLkcTuetlSKG/q8FERFRCBydSvyud/6lM9ETZ+elVAgkBKpCWNto5E/CpCQCQIS3Q3PBS5elpMrlJSeH0WTCdp996+4Rq0176M7aDX6XLsmkIUN/0B0ZAQAy6fjQ6CQRIUJOV6V2ezyRtpcsOztp3dodKCvW9t/XtPFjbqlIE8jQ748YABDh0Iuh8JIOoABRE6jrZC/Suq63DQTGE8ncuraVpyfMW7sDVjM++YOmT691KyKByNDXfE1WJAQ++OT53b+/mEgrRKT8HwCU1e+8ecXOLZ/j8RCqd+3ibdzSobXtsXCD17GjtaCqjIhKGXNcmoykDgRC4c9/57mdDFQ7L5o3L7RbUpcqfiELgXCQdnBfr/I19b4y8PjD1qY762ztde/8wPTQZX9tcpkuGvrZrHADs2j/e/WzIvkzs+UbdWp89GVpwOoob6uuuerhrnfNLfYOXw3GzSWgCgWQmk7JalxtzZLB6BGOJ3OHeuZsJn97W2FKdOX5qvNJdmiXHZ3/Yn0rpRCQQzSbNSfedz9RvWrg5HokpmgEApGApc8HiqqirK/3GQMvS/hhYCECGdzj3+3cZWb6avf8LjdGy4oXlkKnv8bPTSg5PxZzKS9p1LHZX/YeKDnw4xeVIkRFSORKZ2EkT8Kkz0N/ZGbB41DY3VNyc3+8upC7MlMRKZKV6EjOE2vk6xK2z7yeEMPlugbpoCdAzDdqdKnvnd49AQBgt9u1Df2Ct1gsqizLeNovNhp0+C2OA8koaQnp9/rO2/aKxSaD7qM0c2S9LTfqSPVL+g7gdLk3R4Lvn7+mOzy/opoQygzl5NQIq0/p5lhK0e2LzTzqGkWc+MfHZXRCoZDRfINvaOwiz6hElTDSJI1SP3BSkjGC9pRbIp/87qt+t6PBqkEYVgcUzn9CAABxl/mB7nH0vYi07reLMYozcWJtqzI2u8KnIVC0eTcdVgg//uLewJtFRfvcQCkKR1Bh/zeUZYrtdqSVVnXmdI4JjvQEzWUyIPV+s7p1f5Zk8HkDQ5a8h0/4lTsdTtiCWifdzGoFQylFZ2t+TNl1tN2TVdJW99ddpChfbuVZGXiXku/4WVeK/MlA9GqOFK+HGgnWaVXKgvnXjSkLjWEYhmEYhmEYhmEYhmH+sz8BzQDdMcl1yrQAAAAASUVORK5CYII="

app = Flask(__name__)
from models.schema import db, Driver, Vehicle, VehicleCustody, Branch, Document, AuditLog, AppSetting
import os
DB_PATH = os.environ.get('SQLITE_PATH', os.path.join(os.path.dirname(__file__), 'database.sqlite'))
_db_url = os.environ.get('DATABASE_URL')
if _db_url:
    if _db_url.startswith("postgres://"):
        _db_url = _db_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.abspath(DB_PATH)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)
from flask_migrate import Migrate
migrate = Migrate(app, db)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Initialize Security Headers
# content_security_policy=None to allow existing inline scripts to continue working
# force_https=False to prevent 302 redirects on internal health checks (Render/Railway)
Talisman(app, content_security_policy=None, force_https=False)

# Initialize Caching
cache = Cache(config={'CACHE_TYPE': 'SimpleCache'})
cache.init_app(app)

# Initialize Rate Limiter
_secret = os.environ.get("SECRET_KEY")
if not _secret:
    # No hardcoded fallback key (that would let anyone forge sessions).
    # Generate a strong ephemeral key so the app still runs; warn loudly.
    _secret = secrets.token_hex(32)
    logger.warning(
        "SECRET_KEY not set — generated a random ephemeral key. "
        "Sessions will reset on restart. Set SECRET_KEY in the environment for production!"
    )
app.secret_key = _secret

init_db(app)

def init_db_on_startup():
    from flask_migrate import upgrade
    with app.app_context():
        try:
            upgrade()
            logger.info("✅ Database migrated successfully (Alembic).")
        except Exception as e:
            logger.error(f"❌ Error migrating database: {e}")

init_db_on_startup()
# CORS: the app serves its own same-origin frontend, so cross-origin is disabled by
# default. Set ALLOWED_ORIGINS (comma-separated) only if external clients are needed.
_allowed_origins = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if _allowed_origins:
    CORS(app, resources={r"/api/*": {"origins": _allowed_origins}}, supports_credentials=True)

# Harden the session cookie
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_DEBUG", "False").lower() != "true"

# Session lifetime: 8 hours
from datetime import timedelta

app.permanent_session_lifetime = timedelta(hours=8)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')

# Security Decorator (defined early so it is available to all routes below)
WS_PREFIX = "/importantworkstation"


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # The /importantworkstation/* namespace is an OPEN, separate workstation — no main login.
        if request.path.startswith(WS_PREFIX):
            return f(*args, **kwargs)
        if not session.get("authenticated"):
            if request.path.startswith("/api/"):
                return jsonify({"success": False, "error": "Session expired. Please refresh and log in again."}), 401
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)

    return decorated_function

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if request.path.startswith(WS_PREFIX):
                return f(*args, **kwargs)
            if not session.get("authenticated"):
                if request.path.startswith("/api/"):
                    return jsonify({"success": False, "error": "Session expired. Please log in."}), 401
                return redirect(url_for("auth.login"))
            
            user_role = session.get("role", "viewer")
            if user_role not in roles and "admin" not in roles and user_role != "admin": # admin has access to everything
                # Actually, let's strictly check:
                if user_role != 'admin' and user_role not in roles:
                    if request.path.startswith("/api/"):
                        return jsonify({"success": False, "error": "غير مصرح لك (Forbidden)"}), 403
                    # Add flash or error handling if needed, or simply abort
                    from flask import abort
                    abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ── GPS Configuration ────────────────────────────────────────────────────────
GPS_USER = os.environ.get("GPS_USER", "")
GPS_PASS = os.environ.get("GPS_PASS", "")
# Default to the actual JSON API endpoint (NOT the web-UI SPA URL which returns HTML).
GPS_ASSET_URL = os.environ.get(
    "GPS_ASSET_URL", "https://fleetmanagement-api-clust03.gpscockpit.com/api/asset"
)
# Accept either GPS_TOKEN (ArabCord) or the legacy GPS_PERMANENT_TOKEN name.
GPS_PERMANENT_TOKEN = os.environ.get("GPS_TOKEN") or os.environ.get("GPS_PERMANENT_TOKEN", "")


def get_gps_token():
    """Return the GPS API token from environment."""
    return GPS_PERMANENT_TOKEN


logger.info("Starting Fleet Management System...")
logger.info("Template Dir: %s", TEMPLATE_DIR)
logger.info("Database Path: %s", DB_PATH)





@app.route("/api/gps")
@login_required
def get_gps_locations():
    token = get_gps_token()
    if not token:
        return (
            jsonify(
                {"error": "خدمة التتبع غير مهيأة — لم يتم ضبط مفتاح GPS (GPS_TOKEN)."}
            ),
            503,
        )

    headers = {
        "Authorization": f"GpsCockpitApiKey {token}",
        "Accept": "application/json",
    }
    try:
        response = requests.get(GPS_ASSET_URL, headers=headers, timeout=20)
        if response.status_code != 200:
            # Do NOT echo provider body to the client (may leak internals); log it instead.
            logger.warning("GPS API non-200 %s: %s", response.status_code, response.text[:300])
            return (
                jsonify({"error": "تعذّر جلب بيانات GPS من المزوّد حالياً."}),
                502,
            )
        # Guard against HTML responses (e.g. misconfigured URL pointing at the web UI).
        ctype = response.headers.get("Content-Type", "")
        if "application/json" not in ctype:
            logger.warning("GPS API returned non-JSON (%s). Check GPS_ASSET_URL.", ctype)
            return (
                jsonify({"error": "استجابة GPS غير صالحة — تأكد من ضبط GPS_ASSET_URL على نقطة API."}),
                502,
            )
        return jsonify(response.json())
    except requests.Timeout:
        return jsonify({"error": "تجاوز وقت الاستجابة من خدمة GPS."}), 504
    except requests.ConnectionError:
        return jsonify({"error": "تعذّر الاتصال بخدمة GPS."}), 503
    except Exception as e:
        logger.error("GPS API error: %s", e)
        return jsonify({"error": "حدث خطأ غير متوقع أثناء جلب بيانات GPS."}), 500


# Flask-Mail Configuration
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "True").lower() == "true"
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
mail = Mail(app)

# OAuth configuration removed


# ── Workstation namespace (/importantworkstation/*) ───────────────────────────
# A COMPLETELY SEPARATE, OPEN entry that mirrors the site under a URL prefix. It is a
# sandbox (edits go to id=2, never touching the real id=1 data) and the
# Cameras/Employees/GPS-Sync tabs are password-locked. The MAIN site (/) is untouched.
WORKSTATION_PASSWORD = os.environ.get("WORKSTATION_PASSWORD")
if not WORKSTATION_PASSWORD:
    WORKSTATION_PASSWORD = secrets.token_hex(16)
    logger.warning("WORKSTATION_PASSWORD not set in env — generated a random secure key.")

# Kiosk account: workshop-report-only access, no nav, full-screen.
KIOSK_USER     = os.environ.get("KIOSK_USER",     "jam")
KIOSK_PASSWORD = os.environ.get("KIOSK_PASSWORD")
if not KIOSK_PASSWORD:
    KIOSK_PASSWORD = secrets.token_hex(16)
    logger.warning("KIOSK_PASSWORD not set in env — generated a random secure key.")
WS_TABS = {
    "": "index", "dashboard": "dashboard", "kpis": "kpis", "invoice": "index", "fleet_dashboard": "fleet_dashboard",
    "schedule": "schedule", "oils": "oils", "purchase": "purchase", "fuel": "fuel",
    "washing": "washing", "workshop": "workshop", "search": "search", "records": "records",
    "incidents": "incidents", "handover": "handover",
    "tracking": "tracking", "employees": "employees", "gps_sync": "gps_sync",
    "gps_dashboard": "gps_dashboard", "gps_devices": "gps_devices", "cameras": "cameras",
}
WS_LOCKED = {"employees", "gps_sync", "cameras", "tracking"}


def is_workstation():
    """Determined purely by the URL prefix — never by the session — so the main site is
    never affected, even in the same browser."""
    return request.path.startswith(WS_PREFIX)


# ── Branches (multi-branch data isolation) ───────────────────────────────────
# Each branch is a distinct storage "mode" (row id) shared across EVERY blob table,
# so switching branch swaps ALL operational data at once. id=1 is الدمام (the existing,
# FROZEN data); id=2 is reserved for the workstation sandbox; 3+ are the other branches
# and start EMPTY. Never renumber 1 or 2.
BRANCHES = [
    {"id": 1, "name": "الدمام"},
    {"id": 3, "name": "جدة"},
    {"id": 4, "name": "الرياض"},
    {"id": 5, "name": "حفر الباطن"},
    {"id": 6, "name": "بالجرشي"},
    {"id": 7, "name": "جيزان"},
]
BRANCH_IDS = {b["id"] for b in BRANCHES}
BRANCH_NAME = {b["id"]: b["name"] for b in BRANCHES}


def current_branch_id():
    """Active branch id from the session (defaults to الدمام = 1). Validated against the
    known set so a stale/forged cookie — or a call outside a request — can never point
    at an unknown store."""
    try:
        bid = int(session.get("branch_id", 1))
    except (TypeError, ValueError, RuntimeError):
        return 1
    return bid if bid in BRANCH_IDS else 1


def current_branch_name():
    return BRANCH_NAME.get(current_branch_id(), "الدمام")


def _row_id():
    # The workstation sandbox (path-based) always wins; otherwise partition by branch.
    return 2 if is_workstation() else current_branch_id()


# Maps each data-tab route to its snapshot table, so every tab can show its own history.
SNAP_TAB_BY_ROUTE = {
    "/schedule": "schedule_data", "/washing": "washing_schedule", "/employees": "employees",
    "/records": "records_data", "/incidents": "incidents_data", "/oils": "oils_data",
    "/purchase": "purchase_data", "/workshop": "workshop_data", "/gps_devices": "gps_devices_data",
    "/fuel": "fuel_data",
}



@app.context_processor
def inject_system_features():
    features = _global_blob_get("system_features") or {
        "audit_enforced": True,
        "ai_assistant": True,
        "email_alerts": True,
        "workstation_mode": False
    }
    return {"features": features}

@app.context_processor
def inject_branch():
    """Make the active branch + current tab's snapshot key available to every template."""
    path = request.path or "/"
    if path.startswith(WS_PREFIX):
        path = path[len(WS_PREFIX):] or "/"
    return {"active_branch": current_branch_name(),
            "active_branch_id": current_branch_id(),
            "branches": BRANCHES,
            "snap_tab": SNAP_TAB_BY_ROUTE.get(path, "")}


def _safe_tbl(table):
    """Table names are interpolated into SQL (identifiers can't be parameterized), so
    validate they are plain identifiers — makes injection via a table name impossible."""
    if not isinstance(table, str) or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", table):
        raise ValueError("invalid table name: %r" % (table,))
    return table


def _loads_blob(data_str):
    """Parse a blob row's JSON, tolerating corrupt/invalid data instead of raising 500."""
    if not data_str:
        return None
    try:
        return json.loads(data_str)
    except (ValueError, TypeError):
        logger.warning("Corrupt JSON blob encountered; returning None")
        return None


def blob_get(table):
    """Read JSON blob for the active branch from SQLAlchemy AppSetting."""
    from models.schema import AppSetting
    table = _safe_tbl(table)
    rid = _row_id()
    key = f"{table}_branch_{rid}"
    setting = AppSetting.query.get(key)
    return _loads_blob(setting.value) if setting else None


def blob_set(table, data_obj):
    """Write JSON blob for the active branch to SQLAlchemy AppSetting."""
    from models.schema import AppSetting
    data_obj = sanitize_data(data_obj)
    table = _safe_tbl(table)
    rid = _row_id()
    key = f"{table}_branch_{rid}"
    data_str = json.dumps(data_obj, ensure_ascii=False)
    
    try:
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = data_str
        else:
            setting = AppSetting(key=key, value=data_str)
            db.session.add(setting)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving blob {key}: {e}")
        
    _snapshot(table, data_str)


# ── Dated version snapshots (auto-saved per data tab; restored from Settings) ──
SNAPSHOT_TABLES = {"schedule_data", "washing_schedule", "records_data", "employees",
                   "incidents_data", "gps_devices_data", "oils_data", "purchase_data", "workshop_data",
                   "fuel_data"}
SNAP_KEEP = 30
SNAP_LABELS = {
    "schedule_data": "الجدول الأسبوعي", "washing_schedule": "الغسيل", "records_data": "التوثيق",
    "employees": "الموظفون", "incidents_data": "الحوادث والمخالفات", "gps_devices_data": "أجهزة التتبع",
    "oils_data": "الزيوت والفلاتر", "purchase_data": "طلبات الشراء", "workshop_data": "الورشة",
    "fuel_data": "تموين المحروقات",
}


def _snapshot(table, data_str):
    """Append a dated snapshot using SQLAlchemy Snapshot model."""
    from models.schema import Snapshot
    if table not in SNAPSHOT_TABLES:
        return
    mode = _row_id()
    
    try:
        last = Snapshot.query.filter_by(tab=table, branch_id=mode).order_by(Snapshot.id.desc()).first()
        if last and last.data == data_str:
            return  # no change since the last snapshot
            
        new_snap = Snapshot(tab=table, branch_id=mode, data=data_str)
        db.session.add(new_snap)
        
        # Cleanup old snapshots
        snaps = Snapshot.query.filter_by(tab=table, branch_id=mode).order_by(Snapshot.id.desc()).all()
        if len(snaps) > SNAP_KEEP:
            for s in snaps[SNAP_KEEP:]:
                db.session.delete(s)
                
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.warning("snapshot failed for %s: %s", table, e)


# ── GLOBAL (id=1) blobs: login accounts live here regardless of the active branch ─────
def _global_blob_get(table):
    from models.schema import AppSetting
    table = _safe_tbl(table)
    key = f"{table}_global"
    setting = AppSetting.query.get(key)
    return _loads_blob(setting.value) if setting else None



def sanitize_data(data):
    if isinstance(data, dict):
        return {k: sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(v) for v in data]
    elif isinstance(data, str):
        return " ".join(data.split()).strip()
    return data

def _global_blob_set(table, data_obj):
    from models.schema import AppSetting
    data_obj = sanitize_data(data_obj)
    table = _safe_tbl(table)
    key = f"{table}_global"
    data_str = json.dumps(data_obj, ensure_ascii=False)
    
    try:
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = data_str
        else:
            setting = AppSetting(key=key, value=data_str)
            db.session.add(setting)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving global blob {key}: {e}")


# ── System Feature Flags (Admin Control Panel) ────────────────────────────
_DEFAULT_FEATURES = {
    "audit_enforced": True,
    "ai_assistant": True,
    "email_alerts": True,
    "workstation_mode": False,
}

def get_system_features():
    """Read feature flags from DB, falling back to defaults."""
    d = _global_blob_get("system_features")
    if isinstance(d, dict):
        merged = dict(_DEFAULT_FEATURES)
        merged.update(d)
        return merged
    return dict(_DEFAULT_FEATURES)

def set_system_features(features):
    """Persist feature flags to DB."""
    _global_blob_set("system_features", features)


def _persistent_secret_key():
    """A stable session-signing key stored in the DB (survives restarts/redeploys), used
    only when SECRET_KEY isn't set in the environment — so deploys no longer log everyone
    out. Generated once, on first need, then reused forever after."""
    with app.app_context():
        try:
            data = _global_blob_get("app_secret_key")
            key = data.get("key") if isinstance(data, dict) else None
            if key:
                return key
        except Exception:
            logger.exception("Failed to read persistent secret key — falling back to a new one")
        new_key = secrets.token_hex(32)
        try:
            _global_blob_set("app_secret_key", {"key": new_key})
        except Exception:
            logger.exception("Failed to persist secret key — sessions will reset on restart")
        return new_key


# Lock in a stable session key from the persistent DB (unless SECRET_KEY is set in the env),
# so deploys/restarts no longer log everyone out. Runs once at import, before serving requests.
if not (os.environ.get("SECRET_KEY") or "").strip():
    app.secret_key = _persistent_secret_key()
    app.config["SECRET_KEY"] = app.secret_key


# ── Shared login accounts (created from the locked Settings tab) ──────────────
def get_users():
    """List of shared accounts: [{username, name, pw_hash, created}] — global (id=1)."""
    d = _global_blob_get("app_users")
    return d.get("users", []) if isinstance(d, dict) else []


def save_users(users):
    _global_blob_set("app_users", {"users": users})


# ── Per-branch login accounts (username + code → locked to one branch) ────────
def get_branch_accounts():
    """[{branch_id, username, code_hash, created}] — global (id=1)."""
    d = _global_blob_get("branch_accounts")
    return d.get("accounts", []) if isinstance(d, dict) else []


def save_branch_accounts(accounts):
    _global_blob_set("branch_accounts", {"accounts": accounts})


# ── Audit log (append-only; never edited or deleted from the UI) ──────────────
# Records WHO changed WHAT and WHEN. Sandboxed per mode (id=1 main / id=2 workstation).
# Rapid auto-saves of the same target by the same user within a short window are coalesced
# into one row (bumping its time + count) so the trail stays meaningful, not flooded.
AUDIT_MAX = 1000
AUDIT_COALESCE_SEC = 600


def _audit_get():
    rid = _row_id()
    key = f"audit_log_branch_{rid}"
    setting = AppSetting.query.get(key)
    if not setting: return []
    try:
        return json.loads(setting.value) or []
    except Exception:
        return []


def _audit_write(log):
    rid = _row_id()
    key = f"audit_log_branch_{rid}"
    data_str = json.dumps(log[-AUDIT_MAX:], ensure_ascii=False)
    
    try:
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = data_str
        else:
            setting = AppSetting(key=key, value=data_str)
            db.session.add(setting)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving audit log: {e}")


def _audit_add(action, target, count=None, detail=""):
    """Best-effort audit entry. A logging failure must NEVER break the underlying save."""
    try:
        who = "محطة العمل" if is_workstation() else ((session.get("google_user") or {}).get("name") or "النظام")
        now = datetime.now()
        log = _audit_get()
        last = log[-1] if log else None
        if last and last.get("action") == action and last.get("target") == target and last.get("user") == who:
            try:
                delta = (now - datetime.strptime(last["ts"], "%Y-%m-%d %H:%M:%S")).total_seconds()
            except Exception:
                delta = AUDIT_COALESCE_SEC + 1
            if 0 <= delta <= AUDIT_COALESCE_SEC:
                last["ts"] = now.strftime("%Y-%m-%d %H:%M:%S")
                last["count"] = count
                last["detail"] = detail
                last["hits"] = last.get("hits", 1) + 1
                _audit_write(log)
                return
        log.append({
            "ts": now.strftime("%Y-%m-%d %H:%M:%S"),
            "user": who, "action": action, "target": target,
            "count": count, "detail": detail, "hits": 1,
        })
        _audit_write(log)
    except Exception:
        logger.exception("audit_add failed")

def audit_and_verify(action, target, reason):
    """
    تتحقق من وجود سبب وتضيفه لسجل التدقيق. 
    القاعدة الثالثة: لا يُقبل أي تعديل يدوي بدون مسوّغ مكتوب.
    """
    if not reason or len(reason.strip()) < 5:
        raise ValueError("يجب إدخال مسوّغ مكتوب (السبب) لإتمام التعديل.")
    
    _audit_add(action, target, detail=f"السبب: {reason.strip()}")
    return True


# ── Workstation example/demo data (FAKE — for the open sandbox only) ──────────
# Loaded once from ws_example_data.json. The MAIN site never touches any of this.
WS_EXAMPLE_PATH = os.path.join(os.path.dirname(__file__), "ws_example_data.json")
try:
    with open(WS_EXAMPLE_PATH, encoding="utf-8") as _wsf:
        WS_EXAMPLE_DATA = json.load(_wsf)
except Exception as _e:
    logger.warning("ws_example_data.json not loaded: %s", _e)
    WS_EXAMPLE_DATA = {}


def _ws_meta_get(k):
    key = f"ws_meta_{k}"
    setting = AppSetting.query.get(key)
    return setting.value if setting else None


def _ws_meta_set(k, v):
    key = f"ws_meta_{k}"
    try:
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = str(v)
        else:
            setting = AppSetting(key=key, value=str(v))
            db.session.add(setting)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving ws_meta {key}: {e}")


def _ws_get2(table):
    """Read the workstation (id=2) blob for a table, or None."""
    table = _safe_tbl(table)
    key = f"{table}_branch_2"
    setting = AppSetting.query.get(key)
    return _loads_blob(setting.value) if setting else None


def _ws_put2(table, value):
    """Upsert the workstation (id=2) blob for a table."""
    table = _safe_tbl(table)
    key = f"{table}_branch_2"
    data_str = json.dumps(value, ensure_ascii=False)
    try:
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = data_str
        else:
            setting = AppSetting(key=key, value=data_str)
            db.session.add(setting)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving ws blob {key}: {e}")


def _ws_is_empty(value):
    """True if a stored blob carries NO real rows — covers a missing row, [], {}, and an
    'empty shell' object like {title, date:'', main:[], spare:[], vacation:[], summary:{vacation:'0'}}
    that a stray autosave may have written (this is why a tab can look permanently empty).
    Rows live ONLY in LIST fields (main/spare/vacation/oils/filters/parts/…); scalar and dict
    fields (title, date, summary) are metadata and never count as data."""
    if value is None:
        return True
    if isinstance(value, list):
        return len(value) == 0
    if isinstance(value, dict):
        for v in value.values():
            if isinstance(v, list) and len(v) > 0:
                return False
        return True  # no non-empty list field → no rows
    return False


def _ws_write_examples():
    """Overwrite ALL workstation id=2 stores with the FAKE example data (used by 🧪)."""
    for table, value in WS_EXAMPLE_DATA.items():
        _ws_put2(table, value)


def ensure_ws_seeded():
    """Self-healing seed: on every workstation page load, fill any store that is EMPTY
    (missing OR an empty/empty-shell blob) from the example data, while PRESERVING any store
    that holds real rows the user typed. Skipped entirely if the user emptied the sandbox with
    the 🗑️ button (ws_cleared flag), so a deliberate reset stays empty."""
    try:
        if _ws_meta_get("ws_cleared") == "1":
            return
        for table, value in WS_EXAMPLE_DATA.items():
            if _ws_is_empty(_ws_get2(table)):
                _ws_put2(table, value)
    except Exception:
        logger.exception("ensure_ws_seeded error")


@app.route(WS_PREFIX)
@app.route(WS_PREFIX + "/<path:sub>")
def workstation_page(sub=""):
    """Open workstation pages under the prefix. The 3 sensitive tabs require the password."""
    ensure_ws_seeded()  # first visit fills the sandbox with fake example data
    seg = sub.strip("/").split("/")[0] if sub else ""
    if seg == "api":
        return ("", 404)  # API served by the mirrored /importantworkstation/api/* rules
    if seg not in WS_TABS:
        return redirect(WS_PREFIX)
    if seg in WS_LOCKED and not session.get("ws_unlocked"):
        return render_template("tab_lock.html", next=WS_PREFIX + "/" + seg)
    ctx = {"google_user": {"name": "Workstation", "email": "ws@system.local"}, "b64_en": load_logo()}
    if seg == "cameras":
        ctx["cameras_url"] = os.environ.get("CAMERAS_URL", "")
    return render_template(WS_TABS[seg] + ".html", **ctx)


@app.route(WS_PREFIX + "/unlock", methods=["POST"])
def workstation_unlock():
    nxt = request.form.get("next", WS_PREFIX)
    if not nxt.startswith(WS_PREFIX):
        nxt = WS_PREFIX
    if hmac.compare_digest(request.form.get("password", ""), WORKSTATION_PASSWORD):
        session["ws_unlocked"] = True
        resp = redirect(nxt)
        resp.set_cookie("ws_unlocked", "1", path=WS_PREFIX, samesite="Lax",
                        httponly=True,
                        secure=app.config.get("SESSION_COOKIE_SECURE", False))
        return resp
    return render_template("tab_lock.html", next=nxt, error="كلمة المرور غير صحيحة")


@app.route("/tracking")
@login_required
def tracking_page():
    return render_template("tracking.html")


def load_logo():
    b64_en = ""
    txt_path = os.path.join(app.root_path, "b64.txt")
    if os.path.exists(txt_path):
        with open(txt_path, "r") as f:
            content = f.read().strip()
            if not content.startswith("ar='") and not content.startswith("en='"):
                b64_en = content
    return b64_en


@app.route("/tech_updates")
@login_required
def tech_updates():
    return render_template("tech_updates.html", 
                           active_branch_id=session.get("active_branch_id", 1), 
                           active_branch=session.get("active_branch", {}), 
                           snap_tab="tech_updates")

@app.route("/system_commands")
@login_required
def system_commands():
    return render_template("system_commands.html",
                           active_branch_id=session.get("active_branch_id", 1),
                           active_branch=session.get("active_branch", {}),
                           snap_tab="system_commands")

@app.route("/")
@login_required
def index():
    google_user = session.get("google_user")
    b64_en = load_logo()
    # Homepage hides the "نظام الفواتير الذكي" heading (logo takes its place); the /invoice tab shows it.
    return render_template("index.html", google_user=google_user, b64_en=b64_en, show_invoice_title=False)


@app.route("/dashboard")
@login_required
def dashboard():
    # Executive dashboard. Mostly reads existing /api/* data via GET; the two write paths
    # are Update-History "restore" (/api/snapshots/restore) and inline edits inside
    # مركز تنبيهات الوثائق, which go through /api/alerts_center/update.
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("dashboard.html", google_user=google_user, b64_en=b64_en)


@app.route("/kpis")
@login_required
def kpis():
    # Static strategic KPI reference page (descriptive only — no data binding).
    return render_template("kpis.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/handover")
@login_required
def handover():
    # Vehicle delivery/receipt inspection form with touch signature pads (client-side only).
    return render_template("handover.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    # Locked admin tab: re-enter the MASTER_PASSWORD to open, then manage shared accounts.
    master_pass = os.environ.get("MASTER_PASSWORD")
    if not master_pass:
        logger.warning("MASTER_PASSWORD not set in env! Settings page cannot be unlocked.")
    if request.method == "POST" and "password" in request.form:
        if master_pass and hmac.compare_digest(request.form.get("password", ""), master_pass):
            session["settings_unlocked"] = True
            return redirect(url_for("settings"))
        return render_template("tab_lock.html", next="/settings", action="/settings",
                               error="كلمة المرور غير صحيحة")
    if not session.get("settings_unlocked"):
        return render_template("tab_lock.html", next="/settings", action="/settings")
    return render_template("settings.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/system_features", methods=["GET", "POST"])
@login_required
def api_system_features():
    """Read / update system-wide feature flags (admin only)."""
    if not session.get("settings_unlocked"):
        return jsonify({"error": "locked"}), 403
    if request.method == "GET":
        return jsonify({"success": True, "features": get_system_features()})
    try:
        body = request.get_json(silent=True) or {}
        features = get_system_features()
        for key in _DEFAULT_FEATURES:
            if key in body:
                features[key] = bool(body[key])
        set_system_features(features)
        _audit_add("تحديث", "إعدادات النظام", detail="تعديل ميزات النظام")
        return jsonify({"success": True, "features": features})
    except Exception:
        logger.exception("system_features POST error")
        return jsonify({"success": False, "error": "تعذّر حفظ الإعدادات."}), 500





def _snapshot_list(tab, mode):
    from models.schema import Snapshot
    snaps = Snapshot.query.filter_by(tab=tab, branch_id=mode).order_by(Snapshot.id.desc()).all()
    return [{"id": s.id, "ts": s.timestamp.strftime('%Y-%m-%d %H:%M:%S') if s.timestamp else ""} for s in snaps]


def _restore_snapshot(sid, mode, require_tab=None):
    """Restore one dated snapshot into its tab's blob. Returns (ok, info)."""
    from models.schema import Snapshot
    snap = Snapshot.query.filter_by(id=sid, branch_id=mode).first()
    if not snap:
        return False, "not_found"
    if require_tab is not None and snap.tab != require_tab:
        return False, "tab_mismatch"
    try:
        data = json.loads(snap.data)
    except (ValueError, TypeError):
        return False, "corrupt"
    blob_set(snap.tab, data)   # restore (also versioned as the new latest state)
    _audit_add("استعادة نسخة", SNAP_LABELS.get(snap.tab, snap.tab), None, "من سجل النسخ المؤرّخة")
    return True, snap.tab


@app.route("/api/snapshots", methods=["GET"])
@login_required
def api_snapshots():
    # Dated version history per data tab — full all-tabs list from the unlocked Settings tab.
    if not session.get("settings_unlocked"):
        return jsonify({"error": "locked"}), 403
    tab = request.args.get("tab", "")
    tabs = [{"key": k, "label": SNAP_LABELS.get(k, k)} for k in sorted(SNAPSHOT_TABLES)]
    if tab not in SNAPSHOT_TABLES:
        return jsonify({"tabs": tabs, "snapshots": []})
    return jsonify({"tabs": tabs, "label": SNAP_LABELS.get(tab, tab),
                    "snapshots": _snapshot_list(tab, _row_id())})


@app.route("/api/snapshots/restore", methods=["POST"])
@login_required
def api_snapshots_restore():
    if not session.get("settings_unlocked"):
        return jsonify({"error": "locked"}), 403
    ok, info = _restore_snapshot((request.get_json(silent=True) or {}).get("id"), _row_id())
    if not ok:
        return jsonify({"error": info}), (404 if info == "not_found" else 500)
    return jsonify({"success": True})


# Per-tab version history — available INSIDE each data tab to any logged-in editor.
# Scoped to the current branch (via _row_id) AND to the requested tab.
@app.route("/api/tab_history", methods=["GET"])
@login_required
def api_tab_history():
    tab = request.args.get("tab", "")
    if tab not in SNAPSHOT_TABLES:
        return jsonify({"success": False, "snapshots": []}), 400
    return jsonify({"success": True, "label": SNAP_LABELS.get(tab, tab),
                    "snapshots": _snapshot_list(tab, _row_id())})


@app.route("/api/tab_history/restore", methods=["POST"])
@login_required
def api_tab_history_restore():
    body = request.get_json(silent=True) or {}
    tab = body.get("tab") or ""
    if tab not in SNAPSHOT_TABLES:
        return jsonify({"success": False, "reason": "bad_tab"}), 400
    ok, info = _restore_snapshot(body.get("id"), _row_id(), require_tab=tab)
    if not ok:
        return jsonify({"success": False, "reason": info}), (404 if info == "not_found" else 400)
    return jsonify({"success": True})


@app.route("/gps_dashboard")
@login_required
def gps_dashboard():
    # GPS fleet KPI dashboard (sub-tab of the GPS page).
    return render_template("gps_dashboard.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/gps_devices")
@login_required
def gps_devices():
    # GPS tracking-device inventory (active / broken / issue) — sub-tab of the GPS page.
    return render_template("gps_devices.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/invoice")
@login_required
def invoice():
    return render_template("invoice.html", google_user=session.get("google_user"), b64_en=load_logo(), show_invoice_title=True)


@app.route("/fleet_dashboard")
@login_required
def fleet_dashboard():
    # Standalone fleet KPI dashboard (ported from Antigravity).
    branches = []
    is_admin = session.get("role") == "admin"
    try:
        with db_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT id, name FROM erp_branches")
            branches = [{"id": r[0], "name": r[1]} for r in c.fetchall()]
    except Exception as e:
        logger.error(f"Failed to fetch branches: {e}")
    return render_template("fleet_dashboard.html", google_user=session.get("google_user"), b64_en=load_logo(), branches=branches, is_admin=is_admin)


@app.route("/sw.js")
def service_worker():
    # Serve the PWA service worker from the root so its scope covers the whole site.
    resp = app.send_static_file("sw.js")
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.route("/oils")
@login_required
def oils():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("oils.html", google_user=google_user, b64_en=b64_en)


@app.route("/fuel")
@login_required
def fuel():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("fuel.html", google_user=google_user, b64_en=b64_en)


@app.route("/purchase")
@login_required
def purchase():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("purchase.html", google_user=google_user, b64_en=b64_en)


@app.route("/schedule")
@login_required
def schedule():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("schedule.html", google_user=google_user, b64_en=b64_en)


@app.route("/washing")
@login_required
def washing():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("washing.html", google_user=google_user, b64_en=b64_en)

@app.route("/employees")
@login_required
def employees():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("employees.html", google_user=google_user, b64_en=b64_en)


@app.route("/gps_sync")
@login_required
def gps_sync():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("gps_sync.html", google_user=google_user, b64_en=b64_en)


@app.route("/workshop")
@login_required
def workshop():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("workshop.html", google_user=google_user, b64_en=b64_en,
                           kiosk=bool(session.get("kiosk")))


@app.route("/search")
@login_required
def search_page():
    return render_template("search.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/records")
@login_required
def records_page():
    return render_template("records.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/incidents")
@login_required
def incidents_page():
    return render_template("incidents.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/cameras")
@login_required
def cameras_page():
    # Embeds Hik-Connect in an iframe. URL is configurable via the CAMERAS_URL env var
    # (or per-browser in the UI). Note: Hik-Connect may block iframing (X-Frame-Options).
    return render_template(
        "cameras.html",
        google_user=session.get("google_user"),
        b64_en=load_logo(),
        cameras_url=os.environ.get("CAMERAS_URL", ""),
    )


@app.route("/api/vehicles_lookup")
@login_required
def vehicles_lookup():
    """Serve the vehicle/driver registry (authenticated — contains personal data)."""
    path = os.path.join(app.root_path, "vehicles_lookup.json")
    if not os.path.exists(path):
        return jsonify([])
    try:
        with open(path, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception:
        logger.exception("vehicles_lookup read error")
        return jsonify([])


@app.route("/api/download_workshop_template")
@login_required
def download_workshop_template():
    """Download the workshop report Excel template."""
    from flask import send_file
    template_path = os.path.join(os.path.dirname(__file__), "تقرير الورشة.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "قالب تقرير الورشة غير موجود"}), 404
    return send_file(template_path, as_attachment=True, download_name="تقرير_الورشة.xlsx")


# ─── Template Upload & Management ─────────────────────────────
def get_template_path(tab):
    """Return path to user-uploaded template, or default template fallback."""
    user_path = os.path.join(TEMPLATE_DIR, f"{tab}_template.xlsx")
    if os.path.exists(user_path):
        return user_path
    default_name = DEFAULT_TEMPLATES.get(tab)
    if default_name:
        default_path = os.path.join(os.path.dirname(__file__), default_name)
        if os.path.exists(default_path):
            return default_path
    return None


def inject_logo(ws, cell="A1"):
    """Insert company logo into worksheet if logo file exists."""
    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width = 500
            img.height = 120
            ws.add_image(img, cell)
        except Exception:
            pass


MAX_TEMPLATE_SIZE = 16 * 1024 * 1024  # 16 MB


@app.route("/api/upload_template/<tab>", methods=["POST"])
@login_required
def upload_template(tab):
    if is_workstation():
        return jsonify({"error": "غير متاح في محطة العمل"}), 403
    if tab not in VALID_TABS:
        return jsonify({"error": "Invalid tab name"}), 400
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files allowed"}), 400
    # Check file size
    f.seek(0, 2)
    size = f.tell()
    f.seek(0)
    if size > MAX_TEMPLATE_SIZE:
        return jsonify({"error": "File size exceeds 16MB limit"}), 413
    safe_name = secure_filename(f.filename)
    dest = os.path.join(TEMPLATE_DIR, f"{tab}_template.xlsx")
    f.save(dest)
    logger.info("Template uploaded for tab '%s': %s (%d bytes)", tab, safe_name, size)
    return jsonify({"success": True, "tab": tab, "filename": safe_name})


@app.route("/api/templates", methods=["GET"])
@login_required
def list_templates():
    result = {}
    for tab in VALID_TABS:
        user_path = os.path.join(TEMPLATE_DIR, f"{tab}_template.xlsx")
        result[tab] = {"has_custom": os.path.exists(user_path)}
    return jsonify(result)


@app.route("/api/delete_template/<tab>", methods=["DELETE"])
@login_required
def delete_template(tab):
    if is_workstation():
        return jsonify({"error": "غير متاح في محطة العمل"}), 403
    if tab not in VALID_TABS:
        return jsonify({"error": "Invalid tab"}), 400
    user_path = os.path.join(TEMPLATE_DIR, f"{tab}_template.xlsx")
    if os.path.exists(user_path):
        os.remove(user_path)
    return jsonify({"success": True})


# ─── Generate Invoice (Server-Side) ──────────────────────────
@app.route("/api/generate_invoice", methods=["POST"])
@login_required
def generate_invoice():
    try:
        data = request.json
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.title = "فاتورة"

        # Check for custom template
        tpl_path = get_template_path("invoice")
        if tpl_path:
            wb = openpyxl.load_workbook(tpl_path)
            ws = wb.active
        else:
            # Build simple professional invoice from scratch
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            P = "0C2340"
            thin = Side(style="thin", color=P)
            border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # Logo area
            ws.merge_cells("A1:G3")
            inject_logo(ws, "A1")

            # Title
            ws.merge_cells("A4:G4")
            ws["A4"] = "فاتورة صيانة مركبة"
            ws["A4"].font = Font(name="Arial", size=16, bold=True, color=P)
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

            # Info fields
            info = [
                ("A5", "اسم السائق:", "B5", data.get("driver", "")),
                ("D5", "التاريخ:", "E5", data.get("date", "")),
                ("A6", "نوع السيارة:", "B6", data.get("car", "")),
                ("D6", "رقم اللوحة:", "E6", data.get("plate", "")),
                ("A7", "نوع الطلب:", "B7", data.get("requestType", "")),
            ]
            for lbl_cell, lbl, val_cell, val in info:
                ws[lbl_cell] = lbl
                ws[lbl_cell].font = Font(name="Arial", size=11, bold=True, color=P)
                ws[val_cell] = val
                ws[val_cell].font = Font(name="Arial", size=11)

            # Table header
            headers = ["م", "الوصف", "الكمية", "السعر", "المبلغ"]
            for i, h in enumerate(headers):
                c = ws.cell(row=9, column=i + 1, value=h)
                c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=P)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            # Data row
            qty = float(data.get("quantity", 0) or 0)
            price = float(data.get("price", 0) or 0)
            subtotal = qty * price
            row_data = [1, data.get("description", ""), qty, price, subtotal]
            for i, v in enumerate(row_data):
                c = ws.cell(row=10, column=i + 1, value=v)
                c.font = Font(name="Arial", size=11)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            # Totals
            tax = subtotal * 0.15
            total = subtotal + tax
            for ri, (lbl, val) in enumerate(
                [
                    ("المبلغ الإجمالي (قبل الضريبة)", subtotal),
                    ("ضريبة القيمة المضافة (15%)", tax),
                    ("المجموع الكلي (مع الضريبة)", total),
                ]
            ):
                r = 11 + ri
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws.cell(row=r, column=1, value=lbl).font = Font(
                    name="Arial", size=11, bold=True, color=P
                )
                ws.cell(row=r, column=5, value=val).font = Font(
                    name="Arial", size=12, bold=(ri == 2), color=P
                )

            # Signatures
            ws.cell(row=15, column=1, value="توقيع السائق").font = Font(
                name="Arial", bold=True, color=P
            )
            ws.cell(row=15, column=5, value="اعتماد الإدارة").font = Font(
                name="Arial", bold=True, color=P
            )

            # Column widths
            for col, w in [
                (1, 8),
                (2, 30),
                (3, 12),
                (4, 12),
                (5, 15),
                (6, 12),
                (7, 12),
            ]:
                ws.column_dimensions[chr(64 + col)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        driver = data.get("driver", "").replace(" ", "_")[:20]
        plate = data.get("plate", "").replace(" ", "_")[:15]
        fname = f"فاتورة_{driver}_{plate}.xlsx"
        return jsonify({"success": True, "file_b64": b64, "filename": fname})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/generate_schedule", methods=["POST"])
@login_required
def generate_schedule():
    try:
        data = request.json
        template_path = os.path.join(app.root_path, "schedule_base.xlsx")

        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Schedule template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def safe_set(sheet, row, col, val):
            cell = sheet.cell(row=row, column=col)
            if not isinstance(cell, MC):
                cell.value = val

        # Set date in row 6 col 1
        schedule_date = data.get("date", "")
        safe_set(ws, 6, 1, schedule_date)

        main_data = data.get("main", [])
        spare_data = data.get("spare", [])

        # We need to write from bottom up or just write all and delete rows from bottom up

        # 1. Summary Block (Row 53)
        dina = sum(1 for d in main_data if "دينا" in d.get("type", ""))
        lorry = sum(1 for d in main_data if "لوري" in d.get("type", ""))
        delivery = sum(1 for d in main_data if d.get("job", "") == "موصل")
        dist = sum(1 for d in main_data if d.get("job", "") == "موزع")

        safe_set(ws, 53, 2, len(main_data))
        safe_set(ws, 53, 3, dina)
        safe_set(ws, 53, 4, lorry)
        safe_set(ws, 53, 5, delivery)
        safe_set(ws, 53, 6, dist)
        safe_set(ws, 53, 7, len(spare_data))

        # 2. Spare Data (Row 42-49)
        for idx, rd in enumerate(spare_data[:8]):
            r = 42 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 8, rd.get("typemodel", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # 3. Main Data (Row 9-36)
        for idx, rd in enumerate(main_data[:28]):
            r = 9 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 5, rd.get("job", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 7, rd.get("model", ""))
            safe_set(ws, r, 8, rd.get("type", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 11, rd.get("serial", ""))
            safe_set(ws, r, 12, rd.get("inspect", ""))
            safe_set(ws, r, 13, rd.get("license", ""))
            safe_set(ws, r, 14, rd.get("drivercard", ""))
            safe_set(ws, r, 15, rd.get("opcard", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # Delete unused rows from bottom up to avoid shifting index bugs
        spare_deletions = 8 - len(spare_data) if len(spare_data) < 8 else 0
        main_deletions = 28 - len(main_data) if len(main_data) < 28 else 0

        if spare_deletions > 0:
            ws.delete_rows(42 + len(spare_data), spare_deletions)
        if main_deletions > 0:
            ws.delete_rows(9 + len(main_data), main_deletions)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/generate_washing", methods=["POST"])
@login_required
def generate_washing():
    try:
        data = request.json
        vehicles = data.get("vehicles", [])

        template_path = os.path.join(app.root_path, "washing_template.xlsx")
        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Washing template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Insert Logo
        logo_path = os.path.join(
            app.root_path, "templates", "ApplicationFrameHost_KUIZUuJ46O (1).png"
        )
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 300
            img.height = 90
            ws.add_image(img, "F1")

        # Unmerge ALL merged cells to avoid MergedCell write errors
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            ws.unmerge_cells(str(mr))

        # Copy header style from row 4 for reuse
        from copy import copy as shallow_copy
        header_fills = {}
        header_fonts = {}
        header_aligns = {}
        header_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=4, column=c)
            header_fills[c] = shallow_copy(cell.fill)
            header_fonts[c] = shallow_copy(cell.font)
            header_aligns[c] = shallow_copy(cell.alignment)
            header_borders[c] = shallow_copy(cell.border)

        # Copy a data row style (row 5) for styling data rows
        data_fills = {}
        data_fonts = {}
        data_aligns = {}
        data_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=5, column=c)
            data_fills[c] = shallow_copy(cell.fill)
            data_fonts[c] = shallow_copy(cell.font)
            data_aligns[c] = shallow_copy(cell.alignment)
            data_borders[c] = shallow_copy(cell.border)

        # Write vehicle data starting at row 5
        for idx, v in enumerate(vehicles):
            r = 5 + idx
            ws.cell(row=r, column=1, value=v.get("id", idx + 1))
            ws.cell(row=r, column=2, value=v.get("plate", ""))
            ws.cell(row=r, column=3, value=v.get("type", ""))
            ws.cell(row=r, column=4, value=v.get("driver", ""))
            months = v.get("m", [])
            total = sum(months)
            for m_idx in range(12):
                val = "استلم" if m_idx < len(months) and months[m_idx] == 1 else None
                ws.cell(row=r, column=5 + m_idx, value=val)
            ws.cell(row=r, column=17, value=total)
            # Apply data styling
            for c in range(1, 19):
                cell = ws.cell(row=r, column=c)
                cell.fill = shallow_copy(data_fills.get(c, data_fills[1]))
                cell.font = shallow_copy(data_fonts.get(c, data_fonts[1]))
                cell.alignment = shallow_copy(data_aligns.get(c, data_aligns[1]))
                cell.border = shallow_copy(data_borders.get(c, data_borders[1]))

        # Summary row: right after last vehicle
        summary_row = 5 + len(vehicles)
        ws.cell(row=summary_row, column=1, value="إجمالي الغسيل الشهري")
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
        for m_idx in range(12):
            col = 5 + m_idx
            start_cell = ws.cell(row=5, column=col).coordinate.replace("5", "")
            formula = '=COUNTIF(%s5:%s%d,"استلم")' % (start_cell, start_cell, summary_row - 1)
            ws.cell(row=summary_row, column=col, value=formula)
        ws.cell(row=summary_row, column=17, value="=SUM(Q5:Q%d)" % (summary_row - 1))

        # Style summary row bold
        from openpyxl.styles import Font, PatternFill, Alignment
        summary_font = Font(name="Cairo", size=11, bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        summary_align = Alignment(horizontal="center", vertical="center")
        for c in range(1, 19):
            cell = ws.cell(row=summary_row, column=c)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = summary_align

        # Footer row
        footer_row = summary_row + 2
        ws.cell(row=footer_row, column=1, value="تم إعداد هذا الجدول بواسطة قسم الحركة - فرع الدمام")
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=18)

        # Re-merge header rows
        ws.merge_cells("A1:R1")
        ws.merge_cells("A2:D2")
        ws.merge_cells("A3:R3")

        # Update total washes in header
        total_washes = sum(sum(v.get("m", [])) for v in vehicles)
        ws.cell(row=2, column=12, value=total_washes)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/washing_data", methods=["GET", "POST"])
@login_required
def washing_data():
    # Workstation mode writes/reads an isolated copy (id=2); the real data (id=1) is untouched.
    if request.method == "POST":
        try:
            vehicles = (request.json or {}).get("vehicles", [])
            blob_set("washing_schedule", vehicles)
            _audit_add("تحديث", "جدول الغسيل", len(vehicles) if isinstance(vehicles, list) else None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("washing_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ جدول الغسيل."}), 500
    try:
        data = blob_get("washing_schedule")
        if data is not None:
            return jsonify({"success": True, "vehicles": data})
        return jsonify({"success": False, "vehicles": []})
    except Exception:
        logger.exception("washing_data GET error")
        return jsonify({"success": False, "error": "تعذّر جلب جدول الغسيل."}), 500


@app.route("/api/employees", methods=["GET", "POST"])
@login_required
def employees_data():
    """Persist the branch employees grid using the hybrid SQL hr_employees table."""
    if request.method == "POST":
        try:
            req_data = request.json or {}
            reason = req_data.get("reason")
            try:
                audit_and_verify("تحديث", "بيانات الموظفين (SQL)", reason)
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            
            rows = req_data.get("rows", [])
            with db_connection() as db:
                db.execute("DELETE FROM hr_employees")
                for row in rows:
                    if not isinstance(row, list):
                        continue
                    if len(row) < 46:
                        row += [''] * (46 - len(row))
                    empid = str(row[0] or '').strip()
                    iqama = str(row[1] or '').strip()
                    name = str(row[2] or '').strip()
                    plate = str(row[9] or '').strip()
                    phone = str(row[7] or '').strip()
                    job = str(row[6] or '').strip()
                    import json
                    details_json = json.dumps(row)
                    db.execute("INSERT INTO hr_employees (empid, iqama, name, plate, phone, job, details) VALUES (?, ?, ?, ?, ?, ?, ?)",
                               (empid, iqama, name, plate, phone, job, details_json))
                db.commit()
            blob_set("employees", rows)
            return jsonify({"success": True})
        except Exception:
            logger.exception("employees_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ بيانات الموظفين."}), 500
    try:
        with db_connection() as db:
            db_rows = db.execute("SELECT details FROM hr_employees").fetchall()
            import json
            data = [json.loads(r["details"]) for r in db_rows if r["details"]]
            
        if not data:
            fallback = blob_get("employees")
            data = fallback if fallback is not None else []
            
        return jsonify({"success": True, "rows": data})
    except Exception:
        logger.exception("employees_data GET error")
        return jsonify({"success": False, "error": "تعذّر جلب بيانات الموظفين."}), 500


# ===== Learned driver registry (per-branch) =====
# Maps national-id (iqama, digits only) → the driver's personal employee fields, harvested
# from the weekly schedule so بطاقة السائق / الوظيفة / الجوال auto-fill on every future pick.
_REG_FIELDS = ("empid", "phone", "job", "drivercard", "empNotes")
DRIVER_REG_MAX = 4000  # safety cap on distinct drivers remembered per branch
# Per-field length caps so a huge paste into الملاحظات can't bloat the blob that ships to every client.
_REG_FIELD_MAX = {"empid": 40, "phone": 40, "job": 40, "drivercard": 40, "empNotes": 200}
# Serialize the registry read-modify-write. The Procfile pins gunicorn to --workers 1 --threads 8,
# so one in-process lock fully prevents lost updates when two admins on the same branch save at once.
_DRIVER_REG_LOCK = threading.Lock()

# Serializes the alerts-center inline-edit read-modify-write (and the Document Archive
# delete) against itself, for the same reason as _DRIVER_REG_LOCK above.
_ALERTS_CENTER_LOCK = threading.Lock()


def _norm_iqama(v):
    return re.sub(r"\D", "", str(v or ""))


def _harvest_driver_registry(sd):
    """Best-effort: fold any non-empty personal fields from the schedule rows into the
    per-branch driver registry, keyed by iqama. Latest non-empty value wins; a blank NEVER
    clears a remembered value — so a swap that clears a row keeps the driver's memory intact.
    Never raises into the caller (the schedule save must always succeed)."""
    if not isinstance(sd, dict):
        return
    with _DRIVER_REG_LOCK:  # atomic get→mutate→set within the single gunicorn worker
        reg = blob_get("driver_registry") or {}
        if not isinstance(reg, dict):
            reg = {}
        changed = False
        for section in ("main", "spare", "vacation"):
            for row in (sd.get(section) or []):
                if not isinstance(row, dict):
                    continue
                key = _norm_iqama(row.get("iqama"))
                if not key:
                    continue
                existing = key in reg
                if not existing and len(reg) >= DRIVER_REG_MAX:
                    continue  # cap reached: stop remembering NEW drivers, still update existing ones
                entry = reg.get(key) if existing else {}
                if not isinstance(entry, dict):
                    entry = {}
                local_changed = False
                for f in _REG_FIELDS:
                    val = str(row.get(f, "") or "").strip()[:_REG_FIELD_MAX.get(f, 80)]
                    if val and entry.get(f) != val:
                        entry[f] = val
                        local_changed = True
                if local_changed and entry:
                    reg[key] = entry
                    changed = True
        if changed:
            blob_set("driver_registry", reg)


# ===== Learned vehicle registry (per-branch) — symmetric to the driver registry above =====
# Maps normalized plate → the vehicle's technical fields, harvested from the weekly schedule
# so a known vehicle's spec auto-fills on every future selection, mirroring the driver side:
# selecting a driver never touches the vehicle box, and selecting a vehicle never touches the
# driver box — each "swap" only ever replaces its own data.
_VEH_REG_FIELDS = ("model", "vtype", "pallets", "load", "vserial", "inspect", "license", "opcard", "notes")
VEHICLE_REG_MAX = 4000
_VEH_REG_FIELD_MAX = {"model": 20, "vtype": 40, "pallets": 10, "load": 20, "vserial": 40,
                      "inspect": 40, "license": 40, "opcard": 40, "notes": 300}
_VEHICLE_REG_LOCK = threading.Lock()
_AR_DIGITS = "٠١٢٣٤٥٦٧٨٩"


def _norm_plate(v):
    """Digits-then-letters, Arabic-Indic digits folded to Latin — matches window.normalizePlate
    in app_ux.js exactly, so the same plate always resolves to the same registry key regardless
    of digit script or letter/digit typing order."""
    s = str(v or "")
    for i, ch in enumerate(_AR_DIGITS):
        s = s.replace(ch, str(i))
    s = re.sub(r"\s+", "", s)
    digits = "".join(re.findall(r"\d+", s))
    letters = "".join(re.findall(r"[^\d]+", s))
    return digits + letters


def _harvest_vehicle_registry(sd):
    """Best-effort: fold any non-empty technical-spec fields from schedule rows (main/spare —
    vacation rows carry no vehicle) into the per-branch vehicle registry, keyed by normalized
    plate. Latest non-empty value wins; a blank NEVER clears a remembered value. Never raises
    into the caller (the schedule save must always succeed)."""
    if not isinstance(sd, dict):
        return
    with _VEHICLE_REG_LOCK:
        reg = blob_get("vehicle_registry") or {}
        if not isinstance(reg, dict):
            reg = {}
        changed = False
        for section in ("main", "spare"):
            for row in (sd.get(section) or []):
                if not isinstance(row, dict):
                    continue
                key = _norm_plate(row.get("plate"))
                if not key:
                    continue
                existing = key in reg
                if not existing and len(reg) >= VEHICLE_REG_MAX:
                    continue
                entry = reg.get(key) if existing else {}
                if not isinstance(entry, dict):
                    entry = {}
                local_changed = False
                for f in _VEH_REG_FIELDS:
                    val = str(row.get(f, "") or "").strip()[:_VEH_REG_FIELD_MAX.get(f, 80)]
                    if val and entry.get(f) != val:
                        entry[f] = val
                        local_changed = True
                if local_changed and entry:
                    reg[key] = entry
                    changed = True
        if changed:
            blob_set("vehicle_registry", reg)


@app.route("/api/schedule_data", methods=["GET", "POST"])
@login_required
def schedule_data():
    """Persist the weekly schedule (main/spare/vacation/summary). Sandboxed for workstation."""
    if request.method == "POST":
        try:
            sd = request.json or {}
            reason = sd.get("reason")
            try:
                audit_and_verify("تحديث", "الجدول الأسبوعي", reason)
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            
            blob_set("schedule_data", sd)
            try:
                _harvest_driver_registry(sd)  # learn each driver's card-expiry/job for future autofill
            except Exception:
                logger.warning("driver_registry harvest failed (non-fatal)")
            try:
                _harvest_vehicle_registry(sd)  # learn each vehicle's spec for future autofill
            except Exception:
                logger.warning("vehicle_registry harvest failed (non-fatal)")
            _n = (len(sd.get("main", []) or []) + len(sd.get("spare", []) or [])) if isinstance(sd, dict) else None
            return jsonify({"success": True})
        except Exception:
            logger.exception("schedule_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ الجدول الأسبوعي."}), 500
    try:
        return jsonify({"success": True, "data": blob_get("schedule_data")})
    except Exception:
        logger.exception("schedule_data GET error")
        return jsonify({"success": False, "error": "تعذّر جلب الجدول الأسبوعي."}), 500


@app.route("/api/driver_registry", methods=["GET"])
@login_required
def driver_registry():
    """Per-branch learned driver memory (iqama → personal fields) used by schedule autofill."""
    try:
        reg = blob_get("driver_registry") or {}
        return jsonify({"success": True, "data": reg if isinstance(reg, dict) else {}})
    except Exception:
        logger.exception("driver_registry GET error")
        return jsonify({"success": False, "error": "تعذّر جلب سجل السائقين."}), 500


@app.route("/api/vehicle_registry", methods=["GET"])
@login_required
def vehicle_registry():
    """Per-branch learned vehicle memory (plate → technical-spec fields) used by schedule autofill."""
    try:
        reg = blob_get("vehicle_registry") or {}
        return jsonify({"success": True, "data": reg if isinstance(reg, dict) else {}})
    except Exception:
        logger.exception("vehicle_registry GET error")
        return jsonify({"success": False, "error": "تعذّر جلب سجل المركبات."}), 500


@app.route("/api/records", methods=["GET", "POST"])
@login_required
def records_data():
    """Persist the documentation/records log. Sandboxed for workstation."""
    if request.method == "POST":
        try:
            rows = (request.json or {}).get("rows", [])
            blob_set("records_data", rows)
            _audit_add("تحديث", "سجل التوثيق", len(rows) if isinstance(rows, list) else None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("records_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ السجلات."}), 500
    try:
        data = blob_get("records_data")
        return jsonify({"success": True, "rows": data if data is not None else []})
    except Exception:
        logger.exception("records_data GET error")
        return jsonify({"success": False, "rows": []})


@app.route("/api/incidents", methods=["GET", "POST"])
@login_required
def incidents_data():
    """Persist the incidents/violations log (with in-DB base64 attachments). Sandboxed for workstation."""
    if request.method == "POST":
        try:
            rows = (request.json or {}).get("rows", [])
            blob_set("incidents_data", rows)
            atts = sum(len(r.get("attachments", []) or []) for r in rows if isinstance(r, dict))
            _audit_add("تحديث", "سجل الحوادث والمخالفات",
                       len(rows) if isinstance(rows, list) else None,
                       ("%d مرفق" % atts) if atts else "")
            return jsonify({"success": True})
        except Exception:
            logger.exception("incidents_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ سجل الحوادث."}), 500
    try:
        data = blob_get("incidents_data")
        return jsonify({"success": True, "rows": data if data is not None else []})
    except Exception:
        logger.exception("incidents_data GET error")
        return jsonify({"success": False, "rows": []})


@app.route("/api/gps_devices", methods=["GET", "POST"])
@login_required
def gps_devices_data():
    """Persist the GPS tracking-device inventory (editable). Sandboxed for workstation."""
    if request.method == "POST":
        try:
            rows = (request.json or {}).get("rows", [])
            blob_set("gps_devices_data", rows)
            _audit_add("تحديث", "كشف أجهزة التتبع GPS", len(rows) if isinstance(rows, list) else None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("gps_devices_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ كشف الأجهزة."}), 500
    try:
        data = blob_get("gps_devices_data")
        return jsonify({"success": True, "rows": data if data is not None else []})
    except Exception:
        logger.exception("gps_devices_data GET error")
        return jsonify({"success": False, "rows": []})


@app.route("/api/audit_log", methods=["GET"])
@login_required
def audit_log_data():
    """Read-only audit trail. Append-only by design — there is no edit/delete endpoint."""
    try:
        return jsonify({"success": True, "rows": _audit_get()})
    except Exception:
        logger.exception("audit_log GET error")
        return jsonify({"success": True, "rows": []})


@app.route("/api/whoami", methods=["GET"])
@login_required
def whoami():
    """The signed-in identity for the top bar (real session user — no fabricated name)."""
    gu = session.get("google_user") or {}
    name = gu.get("name") or ("محطة العمل" if is_workstation() else "مستخدم النظام")
    return jsonify({"name": name, "role": "إدارة الأسطول والتوثيق"})


# ── Automatic document-expiry email alerts ────────────────────────────────────
# Read-only scan of the schedule + employees blobs for documents that expired or expire
# soon, formatted into a branded email. Recipients come from the request (manual "send now")
# or the ALERT_RECIPIENTS env var (scheduled cron). Sending uses the existing Flask-Mail
# config; if SMTP isn't configured the endpoint reports the reason instead of failing.
ALERT_RECIPIENTS = [e.strip() for e in os.environ.get("ALERT_RECIPIENTS", "").split(",") if e.strip()]
ALERT_CRON_KEY = os.environ.get("ALERT_CRON_KEY", "")


def _parse_iso_date(s):
    if not isinstance(s, str):
        return None
    m = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})\s*$", s)
    if not m:
        return None
    try:
        y, mo, da = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 2000 or y > 2100:
            return None
        return datetime(y, mo, da)
    except Exception:
        return None


def _collect_expiry_alerts(window_days=90, rid=None):
    """Return the list of documents expiring within `window_days` (or already expired).
    Pass `rid` to compute for a SPECIFIC branch (used by the HQ aggregation); otherwise
    the active branch's data is used."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    out = []

    def _bg(table):
        return _blob_get_at(table, rid) if rid is not None else blob_get(table)

    def add(name, plate, doc, dstr):
        d = _parse_iso_date(dstr)
        if not d:
            return
        days = (d - today).days
        if days > window_days:
            return
        if days < 0:
            key, label = "expired", "منتهي"
        elif days <= 30:
            key, label = "d30", "≤ 30 يوم"
        else:
            key, label = "d90", "≤ 90 يوم"
        out.append({"name": name or "—", "plate": plate or "", "doc": doc,
                    "date": dstr, "days": days, "key": key, "label": label})

    sd = _bg("schedule_data") or {}
    if isinstance(sd, dict):
        for sect in ("main", "spare"):
            for r in (sd.get(sect) or []):
                if not isinstance(r, dict):
                    continue
                nm, pl = r.get("name"), r.get("plate")
                add(nm, pl, "الفحص الدوري", r.get("inspect"))
                add(nm, pl, "رخصة السير", r.get("license"))
                add(nm, pl, "بطاقة التشغيل", r.get("opcard"))
                add(nm, pl, "بطاقة السائق", r.get("drivercard"))

    emps = _bg("employees") or []
    if isinstance(emps, list):
        for row in emps:
            if not isinstance(row, list):
                continue
            nm = (row[2] if len(row) > 2 else "") or (row[3] if len(row) > 3 else "")
            pl = row[9] if len(row) > 9 else ""
            if len(row) > 10:
                add(nm, pl, "انتهاء الإقامة", row[10])
            if len(row) > 11:
                add(nm, pl, "انتهاء الجواز", row[11])

    # Uploaded documents (the Document Archive) — their expiry feeds the same alerts + digest.
    docs = _bg("documents_data") or {}
    drows = docs.get("rows") if isinstance(docs, dict) else (docs if isinstance(docs, list) else [])
    for dr in (drows or []):
        if not isinstance(dr, dict):
            continue
        ref = dr.get("entity_ref") or "—"
        plate = ref if dr.get("entity_type") == "vehicle" else ""
        add(ref, plate, dr.get("doc_type") or "وثيقة", dr.get("expiry"))

    # Odometer-Based Maintenance Alerts
    latest_oil = {}
    oils = _bg("oils_data") or []
    orows = oils.get("rows") if isinstance(oils, dict) else (oils if isinstance(oils, list) else [])
    for row in orows:
        if isinstance(row, list) and len(row) > 3:
            plate = str(row[0] or "").strip()
            if not plate: continue
            try: odo = int(row[3])
            except (ValueError, TypeError): continue
            if plate not in latest_oil or odo > latest_oil[plate]:
                latest_oil[plate] = odo

    latest_fuel = {}
    fuel = _bg("fuel_data") or []
    frows = fuel.get("rows") if isinstance(fuel, dict) else (fuel if isinstance(fuel, list) else [])
    for row in frows:
        if isinstance(row, dict):
            plate = str(row.get("plate") or "").strip()
            name = str(row.get("driver") or "").strip()
            try: cur_odo = int(row.get("curOdo") or 0)
            except (ValueError, TypeError): continue
            if not plate: continue
            if plate not in latest_fuel or cur_odo > latest_fuel[plate]['odo']:
                latest_fuel[plate] = {'odo': cur_odo, 'name': name}

    INTERVAL = 10000
    WARNING_THRESHOLD = 9000
    for plate, fdata in latest_fuel.items():
        if plate in latest_oil:
            diff = fdata['odo'] - latest_oil[plate]
            if diff >= WARNING_THRESHOLD:
                remaining_km = INTERVAL - diff
                pseudo_days = max(0, int(remaining_km / 100))
                if diff >= INTERVAL:
                    key, label = "expired", "تجاوز العداد"
                    pseudo_days = -1
                elif remaining_km <= 1000:
                    key, label = "d30", f"باقي {remaining_km} كم"
                else:
                    key, label = "d90", f"باقي {remaining_km} كم"
                
                out.append({
                    "name": fdata['name'] or "—", "plate": plate, "doc": "تغيير زيت وصيانة",
                    "date": f"العداد: {fdata['odo']}", "days": pseudo_days,
                    "key": key, "label": label
                })

    # Attach driver phone numbers for WhatsApp Alerts
    from models.schema import Driver
    phone_map = {}
    db_drivers = Driver.query.filter_by(branch_id=rid).all() if rid is not None else Driver.query.all()
    for d in db_drivers:
        if d.phone:
            phone_map[d.name] = d.phone
            if d.vehicle and d.vehicle.plate_number:
                phone_map[d.vehicle.plate_number] = d.phone

    for a in out:
        a["phone"] = phone_map.get(a["name"]) or phone_map.get(a["plate"]) or ""

    rank = {"expired": 0, "d30": 1, "d90": 2}
    out.sort(key=lambda a: (rank.get(a["key"], 9), a["days"]))
    return out


def _build_alert_email_html(alerts, filter_label=""):
    today_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    def days_txt(a):
        d = a.get("days")
        if d is None:
            return "—"
        if d < 0:
            return "منذ %d يوم" % abs(d)
        if d == 0:
            return "اليوم"
        return "خلال %d يوم" % d

    colors = {"expired": "#dc2626", "d30": "#d97706", "d90": "#ca8a04",
              "valid": "#16a34a", "unknown": "#6b7280"}
    rows = ""
    for a in alerts[:300]:
        c = colors.get(a.get("key"), "#6b7280")
        rows += (
            "<tr>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;direction:ltr;font-family:monospace;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;white-space:nowrap;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;color:%s;font-weight:700;white-space:nowrap;'>%s · %s</td>"
            "</tr>"
        ) % (html.escape(str(a.get("name") or "—")), html.escape(str(a.get("plate") or "—")),
             html.escape(str(a.get("doc") or "")), html.escape(str(a.get("date") or "")),
             c, html.escape(str(a.get("label") or "")), days_txt(a))

    scope = html.escape(str(filter_label or "الكل"))
    return (
        "<div style='font-family:Tahoma,Arial,sans-serif;direction:rtl;text-align:right;background:#f4f6fb;padding:22px;'>"
        "<div style='max-width:760px;margin:auto;background:#fff;border-radius:14px;overflow:hidden;border:1px solid #e6e9f0;'>"
        "<div style='background:#0C2340;padding:18px 24px;border-bottom:4px solid #c9a227;'>"
        "<div style='color:#fff;font-size:18px;font-weight:800;'>BIN ZOMAH INTL. — شركة بن زومة</div>"
        "<div style='color:#c9a227;font-size:13px;font-weight:700;'>تنبيه وثائق الأسطول — فرع الدمام</div></div>"
        "<div style='padding:20px 24px;'>"
        "<p style='color:#334;font-size:14px;margin:0 0 16px;'>الفرز المُرسَل: <b>%s</b> &nbsp;•&nbsp; العدد: <b>%d</b> &nbsp;•&nbsp; حتى تاريخ <b>%s</b></p>"
        "<table style='width:100%%;border-collapse:collapse;font-size:13px;color:#222;'>"
        "<thead><tr style='background:#0C2340;color:#fff;'>"
        "<th style='padding:10px 12px;text-align:right;'>الاسم</th>"
        "<th style='padding:10px 12px;text-align:right;'>اللوحة</th>"
        "<th style='padding:10px 12px;text-align:right;'>الوثيقة</th>"
        "<th style='padding:10px 12px;text-align:right;'>تاريخ الانتهاء</th>"
        "<th style='padding:10px 12px;text-align:right;'>الحالة</th></tr></thead>"
        "<tbody>%s</tbody></table>"
        "<p style='color:#94a3b8;font-size:11px;margin-top:18px;'>هذه رسالة آلية من نظام إدارة الأسطول — شركة بن زومة. الرجاء عدم الرد عليها.</p>"
        "</div></div></div>"
    ) % (scope, len(alerts), today_str, rows)


def _send_expiry_alert_email(recipients, rows=None, filter_label=""):
    """Send the alert email. If `rows` is given (from the dashboard's selected filter),
    email EXACTLY those rows; otherwise fall back to the 'need-action' set."""
    if rows is not None:
        alerts = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            try:
                days = int(r.get("days")) if r.get("days") not in (None, "") else None
            except (TypeError, ValueError):
                days = None
            alerts.append({"name": r.get("name") or "—", "plate": r.get("plate") or "",
                           "doc": r.get("doc") or "", "date": r.get("date") or "",
                           "days": days, "key": r.get("key") or "", "label": r.get("label") or ""})
        if not filter_label:
            filter_label = "مُختار"
    else:
        alerts = [a for a in _collect_expiry_alerts() if a["key"] in ("expired", "d30", "d90")]
        filter_label = filter_label or "تحتاج إجراء"
    if not alerts:
        return {"sent": False, "reason": "no_alerts", "count": 0}
    if not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
        return {"sent": False, "reason": "mail_not_configured", "count": len(alerts)}
    recipients = [r for r in (recipients or []) if r]
    if not recipients:
        return {"sent": False, "reason": "no_recipients", "count": len(alerts)}
    try:
        subject = "🚨 تنبيه وثائق الأسطول — شركة بن زومة (فرع الدمام)"
        if filter_label:
            subject += " — " + filter_label
        msg = Message(
            subject=subject,
            recipients=recipients,
            html=_build_alert_email_html(alerts, filter_label),
            sender=app.config.get("MAIL_DEFAULT_SENDER") or app.config.get("MAIL_USERNAME"),
        )
        mail.send(msg)
        return {"sent": True, "count": len(alerts), "recipients": recipients}
    except Exception as e:
        logger.exception("send expiry alert email failed")
        return {"sent": False, "reason": "send_error", "error": str(e), "count": len(alerts)}


@app.route("/api/expiry_alerts_preview", methods=["GET"])
@login_required
def expiry_alerts_preview():
    """Counts the server sees. The HQ/admin bell aggregates EVERY branch; a branch user
    sees only its own (data is already partitioned by branch)."""
    try:
        def _counts(alerts):
            c = {"expired": 0, "d30": 0, "d90": 0}
            for a in alerts:
                if a["key"] in c:
                    c[a["key"]] += 1
            return c
        mail_cfg = bool(app.config.get("MAIL_USERNAME") and app.config.get("MAIL_PASSWORD"))
        if session.get("is_admin"):
            counts = {"expired": 0, "d30": 0, "d90": 0}
            per_branch = []
            for b in BRANCHES:
                bc = _counts(_collect_expiry_alerts(rid=b["id"]))
                for k in counts:
                    counts[k] += bc[k]
                per_branch.append({"id": b["id"], "name": b["name"], "counts": bc, "total": sum(bc.values())})
            return jsonify({"success": True, "counts": counts, "total": sum(counts.values()),
                            "aggregated": True, "per_branch": per_branch,
                            "mail_configured": mail_cfg, "default_recipients": ALERT_RECIPIENTS})
        counts = _counts(_collect_expiry_alerts())
        return jsonify({"success": True, "counts": counts, "total": sum(counts.values()),
                        "aggregated": False,
                        "mail_configured": mail_cfg, "default_recipients": ALERT_RECIPIENTS})
    except Exception:
        logger.exception("expiry preview error")
        return jsonify({"success": False, "counts": {}, "total": 0})


@app.route("/api/send_expiry_alerts", methods=["POST"])
@login_required
def send_expiry_alerts():
    """Manual 'send now' from the dashboard. Recipients from the request or ALERT_RECIPIENTS."""
    body = request.json or {}
    # Lock: require the access code before sending (matches the workstation password).
    if not hmac.compare_digest(str(body.get("lock", "")), WORKSTATION_PASSWORD):
        return jsonify({"success": False, "reason": "locked"}), 403
    recips = body.get("recipients") or ALERT_RECIPIENTS
    if isinstance(recips, str):
        recips = [e.strip() for e in re.split(r"[,;\s]+", recips) if e.strip()]
    rows = body.get("rows")
    filt = (body.get("filter") or "").strip()
    if isinstance(rows, list) and rows:
        res = _send_expiry_alert_email(recips, rows=rows[:1000], filter_label=filt)
    else:
        res = _send_expiry_alert_email(recips, filter_label=filt)
    if res.get("sent"):
        _audit_add("إرسال", "تنبيهات الوثائق بالبريد", res.get("count"),
                   (("الفرز: " + filt + " — ") if filt else "") + "إلى: " + ", ".join(res.get("recipients", [])))
    return jsonify({"success": res.get("sent", False), **res})


@app.route("/api/cron/expiry_alerts", methods=["GET", "POST"])
def cron_expiry_alerts():
    """Token-protected trigger for an external daily scheduler (ArabCord cron / cron-job.org).
    Not login-protected; guarded by ALERT_CRON_KEY. Uses ALERT_RECIPIENTS for the recipient list."""
    key = request.args.get("key", "")
    if not key and request.is_json:
        key = (request.json or {}).get("key", "")
    if not ALERT_CRON_KEY or not hmac.compare_digest(str(key), ALERT_CRON_KEY):
        return jsonify({"success": False, "error": "unauthorized"}), 401
    res = _send_expiry_alert_email(ALERT_RECIPIENTS)
    if res.get("sent"):
        _audit_add("إرسال تلقائي", "تنبيهات الوثائق بالبريد", res.get("count"),
                   "مجدول — إلى: " + ", ".join(res.get("recipients", [])))
    return jsonify({"success": res.get("sent", False), **res})


# ════════════════════════════════════════════════════════════════════════════════════
# AUTOMATIC scheduled document-expiry alerts — configured IN-APP (no env/external cron),
# driven by an in-process daily scheduler. Sends one branded all-branches digest by email.
# ════════════════════════════════════════════════════════════════════════════════════
def _alert_cfg():
    """Alert settings (global id=1). Disabled by default until the admin configures it."""
    d = _global_blob_get("alert_settings")
    d = d if isinstance(d, dict) else {}
    def _int(v, default):
        try:
            return int(v)
        except (TypeError, ValueError):
            return default
    return {
        "enabled": bool(d.get("enabled", False)),
        "recipients": [e for e in (d.get("recipients") or []) if isinstance(e, str) and e.strip()],
        "hour": max(0, min(23, _int(d.get("hour", 7), 7))),
        "window_days": max(1, min(180, _int(d.get("window_days", 30), 30))),
        "last_sent": d.get("last_sent", ""),
    }


def _alert_cfg_set(cfg):
    _global_blob_set("alert_settings", cfg)


def _collect_all_branches_alerts(window_days):
    """Aggregate expiry alerts across ALL branches, each tagged with the branch name."""
    out = []
    for b in BRANCHES:
        try:
            for a in _collect_expiry_alerts(window_days=window_days, rid=b["id"]):
                if a.get("key") in ("expired", "d30", "d90"):
                    a2 = dict(a)
                    a2["branch"] = b["name"]
                    out.append(a2)
        except Exception:
            logger.warning("expiry collect failed for branch %s", b.get("id"))
    rank = {"expired": 0, "d30": 1, "d90": 2}
    out.sort(key=lambda a: (rank.get(a.get("key"), 9), a.get("days", 0)))
    return out


def _build_digest_html(alerts, window_days):
    today_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    colors = {"expired": "#dc2626", "d30": "#d97706", "d90": "#ca8a04"}

    def days_txt(a):
        d = a.get("days")
        if d is None:
            return "—"
        if d < 0:
            return "منذ %d يوم" % abs(d)
        return "اليوم" if d == 0 else ("خلال %d يوم" % d)

    rows = ""
    for a in alerts[:400]:
        c = colors.get(a.get("key"), "#6b7280")
        rows += (
            "<tr>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;direction:ltr;font-family:monospace;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;'>%s</td>"
            "<td style='padding:9px 12px;border-bottom:1px solid #eee;color:%s;font-weight:700;white-space:nowrap;'>%s</td>"
            "</tr>"
        ) % (html.escape(str(a.get("branch") or "—")), html.escape(str(a.get("name") or "—")),
             html.escape(str(a.get("plate") or "—")), html.escape(str(a.get("doc") or "")),
             html.escape(str(a.get("date") or "")), c, days_txt(a))
    return (
        "<div style='font-family:Tahoma,Arial,sans-serif;direction:rtl;text-align:right;background:#f4f6fb;padding:22px;'>"
        "<div style='max-width:840px;margin:auto;background:#fff;border-radius:14px;overflow:hidden;border:1px solid #e6e9f0;'>"
        "<div style='background:#0C2340;padding:18px 24px;border-bottom:4px solid #c9a227;'>"
        "<div style='color:#fff;font-size:18px;font-weight:800;'>BIN ZOMAH INTL. — شركة بن زومة</div>"
        "<div style='color:#c9a227;font-size:13px;font-weight:700;'>التنبيه اليومي لوثائق الأسطول — جميع الفروع</div></div>"
        "<div style='padding:20px 24px;'>"
        "<p style='color:#334;font-size:14px;margin:0 0 16px;'>وثائق منتهية أو تنتهي خلال <b>%d يوماً</b> &nbsp;•&nbsp; العدد: <b>%d</b> &nbsp;•&nbsp; %s</p>"
        "<table style='width:100%%;border-collapse:collapse;font-size:13px;color:#222;'>"
        "<thead><tr style='background:#0C2340;color:#fff;'>"
        "<th style='padding:10px 12px;text-align:right;'>الفرع</th>"
        "<th style='padding:10px 12px;text-align:right;'>الاسم</th>"
        "<th style='padding:10px 12px;text-align:right;'>اللوحة</th>"
        "<th style='padding:10px 12px;text-align:right;'>الوثيقة</th>"
        "<th style='padding:10px 12px;text-align:right;'>تاريخ الانتهاء</th>"
        "<th style='padding:10px 12px;text-align:right;'>المتبقّي</th></tr></thead>"
        "<tbody>%s</tbody></table>"
        "<p style='color:#94a3b8;font-size:11px;margin-top:18px;'>رسالة آلية يومية من نظام إدارة الأسطول — شركة بن زومة. الرجاء عدم الرد عليها.</p>"
        "</div></div></div>"
    ) % (window_days, len(alerts), today_str, rows)


def _send_scheduled_digest(recipients, window_days):
    recipients = [r for r in (recipients or []) if r]
    if not recipients:
        return {"sent": False, "reason": "no_recipients", "count": 0}
    if not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
        return {"sent": False, "reason": "mail_not_configured", "count": 0}
    alerts = _collect_all_branches_alerts(window_days)
    if not alerts:
        return {"sent": False, "reason": "no_alerts", "count": 0}
    try:
        msg = Message(
            subject="🚨 التنبيه اليومي لوثائق الأسطول — شركة بن زومة (%d)" % len(alerts),
            recipients=recipients,
            html=_build_digest_html(alerts, window_days),
            sender=app.config.get("MAIL_DEFAULT_SENDER") or app.config.get("MAIL_USERNAME"),
        )
        mail.send(msg)
        return {"sent": True, "count": len(alerts), "recipients": recipients}
    except Exception as e:
        logger.exception("scheduled digest failed")
        return {"sent": False, "reason": "send_error", "error": str(e), "count": len(alerts)}


_alert_sched_started = False


def _start_alert_scheduler():
    """One in-process daemon thread; checks every 10 min and sends the daily digest once,
    at/after the configured hour, if enabled. last_sent dedups across restarts."""
    global _alert_sched_started
    if _alert_sched_started:
        return
    _alert_sched_started = True

    def _loop():
        while True:
            try:
                with app.app_context():
                    cfg = _alert_cfg()
                    today = datetime.now().strftime("%Y-%m-%d")
                    if (cfg["enabled"] and cfg["recipients"]
                            and datetime.now().hour >= cfg["hour"] and cfg["last_sent"] != today):
                        alerts = _collect_all_branches_alerts(cfg["window_days"])
                        alert_count = sum(len(x.get("alerts", [])) for x in alerts)
                        if alert_count > 0:
                            send_push_notification("تنبيه أسطول بن زومة", f"يوجد {alert_count} مستند على وشك الانتهاء، يرجى مراجعة لوحة القيادة.")
                        res = _send_scheduled_digest(cfg["recipients"], cfg["window_days"])
                        if res.get("reason") != "mail_not_configured":   # else retry when SMTP is ready
                            cfg["last_sent"] = today
                            _alert_cfg_set(cfg)
                            logger.info("Scheduled expiry digest run: %s", res)
            except Exception:
                logger.exception("alert scheduler tick failed")
            time.sleep(600)

    threading.Thread(target=_loop, name="bz-alert-scheduler", daemon=True).start()
    logger.info("Document-expiry alert scheduler started.")


@app.route("/alerts")
@login_required
def alerts_page():
    """Admin page to configure the automatic daily document-expiry email digest."""
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("alerts.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/alert_settings", methods=["GET", "POST"])
@login_required
def api_alert_settings():
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        cfg = _alert_cfg()
        cfg["mail_configured"] = bool(app.config.get("MAIL_USERNAME") and app.config.get("MAIL_PASSWORD"))
        cfg["preview_count"] = len(_collect_all_branches_alerts(cfg["window_days"]))
        return jsonify({"success": True, "settings": cfg})
    body = request.get_json(silent=True) or {}

    def _recips(v):
        if isinstance(v, str):
            v = re.split(r"[,;\s]+", v)
        return [e.strip() for e in (v or []) if isinstance(e, str) and "@" in e.strip()][:50]

    if body.get("action") == "test":     # send a one-off test now
        cfg = _alert_cfg()
        recips = _recips(body.get("recipients")) or cfg["recipients"]
        try:
            wd = max(1, min(180, int(body.get("window_days") or cfg["window_days"])))
        except (TypeError, ValueError):
            wd = cfg["window_days"]
        res = _send_scheduled_digest(recips, wd)
        if res.get("sent"):
            _audit_add("إرسال تجريبي", "تنبيهات الوثائق التلقائية", res.get("count"),
                       "إلى: " + ", ".join(res.get("recipients", [])))
        return jsonify({"success": res.get("sent", False), **res})

    try:
        hour = max(0, min(23, int(body.get("hour", 7))))
    except (TypeError, ValueError):
        hour = 7
    try:
        wd = max(1, min(180, int(body.get("window_days", 30))))
    except (TypeError, ValueError):
        wd = 30
    cfg = {"enabled": bool(body.get("enabled")), "recipients": _recips(body.get("recipients")),
           "hour": hour, "window_days": wd, "last_sent": _alert_cfg().get("last_sent", "")}
    _alert_cfg_set(cfg)
    _audit_add("إعداد", "تنبيهات الوثائق التلقائية", len(cfg["recipients"]),
               ("مُفعّلة" if cfg["enabled"] else "مُعطّلة") + " · الساعة %d · خلال %d يوم" % (hour, wd))
    return jsonify({"success": True, "settings": cfg})


# ════════════════════════════════════════════════════════════════════════════════════
# DOCUMENT ARCHIVE — upload vehicle/employee documents (image/PDF stored as base64), per
# branch, with expiry that FEEDS the expiry alerts. Additive user data (frozen-seed safe).
# The list returns metadata only (fast); files are streamed by id from a separate endpoint.
# ════════════════════════════════════════════════════════════════════════════════════
DOC_MAX_FILE_BYTES = int(2.6 * 1024 * 1024)         # ~2.5 MB per file (decoded)
DOC_MAX_ROWS = 500                                   # per branch
DOC_MAX_BLOB_BYTES = 45 * 1024 * 1024                # ~45 MB total per branch
DOC_ALLOWED_MIME = {"image/jpeg", "image/png", "image/webp", "image/gif", "application/pdf"}
DOC_TYPES = ["استمارة", "تأمين", "الفحص الدوري", "رخصة السير", "بطاقة التشغيل", "بطاقة السائق",
             "الإقامة", "جواز السفر", "رخصة قيادة", "عقد", "أخرى"]


def _documents_rows():
    d = blob_get("documents_data") or {}
    rows = d.get("rows") if isinstance(d, dict) else (d if isinstance(d, list) else [])
    return [r for r in (rows or []) if isinstance(r, dict)]


def _b64_size(data_uri):
    """Decoded byte size of a data: URI, computed from the base64 length (no full decode)."""
    try:
        s = str(data_uri)
        b64 = s.split(",", 1)[1] if "," in s else s
        b64 = b64.strip()
        return (len(b64) * 3) // 4 - b64[-2:].count("=")
    except Exception:
        return 0


def _doc_meta(r):
    m = {k: r.get(k) for k in ("id", "entity_type", "entity_ref", "doc_type", "number", "expiry", "notes", "ts")}
    fl = r.get("file") or {}
    m["file"] = {"name": fl.get("name"), "mime": fl.get("mime"), "size": fl.get("size")}
    return m


def _sniff_ok(data_uri, mime):
    """Defense-in-depth: the actual decoded bytes must match the CLAIMED mime (blocks
    polyglots / disguised files). Header check only — cheap."""
    try:
        s = str(data_uri)
        b64 = (s.split(",", 1)[1] if "," in s else s).strip()
        chunk = b64[:64]
        chunk += "=" * (-len(chunk) % 4)
        head = base64.b64decode(chunk)[:16]
    except Exception:
        return False
    if mime == "image/jpeg":
        return head[:3] == b"\xff\xd8\xff"
    if mime == "image/png":
        return head[:8] == b"\x89PNG\r\n\x1a\n"
    if mime == "image/gif":
        return head[:6] in (b"GIF87a", b"GIF89a")
    if mime == "image/webp":
        return head[:4] == b"RIFF" and head[8:12] == b"WEBP"
    if mime == "application/pdf":
        return head[:5] == b"%PDF-"
    return False


@app.route("/documents")
@login_required
def documents_page():
    return render_template("documents.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/documents", methods=["GET", "POST"])
@login_required
def api_documents():
    branch_id = current_branch_id()
    
    if request.method == "GET":
        docs = Document.query.filter_by(branch_id=branch_id).all()
        rows = []
        for d in docs:
            rows.append({
                "id": str(d.id),
                "entity_type": d.entity_type,
                "entity_ref": d.entity_ref,
                "doc_type": d.doc_type,
                "number": d.number,
                "expiry": d.expiry.strftime('%Y-%m-%d') if d.expiry else "",
                "notes": d.notes,
                "ts": d.upload_date.strftime('%Y-%m-%d') if d.upload_date else "",
                "file": {"name": "ملف", "mime": d.mime_type, "size": d.file_size}
            })
        return jsonify({"success": True, "rows": rows, "doc_types": DOC_TYPES})
        
    body = request.get_json(silent=True) or {}
    
    current_count = Document.query.filter_by(branch_id=branch_id).count()
    if current_count >= DOC_MAX_ROWS:
        return jsonify({"success": False, "error": "بلغت الحد الأقصى لعدد الوثائق في هذا الفرع (%d)." % DOC_MAX_ROWS}), 400
        
    f = body.get("file") or {}
    data = f.get("data") or ""
    mime = (f.get("mime") or "").lower().strip()
    
    if not data or not str(data).startswith("data:"):
        return jsonify({"success": False, "error": "الملف مطلوب."}), 400
    if mime not in DOC_ALLOWED_MIME:
        return jsonify({"success": False, "error": "نوع الملف غير مدعوم (صور JPG/PNG/WebP/GIF أو PDF فقط)."}), 400
        
    size = _b64_size(data)
    if size <= 0 or size > DOC_MAX_FILE_BYTES:
        return jsonify({"success": False, "error": "حجم الملف يتجاوز 2.5 ميجابايت."}), 400
    if not _sniff_ok(data, mime):
        return jsonify({"success": False, "error": "محتوى الملف لا يطابق نوعه المُعلَن."}), 400
        
    if not body.get("entity_ref"):
        return jsonify({"success": False, "error": "حدّد المركبة/الموظف (المرجع)."}), 400

    def parse_date(dstr):
        if not dstr: return None
        try: return datetime.strptime(str(dstr)[:10], '%Y-%m-%d').date()
        except: return None

    try:
        new_doc = Document(
            branch_id=branch_id,
            doc_type=str(body.get("doc_type") or "أخرى")[:60],
            entity_type="vehicle" if body.get("entity_type") == "vehicle" else "employee",
            entity_ref=str(body.get("entity_ref") or "")[:120],
            number=str(body.get("number") or "")[:80],
            expiry=parse_date(body.get("expiry")),
            notes=str(body.get("notes") or "")[:300],
            file_data=data,
            mime_type=mime,
            file_size=size,
            file_path="base64", # legacy dummy path
            upload_date=datetime.now().date()
        )
        db.session.add(new_doc)
        db.session.flush() # To get ID
        
        audit = AuditLog(
            user_id=getattr(g, 'user', None),
            branch_id=branch_id,
            action="إضافة مستند",
            target_table="erp_documents",
            target_id=str(new_doc.id)
        )
        db.session.add(audit)
        db.session.commit()
        
        row_res = {
            "id": str(new_doc.id),
            "entity_type": new_doc.entity_type,
            "entity_ref": new_doc.entity_ref,
            "doc_type": new_doc.doc_type,
            "number": new_doc.number,
            "expiry": new_doc.expiry.strftime('%Y-%m-%d') if new_doc.expiry else "",
            "notes": new_doc.notes,
            "ts": new_doc.upload_date.strftime('%Y-%m-%d') if new_doc.upload_date else "",
            "file": {"name": "ملف", "mime": new_doc.mime_type, "size": new_doc.file_size}
        }
        return jsonify({"success": True, "row": row_res})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error saving document: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/documents/<int:doc_id>", methods=["DELETE"])
@login_required
def api_documents_delete(doc_id):
    try:
        doc = Document.query.filter_by(id=doc_id, branch_id=current_branch_id()).first()
        if not doc:
            return jsonify({"success": False, "error": "غير موجود."}), 404
            
        db.session.delete(doc)
        audit = AuditLog(
            user_id=getattr(g, 'user', None),
            branch_id=current_branch_id(),
            action="حذف مستند",
            target_table="erp_documents",
            target_id=str(doc_id)
        )
        db.session.add(audit)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting document: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/documents/<int:doc_id>/file")
@login_required
def api_document_file(doc_id):
    doc = Document.query.filter_by(id=doc_id, branch_id=current_branch_id()).first()
    if not doc or not doc.file_data:
        return ("", 404)
        
    try:
        s = str(doc.file_data)
        raw = base64.b64decode(s.split(",", 1)[1] if "," in s else s)
    except Exception:
        return ("", 404)
        
    from flask import Response
    mime = doc.mime_type or "application/octet-stream"
    resp = Response(raw, mimetype=mime)
    # Hardening: never sniff; preview images inline, force everything else (e.g. PDF) to
    # download so no embedded script can run same-origin. Bytes are magic-validated on upload.
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Content-Disposition"] = "inline" if mime.startswith("image/") else "attachment"
    resp.headers["Cache-Control"] = "private, max-age=600"
    return resp
    return ("", 404)


# ── مركز تنبيهات الوثائق — inline edit from the dashboard ─────────────────────
# The dashboard's alert table is BUILT from 3 different sources (weekly-schedule rows,
# employees rows, uploaded Document Archive rows). This endpoint writes an edited value
# straight back to whichever source produced that alert row — every column shown in the
# table (الاسم / رقم اللوحة / الوثيقة / تاريخ الانتهاء) is editable, not just the date —
# using the SAME branch-scoped blob_get/blob_set every other tab uses, so it respects the
# active session's branch and the workstation sandbox exactly like editing the source tab
# directly would.
#
# `targetField` (logical, same vocabulary across all 3 sources) says WHAT is being edited:
#   'date'    → the specific date field the alert is about (dateField: inspect/license/
#               opcard/drivercard for schedule, 10/11 for employee, always expiry for document)
#   'name'    → schedule row.name / employee col 2-or-3 (whichever is populated) / document entity_ref
#   'plate'   → schedule row.plate / employee col 9 / document entity_ref (vehicle docs only)
#   'doctype' → document row.doc_type only (schedule/employee alerts have a fixed label, not
#               a real field, so 'doctype' is rejected for those two sources)
#
# Schedule/employee rows have no stable id, so they're addressed positionally (idx) and
# double-checked against a client-supplied identity snapshot (plate and/or iqama) before
# the write is allowed — anchored on whichever identity field is NOT the one being edited
# (e.g. editing plate anchors on iqama, editing a driver-scoped field anchors on plate) so
# "fix the wrong plate/name" is exactly the case this guard is designed to allow safely. A
# BLANK anchor never counts as a match (two blank rows would otherwise look identical); a
# mismatch (or a row that no longer exists at that index) means the underlying list shifted
# since the page loaded, so the write is refused with 409 and the client is told to reload.
# Document Archive rows already carry a real unique `id` (_documents_rows) so no
# positional/identity check is needed there regardless of which field is being edited.
#
# _ALERTS_CENTER_LOCK (defined near _DRIVER_REG_LOCK above) serializes this route's own
# read-modify-write against itself and against the Document Archive's delete endpoint
# (both threaded under gunicorn --workers 1 --threads 8). It does NOT serialize against
# the sibling whole-blob saves (POST /api/schedule_data, POST /api/employees) — those
# still follow this app's existing last-write-wins model for every tab, same as before
# this endpoint existed.
_ALERTS_FIELD_CAPS = {"date": 10, "name": 100, "plate": 30, "doctype": 60}


def _alerts_center_clean_value(target_field, raw):
    value = str(raw or "").strip()[:_ALERTS_FIELD_CAPS.get(target_field, 60)]
    if target_field == "date" and value and not _parse_iso_date(value):
        return None, jsonify({"success": False, "error": "تاريخ غير صالح."}), 400
    return value, None, None


@app.route("/api/alerts_center/update", methods=["POST"])
@login_required
def alerts_center_update():
    try:
        body = request.get_json(silent=True) or {}
        src = body.get("src")
        target_field = body.get("targetField")
        if target_field not in ("date", "name", "plate", "doctype"):
            return jsonify({"success": False, "error": "طلب غير صالح."}), 400
        value, err_resp, err_code = _alerts_center_clean_value(target_field, body.get("value"))
        if err_resp is not None:
            return err_resp, err_code

        if src == "schedule":
            if target_field == "doctype":
                return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            section = body.get("section")
            if section not in ("main", "spare"):
                return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            try:
                idx = int(body.get("idx"))
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            if target_field == "date":
                date_field = body.get("dateField")
                if date_field not in ("inspect", "license", "opcard", "drivercard"):
                    return jsonify({"success": False, "error": "طلب غير صالح."}), 400
                store_key = date_field
            else:
                store_key = target_field  # 'name' or 'plate'
            with _ALERTS_CENTER_LOCK:
                sd = blob_get("schedule_data")
                rows = sd.get(section) if isinstance(sd, dict) else None
                if not isinstance(rows, list) or idx < 0 or idx >= len(rows) or not isinstance(rows[idx], dict):
                    return jsonify({"success": False, "error": "لم يعد هذا الصف موجوداً — أعد تحميل الصفحة."}), 409
                row = rows[idx]
                # Anchor identity on whichever field ISN'T being written: plate identifies the
                # row/vehicle slot itself; iqama identifies the driver currently assigned to it.
                if store_key == "plate":
                    akey = _norm_iqama(row.get("iqama")); ok = bool(akey) and akey == _norm_iqama(body.get("iqama"))
                else:  # name, drivercard, inspect, license, opcard — all anchor on the row's plate
                    akey = str(row.get("plate") or "").strip(); ok = bool(akey) and akey == str(body.get("plate") or "").strip()
                if not ok:
                    return jsonify({"success": False, "error": "تغيّرت بيانات الصف — أعد تحميل الصفحة."}), 409
                if store_key in ("name", "plate") and not value:
                    return jsonify({"success": False, "error": "هذا الحقل لا يمكن تفريغه."}), 400
                row[store_key] = value
                blob_set("schedule_data", sd)
            try:
                _harvest_driver_registry(sd)
            except Exception:
                logger.warning("driver_registry harvest failed (non-fatal)")
            try:
                _harvest_vehicle_registry(sd)
            except Exception:
                logger.warning("vehicle_registry harvest failed (non-fatal)")
            _audit_add("تعديل", "مركز تنبيهات الوثائق", None, store_key + ": " + (value or "تفريغ") + " — " + (row.get("name") or row.get("plate") or ""))
            return jsonify({"success": True, "value": value})

        if src == "employee":
            if target_field == "doctype":
                return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            try:
                idx = int(body.get("idx"))
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            store_idx = None  # 'name' is resolved below, once the current row is known
            if target_field == "date":
                try:
                    store_idx = int(body.get("dateField"))
                except (TypeError, ValueError):
                    return jsonify({"success": False, "error": "طلب غير صالح."}), 400
                if store_idx not in (10, 11):  # matches employees.html COLS[10]/[11] = إقامة/جواز expiry
                    return jsonify({"success": False, "error": "طلب غير صالح."}), 400
            elif target_field == "plate":
                store_idx = 9  # COLS[9] = رقم اللوحة
            with _ALERTS_CENTER_LOCK:
                rows = blob_get("employees")
                if not isinstance(rows, list) or idx < 0 or idx >= len(rows) or not isinstance(rows[idx], list):
                    return jsonify({"success": False, "error": "لم يعد هذا الصف موجوداً — أعد تحميل الصفحة."}), 409
                row = rows[idx]
                actual_iqama = str(row[1] if len(row) > 1 else "").strip()
                if not actual_iqama or actual_iqama != str(body.get("iqama") or "").strip():
                    return jsonify({"success": False, "error": "تغيّرت بيانات الصف — أعد تحميل الصفحة."}), 409
                if store_idx is None:  # target_field == "name": whichever of COLS[2]/[3] is
                    # actually populated on the CURRENT row — decided server-side, not client-supplied.
                    store_idx = 2 if (row[2] if len(row) > 2 else "") or not (row[3] if len(row) > 3 else "") else 3
                if target_field in ("name", "plate") and not value:
                    return jsonify({"success": False, "error": "هذا الحقل لا يمكن تفريغه."}), 400
                if len(row) <= store_idx:
                    row.extend([""] * (store_idx + 1 - len(row)))
                row[store_idx] = value
                blob_set("employees", rows)
            _audit_add("تعديل", "مركز تنبيهات الوثائق", None, target_field + ": " + (value or "تفريغ") + " — " + (row[2] if len(row) > 2 and row[2] else (row[3] if len(row) > 3 else "")))
            return jsonify({"success": True, "value": value})

        if src == "document":
            store_key = {"date": "expiry", "name": "entity_ref", "plate": "entity_ref", "doctype": "doc_type"}[target_field]
            doc_id = str(body.get("id") or "")
            with _ALERTS_CENTER_LOCK:
                rows = _documents_rows()
                found = next((r for r in rows if r.get("id") == doc_id), None)
                if not found:
                    return jsonify({"success": False, "error": "الوثيقة لم تعد موجودة — أعد تحميل الصفحة."}), 409
                if store_key != "expiry" and not value:
                    return jsonify({"success": False, "error": "هذا الحقل لا يمكن تفريغه."}), 400
                found[store_key] = value
                blob_set("documents_data", {"rows": rows})
            _audit_add("تعديل", "مركز تنبيهات الوثائق", None, store_key + ": " + (value or "تفريغ") + " — " + (found.get("entity_ref") or ""))
            return jsonify({"success": True, "value": value})

        return jsonify({"success": False, "error": "مصدر غير معروف."}), 400
    except Exception:
        logger.exception("alerts_center_update error")
        return jsonify({"success": False, "error": "تعذّر حفظ التعديل."}), 500


# ── Workstation-only tab state (oils / purchase / workshop) ───────────────────
# These tabs have NO server storage on the main site (hardcoded rows + localStorage).
# For the /importantworkstation sandbox they persist a single JSON object on the server
# (id=2 via blob_get/blob_set). On the main site (id=1) these rows stay empty/unused —
# the main pages never POST here, so the main site keeps its original behavior.
@app.route("/api/oils_data", methods=["GET", "POST"])
@login_required
def oils_data():
    if request.method == "POST":
        try:
            blob_set("oils_data", request.json or {})
            return jsonify({"success": True})
        except Exception:
            logger.exception("oils_data POST error")
            return jsonify({"success": False}), 500
    try:
        return jsonify({"success": True, "data": blob_get("oils_data")})
    except Exception:
        logger.exception("oils_data GET error")
        return jsonify({"success": False, "data": None})


# ── Fuel/diesel supply tracking (تموين المحروقات) — fully server-synced on every
# branch and the workstation sandbox alike (unlike oils/purchase/workshop above,
# this tab has no legacy hardcoded-row/localStorage-only history to preserve).
@app.route("/api/fuel_data", methods=["GET", "POST"])
@login_required
def fuel_data():
    if request.method == "POST":
        try:
            entries = (request.json or {}).get("entries", [])
            blob_set("fuel_data", entries)
            _audit_add("تحديث", "تموين المحروقات", len(entries) if isinstance(entries, list) else None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("fuel_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ بيانات التموين."}), 500
    try:
        data = blob_get("fuel_data")
        return jsonify({"success": True, "entries": data if isinstance(data, list) else []})
    except Exception:
        logger.exception("fuel_data GET error")
        return jsonify({"success": False, "entries": []})


# ── Vehicle custody transfer (نقل عهدة) — the actual row reassignment already
# happens client-side via the normal POST /api/schedule_data save; this endpoint only
# records a distinctly-labeled audit entry so the handover shows up clearly in
# سجل التدقيق instead of blending into a generic "تعديل" entry.
@app.route("/api/vehicle_custody_transfer", methods=["POST"])
@login_required
def vehicle_custody_transfer():
    try:
        body = request.json or {}
        plate = str(body.get("plate") or "").strip()
        old_driver = str(body.get("oldDriver") or "").strip()
        new_driver = str(body.get("newDriver") or "").strip()
        note = str(body.get("note") or "").strip()
        if not new_driver:
            return jsonify({"success": False, "error": "اسم السائق المستلم مطلوب."}), 400
        detail = f"{plate or 'بدون لوحة'}: {old_driver or 'بدون سائق'} ← {new_driver}"
        if note:
            detail += f" — {note}"
        _audit_add("نقل عهدة", "الجدول الأسبوعي", None, detail)
        return jsonify({"success": True})
    except Exception:
        logger.exception("vehicle_custody_transfer error")
        return jsonify({"success": False, "error": "تعذّر توثيق نقل العهدة."}), 500


@app.route("/api/purchase_data", methods=["GET", "POST"])
@login_required
def purchase_data():
    if request.method == "POST":
        try:
            body = request.json or {}
            blob_set("purchase_data", body)
            _n = sum(len(v) for v in body.values() if isinstance(v, list)) if isinstance(body, dict) else None
            _audit_add("تحديث", "طلبات الشراء", _n or None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("purchase_data POST error")
            return jsonify({"success": False}), 500
    try:
        return jsonify({"success": True, "data": blob_get("purchase_data")})
    except Exception:
        logger.exception("purchase_data GET error")
        return jsonify({"success": False, "data": None})


@app.route("/api/workshop_data", methods=["GET", "POST"])
@login_required
def workshop_data():
    if request.method == "POST":
        try:
            blob_set("workshop_data", request.json or {})
            return jsonify({"success": True})
        except Exception:
            logger.exception("workshop_data POST error")
            return jsonify({"success": False}), 500
    try:
        return jsonify({"success": True, "data": blob_get("workshop_data")})
    except Exception:
        logger.exception("workshop_data GET error")
        return jsonify({"success": False, "data": None})


# ── Vehicle handover/receipt form: save the signed record + email it ──────────
HANDOVER_KEEP = 200            # cap stored submissions per mode to bound DB growth
_DATAURI_RE = re.compile(r"^data:(image/[A-Za-z0-9.+-]+);base64,(.+)$", re.S)


def _decode_data_uri(uri):
    """Return (mime, bytes) from a data: URI, or (None, None) if it isn't one."""
    if not isinstance(uri, str):
        return None, None
    m = _DATAURI_RE.match(uri.strip())
    if not m:
        return None, None
    try:
        return m.group(1), base64.b64decode(m.group(2))
    except Exception:
        return None, None


def _h_esc(s):
    """Minimal HTML escaping for values embedded in the handover email."""
    return (str("" if s is None else s)
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))


def _build_handover_email_html(rec):
    """Branded HTML summary of one handover form; signatures referenced via cid:.
    Built by concatenation (no %/format) and every user value is HTML-escaped."""
    g = rec.get
    full_from = (g("from_first", "") + " " + g("from_last", "")).strip()
    full_to = (g("to_first", "") + " " + g("to_last", "")).strip()

    def row(lbl, val):
        v = _h_esc(val) if (val not in (None, "")) else "—"
        return ("<tr><td style='padding:7px 12px;color:#6b7280;font-weight:700;white-space:nowrap;"
                "border-bottom:1px solid #eee'>" + _h_esc(lbl) + "</td>"
                "<td style='padding:7px 12px;color:#111;border-bottom:1px solid #eee'>" + v + "</td></tr>")

    rows = (
        row("رقم اللوحة", g("plate", "")) +
        row("تاريخ التسليم", g("date", "")) +
        row("المسلِّم", full_from) +
        row("المستلِم", full_to) +
        row("العداد (كم)", g("odo", "")) +
        row("الوقود", g("fuel", "")) +
        row("مسؤول الحركة", g("officer", "")) +
        row("اسم المستلم", g("recipient", ""))
    )
    damage = _h_esc(g("damage", "") or "لا توجد أضرار مذكورة.")
    return (
        '<div style="font-family:Tahoma,Arial,sans-serif;direction:rtl;max-width:640px;margin:0 auto;'
        'border:1px solid #e5e7eb;border-radius:14px;overflow:hidden">'
        '<div style="background:#1a1a1a;color:#fff;padding:18px 20px;text-align:center">'
        '<div style="font-size:1.15rem;font-weight:800;color:#C5A059">نموذج تسليم واستلام مركبة</div>'
        '<div style="font-size:.8rem;opacity:.85;margin-top:4px">شركة بن زومة للتجارة الدولية والإنماء — فرع الدمام</div>'
        '</div>'
        '<table style="width:100%;border-collapse:collapse;font-size:.9rem">' + rows + '</table>'
        '<div style="padding:16px 20px">'
        '<div style="font-weight:800;color:#111;margin-bottom:6px">تفاصيل الأضرار</div>'
        '<div style="color:#374151;font-size:.88rem;white-space:pre-wrap">' + damage + '</div>'
        '</div>'
        '<div style="display:flex;gap:16px;flex-wrap:wrap;padding:0 20px 18px">'
        '<div style="flex:1;min-width:240px">'
        '<div style="font-weight:700;color:#6b7280;font-size:.82rem;margin-bottom:4px">توقيع مسؤول الحركة (' + _h_esc(g("officer", "")) + ')</div>'
        '<img src="cid:sig_officer" style="max-width:100%;border:1px solid #ddd;border-radius:8px;background:#fff">'
        '</div>'
        '<div style="flex:1;min-width:240px">'
        '<div style="font-weight:700;color:#6b7280;font-size:.82rem;margin-bottom:4px">توقيع المستلم (' + _h_esc(g("recipient", "")) + ')</div>'
        '<img src="cid:sig_recipient" style="max-width:100%;border:1px solid #ddd;border-radius:8px;background:#fff">'
        '</div>'
        '</div>'
        '<div style="background:#faf7f0;color:#8a7a55;text-align:center;font-size:.72rem;padding:10px">'
        'أُرسل تلقائياً من نظام إدارة الأسطول · ' + _h_esc(rec.get("ts", "")) +
        '</div>'
        '</div>'
    )


def _send_handover_email(rec, recipients):
    """Email one handover record (signatures inline via cid, photos attached)."""
    if not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
        return {"sent": False, "reason": "mail_not_configured"}
    recipients = [r for r in (recipients or []) if r]
    if not recipients:
        return {"sent": False, "reason": "no_recipients"}
    try:
        plate = rec.get("plate", "") or "—"
        msg = Message(
            subject="🚗 نموذج تسليم/استلام مركبة — لوحة %s (فرع الدمام)" % plate,
            recipients=recipients,
            html=_build_handover_email_html(rec),
            sender=app.config.get("MAIL_DEFAULT_SENDER") or app.config.get("MAIL_USERNAME"),
        )
        # signatures inline (cid)
        for key, cid in (("sig_officer", "sig_officer"), ("sig_recipient", "sig_recipient")):
            mime, data = _decode_data_uri(rec.get(key))
            if data:
                msg.attach("%s.png" % cid, mime or "image/png", data,
                           disposition="inline", headers=[("Content-ID", "<%s>" % cid)])
        # uploaded photos as regular attachments
        for i, img in enumerate(rec.get("images", []) or []):
            mime, data = _decode_data_uri(img.get("data") if isinstance(img, dict) else img)
            if data:
                ext = (mime or "image/jpeg").split("/")[-1]
                name = (isinstance(img, dict) and img.get("name")) or ("photo_%d.%s" % (i + 1, ext))
                msg.attach(name, mime or "image/jpeg", data)
        mail.send(msg)
        return {"sent": True, "recipients": recipients}
    except Exception as e:
        logger.exception("handover email failed")
        return {"sent": False, "reason": "send_error", "error": str(e)}


@app.route("/api/handover/submit", methods=["POST"])
@login_required
def handover_submit():
    """Save a signed handover form (always) and email it (if SMTP is configured)."""
    body = request.get_json(silent=True) or {}
    plate = (body.get("plate") or "").strip()
    if not plate:
        return jsonify({"success": False, "reason": "missing_plate"}), 400
    if not _decode_data_uri(body.get("sig_officer"))[1] or not _decode_data_uri(body.get("sig_recipient"))[1]:
        return jsonify({"success": False, "reason": "missing_signature"}), 400

    rec = {
        "id": int(datetime.now().timestamp() * 1000),
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "plate": plate,
        "date": (body.get("date") or "").strip(),
        "from_first": (body.get("from_first") or "").strip(),
        "from_last": (body.get("from_last") or "").strip(),
        "to_first": (body.get("to_first") or "").strip(),
        "to_last": (body.get("to_last") or "").strip(),
        "odo": (str(body.get("odo")) if body.get("odo") not in (None, "") else "").strip(),
        "fuel": (body.get("fuel") or "").strip(),
        "damage": (body.get("damage") or "").strip(),
        "officer": (body.get("officer") or "").strip(),
        "recipient": (body.get("recipient") or "").strip(),
        "sig_officer": body.get("sig_officer") or "",
        "sig_recipient": body.get("sig_recipient") or "",
        "images": [im for im in (body.get("images") or []) if isinstance(im, dict) and im.get("data")][:8],
    }

    # ── Save (must succeed independently of email) ──
    try:
        log = blob_get("handover_data")
        if not isinstance(log, list):
            log = []
        log.append(rec)
        if len(log) > HANDOVER_KEEP:
            log = log[-HANDOVER_KEEP:]
        blob_set("handover_data", log)
    except Exception:
        logger.exception("handover save error")
        return jsonify({"success": False, "reason": "save_error"}), 500
    _audit_add("تسليم/استلام مركبة", plate, None,
               "من: %s — إلى: %s" % (rec["from_first"], rec["to_first"]))

    # ── Email (best-effort) ──
    recips = body.get("email_to")
    if isinstance(recips, str):
        recips = [e.strip() for e in re.split(r"[,;\s]+", recips) if e.strip()]
    if not recips:
        recips = ALERT_RECIPIENTS or ([app.config.get("MAIL_USERNAME")] if app.config.get("MAIL_USERNAME") else [])
    mail_result = _send_handover_email(rec, recips)

    return jsonify({"success": True, "saved": True, "id": rec["id"], "mail": mail_result})


@app.route("/api/handover/list", methods=["GET"])
@login_required
def handover_list():
    """Recent saved handover forms (lightweight: text fields only, no images/signatures)."""
    try:
        log = blob_get("handover_data")
        if not isinstance(log, list):
            log = []
        out = [{
            "id": r.get("id"), "ts": r.get("ts"), "plate": r.get("plate"),
            "date": r.get("date"),
            "from": (r.get("from_first", "") + " " + r.get("from_last", "")).strip(),
            "to": (r.get("to_first", "") + " " + r.get("to_last", "")).strip(),
            "officer": r.get("officer", ""), "recipient": r.get("recipient", ""),
            "photos": len(r.get("images", []) or []),
        } for r in reversed(log[-50:])]
        return jsonify({"success": True, "records": out,
                        "mail_configured": bool(app.config.get("MAIL_USERNAME") and app.config.get("MAIL_PASSWORD"))})
    except Exception:
        logger.exception("handover list error")
        return jsonify({"success": False, "records": []})


@app.route("/api/branch", methods=["GET", "POST"])
@login_required
def api_branch():
    """GET: active branch + list + role flags. POST {id}: switch the session branch (swaps
    every blob store via _row_id). Branch-locked accounts may NOT switch."""
    can_switch = not session.get("is_branch_user")
    if request.method == "POST":
        if not can_switch:
            return jsonify({"success": False, "reason": "locked_branch"}), 403
        body = request.get_json(silent=True) or {}
        try:
            bid = int(body.get("id"))
        except (TypeError, ValueError):
            return jsonify({"success": False, "reason": "bad_id"}), 400
        if bid not in BRANCH_IDS:
            return jsonify({"success": False, "reason": "unknown_branch"}), 400
        session["branch_id"] = bid
        session.permanent = True
        _audit_add("تبديل الفرع", BRANCH_NAME.get(bid, str(bid)))
        return jsonify({"success": True, "id": bid, "name": BRANCH_NAME[bid]})
    return jsonify({"success": True, "id": current_branch_id(),
                    "name": current_branch_name(), "branches": BRANCHES,
                    "can_switch": can_switch, "is_admin": bool(session.get("is_admin"))})


def _audit_get_at(rid):
    """Read the audit list for a SPECIFIC mode/branch id (used by the HQ overview)."""
    key = f"audit_log_branch_{rid}"
    from models.schema import AppSetting
    setting = AppSetting.query.get(key)
    if not setting: return []
    try:
        d = json.loads(setting.value)
        return d if isinstance(d, list) else []
    except Exception:
        return []


@app.route("/api/branch_accounts", methods=["GET", "POST", "DELETE"])
@login_required
def api_branch_accounts():
    """Manage per-branch login accounts — only from the unlocked Settings tab."""
    if not session.get("settings_unlocked"):
        return jsonify({"error": "locked"}), 403
    if request.method == "GET":
        accts = {a.get("branch_id"): a for a in get_branch_accounts()}
        return jsonify({"success": True, "rows": [
            {"branch_id": b["id"], "branch_name": b["name"],
             "username": (accts.get(b["id"], {}).get("username") or ""),
             "has_code": bool(accts.get(b["id"], {}).get("code_hash"))}
            for b in BRANCHES
        ]})
    body = request.get_json(silent=True) or {}
    try:
        bid = int(body.get("branch_id"))
    except (TypeError, ValueError):
        return jsonify({"success": False, "reason": "bad_branch"}), 400
    if bid not in BRANCH_IDS:
        return jsonify({"success": False, "reason": "unknown_branch"}), 400
    accounts = get_branch_accounts()
    if request.method == "DELETE":
        accounts = [a for a in accounts if a.get("branch_id") != bid]
        save_branch_accounts(accounts)
        _audit_add("حذف حساب فرع", BRANCH_NAME.get(bid, str(bid)))
        return jsonify({"success": True})
    # POST: create/replace this branch's account
    username = (body.get("username") or "").strip()
    code = body.get("code") or ""
    if not re.match(r"^[A-Za-z0-9_.@-]{2,40}$", username):
        return jsonify({"success": False, "reason": "اسم المستخدم: حروف/أرقام إنجليزية و . _ @ - فقط"}), 400
    if len(code) < 4:
        return jsonify({"success": False, "reason": "الرمز قصير جداً (4 أحرف على الأقل)"}), 400
    if username == os.environ.get("ADMIN_USERNAME", "admin"):
        return jsonify({"success": False, "reason": "اسم المستخدم محجوز للمدير"}), 400
    if any(a.get("username") == username and a.get("branch_id") != bid for a in accounts):
        return jsonify({"success": False, "reason": "اسم المستخدم مستخدم في فرع آخر"}), 400
    if any(u.get("username") == username for u in get_users()):
        return jsonify({"success": False, "reason": "اسم المستخدم موجود ضمن الحسابات المشتركة"}), 400
    accounts = [a for a in accounts if a.get("branch_id") != bid]
    accounts.append({"branch_id": bid, "username": username,
                     "code_hash": generate_password_hash(code),
                     "created": datetime.now().strftime("%Y-%m-%d %H:%M")})
    save_branch_accounts(accounts)
    _audit_add("ضبط حساب فرع", "%s (%s)" % (BRANCH_NAME.get(bid, str(bid)), username))
    return jsonify({"success": True})


@app.route("/overview")
@login_required
def overview():
    """HQ-only page aggregating every branch's recent activity."""
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("overview.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/overview", methods=["GET"])
@login_required
def api_overview():
    """Cross-branch activity feed + per-branch status. HQ/admin only."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    today = datetime.now().strftime("%Y-%m-%d")
    acct_by_branch = {a.get("branch_id"): a.get("username") for a in get_branch_accounts()}
    summary, feed = [], []
    alert_totals = {"expired": 0, "d30": 0, "d90": 0}
    for b in BRANCHES:
        entries = _audit_get_at(b["id"])
        last_ts = max((e.get("ts", "") for e in entries), default="")
        today_n = sum(1 for e in entries if str(e.get("ts", "")).startswith(today))
        ac = {"expired": 0, "d30": 0, "d90": 0}
        for a in _collect_expiry_alerts(rid=b["id"]):
            if a["key"] in ac:
                ac[a["key"]] += 1
        for k in alert_totals:
            alert_totals[k] += ac[k]
        summary.append({"id": b["id"], "name": b["name"],
                        "username": acct_by_branch.get(b["id"], ""),
                        "last_ts": last_ts, "today": today_n, "total": len(entries),
                        "alerts": ac, "alerts_total": sum(ac.values())})
        for e in entries:
            feed.append({"branch_id": b["id"], "branch": b["name"],
                         "ts": e.get("ts", ""), "user": e.get("user", ""),
                         "action": e.get("action", ""), "target": e.get("target", ""),
                         "detail": e.get("detail", ""), "hits": e.get("hits", 1)})
    feed.sort(key=lambda e: e.get("ts", ""), reverse=True)
    return jsonify({"success": True, "branches": summary, "feed": feed[:150],
                    "alert_totals": alert_totals, "alerts_grand": sum(alert_totals.values())})


# أيقونة الإشعار حسب كلمة مفتاحية في هدف التغيير (target)
def _notify_icon(target):
    t = target or ""
    if "موظف" in t:
        return "👥"
    if "الجدول" in t:
        return "📋"
    if "شراء" in t:
        return "🛒"
    if "غسيل" in t:
        return "🚿"
    if "ورشة" in t:
        return "🔧"
    if "زيوت" in t or "فلاتر" in t:
        return "🛢️"
    if "توثيق" in t:
        return "📁"
    if "حوادث" in t or "مخالف" in t:
        return "🚨"
    if "GPS" in t or "تتبع" in t:
        return "🛰️"
    if "تسليم" in t or "مركبة" in t or "أبشر" in t or "سائق" in t or "تفويض" in t:
        return "🚗"
    if "فرع" in t:
        return "🏢"
    return "📝"


from pywebpush import webpush, WebPushException
import json

VAPID_PRIVATE_KEY = os.path.join(os.path.dirname(__file__), "vapid_private.pem")
VAPID_PUBLIC_KEY = "BAzGwlPjwO9421qLp5AckUJgBJ4P2EQTn97mxgWL9K6GkNRy8Ft3ky-0ePFSc2KHMEdW8bDqm5fh0dC1_rVq8i8"
VAPID_CLAIMS = {"sub": "mailto:admin@binzoma.com"}

@app.route("/api/push/vapid_public_key", methods=["GET"])
@login_required
def push_vapid_public_key():
    return jsonify({"public_key": VAPID_PUBLIC_KEY})

@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def push_subscribe():
    sub_info = request.json
    if not sub_info or "endpoint" not in sub_info:
        return jsonify({"error": "Invalid subscription"}), 400
    
    endpoint = sub_info["endpoint"]
    keys = sub_info.get("keys", {})
    p256dh = keys.get("p256dh", "")
    auth = keys.get("auth", "")
    role = session.get("google_user", {}).get("role", "admin")
    
    try:
        subs = _global_blob_get("push_subscriptions") or []
        # Update if exists, else append
        found = False
        for s in subs:
            if s.get("endpoint") == endpoint:
                s["p256dh"] = p256dh
                s["auth"] = auth
                s["user_role"] = role
                found = True
                break
        if not found:
            subs.append({"endpoint": endpoint, "p256dh": p256dh, "auth": auth, "user_role": role})
            
        _global_blob_set("push_subscriptions", subs)
    except Exception as e:
        logger.error(f"Push Subscribe error: {e}")
        return jsonify({"error": str(e)}), 500
        
    return jsonify({"success": True})

def send_push_notification(title, body):
    try:
        if not os.path.exists(VAPID_PRIVATE_KEY):
            return
            
        subs = _global_blob_get("push_subscriptions") or []
        payload = json.dumps({"title": title, "body": body})
        
        for sub in subs:
            try:
                sub_info = {
                    "endpoint": sub["endpoint"],
                    "keys": {
                        "p256dh": sub["p256dh"],
                        "auth": sub["auth"]
                    }
                }
                webpush(
                    subscription_info=sub_info,
                    data=payload,
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims=VAPID_CLAIMS
                )
            except WebPushException as ex:
                logger.error(f"WebPush error: {repr(ex)}")
                # If gone, delete from DB (HTTP 410)
                if ex.response and ex.response.status_code in [404, 410]:
                    try:
                        conn2 = sqlite3.connect(DB_PATH)
                        conn2.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (sub[0],))
                        conn2.commit()
                        conn2.close()
                    except: pass
    except Exception as e:
        logger.error(f"send_push_notification error: {e}")

@app.route("/api/notifications", methods=["GET"])
@login_required
def api_notifications():
    """HQ-only live feed: EVERY recent change in ANY branch (from the audit trail) +
    document-expiry alerts. Each item carries a stable key so the client shows only new ones."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    items = []
    cutoff = (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")
    for b in BRANCHES:
        for e in _audit_get_at(b["id"]):
            ts = e.get("ts", "")
            target = e.get("target", "")
            if not target or ts < cutoff:
                continue
            items.append({
                "key": "ev|%s|%s|%s|%s" % (b["id"], ts, target, e.get("action", "")),
                "kind": "change", "icon": _notify_icon(target),
                "title": "%s — %s" % (e.get("action", "تحديث"), target),
                "branch": b["name"], "user": e.get("user", "") or "النظام", "ts": ts,
            })
        ac = {"expired": 0, "d30": 0, "d90": 0}
        for a in _collect_expiry_alerts(rid=b["id"]):
            if a["key"] in ac:
                ac[a["key"]] += 1
        if sum(ac.values()):
            items.append({
                "key": "docs|%s|%s|%s|%s" % (b["id"], ac["expired"], ac["d30"], ac["d90"]),
                "kind": "docs", "icon": "📄",
                "title": "وثائق قريبة الانتهاء — منتهية %d · ≤30 يوم %d · ≤90 يوم %d" % (ac["expired"], ac["d30"], ac["d90"]),
                "branch": b["name"], "user": "", "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
    items.sort(key=lambda x: x.get("ts", ""), reverse=True)
    return jsonify({"success": True, "items": items[:80]})


def _blob_get_at(table, rid):
    """Read a tab blob for a SPECIFIC branch id (used by the all-branches view)."""
    table = _safe_tbl(table)
    key = f"{table}_branch_{rid}"
    setting = AppSetting.query.get(key)
    return _loads_blob(setting.value) if setting else None


def _blob_count(data):
    """Generic record count across varied tab shapes: list → len; dict → sum of its list
    values (covers {rows:[…]} and {main:[],spare:[],…}); else rows/keys."""
    if data is None:
        return 0
    if isinstance(data, list):
        return len(data)
    if isinstance(data, dict):
        lists = [v for v in data.values() if isinstance(v, list)]
        if lists:
            return sum(len(v) for v in lists)
        if isinstance(data.get("rows"), list):
            return len(data["rows"])
        return len(data)
    return 0


BRANCH_DATA_CATEGORIES = [
    {"key": "drivers", "label": "السائقون", "table": None},   # special: الدمام = SQL drivers
    {"key": "employees", "label": "الموظفون", "table": "employees"},
    {"key": "schedule_data", "label": "الجدول الأسبوعي", "table": "schedule_data"},
    {"key": "washing_schedule", "label": "الغسيل", "table": "washing_schedule"},
    {"key": "workshop_data", "label": "الورشة", "table": "workshop_data"},
    {"key": "oils_data", "label": "الزيوت والفلاتر", "table": "oils_data"},
    {"key": "purchase_data", "label": "طلبات الشراء", "table": "purchase_data"},
    {"key": "incidents_data", "label": "الحوادث", "table": "incidents_data"},
    {"key": "gps_devices_data", "label": "أجهزة التتبع", "table": "gps_devices_data"},
    {"key": "records_data", "label": "التوثيق", "table": "records_data"},
    {"key": "handover_data", "label": "تسليم/استلام", "table": "handover_data"},
]


def _branch_driver_count(bid):
    from models.schema import Driver
    return Driver.query.filter_by(branch_id=bid).count()


# ════════════════════════════════════════════════════════════════════════════════════
# SMART INSIGHTS (التحليلات الذكية) — honest, rule-based analytics computed ONLY from data
# already in the system for the active branch. No machine learning, no fabricated numbers:
# every metric below traces directly to a source tab, so it respects the frozen-data rule.
# ════════════════════════════════════════════════════════════════════════════════════
def _rows_as_dicts(data):
    """Normalize a tab blob to a list of dict rows, tolerating list / {rows:[…]} / {sec:[…]} shapes."""
    if isinstance(data, list):
        src = data
    elif isinstance(data, dict):
        src = data["rows"] if isinstance(data.get("rows"), list) else \
            [x for v in data.values() if isinstance(v, list) for x in v]
    else:
        src = []
    return [r for r in src if isinstance(r, dict)]


def _compute_insights(rid=None):
    """Return the insight cards for the active (or given) branch. Every value is derived from a
    real source tab; nothing is invented."""
    from models.schema import Driver
    def bg(table):
        return _blob_get_at(table, rid) if rid is not None else blob_get(table)

    bid = rid if rid is not None else _row_id()   # workstation-aware: matches what blob_get/_audit read

    # 1) document-expiry risk (registration / license / op-card / driver-card / iqama / passport)
    alerts = _collect_expiry_alerts(rid=rid) if rid is not None else _collect_expiry_alerts()
    doc = {"expired": 0, "d30": 0, "d90": 0}
    for a in alerts:
        if a["key"] in doc:
            doc[a["key"]] += 1
    doc_top = [{"name": a["name"], "plate": a["plate"], "doc": a["doc"],
                "date": a["date"], "days": a["days"], "key": a["key"]} for a in alerts[:8]]

    # 2) fleet composition (drivers + distinct vehicles)
    from models.schema import Driver
    db_drivers = Driver.query.filter_by(branch_id=bid).all()
    drivers = []
    for d in db_drivers:
        active = next((c for c in d.custodies if c.status == "active"), None)
        drivers.append({"name": d.name, "plate": active.vehicle.plate_number if active and active.vehicle else ""})
    with_vehicle = sum(1 for d in drivers if str(d.get("plate") or "").strip())
    plates = {str(d.get("plate")).strip() for d in drivers if str(d.get("plate") or "").strip()}
    fleet = {"drivers": len(drivers), "with_vehicle": with_vehicle,
             "without_vehicle": max(0, len(drivers) - with_vehicle), "vehicles": len(plates)}

    # 3) incidents (open cases + repeat-incident vehicles + breakdown)
    inc_rows = _rows_as_dicts(bg("incidents_data"))
    _closed = {"مغلق", "مغلقة", "منتهي", "منتهية", "مكتمل", "مكتملة", "تم", "closed"}
    open_inc, by_plate, sev = 0, {}, {}
    for r in inc_rows:
        if str(r.get("status") or "").strip().lower() not in _closed and str(r.get("status") or "").strip():
            open_inc += 1
        pl = str(r.get("plate") or "").strip()
        if pl:
            by_plate[pl] = by_plate.get(pl, 0) + 1
        s = str(r.get("severity") or r.get("type") or "").strip()
        if s:
            sev[s] = sev.get(s, 0) + 1
    repeat = sorted(((p, n) for p, n in by_plate.items() if n >= 2), key=lambda x: -x[1])[:6]
    incidents = {"total": len(inc_rows), "open": open_inc,
                 "repeat": [{"plate": p, "count": n} for p, n in repeat], "severity": sev}

    # 4) operations volume (honest record counts for tabs we don't model field-by-field)
    volume = {k: _blob_count(bg(t)) for k, t in (
        ("workshop", "workshop_data"), ("oils", "oils_data"), ("purchase", "purchase_data"),
        ("washing", "washing_schedule"), ("records", "records_data"),
        ("gps_devices", "gps_devices_data"))}
    
    # hr_employees is now a real SQL table, not a JSON blob, so we must count it from DB directly.
    with db_connection() as cdb:
        emp_count = cdb.execute("SELECT COUNT(*) FROM hr_employees").fetchone()[0]
        volume["employees"] = emp_count

    # 5) activity (from the audit trail)
    entries = _audit_get_at(bid)
    today = datetime.now().strftime("%Y-%m-%d")
    wk = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d %H:%M:%S")
    activity = {"today": sum(1 for e in entries if str(e.get("ts", "")).startswith(today)),
                "week": sum(1 for e in entries if str(e.get("ts", "")) >= wk),
                "total": len(entries), "last_ts": max((e.get("ts", "") for e in entries), default="")}

    # 6) transparent fleet-health score (0–100). Explainable rule: 100 minus weighted, visible penalties.
    penalty = doc["expired"] * 6 + doc["d30"] * 2 + doc["d90"] * 0.5 + open_inc * 3 + len(repeat) * 2
    score = int(max(0, min(100, round(100 - penalty))))
    grade, label = (("good", "ممتاز") if score >= 85 else ("ok", "جيد") if score >= 65
                    else ("warn", "يحتاج انتباه") if score >= 45 else ("bad", "حرِج"))

    branch_label = "محطة العمل" if (rid is None and is_workstation()) else BRANCH_NAME.get(bid, "")
    return {"branch_id": bid, "branch": branch_label,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "documents": doc, "documents_top": doc_top, "fleet": fleet,
            "incidents": incidents, "volume": volume, "activity": activity,
            "score": {"value": score, "grade": grade, "label": label, "penalty": round(penalty, 1)}}


@app.route("/insights")
@login_required
def insights_page():
    """التحليلات الذكية — مؤشرات حقيقية محسوبة من بيانات الفرع النشط."""
    return render_template("insights.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/insights", methods=["GET"])
@login_required
def api_insights():
    try:
        return jsonify({"success": True, "insights": _compute_insights()})
    except Exception as e:
        logger.exception("insights failed")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/platform")
@login_required
def platform_page():
    """مركز المنصة — يربط كل وحدات النظام (الأسطول/GPS/الصيانة/الحوادث/التحليلات) في صفحة واحدة."""
    return render_template("platform.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/branches")
@login_required
def branches_all():
    """HQ-only page showing ALL branches' data side by side."""
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("branches_all.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/branches_overview", methods=["GET"])
@login_required
def api_branches_overview():
    """Per-branch record counts across every data category. HQ/admin only."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    acct_by_branch = {a.get("branch_id"): a.get("username") for a in get_branch_accounts()}
    cats = [{"key": c["key"], "label": c["label"]} for c in BRANCH_DATA_CATEGORIES]
    branches, totals = [], {c["key"]: 0 for c in BRANCH_DATA_CATEGORIES}
    for b in BRANCHES:
        counts = {}
        for c in BRANCH_DATA_CATEGORIES:
            n = _branch_driver_count(b["id"]) if c["key"] == "drivers" else _blob_count(_blob_get_at(c["table"], b["id"]))
            counts[c["key"]] = n
            totals[c["key"]] += n
        entries = _audit_get_at(b["id"])
        branches.append({"id": b["id"], "name": b["name"],
                         "username": acct_by_branch.get(b["id"], ""),
                         "last_ts": max((e.get("ts", "") for e in entries), default=""),
                         "counts": counts, "total": sum(counts.values())})
    return jsonify({"success": True, "categories": cats, "branches": branches,
                    "totals": totals, "grand_total": sum(totals.values())})


# All id=2 stores used by the workstation sandbox. Used by the reset endpoint below.
WS_BLOB_TABLES = [
    "employees", "schedule_data", "washing_schedule", "records_data",
    "drivers_ws", "oils_data", "purchase_data", "workshop_data",
    "gps_devices_data", "handover_data", "deauthorized_data", "drivers_backup",
]


def _enrich_deauth(rows):
    """يكمل الرقم الوظيفي/الجوال الناقص لكل اسم بمطابقته (بالإقامة ثم الاسم) مع بيانات
    الموظفين وسجل السائقين للفرع النشط."""
    if not rows:
        return rows
    by_iq, by_name = {}, {}

    def _add(iq, nm, empid, phone):
        rec = {"empid": (str(empid).strip() if empid else ""), "phone": (str(phone).strip() if phone else "")}
        if iq and iq not in by_iq:
            by_iq[iq] = rec
        if nm and nm.strip() and nm.strip() not in by_name:
            by_name[nm.strip()] = rec

    emp = blob_get("employees")
    emp_rows = emp if isinstance(emp, list) else (emp.get("rows") if isinstance(emp, dict) else [])
    for r in (emp_rows or []):
        if isinstance(r, list) and len(r) > 2:
            _add(absher_sync.norm_id(r[1] if len(r) > 1 else ""),
                 str(r[2]).strip() if r[2] else "",
                 r[0] if r[0] else "",
                 r[7] if len(r) > 7 and r[7] else "")
    try:
        _, drv = _drivers_list_for_sync()
        for d in drv:
            _add(absher_sync.norm_id(d.get("iqama")), (d.get("name") or "").strip(), d.get("empid"), d.get("phone"))
    except Exception:
        pass

    for row in rows:
        if (row.get("empid") or "") and (row.get("phone") or ""):
            continue
        src = by_iq.get(absher_sync.norm_id(row.get("iqama"))) or by_name.get((row.get("name") or "").strip())
        if src:
            if not row.get("empid") and src.get("empid"):
                row["empid"] = src["empid"]
            if not row.get("phone") and src.get("phone"):
                row["phone"] = src["phone"]
    return rows


@app.route("/api/deauthorized", methods=["GET", "POST"])
@login_required
def api_deauthorized():
    """قائمة «تم إلغاء تفويضهم» (من مزامنة أبشر) — تظهر أسفل الجدول الأسبوعي.
    قابلة للعرض/التعديل/الحذف لأي مستخدم مسجّل، مقسومة حسب الفرع النشط."""
    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        rows = body.get("rows", body if isinstance(body, list) else [])
        if not isinstance(rows, list):
            rows = []
        blob_set("deauthorized_data", rows)
        _audit_add("تحديث", "قائمة إلغاء التفويض", len(rows))
        return jsonify({"success": True, "count": len(rows)})
    data = blob_get("deauthorized_data")
    rows = data if isinstance(data, list) else (data.get("rows", []) if isinstance(data, dict) else [])
    return jsonify({"success": True, "rows": _enrich_deauth(rows)})


# ── مستورِد أبشر داخل الموقع: ارفع ملفات أبشر → الخادم يحدّث سجل السائقين تلقائياً ──
def _drivers_list_for_sync():
    """(store, قائمة سائقي الفرع النشط) — SQL للدمام، blob لبقية الفروع/المحطة."""
    store = _driver_store()
    if store == "sql":
        with db_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM drivers ORDER BY id")
            return store, [dict(r) for r in c.fetchall()]
    return store, (blob_get(_driver_blob_table(store)) or [])


def _employee_iqamas():
    """أرقام إقامات الموظفين المعتمدين (تبويب الموظفين) للفرع النشط — عمود «الإقامة - البطاقة» (الفهرس 1)."""
    data = blob_get("employees")
    rows = data if isinstance(data, list) else (data.get("rows") if isinstance(data, dict) else [])
    out = set()
    for r in (rows or []):
        iq = None
        if isinstance(r, list) and len(r) > 1:
            iq = absher_sync.norm_id(r[1])
        elif isinstance(r, dict):
            iq = absher_sync.norm_id(r.get("iqama") or r.get("الإقامة - البطاقة") or r.get("الإقامة"))
        if iq:
            out.add(iq)
    return out


def _absher_parse_uploads(files):
    recs = []
    for f in files:
        if not f or not getattr(f, "filename", ""):
            continue
        recs += absher_sync.parse_file(io.BytesIO(f.read()))
    return recs


def _absher_apply(diff, store, remove_non_emp=False):
    # نسخة احتياطية كاملة للفرع الحالي قبل أي تعديل (للرجوع عند الحاجة)
    _, current = _drivers_list_for_sync()
    blob_set("drivers_backup", {"ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "rows": current})
    non_emp = diff.get("non_employees", []) if remove_non_emp else []
    non_emp_ids = set(x.get("id") for x in non_emp if x.get("id") is not None)
    non_emp_iqamas = set(x.get("iqama") for x in non_emp if x.get("iqama"))
    if store == "sql":
        with db_connection() as conn:
            c = conn.cursor()
            for u in diff["updates"]:
                sets, params = [], []
                for f, v in u["changes"].items():
                    sets.append("%s = ?" % f)
                    params.append(v)
                if not sets:
                    continue
                params.append(u["id"])
                c.execute("UPDATE drivers SET %s WHERE id = ?" % ", ".join(sets), params)
            for n in diff["new"]:
                c.execute("INSERT INTO drivers (name, empid, plate, car, iqama, phone, drivercard) VALUES (?,?,?,?,?,?,?)",
                          (n["name"], "", n["plate"], n["car"], n["iqama"], "", ""))
            for nid in non_emp_ids:                                  # حذف من ليسوا في بيانات الموظفين
                c.execute("DELETE FROM drivers WHERE id = ?", (nid,))
            conn.commit()
    else:
        tbl = _driver_blob_table(store)
        lst = blob_get(tbl) or []
        idx = {}
        for i, d in enumerate(lst):
            iq = absher_sync.norm_id(d.get("iqama"))
            if iq:
                idx[iq] = i
        for u in diff["updates"]:
            i = idx.get(u["iqama"])
            if i is not None:
                lst[i].update(u["changes"])
        nid = max([d.get("id", 0) for d in lst], default=0) + 1
        for n in diff["new"]:
            lst.append({"id": nid, "name": n["name"], "empid": "", "plate": n["plate"],
                        "car": n["car"], "iqama": n["iqama"], "phone": "", "drivercard": ""})
            nid += 1
        if non_emp_iqamas:
            lst = [d for d in lst if absher_sync.norm_id(d.get("iqama")) not in non_emp_iqamas]
        blob_set(tbl, lst)
    blob_set("deauthorized_data", diff["deauthorized"])
    _audit_add("مزامنة أبشر", "سجل السائقين", len(diff["new"]) + len(diff["updates"]),
               "إضافة %d · تحديث %d · حذف غير موظفين %d · إلغاء تفويض %d" % (
                   len(diff["new"]), len(diff["updates"]), len(non_emp), len(diff["deauthorized"])))


def _absher_summary(diff):
    return {
        "counts": {"new": len(diff["new"]), "updates": len(diff["updates"]),
                   "deauthorized": len(diff["deauthorized"]), "no_vehicle": len(diff["no_vehicle"]),
                   "unchanged": diff["unchanged"],
                   "non_employees": len(diff.get("non_employees", [])),
                   "ignored_non_employee": diff.get("ignored_non_employee", 0)},
        "deauthorized": diff["deauthorized"][:300],
        "non_employees": diff.get("non_employees", [])[:300],
        "new_sample": diff["new"][:60],
        "updates_sample": diff["updates"][:60],
        "used_employee_filter": diff.get("used_employee_filter", False),
    }


@app.route("/absher_import")
@login_required
def absher_import():
    """صفحة المستورِد — للمدير الرئيسي فقط."""
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("absher_import.html", google_user=session.get("google_user"), b64_en=load_logo())


def _absher_run(apply_it):
    if not session.get("is_admin"):
        return jsonify({"success": False, "reason": "forbidden"}), 403
    files = request.files.getlist("files")
    if not files or not any(getattr(f, "filename", "") for f in files):
        return jsonify({"success": False, "reason": "no_files"}), 400
    try:
        recs = _absher_parse_uploads(files)
    except Exception as e:
        logger.exception("absher parse error")
        return jsonify({"success": False, "reason": "parse_error", "error": str(e)}), 400
    if not recs:
        return jsonify({"success": False, "reason": "empty"}), 400
    by = absher_sync.index_by_iqama(recs)
    store, db = _drivers_list_for_sync()
    emp_iqamas = _employee_iqamas()
    diff = absher_sync.compute_diff(db, by, update_names=bool(request.form.get("update_names")),
                                    employee_iqamas=emp_iqamas)
    if apply_it:
        try:
            _absher_apply(diff, store, remove_non_emp=bool(request.form.get("remove_non_emp")))
        except Exception as e:
            logger.exception("absher apply error")
            return jsonify({"success": False, "reason": "apply_error", "error": str(e)}), 500
    out = _absher_summary(diff)
    out.update({"success": True, "applied": apply_it, "branch": current_branch_name(),
                "vehicles": len(recs), "authorized": len(by), "db": len(db), "employees": len(emp_iqamas)})
    return jsonify(out)


@app.route("/api/absher/preview", methods=["POST"])
@login_required
def absher_preview():
    return _absher_run(False)


@app.route("/api/absher/apply", methods=["POST"])
@login_required
def absher_apply():
    return _absher_run(True)


@app.route("/api/ws_reset", methods=["POST"])
@login_required
def ws_reset():
    """Wipe EVERY workstation (id=2) store so /importantworkstation starts truly empty.
    Workstation-only: the main site can never reach this (guard + path-based mirror)."""
    if not is_workstation():
        return jsonify({"success": False, "error": "workstation only"}), 404
    try:
        AppSetting.query.filter(AppSetting.key.like('%_branch_2')).delete(synchronize_session=False)
        db.session.commit()
        _ws_meta_set("ws_cleared", "1")  # user emptied it on purpose → don't auto-reseed
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"ws_reset error: {e}")
        return jsonify({"success": False}), 500


@app.route("/api/ws_seed", methods=["POST"])
@login_required
def ws_seed():
    """Fill EVERY workstation (id=2) store with the FAKE example data (demo).
    Workstation-only: the main site can never reach this."""
    if not is_workstation():
        return jsonify({"success": False, "error": "workstation only"}), 404
    try:
        _ws_write_examples()
        _ws_meta_set("ws_cleared", "0")  # examples are wanted again → allow auto-seed
        return jsonify({"success": True})
    except Exception:
        logger.exception("ws_seed error")
        return jsonify({"success": False}), 500


def _driver_store():
    """Where the current context's drivers live:
    'ws'   → workstation blob (drivers_ws, id=2)
    'sql'  → الدمام's real drivers table (id=1, FROZEN)
    'blob' → another branch's drivers blob (drivers_branch, partitioned by branch id)"""
    if is_workstation():
        return "ws"
    return "sql" if current_branch_id() == 1 else "blob"


def _driver_blob_table(store):
    return "drivers_ws" if store == "ws" else "drivers_branch"


@app.route("/api/legacy/drivers", methods=["GET"])
@login_required
def get_drivers():
    branch_id = current_branch_id()
    drivers = Driver.query.filter_by(branch_id=branch_id).order_by(Driver.id.desc()).all()
    
    result = []
    for d in drivers:
        # Reconstruct the flat dictionary expected by the frontend
        item = {
            "id": d.id,
            "name": d.name,
            "empid": d.employee_id,
            "iqama": d.iqama_number,
            "phone": d.phone,
            "job": d.job_title,
            "iqama_exp": d.iqama_expiry.strftime('%Y-%m-%d') if d.iqama_expiry else "",
            "license": d.license_expiry.strftime('%Y-%m-%d') if d.license_expiry else "",
            "status": d.status,
            # Fallbacks for empty vehicle
            "plate": "", "car": "", "model": "", "vserial": "", "inspect": "", "notes": ""
        }
        
        # Check if driver has an active vehicle custody
        active_custody = VehicleCustody.query.filter_by(driver_id=d.id, status="active").first()
        if active_custody and active_custody.vehicle:
            v = active_custody.vehicle
            item.update({
                "plate": v.plate_number,
                "car": v.v_type,
                "model": v.model,
                "vserial": v.serial_number,
                "inspect": v.inspection_expiry.strftime('%Y-%m-%d') if v.inspection_expiry else "",
                "notes": active_custody.notes or ""
            })
            
        result.append(item)
        
    return jsonify(result)


def normalize_plate(plate):
    if plate is None:
        return ""
    plate = str(plate)
    plate = re.sub(r"\s+", "", plate)
    arabic_digits = "٠١٢٣٤٥٦٧٨٩"
    english_digits = "0123456789"
    for a, e in zip(arabic_digits, english_digits):
        plate = plate.replace(a, e)
    digits = "".join(re.findall(r"\d+", plate))
    letters = "".join(re.findall(r"[^\d]+", plate))
    return digits + letters

def tafqeet(amount):
    try:
        amount = float(amount)
    except:
        return ""
    if amount == 0: return "صفر ريال فقط لا غير"
    
    units = ["", "واحد", "اثنان", "ثلاثة", "أربعة", "خمسة", "ستة", "سبعة", "ثمانية", "تسعة", "عشرة",
             "أحد عشر", "اثنا عشر", "ثلاثة عشر", "أربعة عشر", "خمسة عشر", "ستة عشر", "سبعة عشر", "ثمانية عشر", "تسعة عشر"]
    tens = ["", "عشرة", "عشرون", "ثلاثون", "أربعون", "خمسون", "ستون", "سبعون", "ثمانون", "تسعون"]
    hundreds = ["", "مائة", "مائتان", "ثلاثمائة", "أربعمائة", "خمسمائة", "ستمائة", "سبعمائة", "ثمانمائة", "تسعمائة"]
    
    def convert_999(n):
        if n == 0: return ""
        h = n // 100
        rem = n % 100
        words = []
        if h > 0:
            words.append(hundreds[h])
        if rem > 0:
            if rem < 20:
                words.append(units[rem])
            else:
                t = rem // 10
                u = rem % 10
                if u > 0:
                    words.append(units[u] + " و" + tens[t])
                else:
                    words.append(tens[t])
        return " و".join(words)

    int_part = int(amount)
    dec_part = int(round((amount - int_part) * 100))
    
    mils = int_part // 1000000
    thous = (int_part % 1000000) // 1000
    ones = int_part % 1000
    
    parts = []
    if mils > 0:
        if mils == 1: parts.append("مليون")
        elif mils == 2: parts.append("مليونان")
        else: parts.append(convert_999(mils) + " مليون")
    if thous > 0:
        if thous == 1: parts.append("ألف")
        elif thous == 2: parts.append("ألفان")
        else: parts.append(convert_999(thous) + " ألف")
    if ones > 0:
        parts.append(convert_999(ones))
        
    res = " و".join(parts) + " ريال"
    if dec_part > 0:
        res += " و" + convert_999(dec_part) + " هللة"
    return res + " فقط لا غير"


@app.route("/api/gps_sync", methods=["POST"])
@login_required
def api_gps_sync():
    if "source_file" not in request.files or "target_file" not in request.files:
        return jsonify({"success": False, "error": "الرجاء رفع الملفين المطلوبة"}), 400

    src_file = request.files["source_file"]
    tgt_file = request.files["target_file"]

    try:
        wb_src = openpyxl.load_workbook(src_file, data_only=True)
        ws_src = wb_src.active

        headers = {}
        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(row=4, column=c).value
            if val:
                headers[str(val).strip()] = c

        if "رقم اللوحة" not in headers:
            headers = {}
            for c in range(1, ws_src.max_column + 1):
                val = ws_src.cell(row=1, column=c).value
                if val:
                    headers[str(val).strip()] = c

        lookup = {}
        plate_src_col = headers.get("رقم اللوحة")
        if plate_src_col:
            # Data starts after header
            start_row = (
                5
                if "رقم اللوحة"
                in [
                    ws_src.cell(row=4, column=c).value
                    for c in range(1, ws_src.max_column + 1)
                ]
                else 2
            )
            for r in range(start_row, ws_src.max_row + 1):
                plate_val = ws_src.cell(row=r, column=plate_src_col).value
                norm = normalize_plate(plate_val)
                if norm:
                    row_data = {}
                    for col_name, c_idx in headers.items():
                        row_data[col_name] = ws_src.cell(row=r, column=c_idx).value
                    lookup[norm] = row_data

        wb = openpyxl.load_workbook(tgt_file)
        ws = wb.active

        plate_col = 9
        vin_col = 1
        sn_col = 2
        year_col = 3
        model_col = 4
        make_col = 5
        reg_col = 6
        branch_col = 7

        match_count = 0
        update_count = 0

        for r in range(6, ws.max_row + 1):
            plate_val = ws.cell(row=r, column=plate_col).value
            if not plate_val:
                continue
            norm = normalize_plate(plate_val)
            if norm in lookup:
                src = lookup[norm]
                match_count += 1
                updates = [
                    (vin_col, "رقم الهيكل"),
                    (sn_col, "الرقم التسلسلي"),
                    (year_col, "سنة الصنع"),
                    (model_col, "الطراز"),
                    (make_col, "الماركة"),
                    (reg_col, "نوع التسجيل"),
                    (branch_col, "الفرع"),
                ]
                for col_idx, src_col_name in updates:
                    if src_col_name in src:
                        new_val = src[src_col_name]
                        if (
                            new_val is not None
                            and str(new_val).strip() != ""
                            and str(new_val).lower() != "nan"
                        ):
                            ws.cell(row=r, column=col_idx).value = new_val
                            update_count += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")

        return jsonify(
            {
                "success": True,
                "matches": match_count,
                "updates": update_count,
                "file_b64": b64,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/drivers", methods=["POST"])
@login_required
def add_driver():
    data = request.json or {}
    fields = ['name', 'empid', 'plate', 'car', 'iqama', 'phone', 'drivercard',
              'job', 'empNotes', 'model', 'pallets', 'load', 'vserial', 
              'inspect', 'license', 'opcard', 'notes', 'fuel_card', 'medical_exp', 'contract_exp']
    
    vals = {f: data.get(f, "").strip() for f in fields}
    
    if not vals['name']:
        return jsonify({"error": "Name is required"}), 400
        
    branch_id = current_branch_id()
    
    def parse_date(dstr):
        if not dstr: return None
        try:
            return datetime.strptime(dstr, '%Y-%m-%d').date()
        except:
            return None

    try:
        driver = Driver(
            branch_id=branch_id,
            name=vals['name'],
            employee_id=vals['empid'] or f"EMP-{datetime.now().timestamp()}",
            iqama_number=vals['iqama'] or None,
            phone=vals['phone'],
            job_title=vals['job'],
            iqama_expiry=parse_date(vals.get('iqama_exp')), # Frontend uses iqama_exp sometimes but sends iqama
            license_expiry=parse_date(vals['license']),
            status="متاح"
        )
        db.session.add(driver)
        db.session.flush() # Get driver.id

        if vals['plate']:
            # See if vehicle exists in this branch
            vehicle = Vehicle.query.filter_by(plate_number=vals['plate']).first()
            if not vehicle:
                vehicle = Vehicle(
                    branch_id=branch_id,
                    plate_number=vals['plate'],
                    v_type=vals['car'],
                    model=vals['model'],
                    serial_number=vals['vserial'],
                    inspection_expiry=parse_date(vals['inspect'])
                )
                db.session.add(vehicle)
                db.session.flush()
                
            custody = VehicleCustody(
                driver_id=driver.id,
                vehicle_id=vehicle.id,
                received_date=datetime.now().date(),
                notes=vals['notes']
            )
            db.session.add(custody)

        db.session.commit()
        logger.info("Driver added via SQLAlchemy: %s (id=%s)", vals['name'], driver.id)
        return jsonify({"success": True, "id": driver.id, **vals})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error adding driver: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/drivers/<int:driver_id>", methods=["DELETE"])
@login_required
@role_required("admin", "branch_manager")
def delete_driver(driver_id):
    try:
        driver = Driver.query.get(driver_id)
        if driver:
            # Delete custodies first
            VehicleCustody.query.filter_by(driver_id=driver_id).delete()
            # Then delete driver
            db.session.delete(driver)
            db.session.commit()
            logger.info("Driver deleted via SQLAlchemy: id=%d", driver_id)
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting driver: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/drivers/bulk_delete", methods=["POST"])
@login_required
@role_required("admin", "branch_manager")
def bulk_delete_drivers():
    body = request.get_json(silent=True) or {}
    if not hmac.compare_digest(str(body.get("lock", "")), WORKSTATION_PASSWORD):
        return jsonify({"success": False, "reason": "locked", "error": "الرمز السري غير صحيح."}), 403
    ids = []
    for x in (body.get("ids") or [])[:2000]:
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            continue
    ids = list(dict.fromkeys(ids))
    if not ids:
        return jsonify({"success": False, "error": "لم تُحدَّد أي عناصر صالحة."}), 400

    try:
        # Delete custodies first
        VehicleCustody.query.filter(VehicleCustody.driver_id.in_(ids)).delete(synchronize_session=False)
        # Delete drivers
        deleted = Driver.query.filter(Driver.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        
        # Log to SQLAlchemy audit
        audit = AuditLog(
            user_id=getattr(g, 'user', None),
            branch_id=current_branch_id(),
            action="حذف جماعي",
            target_table="erp_drivers",
            target_id=str(len(ids))
        )
        db.session.add(audit)
        db.session.commit()
        
        logger.info("Bulk driver delete via SQLAlchemy: requested=%d deleted=%d", len(ids), deleted)
        return jsonify({"success": True, "deleted": deleted})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in bulk delete: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/legacy/drivers/<int:driver_id>", methods=["PUT"])
@login_required
def update_driver(driver_id):
    data = request.json or {}
    fields = ['name', 'empid', 'plate', 'car', 'iqama', 'phone', 'drivercard',
              'job', 'empNotes', 'model', 'pallets', 'load', 'vserial', 
              'inspect', 'license', 'opcard', 'notes', 'fuel_card', 'medical_exp', 'contract_exp']
              
    vals = {f: data.get(f, "").strip() for f in fields}

    if not vals['name']:
        return jsonify({"error": "Name is required"}), 400

    def parse_date(dstr):
        if not dstr: return None
        try:
            return datetime.strptime(dstr, '%Y-%m-%d').date()
        except:
            return None

    try:
        driver = Driver.query.get(driver_id)
        if not driver:
            return jsonify({"error": "Driver not found"}), 404

        old_name = driver.name
        old_plate = ""
        
        # Get active custody
        active_custody = VehicleCustody.query.filter_by(driver_id=driver_id, status="active").first()
        if active_custody and active_custody.vehicle:
            old_plate = active_custody.vehicle.plate_number

        # Update Driver
        driver.name = vals['name']
        driver.employee_id = vals['empid'] or driver.employee_id
        driver.iqama_number = vals['iqama'] or driver.iqama_number
        driver.phone = vals['phone']
        driver.job_title = vals['job']
        driver.iqama_expiry = parse_date(vals.get('iqama_exp')) or driver.iqama_expiry
        driver.license_expiry = parse_date(vals['license']) or driver.license_expiry

        # Handle Vehicle Update
        if vals['plate']:
            vehicle = Vehicle.query.filter_by(plate_number=vals['plate']).first()
            if not vehicle:
                # Create new vehicle if plate changed to a non-existent one
                vehicle = Vehicle(
                    branch_id=current_branch_id(),
                    plate_number=vals['plate'],
                    v_type=vals['car'],
                    model=vals['model'],
                    serial_number=vals['vserial'],
                    inspection_expiry=parse_date(vals['inspect'])
                )
                db.session.add(vehicle)
                db.session.flush()
            else:
                # Update existing vehicle fields
                vehicle.v_type = vals['car']
                vehicle.model = vals['model']
                vehicle.serial_number = vals['vserial']
                if vals['inspect']:
                    vehicle.inspection_expiry = parse_date(vals['inspect'])
            
            # Manage Custody
            if not active_custody or active_custody.vehicle_id != vehicle.id:
                if active_custody:
                    active_custody.status = "returned"
                    active_custody.returned_date = datetime.now().date()
                
                new_custody = VehicleCustody(
                    driver_id=driver.id,
                    vehicle_id=vehicle.id,
                    received_date=datetime.now().date(),
                    notes=vals['notes']
                )
                db.session.add(new_custody)
            else:
                active_custody.notes = vals['notes']
        elif active_custody:
            # Plate was cleared, return custody
            active_custody.status = "returned"
            active_custody.returned_date = datetime.now().date()

        db.session.commit()

        # Legacy sync (keep for tabs not yet rewritten)
        if old_name != vals['name'] or old_plate != vals['plate']:
            try:
                _sync_all_tabs_from_drivers(
                    old_name=old_name, old_plate=old_plate,
                    new_name=vals['name'], new_plate=vals['plate'], new_car=vals['car']
                )
            except Exception as e:
                logger.error(f"Legacy sync failed: {e}")

        logger.info("Driver updated via SQLAlchemy: id=%s name=%s plate=%s", driver_id, vals['name'], vals['plate'])
        return jsonify({"success": True, "id": driver_id, **vals})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating driver: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


def _rebuild_fleet_json():
    # Deprecated: We now use dynamic endpoint /api/fleet_data
    pass


# --- نبض النظام (Health Check) ---
@app.route('/health', methods=['GET'])
def health_check():
    import psutil
    cpu_usage = psutil.cpu_percent(interval=1)
    memory_usage = psutil.virtual_memory().percent
    return jsonify({
        "status": "online",
        "system_health": "good" if memory_usage < 90 else "warning",
        "resources": {"cpu": f"{cpu_usage}%", "memory": f"{memory_usage}%"}
    }), 200

@app.route('/update-driver', methods=['POST'])
@login_required
def update_driver_status():
    if session.get('role') != 'admin':
        return jsonify({"error": "غير مصرح لك"}), 403
    data = request.json
    driver_id = data.get('id')
    status = data.get('status')
    if driver_id:
        try:
            with db_connection() as conn:
                c = conn.cursor()
                if USE_POSTGRES:
                    c.execute("UPDATE drivers SET status = %s WHERE id = %s", (status, driver_id))
                else:
                    c.execute("UPDATE drivers SET status = ? WHERE id = ?", (status, driver_id))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    return jsonify({"error": "Invalid data"}), 400

@app.route('/api/update_driver_branch', methods=['POST'])
@login_required
def update_driver_branch():
    if session.get('role') != 'admin':
        return jsonify({"success": False, "error": "غير مصرح لك"}), 403
    data = request.json
    driver_id = data.get('id')
    branch_id = data.get('branch_id')
    if driver_id and branch_id is not None:
        try:
            with db_connection() as conn:
                c = conn.cursor()
                if USE_POSTGRES:
                    c.execute("UPDATE drivers SET branch_id = %s WHERE id = %s", (branch_id, driver_id))
                else:
                    c.execute("UPDATE drivers SET branch_id = ? WHERE id = ?", (branch_id, driver_id))
                conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": False, "error": "بيانات غير صالحة"}), 400

@app.route("/api/sync_excel", methods=["POST"])
@login_required
def api_sync_excel():
    try:

        df = pd.read_excel('DB-WORK/dammam_employees_data.xlsx', header=[0, 1])
        df.columns = [f"{c[0]}_{c[1]}" if 'Unnamed' not in str(c[0]) else str(c[1]) for c in df.columns]
        
        # --- مصفاة البيانات (Data Cleaner) ---
        df = df.dropna(how='all') # حذف الصفوف الفارغة
        for col in df.columns:
            if 'تاريخ انتهاء الاقامة' in str(col):
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        df = df.fillna("-")
        
        updated_count = 0
        with db_connection() as conn:
            c = conn.cursor()
            for index, row in df.iterrows():
                empid = None
                name = None
                iqama = None
                for col in df.columns:
                    if 'الرقم الوظيفي' in str(col): empid = str(row[col]).strip()
                    if 'اسم العامل' in str(col) or 'الاسم' in str(col): name = str(row[col]).strip()
                    if 'الإقامة' in str(col) or 'البطاقة' in str(col): iqama = str(row[col]).strip()
                
                if not name or str(name) == 'nan':
                    continue
                
                if str(empid) == 'nan': empid = None
                if str(iqama) == 'nan': iqama = None

                c.execute("SELECT id FROM drivers WHERE name=? OR empid=?", (name, empid))
                res = c.fetchone()
                if res:
                    driver_id = res['id']
                    if iqama:
                        c.execute("UPDATE drivers SET iqama=? WHERE id=?", (iqama, driver_id))
                    updated_count += 1
                else:
                    c.execute("INSERT INTO drivers (name, empid, iqama) VALUES (?, ?, ?)", (name, empid, iqama))
                    updated_count += 1
            conn.commit()
            
        return jsonify({"success": True, "message": "تم تحديث البيانات من ملف الإكسيل بنجاح", "updated_count": updated_count})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/fleet_data")

@login_required
def api_fleet_data():
    from models.schema import Driver, VehicleCustody
    fleet = []
    try:
        drivers = Driver.query.all()
        for d in drivers:
            custody = VehicleCustody.query.filter_by(driver_id=d.id, status='active').first()
            v = custody.vehicle if custody and custody.vehicle else None

            fleet.append({
                "id": d.id,
                "name": d.name or "",
                "empid": d.employee_id or "",
                "iqama": d.iqama_number or "",
                "plate": v.plate_number if v else "",
                "car": v.v_type if v else "",
                "phone": d.phone or "",
                "drivercard": "",
                "job": d.job_title or "",
                "empNotes": "",
                "model": v.model if v else "",
                "pallets": "",
                "load": "",
                "vserial": "",
                "inspect": "",
                "license": "",
                "opcard": "",
                "notes": ""
            })
    except Exception as e:
        # The SQLAlchemy tables this route reads are created by a separate manual
        # migration (migrate_db.py), not by the app's own startup init_db(). On a
        # fresh deploy where that migration hasn't run yet, degrade to an empty
        # fleet list instead of a hard 500 that would take down the whole fleet
        # dashboard page.
        logger.error(f"api_fleet_data error (SQLAlchemy tables may be missing): {e}")
        fleet = []
    response = jsonify(fleet)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


def _normalize_plate_py(plate):
    """Python نظير normalizePlate في JS — لمطابقة اللوحات بشكل موحد."""
    p = str(plate or "")
    ar_digits = "\u0660\u0661\u0662\u0663\u0664\u0665\u0666\u0667\u0668\u0669"
    for i, c in enumerate(ar_digits):
        p = p.replace(c, str(i))
    p = "".join(p.split())  # remove spaces
    import re as _re
    digits = "".join(_re.findall(r"\d+", p))
    letters = "".join(_re.findall(r"[^\d]+", p))
    return digits + letters


def _sync_all_tabs_from_drivers(old_name=None, old_plate=None, new_name=None, new_plate=None, new_car=None):
    """يحدّث جدول الغسيل والجدول الأسبوعي تلقائياً بناءً على السائق المحدّث.
    إذا تغيّر الاسم أو اللوحة، يُصحح القيمة في جميع الصفوف التي تحمل القيمة القديمة."""
    changed = False

    # ―― جدول الغسيل ――
    try:
        washing = blob_get("washing_schedule")
        if isinstance(washing, list):
            for v in washing:
                # تصحيح السائق بمطابقة اللوحة أولاً (الأدق)
                plate_match = (old_plate and v.get("plate") and
                               _normalize_plate_py(v["plate"]) == _normalize_plate_py(old_plate))
                name_match = (old_name and v.get("driver", "").strip() == old_name.strip())
                if plate_match or name_match:
                    if new_name and new_name != v.get("driver"):
                        v["driver"] = new_name
                        changed = True
                    if new_car and new_car != v.get("type"):
                        v["type"] = new_car
                        changed = True
                    if new_plate and new_plate != v.get("plate"):
                        v["plate"] = new_plate
                        changed = True
            if changed:
                blob_set("washing_schedule", washing)
                logger.info("washing_schedule synced after driver update")
    except Exception:
        logger.exception("sync washing_schedule failed")

    # ―― الجدول الأسبوعي ――
    try:
        sd = blob_get("schedule_data")
        if isinstance(sd, dict):
            sched_changed = False
            for section in ("main", "spare", "vacation"):
                for row in (sd.get(section) or []):
                    if not isinstance(row, dict):
                        continue
                    plate_match = (old_plate and row.get("plate") and
                                   _normalize_plate_py(row["plate"]) == _normalize_plate_py(old_plate))
                    name_match = (old_name and row.get("name", "").strip() == old_name.strip())
                    if plate_match or name_match:
                        if new_name and new_name != row.get("name"):
                            row["name"] = new_name
                            sched_changed = True
                        if new_plate and new_plate != row.get("plate"):
                            row["plate"] = new_plate
                            sched_changed = True
                        if new_car and new_car != row.get("vtype"):
                            row["vtype"] = new_car
                            sched_changed = True
            if sched_changed:
                blob_set("schedule_data", sd)
                logger.info("schedule_data synced after driver update")
    except Exception:
        logger.exception("sync schedule_data failed")


def _do_sync_all_from_drivers():
    _, drivers = _drivers_list_for_sync()
    # بناء فهرس لوحة→سائق من قاعدة البيانات
    plate_map = {}  # normalized_plate → driver_record
    for d in drivers:
        np = _normalize_plate_py(d.get("plate", ""))
        if np:
            plate_map[np] = d

    washing_updated = 0
    washing_added = 0
    washing_removed = 0
    sched_updated = 0

    # ―― تحديث جدول الغسيل ――
    try:
        washing = blob_get("washing_schedule")
        if not isinstance(washing, list):
            washing = []

        # صحّح الموجود واحذف الملغية
        valid_washing = []
        for v in washing:
            np = _normalize_plate_py(v.get("plate", ""))
            rec = plate_map.get(np)
            if rec:
                if (rec.get("name") or "").strip() and rec["name"] != v.get("driver"):
                    v["driver"] = rec["name"]
                    washing_updated += 1
                if (rec.get("car") or "").strip() and rec["car"] != v.get("type"):
                    v["type"] = rec["car"]
                    washing_updated += 1
                valid_washing.append(v)
            else:
                # لوحة ليست في قاعدة البيانات → احتفظ بها (قد تكون مدخلة يدوياً)
                valid_washing.append(v)

        # أضف المركبات الناقصة
        existing_plates = {_normalize_plate_py(v.get("plate", "")) for v in valid_washing}
        max_id = max((v.get("id") or 0 for v in valid_washing), default=0)
        for np, d in plate_map.items():
            if np and np not in existing_plates and (d.get("name") or "").strip():
                max_id += 1
                valid_washing.append({
                    "id": max_id,
                    "plate": d.get("plate", ""),
                    "type": d.get("car", ""),
                    "driver": d.get("name", ""),
                    "m": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                })
                washing_added += 1

        blob_set("washing_schedule", valid_washing)
    except Exception:
        logger.exception("sync_all_from_drivers: washing_schedule update failed")

    # ―― تحديث الجدول الأسبوعي ――
    try:
        sd = blob_get("schedule_data")
        if isinstance(sd, dict):
            for section in ("main", "spare", "vacation"):
                for row in (sd.get(section) or []):
                    if not isinstance(row, dict):
                        continue
                    np = _normalize_plate_py(row.get("plate", ""))
                    rec = plate_map.get(np)
                    if rec:
                        if (rec.get("name") or "").strip() and rec["name"] != row.get("name"):
                            row["name"] = rec["name"]
                            sched_updated += 1
                        if (rec.get("car") or "").strip() and rec["car"] != row.get("vtype"):
                            row["vtype"] = rec["car"]
                            sched_updated += 1
            blob_set("schedule_data", sd)
    except Exception:
        logger.exception("sync_all_from_drivers: schedule_data update failed")

    # ―― تحديث جميع الأنظمة الأخرى ――
    try:
        def _sync_misc(blob_key, plate_field="plate", name_fields=None, car_fields=None):
            if name_fields is None: name_fields = ["name", "driverName", "driver_name", "driver", "employee", "سائق", "الاسم"]
            if car_fields is None: car_fields = ["car", "vtype", "type", "vehicle_type", "vehicle"]
            data = blob_get(blob_key)
            if not data: return 0
            rows = []
            if isinstance(data, list): rows = data
            elif isinstance(data, dict):
                for k in ["data", "rows", "records", "incidents", "handovers", "drivers"]:
                    if k in data and isinstance(data[k], list):
                        rows = data[k]
                        break
            updated_count = 0
            for row in rows:
                if not isinstance(row, dict): continue
                np = _normalize_plate_py(row.get(plate_field, ""))
                rec = plate_map.get(np)
                if rec:
                    for nf in name_fields:
                        if nf in row and rec.get("name") and row[nf] != rec["name"]:
                            row[nf] = rec["name"]
                            updated_count += 1
                    for cf in car_fields:
                        if cf in row and rec.get("car") and row[cf] != rec["car"]:
                            row[cf] = rec["car"]
                            updated_count += 1
            if updated_count > 0:
                blob_set(blob_key, data)
            return updated_count
        
        misc_updated = 0
        misc_updated += _sync_misc("oils_data")
        misc_updated += _sync_misc("purchase_data")
        misc_updated += _sync_misc("workshop_data")
        misc_updated += _sync_misc("incidents_data")
        misc_updated += _sync_misc("records_data")
        misc_updated += _sync_misc("handover_data")
        logger.info("sync_all_from_drivers: synced %d fields in misc databases", misc_updated)
    except Exception:
        logger.exception("sync_all_from_drivers: misc update failed")

    # إعادة بناء fleet_data.json
    _rebuild_fleet_json()

    logger.info("sync_all_from_drivers: washing upd=%d add=%d | sched upd=%d",
                washing_updated, washing_added, sched_updated)
    return {
        "washing_updated": washing_updated,
        "washing_added": washing_added,
        "schedule_updated": sched_updated,
        "total_drivers": len(drivers)
    }

@app.route("/api/sync_all_from_drivers", methods=["POST"])
@login_required
def api_sync_all_from_drivers():
    """يمزامن جميع التبويبات (غسيل + أسبوعي) مع أحدث بيانات السائقين.
    يحذف المركبات التي لا سائق لها (plate تبيّن أنها ليست في قاعدة البيانات)، ويضيف الناقصة،
    ويصحح الأسماء واللوحات."""
    try:
        res = _do_sync_all_from_drivers()
        res["success"] = True
        return jsonify(res)
    except Exception:
        logger.exception("sync_all_from_drivers failed")
        return jsonify({"success": False, "error": "فشل التزامن"}), 500


@app.route("/api/generate_po", methods=["POST"])
@login_required
def generate_po():
    """Generate PO Excel from original template using openpyxl server-side"""
    try:
        data = request.json or {}

        template_path = os.path.join(os.path.dirname(__file__), "po_template.xlsx")
        if not os.path.exists(template_path):
            logger.error("PO template not found at: %s", template_path)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "قالب طلب الشراء غير موجود (po_template.xlsx)",
                    }
                ),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        ws["D7"] = data.get("driver", "")
        ws["G7"] = data.get("branch", "الدمام")
        ws["I7"] = data.get("job", "سائق")
        ws["D8"] = data.get("car", "")
        ws["G8"] = data.get("plate", "")
        
        model_val = data.get("model", "")
        if not model_val and data.get("car"):
            import re
            match = re.search(r'(\d{4}|[٠١٢٣٤٥٦٧٨٩]{4})', data.get("car", ""))
            if match:
                model_val = match.group(1)
        ws["I8"] = model_val
        
        ws["D9"] = data.get("odometer", "")
        
        try:
            ws.unmerge_cells(start_row=9, start_column=5, end_row=9, end_column=9)
        except Exception:
            pass

        ws["F8"] = "رقم اللوحة:"
        ws["H8"] = "الموديل:"

        ws["F9"] = "رقم الجوال:"
        ws["G9"] = data.get("phone", "")
        ws["H9"] = "الرقم الوظيفي:"
        ws["I9"] = data.get("empid", "")
        
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # neutral gray, B&W-print-safe
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Labels to format: F8(Plate), H8(Model), F9(Phone), H9(EmpID)
        for r, c in [(8, 6), (8, 8), (9, 6), (9, 8)]:
            cell = ws.cell(row=r, column=c)
            # Copy font from A8, but hardcode the rest to ensure it matches perfectly
            try:
                cell.font = copy(ws["A8"].font)
            except:
                cell.font = Font(name="Arial", size=12, bold=True)
            cell.fill = label_fill
            cell.border = thin_border
            cell.alignment = center_align
            
        # Values to format: G8(PlateVal), I8(ModelVal), G9(PhoneVal), I9(EmpIDVal)
        for r, c in [(8, 7), (8, 9), (9, 7), (9, 9)]:
            cell = ws.cell(row=r, column=c)
            try:
                cell.font = copy(ws["D8"].font)
            except:
                pass
            cell.border = thin_border
            cell.alignment = center_align

        date_val = data.get("date", "")
        if date_val:
            ws["C4"] = date_val
        serial_val = data.get("serial", "")
        if serial_val:
            ws["J3"] = f"NO. {serial_val}"

        from openpyxl.styles import Alignment
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def set_formatted(row, col, val, is_currency=False):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MC):
                if val is not None and val != "":
                    if is_currency:
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = val
                    else:
                        cell.value = val
                else:
                    cell.value = None
                cell.alignment = center_align

        # Fill parts (rows 12-14)
        parts = data.get("parts", [])
        for i, p in enumerate(parts[:3]):
            r = 12 + i
            set_formatted(r, 4, p.get("desc", ""))
            set_formatted(r, 7, p.get("qty"))
            set_formatted(r, 8, p.get("price"), is_currency=True)
            set_formatted(r, 9, p.get("val"), is_currency=True)
            set_formatted(r, 10, p.get("notes", ""))

        # Fill repairs (rows 18-20)
        repairs = data.get("repairs", [])
        for i, rep in enumerate(repairs[:3]):
            r = 18 + i
            set_formatted(r, 4, rep.get("desc", ""))
            set_formatted(r, 8, rep.get("val"), is_currency=True)
            set_formatted(r, 9, rep.get("notes", ""))

        # Fill tires (rows 24-26)
        tires = data.get("tires", [])
        for i, t in enumerate(tires[:3]):
            r = 24 + i
            set_formatted(r, 4, t.get("date", ""))
            set_formatted(r, 5, t.get("count"))
            set_formatted(r, 6, t.get("front"))
            set_formatted(r, 7, t.get("back"))
            set_formatted(r, 8, t.get("prev"))
            set_formatted(r, 9, t.get("curr"))
            set_formatted(r, 10, t.get("dist"))

        # Fill batteries (rows 30-32)
        batteries = data.get("batteries", [])
        for i, b in enumerate(batteries[:3]):
            r = 30 + i
            set_formatted(r, 4, b.get("desc", ""))
            set_formatted(r, 6, b.get("count"))
            set_formatted(r, 7, b.get("size", ""))
            set_formatted(r, 8, b.get("amp", ""))
            set_formatted(r, 9, b.get("price"), is_currency=True)
            set_formatted(r, 10, b.get("date", ""))

        # Fill summary totals (row 34: per-category, row 35-36: subtotals, row 37: grand total)
        summary = data.get("summary", {})

        # Row 34: individual category totals
        set_formatted(35, 3, summary.get("parts_total"), is_currency=True)
        set_formatted(35, 4, summary.get("repairs_total"), is_currency=True)
        set_formatted(35, 5, summary.get("tires_total"), is_currency=True)
        set_formatted(35, 6, summary.get("batteries_total"), is_currency=True)
        # Row 34 col 7-8: "الإجمالي شامل الضريبة" (label already in template)
        set_formatted(35, 7, summary.get("grand_total"), is_currency=True)
        # Row 34 col 9: Notes
        notes = data.get("notes", "")
        if notes:
            set_formatted(35, 9, notes)

        # Row 37: الإجمالي شامل الضريبة (grand total line) - Replaced with Tafqeet text
        grand_total_val = summary.get("grand_total")
        
        # Add Tafqeet in Col 5 (E) of Row 37 INSTEAD of the numeric value
        ws.cell(row=37, column=5).value = tafqeet(grand_total_val)
        ws.cell(row=37, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Remove the previous logic that put it in Col 1
        ws.cell(row=37, column=1).value = ""

        # Inject Logo
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_path = os.path.join(os.path.dirname(__file__), "logo_excel.png")
            if not os.path.exists(logo_path):
                logo_path = os.path.join(os.path.dirname(__file__), "static", "site_logo.png")
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 450
                img.height = 85
                ws.add_image(img, 'C1') 
        except Exception as e:
            logger.error("Logo injection failed: %s", e)

        # Clean black-&-white printing: fit to one page width, centered, A4 portrait.
        try:
            from openpyxl.worksheet.properties import PageSetupProperties
            from openpyxl.worksheet.page import PageMargins
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.print_options.horizontalCentered = True
            ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        except Exception as e:
            logger.warning("PO page setup skipped: %s", e)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        logger.exception("generate_po error")
        return jsonify({"success": False, "error": "تعذّر توليد طلب الشراء."}), 500


@app.route("/api/generate_oils", methods=["POST"])
@login_required
def generate_oils():
    """Generate Oils/Filters Excel from original template using openpyxl server-side"""
    try:
        data = request.json or {}

        template_path = os.path.join(os.path.dirname(__file__), "oils_template.xlsx")
        if not os.path.exists(template_path):
            logger.error("Oils template not found at: %s", template_path)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "قالب الزيوت غير موجود (oils_template.xlsx)",
                    }
                ),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def safe_set(row, col, val):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MC):
                cell.value = val

        # Update title (row 4)
        title = data.get("title", "")
        if title:
            ws["A4"] = title

        # Fill main oil data (rows 6-62, max 57 rows)
        oils = data.get("oils", [])
        for i, oil in enumerate(oils[:57]):
            r = 6 + i
            safe_set(r, 1, i + 1)  # م
            safe_set(r, 2, oil.get("plate", ""))  # رقم لوحة
            safe_set(r, 3, oil.get("driver", ""))  # المستخدم (merged C:D)
            safe_set(r, 5, oil.get("date", ""))  # تاريخ تغيير الزيت
            safe_set(r, 6, oil.get("counter", ""))  # رقم العداد
            safe_set(r, 7, oil.get("liters", ""))  # عدد اللترات
            safe_set(r, 8, oil.get("filters", ""))  # عدد الفلاتر

        # Fill filter details - LEFT side (A-E, rows 69-86)
        filters_left = data.get("filters_left", [])
        for i, f in enumerate(filters_left[:18]):
            r = 69 + i
            safe_set(r, 1, i + 1)
            safe_set(r, 2, f.get("name", ""))
            safe_set(r, 3, f.get("prev", ""))
            safe_set(r, 4, f.get("used", ""))
            safe_set(r, 5, f.get("remaining", ""))

        # Fill filter details - RIGHT side (F-H, rows 69-86)
        filters_right = data.get("filters_right", [])
        for i, f in enumerate(filters_right[:18]):
            r = 69 + i
            safe_set(r, 6, len(filters_left) + i + 1)
            safe_set(r, 7, f.get("name", ""))
            safe_set(r, 8, f.get("qty", ""))

        # Fill notes section (rows 65-66 only; 63 and 67 are fixed headers in template)
        notes = data.get("notes", {})
        if notes.get("row65"):
            safe_set(65, 1, notes["row65"])
        if notes.get("row66"):
            safe_set(66, 1, notes["row66"])

        # Fill diesel filter notes (rows 88-95)
        diesel_notes = data.get("diesel_notes", [])
        for i, note in enumerate(diesel_notes[:8]):
            safe_set(88 + i, 1, note)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/send_email", methods=["POST"])
@login_required
def send_email():
    data = request.json or {}
    recipient = (data.get("email") or "").strip()
    filename = data.get("filename", "invoice.xlsx")
    file_b64 = data.get("file_b64", "")
    custom_subject = data.get("subject", "").strip()
    cc_list = data.get("cc", [])

    # Validate required fields
    if not recipient:
        return jsonify({"error": "البريد الإلكتروني مطلوب"}), 400
    if not file_b64:
        return jsonify({"error": "بيانات الملف مطلوبة"}), 400
    # Basic email format check
    if "@" not in recipient or "." not in recipient.split("@")[-1]:
        return jsonify({"error": f"عنوان البريد غير صالح: {recipient}"}), 400

    try:
        file_data = base64.b64decode(file_b64)
        sender_email = app.config["MAIL_USERNAME"]

        # Dynamic subject
        if custom_subject:
            subject = custom_subject
        elif "زيوت" in filename or "فلاتر" in filename:
            subject = "بيان استهلاك الزيوت والفلاتر"
        elif "شراء" in filename or "طلب" in filename:
            subject = "طلب شراء وإصلاح"
        else:
            subject = "فاتورة صيانة"

        all_recipients = [recipient]
        cc_emails = [e.strip() for e in cc_list if e.strip()]

        msg = Message(
            subject,
            sender=sender_email,
            recipients=all_recipients,
            cc=cc_emails if cc_emails else [],
        )

        msg.body = _build_plain_text(subject, filename)
        msg.html = _build_html_email(subject, filename)
        msg.attach(
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_data,
        )

        mail.send(msg)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/compose_email", methods=["POST"])
@login_required
def compose_email():
    recipient = request.form.get("email")
    subject = request.form.get("subject", "BIN ZOMAH INTL.")
    body_text = request.form.get("body", "")
    file = request.files.get("attachment")
    cc_list = request.form.getlist("cc")

    if not recipient:
        return jsonify({"error": "Missing email"}), 400

    try:
        cc_emails = [e.strip() for e in cc_list if e.strip()]
        msg = Message(
            subject,
            sender=app.config["MAIL_USERNAME"],
            recipients=[recipient],
            cc=cc_emails if cc_emails else [],
        )

        msg.body = f"{body_text}\n\n{'—' * 40}\nKHALED AL-GHAMDI\nFleet Management Department\nBIN ZOMAH INTL. Trading & Development Co., Ltd.\nM: +966 53 975 7659\nE: damfleet@bz.sa"
        msg.html = _build_html_compose(subject, body_text)

        if file and file.filename:
            msg.attach(
                file.filename,
                file.content_type or "application/octet-stream",
                file.read(),
            )

        mail.send(msg)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _build_plain_text(subject, filename):
    return (
        f"{subject}\n"
        f"{'—' * 40}\n\n"
        f"Attached: {filename}\n\n"
        f"{'—' * 40}\n"
        f"KHALED AL-GHAMDI\n"
        f"Fleet Management Department\n"
        f"BIN ZOMAH INTL. Trading & Development Co., Ltd.\n"
        f"M: +966 53 975 7659\n"
        f"E: damfleet@bz.sa\n"
        f"{'—' * 40}\n"
        f"This email is auto-generated. Please do not reply directly."
    )


def _build_html_email(subject, filename):
    return f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#f4f4f7;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f7;padding:30px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.06);">

<!-- Header -->
<tr><td style="background:#0C2340;padding:20px 40px;">
<table width="100%" cellpadding="0" cellspacing="0"><tr>
<td style="text-align:right;vertical-align:middle;">
<h1 style="margin:0;color:#ffffff;font-size:20px;font-weight:700;letter-spacing:1px;">BIN ZOMAH INTL.</h1>
<p style="margin:4px 0 0;color:rgba(255,255,255,0.6);font-size:11px;letter-spacing:2px;text-transform:uppercase;">Trading & Development Co., Ltd.</p>
</td>
</tr></table>
</td></tr>

<!-- Subject Bar -->
<tr><td style="background:#f8f9fa;padding:16px 40px;border-bottom:1px solid #e9ecef;">
<p style="margin:0;color:#0C2340;font-size:15px;font-weight:600;">{subject}</p>
</td></tr>

<!-- Body -->
<tr><td style="padding:32px 40px;">
<p style="margin:0 0 20px;color:#333;font-size:15px;line-height:1.7;">
يرجى الاطلاع على الملف المرفق أدناه.
</p>

<!-- Attachment Card -->
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f8f9fa;border:1px solid #e9ecef;border-radius:8px;margin:0 0 24px;">
<tr>
<td style="padding:14px 18px;width:40px;vertical-align:middle;">
<div style="width:36px;height:36px;background:#0C2340;border-radius:8px;text-align:center;line-height:36px;font-size:16px;color:#fff;">📎</div>
</td>
<td style="padding:14px 18px 14px 0;">
<p style="margin:0;font-size:14px;font-weight:600;color:#333;">{filename}</p>
<p style="margin:2px 0 0;font-size:12px;color:#888;">Excel Document (.xlsx)</p>
</td>
</tr>
</table>
</td></tr>

<!-- Divider -->
<tr><td style="padding:0 40px;"><div style="border-top:1px solid #e9ecef;"></div></td></tr>

<!-- Signature -->
<tr><td style="padding:24px 40px 20px;">
<table cellpadding="0" cellspacing="0">
<tr>
<td style="padding-left:16px;border-left:3px solid #0C2340;">
<p style="margin:0;font-size:14px;font-weight:700;color:#0C2340;">KHALED AL-GHAMDI</p>
<p style="margin:3px 0 0;font-size:12px;color:#666;">خالد بن محمد الغامدي</p>
<p style="margin:8px 0 0;font-size:12px;color:#888;line-height:1.6;">
Fleet Management Department<br>
BIN ZOMAH INTL. Trading & Development Co., Ltd.
</p>
<p style="margin:8px 0 0;font-size:12px;color:#888;">
M: +966 53 975 7659 &nbsp;|&nbsp; E: damfleet@bz.sa
</p>
</td>
</tr>
</table>
</td></tr>

<!-- Footer -->
<tr><td style="background:#f8f9fa;padding:16px 40px;border-top:1px solid #e9ecef;text-align:center;">
<p style="margin:0;font-size:11px;color:#aaa;line-height:1.5;">
هذا البريد الإلكتروني تم إرساله تلقائياً — يرجى عدم الرد مباشرة على هذه الرسالة.
<br>This is an automated message. Please do not reply directly.
</p>
</td></tr>

</table>
</td></tr></table>
</body></html>"""


def _build_html_compose(subject, body_text):
    body_text = html.escape(body_text or "")  # user-provided → escape for the HTML email
    return f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#f4f4f7;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f7;padding:30px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.06);">

<!-- Header -->
<tr><td style="background:#0C2340;padding:20px 40px;">
<table width="100%" cellpadding="0" cellspacing="0"><tr>
<td style="text-align:right;vertical-align:middle;">
<h1 style="margin:0;color:#ffffff;font-size:20px;font-weight:700;letter-spacing:1px;">BIN ZOMAH INTL.</h1>
<p style="margin:4px 0 0;color:rgba(255,255,255,0.6);font-size:11px;letter-spacing:2px;text-transform:uppercase;">Trading & Development Co., Ltd.</p>
</td>
</tr></table>
</td></tr>

<!-- Body -->
<tr><td style="padding:32px 40px;">
<div style="color:#333;font-size:15px;line-height:1.8;white-space:pre-line;">{body_text}</div>
</td></tr>

<!-- Divider -->
<tr><td style="padding:0 40px;"><div style="border-top:1px solid #e9ecef;"></div></td></tr>

<!-- Signature -->
<tr><td style="padding:24px 40px 20px;">
<table cellpadding="0" cellspacing="0">
<tr>
<td style="padding-left:16px;border-left:3px solid #0C2340;">
<p style="margin:0;font-size:14px;font-weight:700;color:#0C2340;">KHALED AL-GHAMDI</p>
<p style="margin:3px 0 0;font-size:12px;color:#666;">خالد بن محمد الغامدي</p>
<p style="margin:8px 0 0;font-size:12px;color:#888;line-height:1.6;">
Fleet Management Department<br>
BIN ZOMAH INTL. Trading & Development Co., Ltd.
</p>
<p style="margin:8px 0 0;font-size:12px;color:#888;">
M: +966 53 975 7659 &nbsp;|&nbsp; E: damfleet@bz.sa
</p>
</td>
</tr>
</table>
</td></tr>

<!-- Footer -->
<tr><td style="background:#f8f9fa;padding:16px 40px;border-top:1px solid #e9ecef;text-align:center;">
<p style="margin:0;font-size:11px;color:#aaa;">
This message was sent from BIN ZOMAH INTL. Fleet Management System.
</p>
</td></tr>

</table>
</td></tr></table>
</body></html>"""


from routes.auth import auth_bp
from routes.api_fleet import api_fleet_bp
from routes.api_docs import api_docs_bp
from routes.ai import ai_bp
app.register_blueprint(auth_bp)
app.register_blueprint(api_fleet_bp)
app.register_blueprint(api_docs_bp)
app.register_blueprint(ai_bp)


# Mirror every /api/* endpoint under /importantworkstation/api/* (same handler). Because
# is_workstation() is path-based, those calls transparently read/write the id=2 sandbox,
# while the real /api/* endpoints (id=1) stay completely untouched. The AI routes are
# EXCLUDED from the mirror so the assistant is never reachable unauthenticated on the WS path.
for _rule in list(app.url_map.iter_rules()):
    if _rule.rule.startswith("/api/") and _rule.endpoint not in ("ai_chat", "ai_status"):
        app.add_url_rule(
            WS_PREFIX + _rule.rule,
            endpoint="ws_" + _rule.endpoint,
            view_func=app.view_functions[_rule.endpoint],
            methods=sorted(_rule.methods - {"HEAD", "OPTIONS"}),
        )

# Start the in-process daily scheduler for automatic document-expiry email digests.

# ==============================================================================
# CENTRAL FLEET REGISTRY
# ==============================================================================

@app.route("/registry")
@login_required
def registry_page():
    return render_template("registry.html", google_user=session.get("google_user"), b64_en=load_logo())


@app.route("/api/registry_import", methods=["POST"])
@login_required
def api_registry_import():
    if not session.get("is_admin"): return jsonify({"success": False, "error": "forbidden"}), 403
    f = request.files.get("file")
    if not f: return jsonify({"success": False, "error": "No file"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in sheet[1]]
        
        # Map headers to DB columns
        col_map = {
            "plate": ["لوحة", "اللوحة", "رقم اللوحة", "plate"],
            "iqama": ["اقامة", "إقامة", "هوية", "iqama"],
            "name": ["اسم", "السائق", "name"],
            "job": ["وظيفة", "الوظيفة", "المهنة", "job"],
            "empNotes": ["ملاحظات الموظف", "ملاحظات", "notes"],
            "model": ["موديل", "الموديل", "model"],
            "car": ["مركبة", "نوع المركبة", "car"],
            "pallets": ["طبالي", "الطبالي", "pallets"],
            "load": ["حمولة", "الحمولة", "load"],
            "vserial": ["تسلسلي", "الرقم التسلسلي", "serial"],
            "inspect": ["فحص", "الفحص", "inspect"],
            "license": ["سير", "رخصة السير", "license"],
            "opcard": ["تشغيل", "بطاقة التشغيل", "opcard"],
            "notes": ["ملاحظات", "notes"],
            "phone": ["جوال", "هاتف", "phone"],
            "drivercard": ["بطاقة السائق"]
        }
        
        h_idx = {}
        for db_col, aliases in col_map.items():
            for i, h in enumerate(headers):
                if any(a in h for a in aliases):
                    h_idx[db_col] = i
                    break
        
        if "plate" not in h_idx and "iqama" not in h_idx:
            return jsonify({"success": False, "error": "يجب أن يحتوي الملف على عمود (رقم اللوحة) أو (الإقامة)"}), 400
            
        updates = 0
        adds = 0
        
        with db_connection() as conn:
            c = conn.cursor()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = {}
                for db_col, idx in h_idx.items():
                    val = row[idx]
                    if val is None: val = ""
                    if hasattr(val, "strftime"): val = val.strftime("%Y-%m-%d")
                    data[db_col] = str(val).strip()
                
                plate = data.get("plate", "")
                iqama = data.get("iqama", "")
                if not plate and not iqama: continue
                
                if iqama:
                    c.execute("SELECT id FROM drivers WHERE iqama = ?", (iqama,))
                else:
                    c.execute("SELECT id FROM drivers WHERE plate LIKE ?", (f"%{plate}%",))
                
                existing = c.fetchone()
                if existing:
                    sets = []
                    params = []
                    for k, v in data.items():
                        if v: # Only update if excel cell is not empty
                            sets.append(f"{k} = ?")
                            params.append(v)
                    if sets:
                        params.append(existing['id'])
                        c.execute(f"UPDATE drivers SET {', '.join(sets)} WHERE id = ?", params)
                        updates += 1
                else:
                    cols = list(data.keys())
                    vals = list(data.values())
                    if "name" not in cols:
                        cols.append("name")
                        vals.append("غير معروف")
                    placeholders = ",".join(["?"] * len(cols))
                    c.execute(f"INSERT INTO drivers ({','.join(cols)}) VALUES ({placeholders})", vals)
                    adds += 1
            conn.commit()
        return jsonify({"success": True, "updates": updates, "adds": adds})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/registry_data", methods=["GET"])
@login_required
def api_registry_data():
    try:
        # 1. Fetch Drivers (Base)
        _, drivers = _drivers_list_for_sync()
        
        # 2. Fetch Employees (from SQL hr_employees table instead of JSON blob)
        emp_names = set()
        emp_iqamas = set()
        emp_ids = set()
        with db_connection() as db:
            hr_rows = db.execute("SELECT empid, iqama, name FROM hr_employees").fetchall()
            for r in hr_rows:
                if r["empid"]:
                    emp_ids.add(str(r["empid"]).strip())
                if r["iqama"]:
                    emp_iqamas.add(absher_sync.norm_id(r["iqama"]))
                if r["name"]:
                    emp_names.add(str(r["name"]).strip())
                
        # 3. Fetch Washing
        washing = blob_get("washing_schedule")
        washing_plates = set()
        if isinstance(washing, list):
            for v in washing:
                washing_plates.add(_normalize_plate_py(v.get("plate", "")))
                
        # 4. Fetch Schedule
        sd = blob_get("schedule_data")
        sched_plates = set()
        if isinstance(sd, dict):
            for section in ("main", "spare", "vacation"):
                for row in (sd.get(section) or []):
                    if isinstance(row, dict):
                        sched_plates.add(_normalize_plate_py(row.get("plate", "")))

        # 5. Fetch Other Tabs as Strings
        import json
        def _get_blob_str(t):
            b = blob_get(t)
            return json.dumps(b, ensure_ascii=False) if b else ""
        
        str_oils = _get_blob_str("oils_data")
        str_purch = _get_blob_str("purchase_data")
        str_ws = _get_blob_str("workshop_data")
        str_hand = _get_blob_str("handover_data")
        str_inc = _get_blob_str("incidents_data")
        str_rec = _get_blob_str("records_data")

        # Compile
        results = []
        for d in drivers:
            plate_norm = _normalize_plate_py(d.get("plate", ""))
            plate_raw = str(d.get("plate", "")).strip()
            iqama_norm = absher_sync.norm_id(d.get("iqama", ""))
            name = str(d.get("name", "")).strip()
            empid_val = str(d.get("empid", "")).strip()
            
            in_emp = (empid_val in emp_ids and bool(empid_val)) or (iqama_norm in emp_iqamas and bool(iqama_norm)) or (name in emp_names and bool(name))
            in_wash = plate_norm in washing_plates and bool(plate_norm)
            in_sched = plate_norm in sched_plates and bool(plate_norm)
            
            has_p = bool(plate_raw and len(plate_raw) > 2)
            has_n = bool(name and len(name) > 3)
            
            def check_in_str(s):
                if has_p and plate_raw in s: return True
                if has_p and plate_norm in s: return True
                if has_n and name in s: return True
                return False
            
            results.append({
                "id": d.get("id"),
                "name": d.get("name", ""),
                "empid": d.get("empid", ""),
                "plate": d.get("plate", ""),
                "car": d.get("car", ""),
                "iqama": d.get("iqama", ""),
                "phone": d.get("phone", ""),
                "drivercard": d.get("drivercard", ""),
                "in_emp": in_emp,
                "in_wash": in_wash,
                "in_sched": in_sched,
                "in_oils": check_in_str(str_oils),
                "in_purch": check_in_str(str_purch),
                "in_ws": check_in_str(str_ws),
                "in_hand": check_in_str(str_hand),
                "in_inc": check_in_str(str_inc),
                "in_rec": check_in_str(str_rec)
            })
            
        return jsonify({"success": True, "data": results})
    except Exception as e:
        logger.exception("api_registry_data failed")
        return jsonify({"success": False, "error": str(e)}), 500

# =====================================================================
# SYSTEM MONITORING & API DOCS
# =====================================================================

@app.route("/system_health")
def system_health():
    return render_template("system_health.html")

@app.route("/api/system_metrics", methods=["GET"])
def api_system_metrics():
    try:
        # CPU
        cpu_percent = psutil.cpu_percent(interval=0.1)
        # RAM
        mem = psutil.virtual_memory()
        # DB Size
        db_size_mb = 0
        if not USE_POSTGRES:
            if os.path.exists(DB_PATH):
                db_size_mb = round(os.path.getsize(DB_PATH) / (1024 * 1024), 2)
        else:
            with db_connection() as db:
                cur = db.execute("SELECT pg_database_size(current_database())")
                row = cur.fetchone()
                if row and row[0]:
                    db_size_mb = round(row[0] / (1024 * 1024), 2)

        # Sessions count (rough estimation from file system if sqlite/flask session, or from postgres)
        sessions_count = 0
        if USE_POSTGRES:
            with db_connection() as db:
                # Count distinct user ids if stored in db, or just use a placeholder
                try:
                    cur = db.execute("SELECT count(*) FROM pg_stat_activity")
                    row = cur.fetchone()
                    if row:
                        sessions_count = row[0]
                except:
                    pass

        return jsonify({
            "success": True,
            "metrics": {
                "cpu": cpu_percent,
                "ram_percent": mem.percent,
                "ram_used_gb": round(mem.used / (1024**3), 2),
                "ram_total_gb": round(mem.total / (1024**3), 2),
                "db_size_mb": db_size_mb,
                "active_connections": sessions_count,
                "uptime_hours": round((time.time() - psutil.boot_time()) / 3600, 1),
                "platform": "Cloud Hosting" if USE_POSTGRES else "Local/SQLite"
            }
        })
    except Exception as e:
        logger.exception("Failed to get system metrics")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/docs")
def api_docs():
    return render_template("api_docs.html")

# ── SPARE PARTS INVENTORY ───────────────────────────────────────────
@app.route("/spare_parts")
@login_required
def spare_parts_page():
    return render_template("spare_parts.html")

@app.route("/api/spare_parts", methods=["GET"])
@login_required
def get_spare_parts():
    parts = blob_get("spare_parts")
    if not isinstance(parts, list):
        parts = []
    return jsonify(parts)

@app.route("/api/spare_parts", methods=["POST"])
@login_required
def add_spare_part():
    data = request.json or {}
    parts = blob_get("spare_parts")
    if not isinstance(parts, list): parts = []
    
    new_id = max((p.get("id", 0) for p in parts), default=0) + 1
    new_part = {
        "id": new_id,
        "name": data.get("name", "").strip(),
        "part_number": data.get("part_number", "").strip(),
        "category": data.get("category", "أخرى"),
        "quantity": int(data.get("quantity", 0)),
        "unit_price": float(data.get("unit_price", 0.0)),
        "supplier": data.get("supplier", "").strip(),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    parts.append(new_part)
    blob_set("spare_parts", parts)
    return jsonify({"success": True, "part": new_part})

@app.route("/api/spare_parts/<int:part_id>", methods=["PUT"])
@login_required
def update_spare_part(part_id):
    data = request.json or {}
    parts = blob_get("spare_parts")
    if not isinstance(parts, list): return jsonify({"success": False}), 404
    
    for p in parts:
        if p.get("id") == part_id:
            if "name" in data: p["name"] = data["name"].strip()
            if "part_number" in data: p["part_number"] = data["part_number"].strip()
            if "category" in data: p["category"] = data["category"]
            if "quantity" in data: p["quantity"] = int(data["quantity"])
            if "unit_price" in data: p["unit_price"] = float(data["unit_price"])
            if "supplier" in data: p["supplier"] = data["supplier"].strip()
            p["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            blob_set("spare_parts", parts)
            return jsonify({"success": True, "part": p})
            
    return jsonify({"success": False, "error": "Not found"}), 404

@app.route("/api/spare_parts/<int:part_id>", methods=["DELETE"])
@login_required
def delete_spare_part(part_id):
    parts = blob_get("spare_parts")
    if not isinstance(parts, list): return jsonify({"success": False}), 404
    
    new_parts = [p for p in parts if p.get("id") != part_id]
    if len(new_parts) != len(parts):
        blob_set("spare_parts", new_parts)
        return jsonify({"success": True})
        
    return jsonify({"success": False, "error": "Not found"}), 404

@app.route("/api/dispense_part", methods=["POST"])
@login_required
def dispense_part():
    data = request.json or {}
    part_id = data.get("part_id")
    qty_to_dispense = int(data.get("quantity", 0))
    driver_plate = data.get("plate", "")
    driver_name = data.get("driver", "")
    
    if not part_id or qty_to_dispense <= 0:
        return jsonify({"success": False, "error": "Invalid parameters"}), 400
        
    parts = blob_get("spare_parts")
    if not isinstance(parts, list): return jsonify({"success": False, "error": "Inventory empty"}), 404
    
    for p in parts:
        if p.get("id") == part_id:
            current_qty = int(p.get("quantity", 0))
            if current_qty < qty_to_dispense:
                return jsonify({"success": False, "error": f"الكمية المطلوبة ({qty_to_dispense}) أكبر من المتوفر ({current_qty})"}), 400
                
            p["quantity"] = current_qty - qty_to_dispense
            p["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            blob_set("spare_parts", parts)
            
            # Log transaction for refund capability
            txns = blob_get("inventory_transactions")
            if not isinstance(txns, list): txns = []
            txn_id = str(int(datetime.now().timestamp() * 1000))
            txns.append({
                "id": txn_id,
                "part_id": part_id,
                "quantity": qty_to_dispense,
                "plate": driver_plate,
                "driver": driver_name,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            blob_set("inventory_transactions", txns)
            
            return jsonify({"success": True, "part": p, "transaction_id": txn_id})
            
    return jsonify({"success": False, "error": "Part not found"}), 404

@app.route("/api/refund_part", methods=["POST"])
@login_required
def refund_part():
    data = request.json or {}
    txn_id = data.get("transaction_id")
    if not txn_id:
        return jsonify({"success": False, "error": "No transaction ID"}), 400
        
    txns = blob_get("inventory_transactions")
    if not isinstance(txns, list): return jsonify({"success": False}), 200
    
    target_txn = None
    for t in txns:
        if t.get("id") == txn_id:
            target_txn = t
            break
            
    if not target_txn:
        return jsonify({"success": False, "error": "Transaction not found"}), 404
        
    parts = blob_get("spare_parts")
    if isinstance(parts, list):
        for p in parts:
            if p.get("id") == target_txn.get("part_id"):
                p["quantity"] = int(p.get("quantity", 0)) + int(target_txn.get("quantity", 0))
                p["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                blob_set("spare_parts", parts)
                break
                
    txns = [t for t in txns if t.get("id") != txn_id]
    blob_set("inventory_transactions", txns)
    
    return jsonify({"success": True})

@app.route("/vehicle_report")
@login_required
@role_required("admin", "branch_manager")
def vehicle_report_page():
    return render_template("vehicle_report.html", google_user=session.get("google_user", {}))

@app.route("/api/vehicle_report/<path:plate>")
@login_required
@role_required("admin", "branch_manager", "viewer")
def api_vehicle_report(plate):
    try:
        from models.schema import Vehicle
        plate = str(plate).strip()
        vehicle = Vehicle.query.filter_by(plate_number=plate).first()
        if not vehicle:
            return jsonify({"success": False, "error": "المركبة غير موجودة"}), 404

        # Calculate Fuel Cost
        fuel_cost = 0
        fuel = blob_get("fuel_data") or []
        frows = fuel.get("rows") if isinstance(fuel, dict) else (fuel if isinstance(fuel, list) else [])
        for r in frows:
            if isinstance(r, dict) and str(r.get("plate", "")).strip() == plate:
                try: fuel_cost += float(r.get("cost") or 0)
                except ValueError: pass

        # Calculate Workshop & Oils Cost
        workshop_cost = 0
        workshop = blob_get("workshop_data") or []
        wrows = workshop.get("rows") if isinstance(workshop, dict) else (workshop if isinstance(workshop, list) else [])
        for r in wrows:
            if isinstance(r, list) and len(r) > 10 and str(r[0] or "").strip() == plate:
                try: workshop_cost += float(r[8] or 0) # usually index 8 is cost in workshop
                except ValueError: pass
                
        oils = blob_get("oils_data") or []
        orows = oils.get("rows") if isinstance(oils, dict) else (oils if isinstance(oils, list) else [])
        for r in orows:
            if isinstance(r, list) and len(r) > 10 and str(r[0] or "").strip() == plate:
                try: workshop_cost += float(r[9] or 0) + float(r[10] or 0) # usually 9 and 10 are cost
                except ValueError: pass

        # Calculate Washing Cost
        washing_cost = 0
        washing = blob_get("washing_schedule") or []
        wrows = washing.get("rows") if isinstance(washing, dict) else (washing if isinstance(washing, list) else [])
        for r in wrows:
            if isinstance(r, dict) and str(r.get("plate", "")).strip() == plate:
                try: washing_cost += float(r.get("cost") or 0)
                except ValueError: pass

        driver = vehicle.custodies[-1].driver.name if vehicle.custodies else "غير معين"

        return jsonify({
            "success": True,
            "data": {
                "plate": plate,
                "model": vehicle.model or "غير محدد",
                "v_type": vehicle.v_type or "غير محدد",
                "driver": driver,
                "costs": {
                    "fuel": fuel_cost,
                    "workshop": workshop_cost,
                    "washing": washing_cost,
                    "total": fuel_cost + workshop_cost + washing_cost
                }
            }
        })
    except Exception as e:
        logger.exception("api_vehicle_report error")
        return jsonify({"success": False, "error": str(e)}), 500

# Safe under gunicorn --workers 1 (no --preload): runs in the worker, once.
if os.environ.get("ENABLE_BACKGROUND_SCHEDULER") == "true":
    _start_alert_scheduler()
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "False").lower() == "true"
    logger.info("Starting server on port %d (debug=%s)", port, debug)
    app.run(host="0.0.0.0", port=port, debug=debug)



@app.route('/admin/force_seed')
def force_seed():
    try:
        from models.database import db_connection, _is_header_row, _pk_clause
        import json, os
        js_path = os.path.join(os.path.dirname(__file__), 'drivers_data.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            content = f.read().replace('const driversData = ', '').strip()
            if content.endswith(';'): content = content[:-1]
            data = json.loads(content)
        seeded = 0
        with db_connection() as db:
            for d in data:
                if _is_header_row(d): continue
                db.execute(
                    'INSERT INTO drivers (name, empid, plate, car, iqama, phone, drivercard) '
                    'VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (
                        d.get('name', ''), str(d.get('empid', '')), d.get('plate', ''),
                        d.get('car', ''), str(d.get('iqama', '')), str(d.get('phone', '')),
                        d.get('drivercard', '')
                    )
                )
                seeded += 1
            db.commit()
        return f'Success! Seeded {seeded} drivers.'
    except Exception as e:
        import traceback
        return f'''<pre>Error: {str(e)}\n\n{traceback.format_exc()}</pre>'''


@app.route("/api/export_schedule_exact", methods=["POST"])
@login_required
def export_schedule_exact():
    try:
        data = request.json or {}
        template_path = os.path.join(app.root_path, "weekly_schedule_template.xlsx")
        
        if not os.path.exists(template_path):
            return jsonify({"success": False, "error": "قالب التصدير غير موجود"}), 500

        wb = openpyxl.load_workbook(template_path)
        
        def safe_set(sheet, row, col, val):
            cell = sheet.cell(row=row, column=col)
            try:
                cell.value = val
            except AttributeError:
                pass # Merged cell

        def safe_set_num(sheet, row, col, val):
            try:
                val = int(val)
            except (TypeError, ValueError):
                pass
            safe_set(sheet, row, col, val)

        # --- Active Vehicles (المركبات النشطة) ---
        ws_main = wb["المركبات النشطة"]
        main_data = data.get("main", [])
        for idx, rd in enumerate(main_data):
            r = 5 + idx
            safe_set(ws_main, r, 1, idx + 1)
            safe_set(ws_main, r, 2, rd.get("empid", ""))
            safe_set(ws_main, r, 3, rd.get("name", ""))
            safe_set(ws_main, r, 4, rd.get("iqama", ""))
            safe_set(ws_main, r, 5, rd.get("job", ""))
            safe_set(ws_main, r, 6, rd.get("plate", ""))
            safe_set(ws_main, r, 7, rd.get("model", ""))
            safe_set(ws_main, r, 8, rd.get("vtype", ""))
            safe_set(ws_main, r, 9, rd.get("pallets", ""))
            safe_set(ws_main, r, 10, rd.get("load", ""))
            safe_set(ws_main, r, 11, rd.get("vserial", ""))
            safe_set(ws_main, r, 12, rd.get("inspect", ""))
            # Col 13 is formula (rem_days1)
            safe_set(ws_main, r, 14, rd.get("license", ""))
            # Col 15 is formula (rem_days2)
            safe_set(ws_main, r, 16, rd.get("drivercard", ""))
            # Col 17 is formula (rem_days3)
            safe_set(ws_main, r, 18, rd.get("opcard", ""))
            # Col 19 is formula (rem_days4)
            safe_set(ws_main, r, 20, rd.get("empNotes", ""))
            safe_set(ws_main, r, 21, rd.get("phone", ""))

        # --- Spare and Broken (الأسبير والمعطلة) ---
        if "الأسبير والمعطلة" in wb.sheetnames:
            ws_spare = wb["الأسبير والمعطلة"]
            spare_data = data.get("spare", [])
            for idx, rd in enumerate(spare_data):
                r = 4 + idx
                safe_set(ws_spare, r, 1, idx + 1)
                safe_set(ws_spare, r, 2, rd.get("status", "اسبير"))
                safe_set(ws_spare, r, 3, rd.get("plate", ""))
                safe_set(ws_spare, r, 4, rd.get("model", ""))
                safe_set(ws_spare, r, 5, rd.get("vtype", ""))
                safe_set(ws_spare, r, 6, rd.get("pallets", ""))
                safe_set(ws_spare, r, 7, rd.get("load", ""))
                safe_set(ws_spare, r, 8, rd.get("vserial", ""))
                safe_set(ws_spare, r, 9, rd.get("inspect", ""))
                # Col 10 is the formula for remaining days
                safe_set(ws_spare, r, 11, rd.get("license", ""))
                # Col 12 is the formula for remaining days
                safe_set(ws_spare, r, 13, rd.get("opcard", ""))
                # Col 14 is the formula for remaining days
                safe_set(ws_spare, r, 15, rd.get("empNotes", ""))

        # --- Vacation (السائقون في إجازة) ---
        if "السائقون في إجازة" in wb.sheetnames:
            ws_vac = wb["السائقون في إجازة"]
            vac_data = data.get("vacation", [])
            for idx, rd in enumerate(vac_data):
                r = 4 + idx
                safe_set(ws_vac, r, 1, idx + 1)
                safe_set(ws_vac, r, 2, rd.get("empid", ""))
                safe_set(ws_vac, r, 3, rd.get("name", ""))
                safe_set(ws_vac, r, 4, rd.get("iqama", ""))
                safe_set(ws_vac, r, 5, rd.get("job", ""))
                safe_set(ws_vac, r, 6, rd.get("drivercard", ""))
                # Col 7 is the formula
                safe_set(ws_vac, r, 8, rd.get("phone", ""))
                safe_set(ws_vac, r, 9, rd.get("empNotes", ""))

        # --- Dashboard (لوحة المعلومات) — KPIs, expiry alert list, and the two summary
        # tables, matching whatever's currently shown/edited on the site's dashboard tab.
        # G5 (vacation KPI) is left untouched: it's a live COUNTA() formula in the template.
        if "لوحة المعلومات" in wb.sheetnames:
            ws_dash = wb["لوحة المعلومات"]
            dash = data.get("dashboard", {})

            if data.get("date"):
                safe_set(ws_dash, 2, 2, data.get("date"))

            kpis = dash.get("kpis", {})
            safe_set_num(ws_dash, 5, 1, kpis.get("drivers", ""))
            safe_set_num(ws_dash, 5, 3, kpis.get("delivery", ""))
            safe_set_num(ws_dash, 5, 5, kpis.get("distributors", ""))
            safe_set_num(ws_dash, 5, 9, kpis.get("spare", ""))

            expiring = dash.get("expiring", [])[:7]
            for idx, rd in enumerate(expiring):
                r = 9 + idx
                safe_set(ws_dash, r, 1, idx + 1)
                safe_set(ws_dash, r, 2, rd.get("name", ""))
                safe_set(ws_dash, r, 3, rd.get("plate", ""))
                safe_set(ws_dash, r, 4, rd.get("doc", ""))
                safe_set(ws_dash, r, 5, rd.get("date", ""))
                safe_set_num(ws_dash, r, 6, rd.get("days", ""))

            job_split = dash.get("jobSplit", {})
            for idx, (label, val) in enumerate(list(job_split.items())[:2]):
                r = 19 + idx
                safe_set(ws_dash, r, 1, label)
                safe_set_num(ws_dash, r, 2, val)

            def vt_sort_key(kv):
                try:
                    return -int(kv[1])
                except (TypeError, ValueError):
                    return 0
            vehicle_types = sorted(dash.get("vehicleTypes", {}).items(), key=vt_sort_key)[:5]
            for idx, (label, val) in enumerate(vehicle_types):
                r = 21 + idx
                safe_set(ws_dash, r, 1, label)
                safe_set_num(ws_dash, r, 2, val)

        import io, base64
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        import traceback
        logger.exception("export_schedule_exact error")
        return jsonify({"success": False, "error": str(e)}), 500



@app.route("/api/seed_from_template")
@login_required
def seed_from_template():
    import openpyxl
    template_path = os.path.join(app.root_path, "weekly_schedule_template.xlsx")
    if not os.path.exists(template_path):
        return "Template not found!", 404

    wb = openpyxl.load_workbook(template_path, data_only=True)

    def extract_rows(sheet_name, start_row, keys):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        data = []
        for r in range(start_row, ws.max_row + 1):
            row_data = {}
            has_value = False
            for c, key in enumerate(keys, start=2): # Start from col 2 (skip 'م')
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    has_value = True
                row_data[key] = str(val) if val is not None else ""
            if has_value:
                data.append(row_data)
        return data

    main_keys = ["empid", "name", "iqama", "job", "plate", "model", "vtype", "pallets", "load", "vserial", "inspect", "rem_days1", "license", "rem_days2", "drivercard", "rem_days3", "opcard", "rem_days4", "empNotes", "phone"]
    spare_keys = ["status", "plate", "model", "vtype", "pallets", "load", "vserial", "inspect", "rem_days1", "license", "rem_days2", "opcard", "rem_days3", "empNotes"]
    vac_keys = ["empid", "name", "iqama", "job", "drivercard", "rem_days1", "phone", "empNotes"]

    sd = {
        "main": extract_rows("المركبات النشطة", 5, main_keys),
        "spare": extract_rows("الأسبير والمعطلة", 4, spare_keys),
        "vacation": extract_rows("السائقون في إجازة", 4, vac_keys),
        "summary": {}
    }

    blob_set("schedule_data", sd)
    return "تم سحب البيانات من ملف الإكسل وإدخالها في الموقع بنجاح!"

@app.route('/api/weekly_update', methods=['GET', 'POST'])
def weekly_update_api():
    file_path = 'تحديث الاسبوعي - فرع الدمام (محدث).xlsx'
    
    if request.method == 'GET':
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        try:

            xlsx = pd.ExcelFile(file_path)
            data = {}
            for sheet in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet, header=None)
                df = df.replace({np.nan: ""})
                # Convert timestamps to string if any
                for col in df.columns:
                    if pd.api.types.is_datetime64_any_dtype(df[col]):
                        df[col] = df[col].dt.strftime('%Y-%m-%d')
                data[sheet] = df.values.tolist()
            return jsonify({"success": True, "data": data, "sheets": xlsx.sheet_names})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
            
    if request.method == 'POST':
        try:

            data = request.json.get('data', {})
            if not data:
                return jsonify({"error": "No data provided"}), 400
                
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, grid_data in data.items():
                    df = pd.DataFrame(grid_data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    
            return jsonify({"success": True, "message": "تم حفظ التعديلات بنجاح"})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
@app.route('/api/audit/deep_link', methods=['POST'])
@login_required
def api_audit_deep_link():
    body = request.get_json(silent=True) or {}
    action = body.get('action', 'وصول إداري (Deep Link)')
    target = body.get('type', 'غير محدد')
    _audit_add(action, target, 1)
    return jsonify({'success': True})
