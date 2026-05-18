# -*- coding: utf-8 -*-
"""Full audit: cross-reference every driver in DB against all Excel sources."""
import openpyxl, sqlite3, sys, io, os, json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
BASE = os.path.dirname(os.path.abspath(__file__))

# Build master lookup from the most reliable source: ربط_الأسماء_باللوحات
name_to_plate = {}
plate_file = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\ربط_الأسماء_باللوحات.xlsx'
wb = openpyxl.load_workbook(plate_file, data_only=True)
ws = wb.active
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(row=r, column=1).value or '').strip()
    plate = str(ws.cell(row=r, column=2).value or '').strip()
    if name and plate:
        name_to_plate[name] = plate
wb.close()
print(f"Source 1 (ربط_الأسماء_باللوحات): {len(name_to_plate)} entries")

# Build iqama->vehicle lookup from قائمة_المركبات
iqama_to_vehicle = {}
veh_file = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\قائمة_المركبات_بعد_ربط_الأسماء_باللوحة.xlsx'
wb2 = openpyxl.load_workbook(veh_file, data_only=True)
ws2 = wb2.active
for r in range(2, ws2.max_row + 1):
    plate = str(ws2.cell(row=r, column=1).value or '').strip()
    brand = str(ws2.cell(row=r, column=4).value or '').strip()
    model = str(ws2.cell(row=r, column=5).value or '').strip()
    year = str(ws2.cell(row=r, column=6).value or '').strip()
    iqama = str(ws2.cell(row=r, column=14).value or '').strip()
    name = str(ws2.cell(row=r, column=15).value or '').strip()
    if iqama and len(iqama) >= 8:
        car_desc = f"{year} {brand} {model}".strip()
        iqama_to_vehicle[iqama] = {'plate': plate, 'car': car_desc, 'name': name}
wb2.close()
print(f"Source 2 (قائمة_المركبات): {len(iqama_to_vehicle)} entries")

# Build iqama->data from PDF source (تحديث المركبات)
update_file = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\تحديث المركبات والسائقين-الدمام2026.xlsx'
iqama_to_update = {}
if os.path.exists(update_file):
    wb3 = openpyxl.load_workbook(update_file, data_only=True)
    ws3 = wb3.active
    # Find header row
    for r in range(1, min(ws3.max_row+1, 10)):
        for c in range(1, ws3.max_column+1):
            v = str(ws3.cell(row=r, column=c).value or '')
            if 'اسم' in v and 'سائق' in v.replace(' ',''):
                header_row = r
                break
    
    # Map columns
    headers = {}
    for c in range(1, ws3.max_column+1):
        v = str(ws3.cell(row=header_row, column=c).value or '').strip()
        if v:
            headers[c] = v
    
    for r in range(header_row+1, ws3.max_row+1):
        row_data = {}
        for c, h in headers.items():
            v = ws3.cell(row=r, column=c).value
            if v is not None:
                row_data[h] = str(v).strip()
        
        iqama = ''
        name = ''
        plate = ''
        car = ''
        for h, v in row_data.items():
            if 'اقام' in h or 'إقام' in h: iqama = v
            if 'اسم' in h and 'سائق' in h.replace(' ',''): name = v
            if 'لوح' in h: plate = v
            if 'نوع' in h and 'مركب' in h.replace(' ',''): car = v
        
        if iqama and len(iqama) >= 8:
            iqama_to_update[iqama] = {'name': name, 'plate': plate, 'car': car}
    wb3.close()
    print(f"Source 3 (تحديث المركبات): {len(iqama_to_update)} entries")

# Now audit every driver in DB
db = sqlite3.connect(os.path.join(BASE, 'database.sqlite'))
db.row_factory = sqlite3.Row
drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()

print(f"\n{'='*80}")
print(f"  AUDIT REPORT: {len(drivers)} drivers in database")
print(f"{'='*80}\n")

issues = []
fixes = []

for d in drivers:
    name = d['name'] or ''
    iqama = d['iqama'] or ''
    plate = d['plate'] or ''
    car = d['car'] or ''
    phone = d['phone'] or ''
    
    problems = []
    fix_data = {}
    
    # 1. Check against ربط_الأسماء_باللوحات
    if name in name_to_plate:
        correct_plate = name_to_plate[name]
        if plate and plate != correct_plate and correct_plate not in plate and plate not in correct_plate:
            problems.append(f"PLATE MISMATCH: DB='{plate}' vs Source='{correct_plate}'")
            fix_data['plate'] = correct_plate
        elif not plate:
            problems.append(f"MISSING PLATE -> should be '{correct_plate}'")
            fix_data['plate'] = correct_plate
    
    # 2. Check against قائمة_المركبات (by iqama)
    if iqama in iqama_to_vehicle:
        src = iqama_to_vehicle[iqama]
        src_plate = src['plate']
        src_car = src['car']
        
        if src_plate and plate and src_plate != plate and src_plate not in plate and plate not in src_plate:
            # Check if it's just formatting
            if src_plate.replace(' ','') != plate.replace(' ',''):
                problems.append(f"PLATE vs قائمة: DB='{plate}' vs Source='{src_plate}'")
        
        if src_car and not car:
            problems.append(f"MISSING CAR -> should be '{src_car}'")
            fix_data['car'] = src_car
    
    # 3. Check against تحديث المركبات (by iqama)  
    if iqama in iqama_to_update:
        src = iqama_to_update[iqama]
        if src.get('plate') and not plate:
            problems.append(f"MISSING PLATE (تحديث) -> '{src['plate']}'")
            fix_data['plate'] = src['plate']
        if src.get('car') and not car:
            problems.append(f"MISSING CAR (تحديث) -> '{src['car']}'")
            fix_data['car'] = src['car']
    
    # 4. Basic data completeness
    if not plate: problems.append("NO PLATE")
    if not car: problems.append("NO CAR TYPE")
    if not iqama: problems.append("NO IQAMA")
    if not phone: problems.append("NO PHONE")
    
    if problems:
        issues.append((d, problems))
        if fix_data:
            fixes.append((d['id'], name, fix_data))

# Print results
ok_count = len(drivers) - len(issues)
print(f"✅ CORRECT: {ok_count}/{len(drivers)} drivers have complete & verified data\n")

if issues:
    print(f"⚠️  ISSUES FOUND: {len(issues)} drivers\n")
    for d, probs in issues:
        print(f"  [{d['id']}] {d['name']}")
        print(f"      DB: iqama={d['iqama'] or '-'} | plate={d['plate'] or '-'} | car={d['car'] or '-'} | phone={d['phone'] or '-'}")
        for p in probs:
            print(f"      ❌ {p}")
        print()

# Apply fixes
if fixes:
    print(f"\n{'='*60}")
    print(f"  AUTO-FIXING {len(fixes)} drivers...")
    print(f"{'='*60}")
    for did, dname, fix in fixes:
        sets = ', '.join(f"{k} = ?" for k in fix.keys())
        vals = list(fix.values()) + [did]
        db.execute(f"UPDATE drivers SET {sets} WHERE id = ?", vals)
        print(f"  FIXED: {dname} <- {fix}")
    db.commit()

# Final stats
total = db.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
no_plate = db.execute("SELECT COUNT(*) FROM drivers WHERE plate IS NULL OR plate = '' OR plate = 'None'").fetchone()[0]
no_car = db.execute("SELECT COUNT(*) FROM drivers WHERE car IS NULL OR car = '' OR car = 'None'").fetchone()[0]
no_phone = db.execute("SELECT COUNT(*) FROM drivers WHERE phone IS NULL OR phone = '' OR phone = 'None'").fetchone()[0]
print(f"\nFINAL: Total={total} | No plate={no_plate} | No car={no_car} | No phone={no_phone}")

db.close()
