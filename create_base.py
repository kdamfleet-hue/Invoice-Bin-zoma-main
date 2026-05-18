import os
import openpyxl
from openpyxl.utils import get_column_letter

def create_base_template():
    src = None
    for f in os.listdir('.'):
        if f.endswith('.xlsx') and 'تحديث' in f and '(1)' in f:
            src = f
            break
    
    if not src:
        print("Source not found")
        return
        
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    
    # We will keep this as 'schedule_base.xlsx'. It has ALL formatting.
    # We will just clear the data cells.
    
    # Clear main data (row 9 to 36)
    for r in range(9, 37):
        for c in range(1, 18):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
                
    # Clear spare data (row 42 to 49)
    for r in range(42, 50):
        for c in range(1, 18):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None

    # Clear summary data (row 53)
    for c in range(1, 8):
        ws.cell(row=53, column=c).value = None

    # Clear date (row 6)
    ws.cell(row=6, column=1).value = None

    wb.save('schedule_base.xlsx')
    print("Created schedule_base.xlsx")

create_base_template()


