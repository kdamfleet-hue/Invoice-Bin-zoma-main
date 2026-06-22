import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'تقرير الورشة'
ws.sheet_view.rightToLeft = True

thin = Side(style='thin', color='000000')
thick = Side(style='medium', color='000000')
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
font_bold = Font(name='Arial', size=12, bold=True)
font_normal = Font(name='Arial', size=12)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 30

ws.merge_cells('A3:E3')
title_cell = ws['A3']
title_cell.value = 'تقرير فنى الورشة عن السيارة الموضحه بيناتها ادناه لفرع الدمام'
title_cell.font = Font(name='Arial', size=16, bold=True)
title_cell.alignment = align_center

ws.merge_cells('A4:E4')
date_cell = ws['A4']
date_cell.value = 'التاريخ:    /    / 2026'
date_cell.font = font_bold
date_cell.alignment = Alignment(horizontal='right', vertical='center')

ws.merge_cells('A5:B5')
ws['A5'] = 'نوع السيارة (V/NAME)'
ws['C5'] = 'رقم اللوحة (P:NO)'
ws['D5'] = 'الموديل (MODEL)'
ws['E5'] = 'العداد (K/m)'

ws.merge_cells('A6:B6')
for col in ['A', 'C', 'D', 'E']:
    ws[f'{col}5'].font = font_bold
    ws[f'{col}5'].alignment = align_center
    ws[f'{col}5'].border = border_all
    
    cell_val = ws[f'{col}6'] if col != 'A' else ws['A6']
    cell_val.font = font_normal
    cell_val.alignment = align_center
    cell_val.border = border_all
ws.row_dimensions[6].height = 30

ws['B5'].border = border_all
ws['B6'].border = border_all

ws['A7'] = 'م'
ws.merge_cells('B7:D7')
ws['B7'] = 'الملاحظات'
ws['E7'] = 'الحل'

for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}7'].font = font_bold
    ws[f'{col}7'].alignment = align_center
    ws[f'{col}7'].border = border_all

for i in range(1, 11):
    r = 7 + i
    ws[f'A{r}'] = i
    ws.merge_cells(f'B{r}:D{r}')
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{r}'].font = font_normal
        ws[f'{col}{r}'].alignment = align_center
        ws[f'{col}{r}'].border = border_all
    ws.row_dimensions[r].height = 35

r_tech = 18
ws.merge_cells(f'A{r_tech}:E{r_tech}')
ws[f'A{r_tech}'] = 'توقيع الفني / '
ws[f'A{r_tech}'].font = font_bold
ws[f'A{r_tech}'].alignment = Alignment(horizontal='right', vertical='center')
ws[f'A{r_tech}'].border = border_all
ws.row_dimensions[r_tech].height = 35

r_mgr = 19
ws.merge_cells(f'A{r_mgr}:E{r_mgr}')
ws[f'A{r_mgr}'] = 'توقيع مسؤول الحركة / '
ws[f'A{r_mgr}'].font = font_bold
ws[f'A{r_mgr}'].alignment = Alignment(horizontal='right', vertical='center')
ws[f'A{r_mgr}'].border = border_all
ws.row_dimensions[r_mgr].height = 35

try:
    img = XLImage('static/site_logo.png')
    img.width = 300
    img.height = 80
    ws.add_image(img, 'B1')
except:
    pass

wb.save('تقرير الورشة.xlsx')
