# -*- coding: utf-8 -*-
"""Search all Excel files on Desktop/Downloads for the 8 missing drivers."""
import openpyxl, sqlite3, sys, io, os, json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Missing drivers iqamas
MISSING = {
    '2342453558': 'أشرف كمال كامل هريدي',
    '2249251576': 'ابراهيم يوسف عوض علي',
    '2533532608': 'احمد رشاد محمد شوقى عبد المطلب',
    '2596889713': 'احمد محمد زكى شعلان',
    '2482824873': 'عبد الرحمن جوده محمد محمود',
    '2519087791': 'عبد الرحمن محمد سعيد البندارى',
    '2578319739': 'كرم جلال الدين عابد عثمان',
    '2494382019': 'محمد حسن عبد الرحمن السيد',
    '2532825581': 'محمد محمود محمد ابو الحسن',
}

# Key Excel files to search
search_files = [
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\نسخة من تحديث المركبات والسائقين-الدمام2026 -5-3 تم اضافه ارقام الاقامه-1.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\تحديث المركبات والسائقين-الدمام2026.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\جدول الخاص بجميع المركبات والسائقين في الفرع وذلك للأسبوع الأول من شهر مايو.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\مركبات لها مستخدم فعلي 2026-05-03 2.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\مركبات لها مستخدم فعلي 2026-05-03.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\1جدول بيانات السائقين والسيارات وتاريخ انتهاء بطائق السائقين لفرع الدمام.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\vehicle_list_excel_without_qr-code_2026-05-03.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\vehicle_list_excel_without_qr-code_2026-05-06 (1) 2.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\قائمة_المركبات_بعد_ربط_الأسماء_باللوحة.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\مشروع موقع\ربط_الأسماء_باللوحات.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\New\تحديث_المركبات_والسائقين_الدمام_2026_محدث_نهائي.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\أرشيف\التفاويض 2026-05-15.xlsx',
    r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy - Copy (2)\أرشيف\عرض المستخدمين الفعليين 2026-05-15.xlsx',
]

found_data = {}

for fpath in search_files:
    if not os.path.exists(fpath):
        continue
    
    fname = os.path.basename(fpath)
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                row_text = ''
                row_vals = {}
                for c in range(1, min(ws.max_column + 1, 20)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        sv = str(v).strip()
                        row_text += sv + ' | '
                        row_vals[c] = sv
                
                # Check if any missing iqama is in this row
                for iqama, name in MISSING.items():
                    if iqama in row_text or name in row_text:
                        print(f"\n  FOUND in {fname} / {sn} / Row {r}:")
                        for col, val in row_vals.items():
                            print(f"    Col {col}: {val}")
                        
                        if iqama not in found_data:
                            found_data[iqama] = {'name': name}
                        
                        # Try to extract plate/car
                        for col, val in row_vals.items():
                            sv = str(val)
                            # Plate pattern (Arabic letters + numbers)
                            if any(ch in sv for ch in 'ابتثجحخدذرزسشصضطظعغفقكلمنهوي') and any(ch.isdigit() for ch in sv) and len(sv) < 20:
                                if 'plate' not in found_data[iqama]:
                                    found_data[iqama]['plate'] = sv
                            # Car type
                            if any(k in sv for k in ['ايسوزو', 'متسوبيشي', 'دينا', 'لوري', 'هينو', 'تويوتا']):
                                if 'car' not in found_data[iqama]:
                                    found_data[iqama]['car'] = sv
    except Exception as e:
        print(f"  Error reading {fname}: {e}")

print(f"\n{'='*60}")
print(f"Found data for {len(found_data)} missing drivers:")
for iq, data in found_data.items():
    print(f"  {data.get('name','')} ({iq}): plate={data.get('plate','-')} car={data.get('car','-')}")

# Update database
if found_data:
    db = sqlite3.connect(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.sqlite'))
    updated = 0
    for iqama, data in found_data.items():
        plate = data.get('plate', '')
        car = data.get('car', '')
        if plate or car:
            sets = []
            vals = []
            if plate:
                sets.append("plate = ?")
                vals.append(plate)
            if car:
                sets.append("car = ?")
                vals.append(car)
            vals.append(iqama)
            db.execute(f"UPDATE drivers SET {', '.join(sets)} WHERE iqama = ? AND (plate IS NULL OR plate = '' OR plate = 'None')", vals)
            updated += 1
    db.commit()
    print(f"\nUpdated {updated} drivers in database")
    
    total = db.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
    no_plate = db.execute("SELECT COUNT(*) FROM drivers WHERE plate IS NULL OR plate = '' OR plate = 'None'").fetchone()[0]
    print(f"Total: {total} | Still missing plate: {no_plate}")
    db.close()


