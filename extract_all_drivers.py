# -*- coding: utf-8 -*-
"""Extract all driver data from all Excel files and merge into a unified dataset."""
import os, sys, json, re
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))

def read_excel_all_rows(path):
    """Read all rows from all sheets of an Excel file."""
    rows = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            headers = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=c).value
                headers.append(str(v).strip() if v else f"col{c}")
            # Try row 4 as header if row 1 looks like title
            if ws.max_row > 5:
                h4 = []
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=4, column=c).value
                    if v:
                        h4.append(str(v).strip())
                if len(h4) > len([h for h in headers if not h.startswith("col")]):
                    headers = []
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=4, column=c).value
                        headers.append(str(v).strip() if v else f"col{c}")
            
            for r in range(2, ws.max_row + 1):
                row_data = {"_source": os.path.basename(path), "_sheet": sn}
                has_data = False
                for c in range(1, min(len(headers) + 1, ws.max_column + 1)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None and str(v).strip():
                        key = headers[c-1] if c-1 < len(headers) else f"col{c}"
                        row_data[key] = str(v).strip()
                        has_data = True
                if has_data:
                    rows.append(row_data)
    except Exception as e:
        print(f"Error reading {path}: {e}", file=sys.stderr)
    return rows

# Collect all Excel files
excel_files = []
for f in os.listdir(BASE):
    if f.endswith(".xlsx") and not f.startswith("~"):
        excel_files.append(os.path.join(BASE, f))

print(f"Found {len(excel_files)} Excel files")

all_rows = []
for ef in excel_files:
    rows = read_excel_all_rows(ef)
    all_rows.extend(rows)
    print(f"  {os.path.basename(ef)}: {len(rows)} rows")

# Also read drivers_data.js
js_path = os.path.join(BASE, "drivers_data.js")
js_drivers = []
if os.path.exists(js_path):
    with open(js_path, "r", encoding="utf-8") as f:
        content = f.read().replace("const driversData = ", "").strip()
        if content.endswith(";"): content = content[:-1]
        js_drivers = json.loads(content)
    print(f"  drivers_data.js: {len(js_drivers)} drivers")

# Build unified driver map
# Key fields to look for:
NAME_KEYS = ["اسم السائق", "اسم المستخدم", "المستخدم", "الاسم", "name", "اسم السائق:"]
IQAMA_KEYS = ["رقم الاقامه", "رقم الاقامة", "الاقامة", "الاقامه", "iqama", "رقم الإقامة"]
EMPID_KEYS = ["الرقم الوظيفي", "رقم وظيفي", "empid", "م", "الرقم"]
PLATE_KEYS = ["رقم اللوحة", "رقم اللوحه", "اللوحة", "plate"]
CAR_KEYS = ["نوع السيارة", "نوع المركبة", "الماركة", "car", "نوع السياره"]
PHONE_KEYS = ["رقم الجوال", "الجوال", "جوال", "phone", "رقم التواصل", "هاتف"]
LICENSE_KEYS = ["انتهاء الرخصة", "تاريخ انتهاء الرخصة", "رخصة", "رخصه", "انتهاء رخصة القيادة"]
CARD_KEYS = ["انتهاء بطاقة التشغيل", "بطاقة التشغيل", "انتهاء البطاقة"]
DRIVER_CARD_KEYS = ["انتهاء بطاقة السائق", "بطاقة السائق"]
MODEL_KEYS = ["الموديل", "موديل", "model", "سنة الصنع"]
JOB_KEYS = ["الوظيفة", "وظيفة", "المهنة"]

def find_val(row, keys):
    for k in keys:
        for rk, rv in row.items():
            if k in rk.lower() or rk.lower() in k:
                return rv
    return None

# Build driver database
drivers_db = {}

def add_to_db(name, data):
    if not name or len(name) < 2:
        return
    name = name.strip()
    if name in drivers_db:
        for k, v in data.items():
            if v and (k not in drivers_db[name] or not drivers_db[name][k]):
                drivers_db[name][k] = v
    else:
        drivers_db[name] = data

# Process JS drivers first (most reliable)
for d in js_drivers:
    name = d.get("name", "").strip()
    if name:
        add_to_db(name, {
            "empid": d.get("empid", ""),
            "iqama": d.get("iqama", ""),
            "plate": d.get("plate", ""),
            "car": d.get("car", ""),
            "phone": d.get("phone", ""),
        })

# Process Excel rows
for row in all_rows:
    name = find_val(row, NAME_KEYS)
    if not name:
        continue
    
    data = {
        "empid": find_val(row, EMPID_KEYS) or "",
        "iqama": find_val(row, IQAMA_KEYS) or "",
        "plate": find_val(row, PLATE_KEYS) or "",
        "car": find_val(row, CAR_KEYS) or "",
        "phone": find_val(row, PHONE_KEYS) or "",
        "license_expiry": find_val(row, LICENSE_KEYS) or "",
        "op_card_expiry": find_val(row, CARD_KEYS) or "",
        "driver_card_expiry": find_val(row, DRIVER_CARD_KEYS) or "",
        "model": find_val(row, MODEL_KEYS) or "",
        "job": find_val(row, JOB_KEYS) or "",
    }
    add_to_db(name, data)

# Print results
print(f"\n{'='*80}")
print(f"Total unique drivers found: {len(drivers_db)}")
print(f"{'='*80}\n")

# Sort by name
for i, (name, data) in enumerate(sorted(drivers_db.items()), 1):
    print(f"[{i}] {name}")
    for k, v in data.items():
        if v:
            label = {
                "empid": "الرقم الوظيفي",
                "iqama": "رقم الاقامة",
                "plate": "رقم اللوحة",
                "car": "نوع السيارة",
                "phone": "رقم الجوال",
                "license_expiry": "انتهاء الرخصة",
                "op_card_expiry": "انتهاء بطاقة التشغيل",
                "driver_card_expiry": "انتهاء بطاقة السائق",
                "model": "الموديل",
                "job": "الوظيفة",
            }.get(k, k)
            print(f"    {label}: {v}")
    print()

# Save as JSON
output_path = os.path.join(BASE, "all_drivers_merged.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(drivers_db, f, ensure_ascii=False, indent=2)
print(f"Saved to: {output_path}")
