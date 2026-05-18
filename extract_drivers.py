import openpyxl, os, json, re

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New'
driver_file = [f for f in os.listdir(folder) if 'بيانات السائ' in f.replace('ئ', 'ئ') or 'بيانات' in f][0]
vehicle_file = [f for f in os.listdir(folder) if 'دينا' in f and 'لوري' in f][0]

driver_path = os.path.join(folder, driver_file)
vehicle_path = os.path.join(folder, vehicle_file)

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

# Load vehicle types
veh_types = {}
wb_veh = openpyxl.load_workbook(vehicle_path, data_only=True)
ws_veh = wb_veh.active
for r in range(6, ws_veh.max_row + 1):
    plate = ws_veh.cell(row=r, column=11).value
    if plate and str(plate).strip() and str(plate) != 'رقم اللوحة':
        np = normalize_plate(plate)
        car_type = str(ws_veh.cell(row=r, column=12).value or '')
        car_model = str(ws_veh.cell(row=r, column=13).value or '')
        full_car = f"{car_type} {car_model}".strip()
        if not full_car: full_car = "غير محدد"
        veh_types[np] = full_car

# Load drivers
drivers = []
wb_drv = openpyxl.load_workbook(driver_path, data_only=True)
ws_drv = wb_drv.active
for r in range(9, ws_drv.max_row + 1):
    name = str(ws_drv.cell(row=r, column=3).value or '').strip()
    plate = str(ws_drv.cell(row=r, column=7).value or '').strip()
    
    if name and name != 'None':
        np = normalize_plate(plate)
        car = veh_types.get(np, "غير محدد")
        drivers.append({"name": name, "plate": plate, "car": car})

out_path = os.path.join(folder, '..', 'InvoiceApp', 'drivers_data.js')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write("const driversData = " + json.dumps(drivers, ensure_ascii=False) + ";\n")
print(f"Generated {len(drivers)} drivers.")



