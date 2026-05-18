import openpyxl, sys, json
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\InvoiceApp\طلب شرا قسم حركه.xlsx')
ws = wb.active

print("=== MERGED CELLS ===")
for m in sorted(ws.merged_cells.ranges, key=lambda x: x.min_row):
    print(f"  {m}")

print("\n=== COL WIDTHS ===")
for col in range(1, 12):
    l = get_column_letter(col)
    print(f"  {l}: {ws.column_dimensions[l].width}")

print("\n=== ROW HEIGHTS ===")
for row in range(1, 45):
    print(f"  Row {row}: {ws.row_dimensions[row].height}")

print("\n=== CELL DETAILS (non-empty or styled) ===")
for row in range(1, 45):
    for col in range(1, 12):
        c = ws.cell(row=row, column=col)
        val = c.value
        has_fill = c.fill and c.fill.fgColor and c.fill.fgColor.theme is not None or (c.fill and c.fill.fgColor and c.fill.fgColor.rgb and c.fill.fgColor.rgb != '00000000')
        has_border = c.border and (c.border.top.style or c.border.bottom.style or c.border.left.style or c.border.right.style)
        
        if val is not None or has_fill or has_border:
            letter = get_column_letter(col)
            fill_info = ''
            if c.fill and c.fill.fgColor:
                try:
                    fill_info = f"fill_rgb={c.fill.fgColor.rgb} theme={c.fill.fgColor.theme} tint={c.fill.fgColor.tint}"
                except:
                    fill_info = "fill=theme"
            
            font_info = f"font={c.font.name} size={c.font.size} bold={c.font.bold}"
            if c.font.color:
                try:
                    font_info += f" color_rgb={c.font.color.rgb} theme={c.font.color.theme}"
                except:
                    font_info += " color=theme"
            
            border_info = ''
            if has_border:
                sides = []
                for side_name in ['top','bottom','left','right']:
                    s = getattr(c.border, side_name)
                    if s.style:
                        sides.append(f"{side_name}={s.style}")
                border_info = f" borders=[{','.join(sides)}]"
            
            align = ''
            if c.alignment:
                align = f" h={c.alignment.horizontal} v={c.alignment.vertical}"
            
            print(f"  {letter}{row}: val=\"{val}\" {font_info} {fill_info}{border_info}{align}")



