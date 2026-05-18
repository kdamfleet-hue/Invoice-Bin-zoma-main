"""
finalize_gps.py
---------------
ينسخ البيانات من ملف المصدر إلى ملف الكشف الأساسي بالترتيب الصحيح
مع الحفاظ على التنسيق بالكامل ويحفظ الملف النهائي المميز.
"""
import os
import sys
# Fix Windows console encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy

# ── مسارات الملفات ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_SRC    = os.path.join(BASE_DIR, "GPS_source_data.xlsx")
FILE_BASE   = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026-1-26 - Copy.xlsx")
FILE_FINAL  = os.path.join(BASE_DIR, "كشف_اجهزة_GPS_الدمام_2026_نهائي_مميز.xlsx")

# ── التحقق من وجود الملفات ───────────────────────────────────────────────────
for f in [FILE_SRC, FILE_BASE]:
    if not os.path.exists(f):
        print(f"[ERROR] الملف غير موجود: {f}")
        sys.exit(1)

# ── خريطة الأعمدة ─────────────────────────────────────────────────────────────
COL_MAP = {
    'رقم الهيكل':    1,   # A
    'الرقم التسلسلي': 2,  # B
    'سنة الصنع':     3,   # C
    'الطراز':        4,   # D
    'الماركة':       5,   # E
    'نوع التسجيل':   6,   # F
    'الفرع':         7,   # G
    'رقم اللوحة':    9,   # I
}
SERIAL_COL = 10  # J

# ── كتابة آمنة تدعم الخلايا المدمجة ─────────────────────────────────────────
def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
    else:
        cell.value = value

# ── تحميل البيانات ───────────────────────────────────────────────────────────
print("قراءة بيانات المصدر (صفوف 5 إلى 81)...")
df_src_full = pd.read_excel(FILE_SRC, header=3)
df_src_data = df_src_full.iloc[0:77].copy()
print(f"تم قراءة {len(df_src_data)} صف.")

# التحقق من الأعمدة
missing = [c for c in COL_MAP.keys() if c not in df_src_data.columns]
if missing:
    print(f"[WARNING] أعمدة ناقصة في المصدر: {missing}")

print("تحميل ملف الكشف الأساسي...")
wb = load_workbook(FILE_BASE)
ws = wb.active

# ── كتابة البيانات ───────────────────────────────────────────────────────────
print("كتابة البيانات مع الحفاظ على التنسيق...")
START_ROW = 6

for i, (_, src_row) in enumerate(df_src_data.iterrows()):
    target_row = START_ROW + i
    for src_col, col_idx in COL_MAP.items():
        if src_col not in df_src_data.columns:
            continue
        val = src_row[src_col]
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            safe_write(ws, target_row, col_idx, val)
    safe_write(ws, target_row, SERIAL_COL, i + 1)

# ── نسخ التنسيق من الصف الأول ────────────────────────────────────────────────
print("تطبيق التنسيق الاحترافي...")
TEMPLATE_ROW = START_ROW
DATA_END_ROW = START_ROW + len(df_src_data) - 1
MAX_COL = 10  # A:J

for r in range(TEMPLATE_ROW + 1, DATA_END_ROW + 1):
    for c in range(1, MAX_COL + 1):
        src_cell  = ws.cell(row=TEMPLATE_ROW, column=c)
        tgt_cell  = ws.cell(row=r, column=c)
        if src_cell.has_style:
            tgt_cell.font         = copy(src_cell.font)
            tgt_cell.border       = copy(src_cell.border)
            tgt_cell.fill         = copy(src_cell.fill)
            tgt_cell.alignment    = copy(src_cell.alignment)
            tgt_cell.number_format = copy(src_cell.number_format)

# ضبط ارتفاع الصفوف
template_h = ws.row_dimensions[TEMPLATE_ROW].height
if template_h:
    for r in range(TEMPLATE_ROW, DATA_END_ROW + 1):
        ws.row_dimensions[r].height = template_h

wb.save(FILE_FINAL)
print(f"تم الحفظ: {FILE_FINAL}")
