"""
import_weekly_schedule.py
=========================
ينسخ ملف الجدول الأسبوعي (xlsx) ويُنشئ نسخة محدّثة بنفس الشكل والبيانات تماماً.

المبدأ: الملف المصدر هو القالب — نُصلح تلف الـ stylesheet ثم نحفظ بتاريخ جديد.

الاستخدام:
  python import_weekly_schedule.py
  python import_weekly_schedule.py --xlsx "مسار_الملف.xlsx"
  python import_weekly_schedule.py --date "2026-05-26"
  python import_weekly_schedule.py --output "جدول_الاسبوع2.xlsx"
"""
import os
import sys
import re
import zipfile
import shutil
import argparse

# ── إعداد الـ encoding ────────────────────────────────────────────────────────
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── المسارات الافتراضية ───────────────────────────────────────────────────────
DEFAULT_XLSX = os.path.join(
    os.path.expanduser('~'), 'Downloads',
    'جدول الخاص بجميع المركبات والسائقين في الفرع وذلك للأسبوع الأول من شهر مايو ^. خالد الغامدي.xlsx'
)
DEFAULT_OUTPUT = os.path.join(BASE_DIR, 'جدول_الأسبوعي_مايو_2026.xlsx')
DEFAULT_DATE   = '2026-05-19'

# ── إصلاح ملف xlsx التالف ────────────────────────────────────────────────────
def repair_xlsx_to(src_path, dst_path):
    """ينسخ ويُصلح مشاكل الـ stylesheet ثم يحفظ في dst_path."""
    with zipfile.ZipFile(src_path, 'r') as zin:
        file_contents = {name: zin.read(name) for name in zin.namelist()}

    needs_fix = False
    if 'xl/styles.xml' in file_contents:
        xml = file_contents['xl/styles.xml'].decode('utf-8', errors='replace')
        families = re.findall(r'<family val="(\d+)"/>', xml)
        if any(int(v) > 14 for v in families):
            needs_fix = True

    if needs_fix:
        xml = file_contents['xl/styles.xml'].decode('utf-8', errors='replace')
        xml = re.sub(
            r'<family val="(\d+)"/>',
            lambda m: f'<family val="{min(int(m.group(1)), 14)}"/>',
            xml
        )
        xml = re.sub(r'<sz val="0"', '<sz val="11"', xml)
        file_contents['xl/styles.xml'] = xml.encode('utf-8')

    tmp = dst_path + '.__writing__.zip'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in file_contents.items():
            zout.writestr(name, data)
    os.replace(tmp, dst_path)
    return needs_fix

# ── قراءة وإحصاء البيانات ────────────────────────────────────────────────────
def read_stats(xlsx_path):
    """يقرأ الملف ويُعيد إحصائيات مختصرة بدون فتح كامل."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    records = []
    title   = ''
    date_val = ''

    for r_idx, row in enumerate(ws.rows, start=1):
        vals = [cell.value for cell in row]

        # عنوان الجدول (صف 7)
        if r_idx == 7:
            for v in vals:
                if v and 'جدول' in str(v):
                    title = str(v)
                    break
            for v in vals:
                if v and re.search(r'\d{4}', str(v)):
                    date_val = str(v)
                    break

        # بيانات السائقين (صف 9+)
        if r_idx >= 9:
            name = vals[3] if len(vals) > 3 else None  # col D = index 3
            serial = vals[1] if len(vals) > 1 else None  # col B = index 1
            if name and str(name).strip() and str(name).strip() != 'nan':
                if serial is not None and str(serial).strip().isdigit():
                    records.append({
                        'serial': str(serial).strip(),
                        'name':   str(name).strip(),
                        'plate':  str(vals[6]).strip() if len(vals) > 6 and vals[6] else '',
                        'type':   str(vals[8]).strip() if len(vals) > 8 and vals[8] else '',
                        'job':    str(vals[5]).strip() if len(vals) > 5 and vals[5] else '',
                    })

        if r_idx > 200:
            break

    wb.close()
    return records, title, date_val

# ── تحديث التاريخ في الملف ───────────────────────────────────────────────────
def update_date_in_file(xlsx_path, new_date):
    """يُحدّث التاريخ في الملف (الصف 7، عمود C)."""
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # ابحث عن خلية التاريخ في الصف 7
    date_updated = False
    for c in range(1, 20):
        cell = ws.cell(7, c)
        if isinstance(cell, MergedCell):
            continue
        v = cell.value
        if v and re.search(r'\d{4}', str(v)):
            # تحقق إذا كانت تاريخ فقط وليست عنوان
            if re.match(r'^\d{4}[/-]\d{2}[/-]\d{2}$', str(v).strip()):
                cell.value = new_date
                date_updated = True
                print(f"  تحديث التاريخ: {v} → {new_date} (الخلية {cell.coordinate})")
                break

    # إذا لم نجد تاريخاً في الصف 7، ابحث في C7
    if not date_updated:
        cell = ws.cell(7, 3)
        if not isinstance(cell, MergedCell):
            cell.value = new_date
            print(f"  كتابة التاريخ في C7: {new_date}")

    wb.save(xlsx_path)
    wb.close()

# ── الدالة الرئيسية ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Weekly Schedule Copier')
    parser.add_argument('--xlsx',   default=DEFAULT_XLSX,   help='مسار ملف xlsx المصدر')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='مسار الملف الناتج')
    parser.add_argument('--date',   default=DEFAULT_DATE,   help='تاريخ الجدول (YYYY-MM-DD)')
    parser.add_argument('--no-date-update', action='store_true',
                        help='لا تُحدّث التاريخ (انسخ كما هو)')
    args = parser.parse_args()

    LINE = '─' * 55
    print(LINE)
    print(' نسخ الجدول الأسبوعي — بن زومة الدولية')
    print(LINE)

    # التحقق
    if not os.path.exists(args.xlsx):
        print(f"[ERROR] الملف غير موجود: {args.xlsx}")
        sys.exit(1)

    src_kb = os.path.getsize(args.xlsx) / 1024
    print(f"\n  المصدر : {os.path.basename(args.xlsx)} ({src_kb:.0f} KB)")
    print(f"  الناتج : {os.path.basename(args.output)}")
    print(f"  التاريخ: {args.date}")

    # 1. إصلاح ونسخ
    print(f"\n[1] نسخ وإصلاح الملف...")
    fixed = repair_xlsx_to(args.xlsx, args.output)
    if fixed:
        print("  [تصحيح] تم إصلاح تنسيق الخطوط تلقائياً")
    else:
        print("  الملف سليم — تم النسخ مباشرة")

    # 2. تحديث التاريخ
    if not args.no_date_update:
        print(f"\n[2] تحديث التاريخ...")
        update_date_in_file(args.output, args.date)
    else:
        print(f"\n[2] التاريخ غير مُحدَّث (--no-date-update)")

    # 3. قراءة الإحصائيات
    print(f"\n[3] التحقق من البيانات...")
    records, title, date_val = read_stats(args.output)

    out_kb = os.path.getsize(args.output) / 1024

    # عرض عينة
    main_r  = [r for r in records if r.get('job') in ('موصل', 'موزع', '')]
    lorry   = sum(1 for r in records if 'لوري' in r.get('type', ''))
    dina    = sum(1 for r in records if 'دينا' in r.get('type', ''))

    print(f"\n  عينة من السائقين:")
    for rec in records[:4]:
        print(f"  [{rec['serial']:>2}] {rec['name']:<28} | {rec['plate']:<14} | {rec['type']}")

    print(f"\n{LINE}")
    print(f" ✅ تمّ بنجاح!")
    print(f"  السائقون الكلي    : {len(records)}")
    print(f"  لوري               : {lorry}")
    print(f"  دينا               : {dina}")
    print(f"  التاريخ المُدرج   : {args.date}")
    print(f"  الملف الناتج      : {os.path.basename(args.output)} ({out_kb:.0f} KB)")
    print(f"  الموقع            : {args.output}")
    print(LINE)


if __name__ == '__main__':
    main()
