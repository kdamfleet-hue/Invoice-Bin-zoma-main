import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New'

def print_sheet(ws, num_rows=15):
    print(f'Sheet: {ws.title} | Max Row: {ws.max_row} | Max Col: {ws.max_column}')
    for r in range(1, min(ws.max_row+1, num_rows+1)):
        row_data = []
        for c in range(1, ws.max_column+1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_data.append(f'C{c}: {val}')
        if row_data:
            print(f'  Row {r}: | ' + ' | '.join(row_data) + ' |')

files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and 'دينا' in f]

for f in files:
    fpath = os.path.join(folder, f)
    print(f'\n{"="*60}\nFILE: {f}\n{"="*60}')
    if not os.path.exists(fpath):
        print("FILE NOT FOUND")
        continue
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sn in wb.sheetnames:
            print_sheet(wb[sn])
    except Exception as e:
        print(f"Error reading {f}: {e}")


