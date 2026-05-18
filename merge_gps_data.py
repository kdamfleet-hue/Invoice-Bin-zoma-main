"""
merge_gps_data.py
-----------------
يدمج بيانات GPS من الملف المصدر إلى ملف الكشف الأساسي عن طريق مطابقة رقم اللوحة.
"""
import os
import sys
# Fix Windows console encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ── مسارات الملفات ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_SRC    = os.path.join(BASE_DIR, "GPS_source_data.xlsx")
FILE_TARGET = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026-1-26 - Copy.xlsx")
FILE_OUTPUT = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026_محدث.xlsx")

# ── التحقق من وجود الملفات ───────────────────────────────────────────────────
for f in [FILE_SRC, FILE_TARGET]:
    if not os.path.exists(f):
        print(f"[ERROR] الملف غير موجود: {f}")
        sys.exit(1)

# ── تطبيع رقم اللوحة ─────────────────────────────────────────────────────────
def normalize_plate(plate):
    if plate is None or (isinstance(plate, float) and pd.isna(plate)):
        return ""
    plate = str(plate).strip()
    plate = re.sub(r'\s+', '', plate)
    ar = "٠١٢٣٤٥٦٧٨٩"
    en = "0123456789"
    for a, e in zip(ar, en):
        plate = plate.replace(a, e)
    digits  = "".join(re.findall(r'\d+', plate))
    letters = "".join(re.findall(r'[^\d]+', plate))
    return digits + letters

# ── كتابة آمنة تدعم الخلايا المدمجة ─────────────────────────────────────────
def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
    else:
        cell.value = value

# ── تحميل البيانات ───────────────────────────────────────────────────────────
print("تحميل ملف المصدر...")
df_src = pd.read_excel(FILE_SRC, header=3)

# التحقق من الأعمدة المطلوبة
required_cols = ['رقم اللوحة', 'رقم الهيكل', 'الرقم التسلسلي',
                 'سنة الصنع', 'الطراز', 'الماركة', 'نوع التسجيل', 'الفرع']
missing = [c for c in required_cols if c not in df_src.columns]
if missing:
    print(f"[ERROR] أعمدة ناقصة في المصدر: {missing}")
    print(f"الأعمدة الموجودة: {list(df_src.columns)}")
    sys.exit(1)

df_src['norm_plate'] = df_src['رقم اللوحة'].apply(normalize_plate)

# بناء جدول البحث
lookup = {}
for _, row in df_src.iterrows():
    norm = row['norm_plate']
    if norm:
        lookup[norm] = row

print(f"تم تحميل {len(lookup)} سجل من المصدر.")

# ── تحميل ملف الهدف مع الحفاظ على التنسيق ──────────────────────────────────
print("تحميل ملف الهدف...")
wb = load_workbook(FILE_TARGET)
ws = wb.active

# خريطة الأعمدة (اسم العمود في المصدر → رقم العمود في الهدف)
COL_MAP = {
    'رقم الهيكل':    1,   # A
    'الرقم التسلسلي': 2,  # B
    'سنة الصنع':     3,   # C
    'الطراز':        4,   # D
    'الماركة':       5,   # E
    'نوع التسجيل':   6,   # F
    'الفرع':         7,   # G
}
PLATE_COL = 9  # I

# ── المطابقة والكتابة ─────────────────────────────────────────────────────────
match_count  = 0
update_count = 0

for r in range(6, ws.max_row + 1):
    plate_val = ws.cell(row=r, column=PLATE_COL).value
    if not plate_val:
        continue
    norm = normalize_plate(plate_val)
    if norm not in lookup:
        continue

    src = lookup[norm]
    match_count += 1
    for src_col, col_idx in COL_MAP.items():
        new_val = src.get(src_col)
        if new_val is not None and not (isinstance(new_val, float) and pd.isna(new_val)):
            if str(new_val).strip() not in ("", "nan"):
                safe_write(ws, r, col_idx, new_val)
                update_count += 1

print(f"تطابق: {match_count} مركبة | تحديث: {update_count} خلية")

wb.save(FILE_OUTPUT)
print(f"تم الحفظ: {FILE_OUTPUT}")


