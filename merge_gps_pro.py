"""
merge_gps_pro.py
----------------
يدمج بيانات GPS باستخدام مطابقة رقم اللوحة مع الحفاظ على التنسيق الكامل.
"""
import os
import sys
# Fix Windows console encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ── مسارات الملفات ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_SRC    = os.path.join(BASE_DIR, "GPS_source_data.xlsx")
FILE_TARGET = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026-1-26 - Copy.xlsx")
FILE_OUTPUT = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026_محدث_احترافي.xlsx")

# ── التحقق من وجود الملفات ───────────────────────────────────────────────────
for f in [FILE_SRC, FILE_TARGET]:
    if not os.path.exists(f):
        print(f"[ERROR] الملف غير موجود: {f}")
        sys.exit(1)

# ── تطبيع رقم اللوحة ─────────────────────────────────────────────────────────
def normalize_plate(plate):
    if plate is None or (isinstance(plate, float) and pd.isna(plate)):
        return ""
    plate = str(plate).strip()
    plate = re.sub(r'\s+', '', plate)
    ar = "٠١٢٣٤٥٦٧٨٩"
    en = "0123456789"
    for a, e in zip(ar, en):
        plate = plate.replace(a, e)
    digits  = "".join(re.findall(r'\d+', plate))
    letters = "".join(re.findall(r'[^\d]+', plate))
    return digits + letters

# ── كتابة آمنة تدعم الخلايا المدمجة ─────────────────────────────────────────
def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
    else:
        cell.value = value

# ── تحميل بيانات المصدر ──────────────────────────────────────────────────────
print("تحميل ملف المصدر...")
df_src = pd.read_excel(FILE_SRC, header=3)
df_src['norm_plate'] = df_src['رقم اللوحة'].apply(normalize_plate)

# بناء جدول البحث
lookup = {row['norm_plate']: row for _, row in df_src.iterrows() if row['norm_plate']}
print(f"تم تحميل {len(lookup)} سجل من المصدر.")

# ── تحميل ملف الهدف ──────────────────────────────────────────────────────────
print("تحميل ملف الهدف مع الحفاظ على التنسيق...")
wb = load_workbook(FILE_TARGET)
ws = wb.active

# خريطة الأعمدة في ملف الهدف (1-based)
PLATE_COL  = 9   # I
COL_MAP = [
    (1, 'رقم الهيكل'),
    (2, 'الرقم التسلسلي'),
    (3, 'سنة الصنع'),
    (4, 'الطراز'),
    (5, 'الماركة'),
    (6, 'نوع التسجيل'),
    (7, 'الفرع'),
]

# ── المطابقة والتحديث ────────────────────────────────────────────────────────
match_count  = 0
update_count = 0

for r in range(6, ws.max_row + 1):
    plate_val = ws.cell(row=r, column=PLATE_COL).value
    if not plate_val:
        continue
    norm = normalize_plate(plate_val)
    if norm not in lookup:
        continue

    src = lookup[norm]
    match_count += 1

    for col_idx, src_col in COL_MAP:
        if src_col not in df_src.columns:
            continue
        new_val = src.get(src_col)
        if new_val is not None and not (isinstance(new_val, float) and pd.isna(new_val)):
            if str(new_val).strip() not in ("", "nan"):
                safe_write(ws, r, col_idx, new_val)
                update_count += 1

print(f"تطابق: {match_count} مركبة | تحديث: {update_count} خلية")
wb.save(FILE_OUTPUT)
print(f"تم الحفظ: {FILE_OUTPUT}")
