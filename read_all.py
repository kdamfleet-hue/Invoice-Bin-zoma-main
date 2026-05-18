# -*- coding: utf-8 -*-
import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy\New'

for f in os.listdir(folder):
    if not f.endswith('.xlsx'): continue
    fpath = os.path.join(folder, f)
    print(f'\n{"="*60}')
    print(f'FILE: {f}')
    print(f'{"="*60}')
    wb = openpyxl.load_workbook(fpath)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'\nSheet: {sn} | Rows: {ws.max_row} | Cols: {ws.max_column}')
        for r in range(1, min(ws.max_row+1, 55)):
            rd = []
            for c in range(1, ws.max_column+1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    rd.append(f'C{c}={v}')
            sep = ' | '
            if rd:
                print(f'  R{r}: {sep.join(rd)}')
