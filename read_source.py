# -*- coding: utf-8 -*-
import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy'

fpath = None
for f in os.listdir(folder):
    if f.endswith('.xlsx') and 'بيانات' in f:
        fpath = os.path.join(folder, f)
        break

print(f'File: {os.path.basename(fpath)}')
wb = openpyxl.load_workbook(fpath)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\nSheet: {sheet_name}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    print()
    for r in range(1, min(ws.max_row + 1, 80)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_data.append(f'C{c}={val}')
        sep = ' | '
        if row_data:
            print(f'  Row {r}: {sep.join(row_data)}')
        else:
            print(f'  Row {r}: (empty)')


