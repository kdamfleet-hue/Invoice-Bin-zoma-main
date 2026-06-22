# -*- coding: utf-8 -*-
"""Extract driver name, plate, iqama from all specified files and rebuild database."""
import openpyxl, pdfplumber, sqlite3, sys, io, os, json, re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
BASE = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Downloads\Invoice-Bin-zoma-main\New'

all_drivers = {}  # iqama -> {name, plate, car, phone, ...}

def add_driver(name, iqama='', plate='', car='', phone='', source=''):
    """Add or merge driver data."""
    name = (name or '').strip()
    iqama = (iqama or '').strip()
    plate = (plate or '').strip()
    car = (car or '').strip()
    phone = (phone or '').strip()
    
    if not name or len(name) < 2:
        return
    # Skip junk
    if any(j in name for j in ['ددع', 'مقر', 'ةحول', 'لادب']):
        return
    
    key = iqama if iqama and len(iqama) >= 8 else f'name:{name}'
    
    if key not in all_drivers:
        all_drivers[key] = {'name': name, 'iqama': iqama, 'plate': plate, 'car': car, 'phone': phone, 'sources': [source]}
    else:
        d = all_drivers[key]
        # Keep longest name
        if len(name) > len(d['name']):
            d['name'] = name
        # Fill missing fields
        if plate and not d['plate']:
            d['plate'] = plate
        if car and not d['car']:
            d['car'] = car
        if phone and not d['phone']:
            d['phone'] = phone
        if iqama and not d['iqama']:
            d['iqama'] = iqama
        if source not in d['sources']:
            d['sources'].append(source)

# ============================================================
# 1. all_drivers_merged.json
# ============================================================
print("=" * 60)
print("1. Reading all_drivers_merged.json...")
json_path = os.path.join(BASE, 'all_drivers_merged.json')
if os.path.exists(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        jdata = json.load(f)
    if isinstance(jdata, dict):
        for name, data in jdata.items():
            add_driver(name, data.get('iqama',''), data.get('plate',''), data.get('car',''), data.get('phone',''), 'JSON')
    elif isinstance(jdata, list):
        for d in jdata:
            add_driver(d.get('name',''), d.get('iqama',''), d.get('plate',''), d.get('car',''), d.get('phone',''), 'JSON')
    print(f"   Loaded {len(jdata)} entries")

# ============================================================
# 2. التفاويض 2026-05-09.xlsx (plate, iqama, phone)
# ============================================================
print("\n2. Reading التفاويض...")
tf_path = os.path.join(BASE, 'التفاويض 2026-05-09.xlsx')
if os.path.exists(tf_path):
    wb = openpyxl.load_workbook(tf_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            plate = str(ws.cell(row=r, column=1).value or '').strip()
            iqama = str(ws.cell(row=r, column=3).value or '').strip()
            phone = str(ws.cell(row=r, column=4).value or '').strip()
            if iqama and len(iqama) >= 8:
                # Find name by iqama in existing data
                if iqama in all_drivers:
                    if plate and not all_drivers[iqama]['plate']:
                        all_drivers[iqama]['plate'] = plate
                    if phone and not all_drivers[iqama]['phone']:
                        all_drivers[iqama]['phone'] = phone
    wb.close()
    print("   Done")

# ============================================================
# 3. vehicle_list_excel_without_qr-code_2026-05-11.xlsx
# ============================================================
print("\n3. Reading vehicle_list...")
vl_path = os.path.join(BASE, 'vehicle_list_excel_without_qr-code_2026-05-11.xlsx')
if os.path.exists(vl_path):
    wb = openpyxl.load_workbook(vl_path, data_only=True)
    ws = wb.active
    # Detect columns
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=1, column=c).value or '').strip()
        headers[c] = v
    print(f"   Headers: {headers}")
    
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in headers.items():
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row[h] = str(v).strip()
        
        plate = ''
        iqama = ''
        name = ''
        car = ''
        
        for h, v in row.items():
            hl = h.lower().replace(' ','')
            if 'plate' in hl or 'لوح' in h: plate = v
            if 'iqama' in hl or 'اقام' in h or 'إقام' in h or 'driver' in hl.lower(): 
                if v.isdigit() and len(v) >= 8: iqama = v
            if 'user' in hl or 'مستخدم' in h or 'اسم' in h: name = v
            if 'brand' in hl or 'ماركة' in h or 'نوع' in h: car = v
        
        if name or iqama:
            add_driver(name, iqama, plate, car, '', 'VehicleList')
    wb.close()
    print(f"   Total drivers now: {len(all_drivers)}")

# ============================================================
# 4. Processed_GPS_Vehicles (1).xlsx
# ============================================================
print("\n4. Reading GPS Vehicles...")
gps_path = os.path.join(BASE, 'Processed_GPS_Vehicles (1).xlsx')
if os.path.exists(gps_path):
    wb = openpyxl.load_workbook(gps_path, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=1, column=c).value or '').strip()
        headers[c] = v
    print(f"   Headers: {headers}")
    wb.close()

# ============================================================
# 5. كشف_اجهزة_GPS_محدث (2).xlsx
# ============================================================
print("\n5. Reading GPS devices...")
gps2_path = os.path.join(BASE, 'كشف_اجهزة_GPS_محدث (2).xlsx')
if os.path.exists(gps2_path):
    wb = openpyxl.load_workbook(gps2_path, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, min(ws.max_column + 1, 20)):
        v = str(ws.cell(row=1, column=c).value or '').strip()
        headers[c] = v
    print(f"   Headers: {headers}")
    
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in headers.items():
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row[h] = str(v).strip()
        
        plate = ''
        name = ''
        for h, v in row.items():
            if 'لوح' in h or 'plate' in h.lower(): plate = v
            if 'سائق' in h or 'اسم' in h or 'مستخدم' in h: name = v
        
        if name and plate:
            add_driver(name, '', plate, '', '', 'GPS')
    wb.close()

# ============================================================
# 6. PDFs - اخر تحديث لسيارات الواري
# ============================================================
print("\n6. Reading PDF: اخر تحديث لسيارات الواري...")
pdf1 = os.path.join(BASE, 'اخر تحديث لسيارات الواري تاريخ 11-5.pdf')
if os.path.exists(pdf1):
    with pdfplumber.open(pdf1) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                for row in table[1:]:
                    if not row or len(row) < 5:
                        continue
                    clean = [str(c).strip() if c else '' for c in row]
                    # Try to find iqama (10-digit number)
                    iqama = ''
                    plate = ''
                    name = ''
                    car = ''
                    phone = ''
                    for val in clean:
                        if val.isdigit() and len(val) == 10 and val.startswith('2'):
                            iqama = val
                        elif val.startswith('05') and len(val) >= 10:
                            phone = val
                        elif re.match(r'^[\u0621-\u064A\s]+\d+$', val.replace(' ','')):
                            plate = val
                        elif len(val) > 5 and all(c in '\u0621\u0622\u0623\u0624\u0625\u0626\u0627\u0628\u0629\u062a\u062b\u062c\u062d\u062e\u062f\u0630\u0631\u0632\u0633\u0634\u0635\u0636\u0637\u0638\u0639\u063a\u0641\u0642\u0643\u0644\u0645\u0646\u0647\u0648\u064a \u0020' for c in val):
                            if not name or len(val) > len(name):
                                name = val
                    
                    if name or iqama:
                        add_driver(name, iqama, plate, car, phone, 'PDF-لواري')

# ============================================================
# 7. PDF - اخر تحديث للمركبات
# ============================================================
print("\n7. Reading PDF: اخر تحديث للمركبات...")
pdf2 = os.path.join(BASE, 'اخر تحديث للمركبات من تم تاريخ 511.pdf')
if os.path.exists(pdf2):
    with pdfplumber.open(pdf2) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                for row in table[1:]:
                    if not row or len(row) < 5:
                        continue
                    clean = [str(c).strip() if c else '' for c in row]
                    iqama = ''
                    plate = ''
                    name = ''
                    car = ''
                    phone = ''
                    for val in clean:
                        if val.isdigit() and len(val) == 10 and val.startswith('2'):
                            iqama = val
                        elif val.startswith('05') and len(val) >= 10:
                            phone = val
                    
                    if iqama:
                        add_driver('', iqama, '', '', phone, 'PDF-مركبات')

# ============================================================
# 8. PDF - جدول تحديث اسبوعي
# ============================================================
print("\n8. Reading PDF: جدول تحديث اسبوعي...")
pdf3 = os.path.join(BASE, 'جدول تحديث اسبوعي.pdf')
if os.path.exists(pdf3):
    with pdfplumber.open(pdf3) as pdf:
        for pi, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                # Print first few rows to understand structure
                if pi <= 2:
                    print(f"   Page {pi} table ({len(table)} rows), sample:")
                    for ri, row in enumerate(table[:3]):
                        clean = [str(c).strip()[:30] if c else '' for c in row]
                        print(f"     R{ri}: {clean}")

# ============================================================
# FINAL: Build clean database
# ============================================================
print(f"\n{'='*60}")
print(f"Total unique entries: {len(all_drivers)}")

# Clean and deduplicate
final = []
seen_iqamas = set()
for key, d in all_drivers.items():
    iq = d.get('iqama', '')
    if iq and iq in seen_iqamas:
        continue
    if iq:
        seen_iqamas.add(iq)
    
    name = d.get('name', '')
    if not name or len(name) < 2:
        continue
    
    final.append(d)

final.sort(key=lambda x: x.get('name', ''))

print(f"Final clean drivers: {len(final)}")
print(f"\n{'='*60}")
print(f"{'#':>3} | {'الاسم':<35} | {'رقم الإقامة':<12} | {'رقم اللوحة':<14} | {'نوع السيارة':<25} | {'الجوال':<12}")
print(f"{'-'*3}-+-{'-'*35}-+-{'-'*12}-+-{'-'*14}-+-{'-'*25}-+-{'-'*12}")

for i, d in enumerate(final, 1):
    print(f"{i:3d} | {d['name']:<35} | {(d['iqama'] or '-'):<12} | {(d['plate'] or '-'):<14} | {(d['car'] or '-'):<25} | {(d['phone'] or '-'):<12}")

# Save to JSON
out_path = os.path.join(BASE, 'final_drivers_database.json')
with open(out_path, 'w', encoding='utf-8') as f:
    # Remove 'sources' for cleaner output
    clean_final = [{k:v for k,v in d.items() if k != 'sources'} for d in final]
    json.dump(clean_final, f, ensure_ascii=False, indent=2)
print(f"\nSaved to: {out_path}")

# Update SQLite database
db_path = os.path.join(os.path.dirname(BASE), 'database.sqlite')
if os.path.exists(db_path):
    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    
    updated = 0
    added = 0
    
    for d in final:
        iqama = d.get('iqama', '')
        name = d['name']
        plate = d.get('plate', '')
        car = d.get('car', '')
        phone = d.get('phone', '')
        
        # Try find by iqama
        existing = None
        if iqama:
            existing = db.execute("SELECT * FROM drivers WHERE iqama = ?", (iqama,)).fetchone()
        if not existing:
            existing = db.execute("SELECT * FROM drivers WHERE name = ?", (name,)).fetchone()
        
        if existing:
            changes = {}
            if plate and (not existing['plate'] or existing['plate'] in ('', 'None')):
                changes['plate'] = plate
            if car and (not existing['car'] or existing['car'] in ('', 'None')):
                changes['car'] = car
            if phone and (not existing['phone'] or existing['phone'] in ('', 'None')):
                changes['phone'] = phone
            if iqama and (not existing['iqama'] or existing['iqama'] in ('', 'None')):
                changes['iqama'] = iqama
            
            if changes:
                sets = ', '.join(f"{k} = ?" for k in changes.keys())
                vals = list(changes.values()) + [existing['id']]
                db.execute(f"UPDATE drivers SET {sets} WHERE id = ?", vals)
                updated += 1
        # Don't add new - just update existing
    
    db.commit()
    
    # Final stats
    total = db.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
    no_plate = db.execute("SELECT COUNT(*) FROM drivers WHERE plate IS NULL OR plate = '' OR plate = 'None'").fetchone()[0]
    no_car = db.execute("SELECT COUNT(*) FROM drivers WHERE car IS NULL OR car = '' OR car = 'None'").fetchone()[0]
    no_iqama = db.execute("SELECT COUNT(*) FROM drivers WHERE iqama IS NULL OR iqama = ''").fetchone()[0]
    no_phone = db.execute("SELECT COUNT(*) FROM drivers WHERE phone IS NULL OR phone = '' OR phone = 'None'").fetchone()[0]
    
    print(f"\nDatabase updated: {updated} drivers updated")
    print(f"DB Stats: Total={total} | No plate={no_plate} | No car={no_car} | No iqama={no_iqama} | No phone={no_phone}")
    
    # Print final DB state
    print(f"\n{'='*60}")
    print(f"FINAL DATABASE STATE:")
    print(f"{'='*60}")
    drivers = db.execute("SELECT * FROM drivers ORDER BY name").fetchall()
    for i, d in enumerate(drivers, 1):
        status = '✅' if d['plate'] and d['car'] and d['iqama'] and d['phone'] else '🔶'
        print(f"  {status} {i:2d}. {d['name']:<35} | 🪪 {(d['iqama'] or '-'):<12} | 🚗 {(d['plate'] or '-'):<14} | 🚛 {(d['car'] or '-'):<25} | 📱 {(d['phone'] or '-')}")
    
    db.close()

