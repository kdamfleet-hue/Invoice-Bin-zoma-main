# -*- coding: utf-8 -*-
import sys, os, re, copy
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

folder = r'c:\Users\Administrator.MADEBYA-NBLPM4O\Desktop\جدول تحسين - Copy'
new_folder = os.path.join(folder, 'New')

def normalize_plate(p):
    if not p: return ''
    s = re.sub(r'\s+', '', str(p).strip())
    return s.replace('أ','ا').replace('إ','ا').replace('آ','ا')

def parse_date(val):
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if '00:00:00' in s: s = s.split(' ')[0]
    return s

target_file = [f for f in os.listdir(folder) if 'تحديث' in f and 'الدمام' in f][0]
driver_file = [f for f in os.listdir(new_folder) if 'بيانات السائ' in f.replace('ئ', 'ئ') or 'بيانات' in f][0]
vehicle_file = [f for f in os.listdir(new_folder) if 'دينا' in f and 'لوري' in f][0]

target_path = os.path.join(folder, "تحديث المركبات والسائقين-الدمام2026.xlsx")
driver_path = os.path.join(new_folder, driver_file)
vehicle_path = os.path.join(new_folder, vehicle_file)

# 1. Load Driver Info
emp_map = {} 
plate_to_emp = {} 
name_to_emp = {} 

wb_drv = openpyxl.load_workbook(driver_path, data_only=True)
ws_drv = wb_drv.active
for r in range(9, ws_drv.max_row + 1):
    emp_id = str(ws_drv.cell(row=r, column=2).value).strip()
    name = str(ws_drv.cell(row=r, column=3).value).strip()
    iqama = str(ws_drv.cell(row=r, column=4).value).strip()
    plate = ws_drv.cell(row=r, column=7).value
    card_exp = parse_date(ws_drv.cell(row=r, column=9).value)
    
    if emp_id and emp_id != 'None':
        emp_map[emp_id] = {'iqama': iqama if iqama != 'None' else '', 'card_exp': card_exp}
        name_to_emp[name] = emp_id
    if plate:
        plate_to_emp[normalize_plate(plate)] = emp_id

# 2. Load Vehicle Info
veh_map = {} 
wb_veh = openpyxl.load_workbook(vehicle_path, data_only=True)
ws_veh = wb_veh.active
for r in range(6, ws_veh.max_row + 1):
    plate = ws_veh.cell(row=r, column=11).value
    if plate and str(plate).strip() and str(plate) != 'رقم اللوحة':
        np = normalize_plate(plate)
        istimara = parse_date(ws_veh.cell(row=r, column=4).value)
        tasgheel = parse_date(ws_veh.cell(row=r, column=5).value)
        serial = str(ws_veh.cell(row=r, column=15).value).strip()
        serial = '' if serial == 'None' else serial
        fahs_m = parse_date(ws_veh.cell(row=r, column=2).value)
        fahs_h = parse_date(ws_veh.cell(row=r, column=3).value)
        fahs = fahs_m if fahs_m and 'لا' not in fahs_m and 'سيارة' not in fahs_m else fahs_h
        if not fahs or fahs == 'None': fahs = ''
        veh_map[np] = {
            'serial': serial, 'istimara': istimara, 'fahs': fahs, 'tasgheel': tasgheel
        }

# 3. Rewrite Target Excel
wb_target = openpyxl.load_workbook(target_path)
ws_target = wb_target.active

def get_new_col(r, c):
    if 6 <= r <= 34:
        if c <= 10: return c
        if c >= 11: return c + 5 # 11 -> 16
    elif 39 <= r <= 47:
        if c <= 2: return c
        if c >= 3: return c + 2 # 3->5, 7->9
    return c

# Save all cell data
cells_data = {}
for r in range(1, ws_target.max_row + 1):
    for c in range(1, ws_target.max_column + 1):
        cell = ws_target.cell(row=r, column=c)
        if cell.value is not None or cell.has_style:
            cells_data[(r, c)] = {
                'value': cell.value,
                'font': copy.copy(cell.font) if cell.has_style else None,
                'border': copy.copy(cell.border) if cell.has_style else None,
                'fill': copy.copy(cell.fill) if cell.has_style else None,
                'alignment': copy.copy(cell.alignment) if cell.has_style else None,
                'number_format': copy.copy(cell.number_format) if cell.has_style else None,
            }

# Save and transform merged ranges
merged_ranges = list(ws_target.merged_cells.ranges)
new_merges = []
for mr in merged_ranges:
    ws_target.unmerge_cells(str(mr))
    
    min_r, min_c, max_r, max_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
    
    new_min_c = get_new_col(min_r, min_c)
    new_max_c = get_new_col(max_r, max_c)
    
    # Special fix for ranges that span the whole table
    if max_c == 11 and min_r not in range(6, 35) and min_r not in range(39, 48):
        new_max_c = 16
    
    # Special fix for vacation table "الملاحظات"
    if 39 <= min_r <= 47 and max_c == 11:
        new_max_c = 16
        
    # Special fix for summary table
    if 50 <= min_r <= 51 and min_c == 7 and max_c == 11:
        new_max_c = 16
        
    # Special fix for footer F55:K55 -> F55:P55
    if min_r >= 55 and max_c == 11:
        new_max_c = 16

    new_merges.append((min_r, new_min_c, max_r, new_max_c))

# Clear sheet
for r in range(1, ws_target.max_row + 1):
    for c in range(1, ws_target.max_column + 1 + 5):
        cell = ws_target.cell(row=r, column=c)
        cell.value = None
        cell.font = openpyxl.styles.Font()
        cell.border = openpyxl.styles.Border()
        cell.fill = openpyxl.styles.PatternFill()

# Write shifted cells
for (old_r, old_c), data in cells_data.items():
    new_c = get_new_col(old_r, old_c)
    cell = ws_target.cell(row=old_r, column=new_c)
    cell.value = data['value']
    if data['font']: cell.font = data['font']
    if data['border']: cell.border = data['border']
    if data['fill']: cell.fill = data['fill']
    if data['alignment']: cell.alignment = data['alignment']
    if data['number_format']: cell.number_format = data['number_format']

def copy_cell_style(src_r, src_c, dst_r, dst_c):
    src = ws_target.cell(row=src_r, column=src_c)
    dst = ws_target.cell(row=dst_r, column=dst_c)
    dst.font = copy.copy(src.font)
    dst.border = copy.copy(src.border)
    dst.fill = copy.copy(src.fill)
    dst.alignment = copy.copy(src.alignment)

# Fill NEW columns in Main Table (Row 6 to 34)
new_main_headers = [
    "الرقم التسلسلي للمركبة", "تاريخ انتهاء الفحص الدوري",
    "تاريخ انتهاء رخصة السير (الاستمارة)", "تاريخ انتهاء بطاقة السائق",
    "تاريخ انتهاء بطاقة التشغيل"
]
for i, h in enumerate(new_main_headers):
    ws_target.cell(row=6, column=11+i, value=h)
    copy_cell_style(6, 16, 6, 11+i)

for r in range(7, 35):
    emp_id = str(ws_target.cell(row=r, column=2).value).strip()
    plate = str(ws_target.cell(row=r, column=6).value).strip()
    np = normalize_plate(plate)
    
    card_exp = ''
    serial = ''
    fahs = ''
    istimara = ''
    tasgheel = ''
    
    if emp_id in emp_map: card_exp = emp_map[emp_id]['card_exp']
    elif np in plate_to_emp and plate_to_emp[np] in emp_map: card_exp = emp_map[plate_to_emp[np]]['card_exp']
        
    if np in veh_map:
        serial = veh_map[np]['serial']
        fahs = veh_map[np]['fahs']
        istimara = veh_map[np]['istimara']
        tasgheel = veh_map[np]['tasgheel']
    
    vals = [serial, fahs, istimara, card_exp, tasgheel]
    for i, v in enumerate(vals):
        ws_target.cell(row=r, column=11+i, value=v)
        copy_cell_style(r, 16, r, 11+i)

# Fill NEW columns in Vacation Table (Row 39 to 47)
ws_target.cell(row=39, column=3, value="الرقم الوظيفي")
copy_cell_style(39, 5, 39, 3)
ws_target.cell(row=39, column=4, value="رقم الإقامة")
copy_cell_style(39, 5, 39, 4)

for r in range(40, 48):
    name = str(ws_target.cell(row=r, column=2).value).strip()
    plate = str(ws_target.cell(row=r, column=5).value).strip()
    np = normalize_plate(plate)
    
    emp_id = ''
    iqama = ''
    
    if np in plate_to_emp: emp_id = plate_to_emp[np]
    else:
        for n, eid in name_to_emp.items():
            if name in n or n in name:
                emp_id = eid
                break
                
    if emp_id in emp_map: iqama = emp_map[emp_id]['iqama']
        
    ws_target.cell(row=r, column=3, value=emp_id)
    copy_cell_style(r, 5, r, 3)
    ws_target.cell(row=r, column=4, value=iqama)
    copy_cell_style(r, 5, r, 4)

# Apply Merges
for min_r, min_c, max_r, max_c in new_merges:
    if min_r == max_r and min_c == max_c: continue
    range_str = f"{openpyxl.utils.get_column_letter(min_c)}{min_r}:{openpyxl.utils.get_column_letter(max_c)}{max_r}"
    ws_target.merge_cells(range_str)

# Adjust column widths
widths = [
    3.5,  # 1 م
    13.2, # 2 الوظيفي
    22.0, # 3 الاسم
    18.0, # 4 الإقامة
    7.1,  # 5 الوظيفة
    12.1, # 6 اللوحة
    14.4, # 7 الموديل
    11.7, # 8 النوع
    10.0, # 9 الطبالي
    20.8, # 10 حمولة
    15.0, # 11 التسلسلي
    15.0, # 12 فحص
    15.0, # 13 استمارة
    15.0, # 14 بطاقة سائق
    15.0, # 15 كرت تشغيل
    33.7  # 16 الملاحظات
]
for i, w in enumerate(widths):
    ws_target.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w

# Fix image anchor if any
if ws_target._images:
    for img in ws_target._images:
        try:
            if img.anchor._from.col >= 10:
                img.anchor._from.col += 5
                img.anchor.to.col += 5
        except: pass

out_path = os.path.join(new_folder, "تحديث المركبات والسائقين-الدمام2026.xlsx")
wb_target.save(out_path)
print(f"✅ Success! File rewritten to match Image 1 completely: {out_path}")

