import os
import io
import base64
import requests
import openpyxl
import logging
import secrets
import time
import threading
from datetime import datetime, timezone
from urllib.parse import urlsplit
from flask import Blueprint, render_template, session, request, jsonify

from helpers import login_required, load_logo, blob_set, normalize_plate, _global_blob_get, _global_blob_set

logger = logging.getLogger("InvoiceApp")
gps_bp = Blueprint('gps', __name__)

# ── 360Locate (Location Solutions) FleetManagement API ──────────────────────────────────────
# Verified against the provider's own OpenAPI document (fleetmanagement-api-clust03 …/swagger):
#   POST /api/Authentication/RefreshToken  {refreshToken}          -> {token, refreshToken, tokenExpires, refreshTokenExpiryDate}
#   POST /api/Authentication/Login         {username, password, useRefreshToken} -> same shape
#   GET  /api/Asset                                                 -> [AssetDto{id, name, plateNumber, deviceId, brand, model, driver{name}, currentAddress{address, city}, …}]
#   POST /api/StateLookup  {deviceIds, driverIds, geofenceIds, previousLookupTimestamp}
#        -> {deviceStates: {deviceId: DeviceStateDto}, timestamp}
#   DeviceStateDto: currentPosition{latitude, longitude, speed, heading, updateTimestamp},
#        cellPosition{…}, calculatedIgnitionState{isIgnitionOn}, ignition{state},
#        idlingState{isCurrentlyIdling}, calculatedCommunicatingState{isCommunicating},
#        calculatedDeviceState{deviceState}, odometer{gpsOdometer}, communicationState{updateTimestamp}
#
# IMPORTANT — token rotation: RefreshToken returns a NEW refreshToken each time and the old one
# stops working. The previous implementation exchanged the env token on every poll and threw
# the replacement away, so the first poll consumed the token and every later poll got
# "RefreshToken invalid". The rotated token is now persisted (AppSetting blob gps_auth_state)
# and shared by all workers; the env value is only the seed. Username/password login is a
# second path; a raw bearer token is kept as a last-resort legacy mode.
GPS_USER = os.environ.get("GPS_USER", "")
GPS_PASS = os.environ.get("GPS_PASS", "")
GPS_API_BASE = os.environ.get(
    "GPS_API_BASE", "https://fleetmanagement-api-clust03.gpscockpit.com/api"
).rstrip("/")
GPS_DEVICES_URL = os.environ.get("GPS_DEVICES_URL", f"{GPS_API_BASE}/Asset")
GPS_STATE_URL = os.environ.get("GPS_STATE_URL", f"{GPS_API_BASE}/StateLookup")
GPS_REFRESH_URL = os.environ.get("GPS_REFRESH_URL", f"{GPS_API_BASE}/Authentication/RefreshToken")
GPS_LOGIN_URL = os.environ.get("GPS_LOGIN_URL", f"{GPS_API_BASE}/Authentication/Login")
GPS_PERMANENT_TOKEN = (os.environ.get("GPS_TOKEN") or os.environ.get("GPS_PERMANENT_TOKEN") or "").strip()
GPS_AUTH_SCHEME = os.environ.get("GPS_AUTH_SCHEME", "Bearer")
GPS_TIMEOUT = float(os.environ.get("GPS_TIMEOUT", "20"))
FLEET_CACHE_SECONDS = float(os.environ.get("GPS_CACHE_SECONDS", "5"))   # absorbs several tabs polling
ONLINE_WINDOW_SECONDS = 15 * 60   # a device with no communication flag counts as online if it reported within this window

_AUTH_STATE_KEY = "gps_auth_state"
_auth_lock = threading.Lock()
_auth_mem = {}                     # process mirror of the persisted auth state (never logged)
_diag = {"auth_source": None, "last_auth_error": None, "last_success_at": None, "last_error": None, "last_count": 0}
_fleet_cache = {"at": 0.0, "data": None}


class GpsAuthError(Exception):
    pass


def get_gps_token():
    return GPS_PERMANENT_TOKEN


def _gps_headers(token, scheme=None):
    return {
        "Authorization": f"{scheme or GPS_AUTH_SCHEME} {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def _parse_ts(value):
    """ISO-8601 (with or without Z) -> epoch seconds, or None."""
    if not value:
        return None
    try:
        s = str(value).strip().replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.timestamp()
    except (ValueError, TypeError):
        return None


def _gps_json(response, label):
    ctype = response.headers.get("Content-Type", "")
    if "application/json" not in ctype.lower():
        logger.warning("GPS %s returned non-JSON content type %s", label, ctype[:100])
        return None
    try:
        return response.json()
    except ValueError:
        logger.warning("GPS %s returned invalid JSON", label)
        return None


# ── auth state (persisted so the rotated refresh token survives restarts / other workers) ──
def _auth_state_load():
    global _auth_mem
    if _auth_mem:
        return dict(_auth_mem)
    try:
        stored = _global_blob_get(_AUTH_STATE_KEY)
        if isinstance(stored, dict):
            _auth_mem = dict(stored)
    except Exception:
        logger.warning("GPS auth state could not be read; starting from the env token")
    return dict(_auth_mem)


def _auth_state_save(state):
    global _auth_mem
    _auth_mem = dict(state)
    try:
        _global_blob_set(_AUTH_STATE_KEY, state)
    except Exception:
        logger.warning("GPS auth state could not be persisted (rotation kept in memory only)")


def _state_from_tokens(payload, source):
    access = payload.get("token") or payload.get("accessToken")
    if not access:
        return None
    return {
        "access": access,
        "access_exp": _parse_ts(payload.get("tokenExpires")) or (time.time() + 55 * 60),
        "refresh": payload.get("refreshToken") or "",
        "refresh_exp": _parse_ts(payload.get("refreshTokenExpiryDate")),
        "seed": GPS_PERMANENT_TOKEN[-8:] if GPS_PERMANENT_TOKEN else "",   # which env token this chain came from
        "source": source,
        "updated_at": time.time(),
    }


def _post_auth(url, body, label, request_id):
    """Returns (payload or None, short error message). Tokens are never logged."""
    try:
        r = requests.post(url, json=body, headers={"Accept": "application/json", "Content-Type": "application/json"}, timeout=GPS_TIMEOUT)
    except requests.Timeout:
        return None, f"{label}: timeout"
    except requests.ConnectionError:
        return None, f"{label}: connection error"
    except Exception as e:  # pragma: no cover
        return None, f"{label}: {type(e).__name__}"
    if r.status_code == 200:
        payload = _gps_json(r, label)
        if isinstance(payload, dict):
            return payload, ""
        return None, f"{label}: invalid JSON"
    msg = ""
    try:
        msg = (r.json() or {}).get("message") or ""
    except Exception:
        pass
    logger.warning("GPS %s rejected request_id=%s host=%s status=%s msg=%s",
                   label, request_id, urlsplit(url).netloc, r.status_code, msg[:80])
    return None, f"{label}: HTTP {r.status_code} {msg}".strip()


def _acquire_headers(request_id, force=False):
    """Headers carrying a valid access token, obtaining/refreshing one as needed.

    Order: cached access token -> persisted (rotated) refresh token -> env refresh token
    -> username/password login -> legacy raw bearer. Raises GpsAuthError with a message
    that is safe to show (no secrets)."""
    with _auth_lock:
        st = _auth_state_load()
        now = time.time()
        seed_now = GPS_PERMANENT_TOKEN[-8:] if GPS_PERMANENT_TOKEN else ""
        if st.get("seed") and seed_now and st["seed"] != seed_now:
            st = {}   # the operator set a NEW env token: drop the chain that came from the old one
        if not force and st.get("access") and (st.get("access_exp") or 0) - 30 > now:
            _diag["auth_source"] = st.get("source")
            return _gps_headers(st["access"], "Bearer")

        errors = []
        candidates = []
        if st.get("refresh") and ((st.get("refresh_exp") or now + 1) > now):
            candidates.append(("rotated refresh token", st["refresh"]))
        if GPS_PERMANENT_TOKEN and GPS_PERMANENT_TOKEN != st.get("refresh"):
            candidates.append(("env refresh token", GPS_PERMANENT_TOKEN))
        for label, rt in candidates:
            payload, err = _post_auth(GPS_REFRESH_URL, {"refreshToken": rt}, "token refresh", request_id)
            if payload:
                new_state = _state_from_tokens(payload, label)
                if new_state:
                    _auth_state_save(new_state)
                    _diag["auth_source"] = label
                    _diag["last_auth_error"] = None
                    return _gps_headers(new_state["access"], "Bearer")
            errors.append(f"{label} → {err}")

        if GPS_USER and GPS_PASS:
            payload, err = _post_auth(GPS_LOGIN_URL, {"username": GPS_USER, "password": GPS_PASS, "useRefreshToken": True}, "login", request_id)
            if payload:
                new_state = _state_from_tokens(payload, "username/password login")
                if new_state:
                    _auth_state_save(new_state)
                    _diag["auth_source"] = "username/password login"
                    _diag["last_auth_error"] = None
                    return _gps_headers(new_state["access"], "Bearer")
            errors.append(f"login → {err}")

        if GPS_PERMANENT_TOKEN and GPS_AUTH_SCHEME and not candidates:
            # Nothing to exchange and no login: legacy mode, send the env token as-is.
            _diag["auth_source"] = "legacy bearer"
            return _gps_headers(GPS_PERMANENT_TOKEN)

        _diag["last_auth_error"] = "; ".join(errors) or "no GPS credentials configured"
        raise GpsAuthError(_diag["last_auth_error"])


# ── fleet snapshot ──────────────────────────────────────────────────────────────────────────
_ASSET_PARAMS = {
    "isActive": "null",
    "includeGroups": "false",
    "uniqueDevices": "false",
    "includeGroupIds": "true",
    "IncludeHierarchyGroupIds": "false",
}


def _get_with_reauth(method, url, request_id, **kw):
    """One provider call; on 401 the access token is refreshed once and the call retried."""
    headers = _acquire_headers(request_id)
    r = requests.request(method, url, headers=headers, timeout=GPS_TIMEOUT, **kw)
    if r.status_code == 401:
        headers = _acquire_headers(request_id, force=True)
        r = requests.request(method, url, headers=headers, timeout=GPS_TIMEOUT, **kw)
    return r


def _vehicle_status(speed, ignition_on, idling, online):
    if not online:
        return "offline"
    if speed is not None and speed > 2:
        return "moving"
    if ignition_on and idling:
        return "idling"
    if ignition_on:
        return "engine_on"
    return "stopped"


def _build_vehicles(assets, states, now):
    by_device = {}
    for item in assets:
        if not isinstance(item, dict):
            continue
        asset = item.get("asset") if isinstance(item.get("asset"), dict) else item
        device_id = asset.get("deviceId") or item.get("deviceId")
        if device_id is None:
            continue
        driver = asset.get("driver") if isinstance(asset.get("driver"), dict) else {}
        addr = asset.get("currentAddress") if isinstance(asset.get("currentAddress"), dict) else {}
        by_device[str(device_id)] = {
            "id": asset.get("id", item.get("id")),
            "device_id": device_id,
            "name": asset.get("name") or item.get("name") or "مركبة",
            "plate": asset.get("plateNumber") or "",
            "brand": asset.get("brand") or "",
            "model": asset.get("model") or "",
            "asset_type": asset.get("assetTypeName") or "",
            "driver": (driver.get("name") or " ".join(x for x in (driver.get("firstName"), driver.get("lastName")) if x) or "").strip(),
            "address": ", ".join(x for x in (addr.get("address"), addr.get("city")) if x),
            "last_communication": asset.get("lastCommunication"),
        }

    vehicles = []
    for device_id, meta in by_device.items():
        state = states.get(device_id) if isinstance(states, dict) else None
        if not isinstance(state, dict):
            state = {}
        pos = state.get("currentPosition") if isinstance(state.get("currentPosition"), dict) else {}
        cell = state.get("cellPosition") if isinstance(state.get("cellPosition"), dict) else {}
        lat = pos.get("latitude") if pos.get("latitude") is not None else cell.get("latitude")
        lng = pos.get("longitude") if pos.get("longitude") is not None else cell.get("longitude")
        speed = pos.get("speed")
        try:
            speed = round(float(speed), 1) if speed is not None else None
        except (TypeError, ValueError):
            speed = None
        calc_ign = state.get("calculatedIgnitionState") if isinstance(state.get("calculatedIgnitionState"), dict) else {}
        raw_ign = state.get("ignition") if isinstance(state.get("ignition"), dict) else {}
        ignition_on = calc_ign.get("isIgnitionOn")
        if ignition_on is None:
            ignition_on = raw_ign.get("state")
        ignition_on = bool(ignition_on) if ignition_on is not None else None
        idling = bool((state.get("idlingState") or {}).get("isCurrentlyIdling")) if isinstance(state.get("idlingState"), dict) else False
        comm_calc = state.get("calculatedCommunicatingState") if isinstance(state.get("calculatedCommunicatingState"), dict) else {}
        comm = state.get("communicationState") if isinstance(state.get("communicationState"), dict) else {}
        last_update = pos.get("updateTimestamp") or comm.get("updateTimestamp") or cell.get("updateTimestamp") or meta.get("last_communication")
        last_ts = _parse_ts(last_update)
        online = comm_calc.get("isCommunicating")
        if online is None:
            online = bool(last_ts and (now - last_ts) <= ONLINE_WINDOW_SECONDS)
        online = bool(online)
        odo = (state.get("odometer") or {}) if isinstance(state.get("odometer"), dict) else {}
        odometer_m = odo.get("gpsOdometer") if odo.get("gpsOdometer") is not None else odo.get("canBusOdometer")
        dev_state = (state.get("calculatedDeviceState") or {}).get("deviceState") if isinstance(state.get("calculatedDeviceState"), dict) else None
        status = _vehicle_status(speed, ignition_on, idling, online)
        vehicles.append({
            "id": meta["id"],
            "device_id": meta["device_id"],
            "name": meta["name"],
            "plate": meta["plate"],
            "brand": meta["brand"],
            "model": meta["model"],
            "asset_type": meta["asset_type"],
            "driver": meta["driver"],
            "address": meta["address"],
            "lat": lat, "lng": lng, "latitude": lat, "longitude": lng,   # both spellings: older consumers read latitude/longitude
            "speed_kmh": speed,
            "heading": pos.get("heading"),
            "ignition_on": ignition_on,
            "idling": idling,
            "moving": status == "moving",
            "online": online,
            "status": status,
            "device_state": dev_state,
            "odometer_km": round(odometer_m / 1000.0, 1) if isinstance(odometer_m, (int, float)) else None,
            "last_update": last_update,
            "last_update_age_s": int(now - last_ts) if last_ts else None,
        })
    vehicles.sort(key=lambda v: (not v["online"], v["name"]))
    return vehicles


def _fetch_fleet(request_id):
    """Assets + live states from the provider, merged into the tab's vehicle records."""
    devices_response = _get_with_reauth("GET", GPS_DEVICES_URL, request_id, params=_ASSET_PARAMS)
    if devices_response.status_code != 200:
        logger.warning("GPS asset lookup failed request_id=%s host=%s status=%s", request_id, urlsplit(GPS_DEVICES_URL).netloc, devices_response.status_code)
        raise RuntimeError("assets:%s" % devices_response.status_code)
    devices_payload = _gps_json(devices_response, "asset lookup")
    if devices_payload is None:
        raise RuntimeError("assets:invalid")
    assets = devices_payload if isinstance(devices_payload, list) else (
        devices_payload.get("items") or devices_payload.get("devices") or devices_payload.get("assets") or []
    )
    device_ids = []
    for item in assets:
        if not isinstance(item, dict):
            continue
        asset = item.get("asset") if isinstance(item.get("asset"), dict) else item
        d = asset.get("deviceId") or item.get("deviceId")
        if d is not None:
            device_ids.append(d)
    states = {}
    if device_ids:
        state_response = _get_with_reauth(
            "POST", GPS_STATE_URL, request_id,
            json={"deviceIds": device_ids, "driverIds": [], "geofenceIds": [], "previousLookupTimestamp": None},
        )
        if state_response.status_code != 200:
            logger.warning("GPS state lookup failed request_id=%s status=%s", request_id, state_response.status_code)
            raise RuntimeError("states:%s" % state_response.status_code)
        state_payload = _gps_json(state_response, "state lookup")
        if state_payload is None:
            raise RuntimeError("states:invalid")
        raw_states = state_payload.get("deviceStates", {}) if isinstance(state_payload, dict) else {}
        if isinstance(raw_states, dict):
            states = {str(k): v for k, v in raw_states.items()}
        elif isinstance(raw_states, list):
            states = {str(s.get("id")): s for s in raw_states if isinstance(s, dict) and s.get("id") is not None}
    return _build_vehicles(assets, states, time.time())


@gps_bp.route("/api/gps")
@login_required
def get_gps_locations():
    """Live fleet positions for the Live Tracking tab (and the homepage map).

    Returns a JSON list — one record per asset — with name/plate/driver, lat/lng, speed
    (km/h), heading, ignition, idling, online and a derived `status`
    (moving | idling | engine_on | stopped | offline). Errors return {"error", "request_id"}
    with 502/503/504 so the page can show a message without losing the last known markers."""
    request_id = secrets.token_hex(6)
    if not (GPS_PERMANENT_TOKEN or (GPS_USER and GPS_PASS)):
        return jsonify({"error": "خدمة التتبع غير مهيأة — اضبط GPS_TOKEN (أو GPS_USER/GPS_PASS) في إعدادات الاستضافة.", "request_id": request_id}), 503

    now = time.time()
    if _fleet_cache["data"] is not None and now - _fleet_cache["at"] < FLEET_CACHE_SECONDS:
        result = jsonify(_fleet_cache["data"])
        result.headers["X-GPS-Request-ID"] = request_id
        result.headers["X-GPS-Cache"] = "hit"
        return result

    def _fail(message, status):
        _diag["last_error"] = message
        result = jsonify({"error": message, "request_id": request_id})
        result.headers["X-GPS-Request-ID"] = request_id
        return result, status

    try:
        vehicles = _fetch_fleet(request_id)
    except GpsAuthError as e:
        return _fail("رفض مزوّد التتبع المصادقة — مفتاح GPS_TOKEN غير صالح أو مستهلك؛ أنشئ مفتاحاً جديداً من 360Locate. (%s)" % str(e)[:160], 502)
    except requests.Timeout:
        logger.warning("GPS timeout request_id=%s", request_id)
        return _fail("تجاوز وقت الاستجابة من خدمة التتبع.", 504)
    except requests.ConnectionError:
        logger.warning("GPS connection failure request_id=%s", request_id)
        return _fail("تعذّر الاتصال بخدمة التتبع.", 503)
    except RuntimeError as e:
        code = str(e)
        if code.startswith("assets:401") or code.startswith("states:401"):
            return _fail("رفض مزوّد التتبع المصادقة. تحقّق من GPS_TOKEN في إعدادات الاستضافة.", 502)
        return _fail("تعذّر جلب مواقع الأسطول من المزوّد حالياً.", 502)
    except Exception:
        logger.exception("GPS API error request_id=%s", request_id)
        return _fail("حدث خطأ غير متوقع أثناء جلب بيانات التتبع.", 500)

    _fleet_cache["data"] = vehicles
    _fleet_cache["at"] = time.time()
    _diag["last_success_at"] = _fleet_cache["at"]
    _diag["last_count"] = len(vehicles)
    _diag["last_error"] = None
    result = jsonify(vehicles)
    result.headers["X-GPS-Request-ID"] = request_id
    result.headers["X-GPS-Cache"] = "miss"
    return result


@gps_bp.route("/api/gps/status")
@login_required
def gps_status():
    """Admin diagnostics for the tracking integration — no secrets are ever returned."""
    if not session.get("is_admin"):
        return jsonify({"success": False, "error": "forbidden"}), 403
    st = _auth_state_load()
    return jsonify({
        "success": True,
        "configured": bool(GPS_PERMANENT_TOKEN or (GPS_USER and GPS_PASS)),
        "api_base": GPS_API_BASE,
        "env_token_present": bool(GPS_PERMANENT_TOKEN),
        "login_credentials_present": bool(GPS_USER and GPS_PASS),
        "auth_source": _diag.get("auth_source"),
        "access_token_valid_for_s": int((st.get("access_exp") or 0) - time.time()) if st.get("access") else None,
        "refresh_token_rotated": bool(st.get("refresh")) and st.get("refresh") != GPS_PERMANENT_TOKEN,
        "refresh_token_expires": datetime.fromtimestamp(st["refresh_exp"], tz=timezone.utc).isoformat() if st.get("refresh_exp") else None,
        "last_auth_error": _diag.get("last_auth_error"),
        "last_error": _diag.get("last_error"),
        "last_success_at": datetime.fromtimestamp(_diag["last_success_at"], tz=timezone.utc).isoformat() if _diag.get("last_success_at") else None,
        "last_vehicle_count": _diag.get("last_count"),
        "cache_seconds": FLEET_CACHE_SECONDS,
    })


@gps_bp.route("/gps_dashboard")
@login_required
def gps_dashboard():
    return render_template("gps_dashboard.html", google_user=session.get("google_user"), b64_en=load_logo())


@gps_bp.route("/gps_devices")
@login_required
def gps_devices():
    return render_template("gps_devices.html", google_user=session.get("google_user"), b64_en=load_logo())


@gps_bp.route("/gps_sync")
@login_required
def gps_sync():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("gps_sync.html", google_user=google_user, b64_en=b64_en)


@gps_bp.route("/api/gps_devices", methods=["GET", "POST"])
@login_required
def gps_devices_data():
    """Persist the GPS tracking-device inventory (editable). Sandboxed for workstation."""
    if request.method == "POST":
        try:
            rows = (request.json or {}).get("rows", [])
            blob_set("gps_devices_data", rows)
            return jsonify({"success": True})
        except Exception as e:
            from app import db
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
    
    # GET logic is currently handled globally by /api/<table>, but added here for completeness
    from helpers import blob_get
    data = blob_get("gps_devices_data")
    if data:
        return jsonify({"success": True, "rows": data})
    return jsonify({"success": False, "rows": []})


@gps_bp.route("/api/gps_sync", methods=["POST"])
@login_required
def api_gps_sync():
    if "source_file" not in request.files or "target_file" not in request.files:
        return jsonify({"success": False, "error": "الرجاء رفع الملفين المطلوبة"}), 400

    src_file = request.files["source_file"]
    tgt_file = request.files["target_file"]

    try:
        wb_src = openpyxl.load_workbook(src_file, data_only=True)
        ws_src = wb_src.active

        headers = {}
        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(row=4, column=c).value
            if val:
                headers[str(val).strip()] = c

        if "رقم اللوحة" not in headers:
            headers = {}
            for c in range(1, ws_src.max_column + 1):
                val = ws_src.cell(row=1, column=c).value
                if val:
                    headers[str(val).strip()] = c

        lookup = {}
        plate_src_col = headers.get("رقم اللوحة")
        if plate_src_col:
            start_row = 5 if "رقم اللوحة" in [ws_src.cell(row=4, column=c).value for c in range(1, ws_src.max_column + 1)] else 2
            for r in range(start_row, ws_src.max_row + 1):
                plate_val = ws_src.cell(row=r, column=plate_src_col).value
                norm = normalize_plate(plate_val)
                if norm:
                    row_data = {}
                    for col_name, c_idx in headers.items():
                        row_data[col_name] = ws_src.cell(row=r, column=c_idx).value
                    lookup[norm] = row_data

        wb = openpyxl.load_workbook(tgt_file)
        ws = wb.active

        plate_col = 9
        vin_col = 1
        sn_col = 2
        year_col = 3
        model_col = 4
        make_col = 5
        reg_col = 6
        branch_col = 7

        match_count = 0
        update_count = 0

        for r in range(6, ws.max_row + 1):
            plate_val = ws.cell(row=r, column=plate_col).value
            if not plate_val:
                continue
            norm = normalize_plate(plate_val)
            if norm in lookup:
                src = lookup[norm]
                match_count += 1
                updates = [
                    (vin_col, "رقم الهيكل"),
                    (sn_col, "الرقم التسلسلي"),
                    (year_col, "سنة الصنع"),
                    (model_col, "الطراز"),
                    (make_col, "الماركة"),
                    (reg_col, "نوع التسجيل"),
                    (branch_col, "الفرع"),
                ]
                for col_idx, src_col_name in updates:
                    if src_col_name in src:
                        new_val = src[src_col_name]
                        if new_val is not None and str(new_val).strip() != "" and str(new_val).lower() != "nan":
                            ws.cell(row=r, column=col_idx).value = new_val
                            update_count += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")

        return jsonify({
            "success": True,
            "matches": match_count,
            "updates": update_count,
            "file_b64": b64,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
