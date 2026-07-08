from flask import Blueprint, request, jsonify, send_file, current_app, session
import os
import json
import io
import base64
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell as MC
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy as shallow_copy
from datetime import datetime

invoices_bp = Blueprint('invoices', __name__)

# Import dependencies from app (since they are still there)
from app import (
    login_required, _audit_add, blob_get, blob_set, 
    _harvest_driver_registry, _harvest_vehicle_registry, 
    get_template_path, inject_logo, logger
)

def generate_invoice():
    try:
        data = request.json
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True
        ws.title = "فاتورة"

        # Check for custom template
        tpl_path = get_template_path("invoice")
        if tpl_path:
            wb = openpyxl.load_workbook(tpl_path)
            ws = wb.active
        else:
            # Build simple professional invoice from scratch
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            P = "0C2340"
            thin = Side(style="thin", color=P)
            border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # Logo area
            ws.merge_cells("A1:G3")
            inject_logo(ws, "A1")

            # Title
            ws.merge_cells("A4:G4")
            ws["A4"] = "فاتورة صيانة مركبة"
            ws["A4"].font = Font(name="Arial", size=16, bold=True, color=P)
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

            # Info fields
            info = [
                ("A5", "اسم السائق:", "B5", data.get("driver", "")),
                ("D5", "التاريخ:", "E5", data.get("date", "")),
                ("A6", "نوع السيارة:", "B6", data.get("car", "")),
                ("D6", "رقم اللوحة:", "E6", data.get("plate", "")),
                ("A7", "نوع الطلب:", "B7", data.get("requestType", "")),
            ]
            for lbl_cell, lbl, val_cell, val in info:
                ws[lbl_cell] = lbl
                ws[lbl_cell].font = Font(name="Arial", size=11, bold=True, color=P)
                ws[val_cell] = val
                ws[val_cell].font = Font(name="Arial", size=11)

            # Table header
            headers = ["م", "الوصف", "الكمية", "السعر", "المبلغ"]
            for i, h in enumerate(headers):
                c = ws.cell(row=9, column=i + 1, value=h)
                c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=P)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            # Data row
            qty = float(data.get("quantity", 0) or 0)
            price = float(data.get("price", 0) or 0)
            subtotal = qty * price
            row_data = [1, data.get("description", ""), qty, price, subtotal]
            for i, v in enumerate(row_data):
                c = ws.cell(row=10, column=i + 1, value=v)
                c.font = Font(name="Arial", size=11)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            # Totals
            tax = subtotal * 0.15
            total = subtotal + tax
            for ri, (lbl, val) in enumerate(
                [
                    ("المبلغ الإجمالي (قبل الضريبة)", subtotal),
                    ("ضريبة القيمة المضافة (15%)", tax),
                    ("المجموع الكلي (مع الضريبة)", total),
                ]
            ):
                r = 11 + ri
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws.cell(row=r, column=1, value=lbl).font = Font(
                    name="Arial", size=11, bold=True, color=P
                )
                ws.cell(row=r, column=5, value=val).font = Font(
                    name="Arial", size=12, bold=(ri == 2), color=P
                )

            # Signatures
            ws.cell(row=15, column=1, value="توقيع السائق").font = Font(
                name="Arial", bold=True, color=P
            )
            ws.cell(row=15, column=5, value="اعتماد الإدارة").font = Font(
                name="Arial", bold=True, color=P
            )

            # Column widths
            for col, w in [
                (1, 8),
                (2, 30),
                (3, 12),
                (4, 12),
                (5, 15),
                (6, 12),
                (7, 12),
            ]:
                ws.column_dimensions[chr(64 + col)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        driver = data.get("driver", "").replace(" ", "_")[:20]
        plate = data.get("plate", "").replace(" ", "_")[:15]
        fname = f"فاتورة_{driver}_{plate}.xlsx"
        return jsonify({"success": True, "file_b64": b64, "filename": fname})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def generate_schedule():
    try:
        data = request.json
        template_path = os.path.join(current_app.root_path, "schedule_base.xlsx")

        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Schedule template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def safe_set(sheet, row, col, val):
            cell = sheet.cell(row=row, column=col)
            if not isinstance(cell, MC):
                cell.value = val

        # Set date in row 6 col 1
        schedule_date = data.get("date", "")
        safe_set(ws, 6, 1, schedule_date)

        main_data = data.get("main", [])
        spare_data = data.get("spare", [])

        # We need to write from bottom up or just write all and delete rows from bottom up

        # 1. Summary Block (Row 53)
        dina = sum(1 for d in main_data if "دينا" in d.get("type", ""))
        lorry = sum(1 for d in main_data if "لوري" in d.get("type", ""))
        delivery = sum(1 for d in main_data if d.get("job", "") == "موصل")
        dist = sum(1 for d in main_data if d.get("job", "") == "موزع")

        safe_set(ws, 53, 2, len(main_data))
        safe_set(ws, 53, 3, dina)
        safe_set(ws, 53, 4, lorry)
        safe_set(ws, 53, 5, delivery)
        safe_set(ws, 53, 6, dist)
        safe_set(ws, 53, 7, len(spare_data))

        # 2. Spare Data (Row 42-49)
        for idx, rd in enumerate(spare_data[:8]):
            r = 42 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 8, rd.get("typemodel", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # 3. Main Data (Row 9-36)
        for idx, rd in enumerate(main_data[:28]):
            r = 9 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 5, rd.get("job", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 7, rd.get("model", ""))
            safe_set(ws, r, 8, rd.get("type", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 11, rd.get("serial", ""))
            safe_set(ws, r, 12, rd.get("inspect", ""))
            safe_set(ws, r, 13, rd.get("license", ""))
            safe_set(ws, r, 14, rd.get("drivercard", ""))
            safe_set(ws, r, 15, rd.get("opcard", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # Delete unused rows from bottom up to avoid shifting index bugs
        spare_deletions = 8 - len(spare_data) if len(spare_data) < 8 else 0
        main_deletions = 28 - len(main_data) if len(main_data) < 28 else 0

        if spare_deletions > 0:
            ws.delete_rows(42 + len(spare_data), spare_deletions)
        if main_deletions > 0:
            ws.delete_rows(9 + len(main_data), main_deletions)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def generate_washing():
    try:
        data = request.json
        vehicles = data.get("vehicles", [])

        template_path = os.path.join(current_app.root_path, "washing_template.xlsx")
        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Washing template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Insert Logo
        logo_path = os.path.join(
            current_app.root_path, "templates", "ApplicationFrameHost_KUIZUuJ46O (1).png"
        )
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 300
            img.height = 90
            ws.add_image(img, "F1")

        # Unmerge ALL merged cells to avoid MergedCell write errors
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            ws.unmerge_cells(str(mr))

        # Copy header style from row 4 for reuse
        from copy import copy as shallow_copy
        header_fills = {}
        header_fonts = {}
        header_aligns = {}
        header_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=4, column=c)
            header_fills[c] = shallow_copy(cell.fill)
            header_fonts[c] = shallow_copy(cell.font)
            header_aligns[c] = shallow_copy(cell.alignment)
            header_borders[c] = shallow_copy(cell.border)

        # Copy a data row style (row 5) for styling data rows
        data_fills = {}
        data_fonts = {}
        data_aligns = {}
        data_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=5, column=c)
            data_fills[c] = shallow_copy(cell.fill)
            data_fonts[c] = shallow_copy(cell.font)
            data_aligns[c] = shallow_copy(cell.alignment)
            data_borders[c] = shallow_copy(cell.border)

        # Write vehicle data starting at row 5
        for idx, v in enumerate(vehicles):
            r = 5 + idx
            ws.cell(row=r, column=1, value=v.get("id", idx + 1))
            ws.cell(row=r, column=2, value=v.get("plate", ""))
            ws.cell(row=r, column=3, value=v.get("type", ""))
            ws.cell(row=r, column=4, value=v.get("driver", ""))
            months = v.get("m", [])
            total = sum(months)
            for m_idx in range(12):
                val = "استلم" if m_idx < len(months) and months[m_idx] == 1 else None
                ws.cell(row=r, column=5 + m_idx, value=val)
            ws.cell(row=r, column=17, value=total)
            # Apply data styling
            for c in range(1, 19):
                cell = ws.cell(row=r, column=c)
                cell.fill = shallow_copy(data_fills.get(c, data_fills[1]))
                cell.font = shallow_copy(data_fonts.get(c, data_fonts[1]))
                cell.alignment = shallow_copy(data_aligns.get(c, data_aligns[1]))
                cell.border = shallow_copy(data_borders.get(c, data_borders[1]))

        # Summary row: right after last vehicle
        summary_row = 5 + len(vehicles)
        ws.cell(row=summary_row, column=1, value="إجمالي الغسيل الشهري")
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
        for m_idx in range(12):
            col = 5 + m_idx
            start_cell = ws.cell(row=5, column=col).coordinate.replace("5", "")
            formula = '=COUNTIF(%s5:%s%d,"استلم")' % (start_cell, start_cell, summary_row - 1)
            ws.cell(row=summary_row, column=col, value=formula)
        ws.cell(row=summary_row, column=17, value="=SUM(Q5:Q%d)" % (summary_row - 1))

        # Style summary row bold
        from openpyxl.styles import Font, PatternFill, Alignment
        summary_font = Font(name="Cairo", size=11, bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        summary_align = Alignment(horizontal="center", vertical="center")
        for c in range(1, 19):
            cell = ws.cell(row=summary_row, column=c)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = summary_align

        # Footer row
        footer_row = summary_row + 2
        ws.cell(row=footer_row, column=1, value="تم إعداد هذا الجدول بواسطة قسم الحركة - فرع الدمام")
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=18)

        # Re-merge header rows
        ws.merge_cells("A1:R1")
        ws.merge_cells("A2:D2")
        ws.merge_cells("A3:R3")

        # Update total washes in header
        total_washes = sum(sum(v.get("m", [])) for v in vehicles)
        ws.cell(row=2, column=12, value=total_washes)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def generate_po():
    """Generate PO Excel from original template using openpyxl server-side"""
    try:
        data = request.json or {}

        template_path = os.path.join(os.path.dirname(__file__), "po_template.xlsx")
        if not os.path.exists(template_path):
            logger.error("PO template not found at: %s", template_path)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "قالب طلب الشراء غير موجود (po_template.xlsx)",
                    }
                ),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        ws["D7"] = data.get("driver", "")
        ws["G7"] = data.get("branch", "الدمام")
        ws["I7"] = data.get("job", "سائق")
        ws["D8"] = data.get("car", "")
        ws["G8"] = data.get("plate", "")
        
        model_val = data.get("model", "")
        if not model_val and data.get("car"):
            import re
            match = re.search(r'(\d{4}|[٠١٢٣٤٥٦٧٨٩]{4})', data.get("car", ""))
            if match:
                model_val = match.group(1)
        ws["I8"] = model_val
        
        ws["D9"] = data.get("odometer", "")
        
        try:
            ws.unmerge_cells(start_row=9, start_column=5, end_row=9, end_column=9)
        except Exception:
            pass

        ws["F8"] = "رقم اللوحة:"
        ws["H8"] = "الموديل:"

        ws["F9"] = "رقم الجوال:"
        ws["G9"] = data.get("phone", "")
        ws["H9"] = "الرقم الوظيفي:"
        ws["I9"] = data.get("empid", "")
        
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # neutral gray, B&W-print-safe
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Labels to format: F8(Plate), H8(Model), F9(Phone), H9(EmpID)
        for r, c in [(8, 6), (8, 8), (9, 6), (9, 8)]:
            cell = ws.cell(row=r, column=c)
            # Copy font from A8, but hardcode the rest to ensure it matches perfectly
            try:
                cell.font = copy(ws["A8"].font)
            except:
                cell.font = Font(name="Arial", size=12, bold=True)
            cell.fill = label_fill
            cell.border = thin_border
            cell.alignment = center_align
            
        # Values to format: G8(PlateVal), I8(ModelVal), G9(PhoneVal), I9(EmpIDVal)
        for r, c in [(8, 7), (8, 9), (9, 7), (9, 9)]:
            cell = ws.cell(row=r, column=c)
            try:
                cell.font = copy(ws["D8"].font)
            except:
                pass
            cell.border = thin_border
            cell.alignment = center_align

        date_val = data.get("date", "")
        if date_val:
            ws["C4"] = date_val
        serial_val = data.get("serial", "")
        if serial_val:
            ws["J3"] = f"NO. {serial_val}"

        from openpyxl.styles import Alignment
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def set_formatted(row, col, val, is_currency=False):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MC):
                if val is not None and val != "":
                    if is_currency:
                        try:
                            cell.value = float(val)
                            cell.number_format = '#,##0.00'
                        except:
                            cell.value = val
                    else:
                        cell.value = val
                else:
                    cell.value = None
                cell.alignment = center_align

        # Fill parts (rows 12-14)
        parts = data.get("parts", [])
        for i, p in enumerate(parts[:3]):
            r = 12 + i
            set_formatted(r, 4, p.get("desc", ""))
            set_formatted(r, 7, p.get("qty"))
            set_formatted(r, 8, p.get("price"), is_currency=True)
            set_formatted(r, 9, p.get("val"), is_currency=True)
            set_formatted(r, 10, p.get("notes", ""))

        # Fill repairs (rows 18-20)
        repairs = data.get("repairs", [])
        for i, rep in enumerate(repairs[:3]):
            r = 18 + i
            set_formatted(r, 4, rep.get("desc", ""))
            set_formatted(r, 8, rep.get("val"), is_currency=True)
            set_formatted(r, 9, rep.get("notes", ""))

        # Fill tires (rows 24-26)
        tires = data.get("tires", [])
        for i, t in enumerate(tires[:3]):
            r = 24 + i
            set_formatted(r, 4, t.get("date", ""))
            set_formatted(r, 5, t.get("count"))
            set_formatted(r, 6, t.get("front"))
            set_formatted(r, 7, t.get("back"))
            set_formatted(r, 8, t.get("prev"))
            set_formatted(r, 9, t.get("curr"))
            set_formatted(r, 10, t.get("dist"))

        # Fill batteries (rows 30-32)
        batteries = data.get("batteries", [])
        for i, b in enumerate(batteries[:3]):
            r = 30 + i
            set_formatted(r, 4, b.get("desc", ""))
            set_formatted(r, 6, b.get("count"))
            set_formatted(r, 7, b.get("size", ""))
            set_formatted(r, 8, b.get("amp", ""))
            set_formatted(r, 9, b.get("price"), is_currency=True)
            set_formatted(r, 10, b.get("date", ""))

        # Fill summary totals (row 34: per-category, row 35-36: subtotals, row 37: grand total)
        summary = data.get("summary", {})

        # Row 34: individual category totals
        set_formatted(35, 3, summary.get("parts_total"), is_currency=True)
        set_formatted(35, 4, summary.get("repairs_total"), is_currency=True)
        set_formatted(35, 5, summary.get("tires_total"), is_currency=True)
        set_formatted(35, 6, summary.get("batteries_total"), is_currency=True)
        # Row 34 col 7-8: "الإجمالي شامل الضريبة" (label already in template)
        set_formatted(35, 7, summary.get("grand_total"), is_currency=True)
        # Row 34 col 9: Notes
        notes = data.get("notes", "")
        if notes:
            set_formatted(35, 9, notes)

        # Row 37: الإجمالي شامل الضريبة (grand total line) - Replaced with Tafqeet text
        grand_total_val = summary.get("grand_total")
        
        # Add Tafqeet in Col 5 (E) of Row 37 INSTEAD of the numeric value
        ws.cell(row=37, column=5).value = tafqeet(grand_total_val)
        ws.cell(row=37, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Remove the previous logic that put it in Col 1
        ws.cell(row=37, column=1).value = ""

        # Inject Logo
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_path = os.path.join(os.path.dirname(__file__), "logo_excel.png")
            if not os.path.exists(logo_path):
                logo_path = os.path.join(os.path.dirname(__file__), "static", "site_logo.png")
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 450
                img.height = 85
                ws.add_image(img, 'C1') 
        except Exception as e:
            logger.error("Logo injection failed: %s", e)

        # Clean black-&-white printing: fit to one page width, centered, A4 portrait.
        try:
            from openpyxl.worksheet.properties import PageSetupProperties
            from openpyxl.worksheet.page import PageMargins
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.print_options.horizontalCentered = True
            ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        except Exception as e:
            logger.warning("PO page setup skipped: %s", e)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        logger.exception("generate_po error")
        return jsonify({"success": False, "error": "تعذّر توليد طلب الشراء."}), 500

def generate_oils():
    """Generate Oils/Filters Excel from original template using openpyxl server-side"""
    try:
        data = request.json or {}

        template_path = os.path.join(os.path.dirname(__file__), "oils_template.xlsx")
        if not os.path.exists(template_path):
            logger.error("Oils template not found at: %s", template_path)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "قالب الزيوت غير موجود (oils_template.xlsx)",
                    }
                ),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def safe_set(row, col, val):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MC):
                cell.value = val

        # Update title (row 4)
        title = data.get("title", "")
        if title:
            ws["A4"] = title

        # Fill main oil data (rows 6-62, max 57 rows)
        oils = data.get("oils", [])
        for i, oil in enumerate(oils[:57]):
            r = 6 + i
            safe_set(r, 1, i + 1)  # م
            safe_set(r, 2, oil.get("plate", ""))  # رقم لوحة
            safe_set(r, 3, oil.get("driver", ""))  # المستخدم (merged C:D)
            safe_set(r, 5, oil.get("date", ""))  # تاريخ تغيير الزيت
            safe_set(r, 6, oil.get("counter", ""))  # رقم العداد
            safe_set(r, 7, oil.get("liters", ""))  # عدد اللترات
            safe_set(r, 8, oil.get("filters", ""))  # عدد الفلاتر

        # Fill filter details - LEFT side (A-E, rows 69-86)
        filters_left = data.get("filters_left", [])
        for i, f in enumerate(filters_left[:18]):
            r = 69 + i
            safe_set(r, 1, i + 1)
            safe_set(r, 2, f.get("name", ""))
            safe_set(r, 3, f.get("prev", ""))
            safe_set(r, 4, f.get("used", ""))
            safe_set(r, 5, f.get("remaining", ""))

        # Fill filter details - RIGHT side (F-H, rows 69-86)
        filters_right = data.get("filters_right", [])
        for i, f in enumerate(filters_right[:18]):
            r = 69 + i
            safe_set(r, 6, len(filters_left) + i + 1)
            safe_set(r, 7, f.get("name", ""))
            safe_set(r, 8, f.get("qty", ""))

        # Fill notes section (rows 65-66 only; 63 and 67 are fixed headers in template)
        notes = data.get("notes", {})
        if notes.get("row65"):
            safe_set(65, 1, notes["row65"])
        if notes.get("row66"):
            safe_set(66, 1, notes["row66"])

        # Fill diesel filter notes (rows 88-95)
        diesel_notes = data.get("diesel_notes", [])
        for i, note in enumerate(diesel_notes[:8]):
            safe_set(88 + i, 1, note)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

