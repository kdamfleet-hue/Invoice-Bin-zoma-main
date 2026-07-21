import re
import json
import io
import base64
import openpyxl
import threading
import logging
from datetime import datetime, date
from flask import Blueprint, render_template, session, request, jsonify, make_response
from helpers import login_required, role_required, load_logo, blob_get, blob_set, audit_and_verify, current_branch_id

logger = logging.getLogger("InvoiceApp")
schedule_bp = Blueprint('schedule', __name__)

_REG_FIELDS = ("empid", "phone", "job", "drivercard", "empNotes")
DRIVER_REG_MAX = 4000
_REG_FIELD_MAX = {"empid": 40, "phone": 40, "job": 40, "drivercard": 40, "empNotes": 200}
_DRIVER_REG_LOCK = threading.Lock()

_VEH_REG_FIELDS = ("model", "vtype", "pallets", "load", "vserial", "inspect", "license", "opcard", "notes")
VEHICLE_REG_MAX = 4000
_VEH_REG_FIELD_MAX = {"model": 20, "vtype": 40, "pallets": 10, "load": 20, "vserial": 40,
                      "inspect": 40, "license": 40, "opcard": 40, "notes": 300}
_VEHICLE_REG_LOCK = threading.Lock()
_AR_DIGITS = "٠١٢٣٤٥٦٧٨٩"

@schedule_bp.route("/schedule")
@role_required('admin', 'operations')
def schedule():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("schedule.html", google_user=google_user, b64_en=b64_en)

@schedule_bp.route("/washing")
@role_required('admin', 'operations', 'maintenance')
def washing():
    google_user = session.get("google_user")
    b64_en = load_logo()
    return render_template("washing.html", google_user=google_user, b64_en=b64_en)


@schedule_bp.route("/api/schedule_data", methods=["GET", "POST"])
@role_required('admin', 'operations')
def schedule_data():
    """Persist the weekly schedule (main/spare/vacation/summary). Sandboxed for workstation."""
    if request.method == "POST":
        try:
            sd = request.json or {}
            reason = sd.get("reason")
            try:
                audit_and_verify("تحديث", "الجدول الأسبوعي", reason)
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            
            blob_set("schedule_data", sd)
            try:
                _harvest_driver_registry(sd)  # learn each driver's card-expiry/job for future autofill
            except Exception:
                logger.warning("driver_registry harvest failed (non-fatal)")
            try:
                _harvest_vehicle_registry(sd)  # learn each vehicle's spec for future autofill
            except Exception:
                logger.warning("vehicle_registry harvest failed (non-fatal)")
            _n = (len(sd.get("main", []) or []) + len(sd.get("spare", []) or [])) if isinstance(sd, dict) else None
            return jsonify({"success": True})
        except Exception:
            logger.exception("schedule_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ الجدول الأسبوعي."}), 500
    try:
        return jsonify({"success": True, "data": blob_get("schedule_data")})
    except Exception:
        logger.exception("schedule_data GET error")
        return jsonify({"success": False, "error": "تعذّر جلب الجدول الأسبوعي."}), 500

@schedule_bp.route("/api/washing_data", methods=["GET", "POST"])
@role_required('admin', 'operations', 'maintenance')
def washing_data():
    # Workstation mode writes/reads an isolated copy (id=2); the real data (id=1) is untouched.
    if request.method == "POST":
        try:
            vehicles = (request.json or {}).get("vehicles", [])
            blob_set("washing_schedule", vehicles)
            _audit_add("تحديث", "جدول الغسيل", len(vehicles) if isinstance(vehicles, list) else None)
            return jsonify({"success": True})
        except Exception:
            logger.exception("washing_data POST error")
            return jsonify({"success": False, "error": "تعذّر حفظ جدول الغسيل."}), 500
    try:
        data = blob_get("washing_schedule")
        if data is not None:
            return jsonify({"success": True, "vehicles": data})
        return jsonify({"success": False, "vehicles": []})
    except Exception:
        logger.exception("washing_data GET error")
        return jsonify({"success": False, "error": "تعذّر جلب جدول الغسيل."}), 500

@schedule_bp.route("/api/generate_schedule", methods=["POST"])
@role_required('admin', 'operations')
def generate_schedule():
    try:
        data = request.json
        template_path = os.path.join(app.root_path, "schedule_base.xlsx")

        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Schedule template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def safe_set(sheet, row, col, val):
            cell = sheet.cell(row=row, column=col)
            if not isinstance(cell, MC):
                cell.value = val

        # Set date in row 6 col 1
        schedule_date = data.get("date", "")
        safe_set(ws, 6, 1, schedule_date)

        main_data = data.get("main", [])
        spare_data = data.get("spare", [])

        # We need to write from bottom up or just write all and delete rows from bottom up

        # 1. Summary Block (Row 53)
        dina = sum(1 for d in main_data if "دينا" in d.get("type", ""))
        lorry = sum(1 for d in main_data if "لوري" in d.get("type", ""))
        delivery = sum(1 for d in main_data if d.get("job", "") == "موصل")
        dist = sum(1 for d in main_data if d.get("job", "") == "موزع")

        safe_set(ws, 53, 2, len(main_data))
        safe_set(ws, 53, 3, dina)
        safe_set(ws, 53, 4, lorry)
        safe_set(ws, 53, 5, delivery)
        safe_set(ws, 53, 6, dist)
        safe_set(ws, 53, 7, len(spare_data))

        # 2. Spare Data (Row 42-49)
        for idx, rd in enumerate(spare_data[:8]):
            r = 42 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 8, rd.get("typemodel", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # 3. Main Data (Row 9-36)
        for idx, rd in enumerate(main_data[:28]):
            r = 9 + idx
            safe_set(ws, r, 1, idx + 1)
            safe_set(ws, r, 2, rd.get("empid", ""))
            safe_set(ws, r, 3, rd.get("name", ""))
            safe_set(ws, r, 4, rd.get("iqama", ""))
            safe_set(ws, r, 5, rd.get("job", ""))
            safe_set(ws, r, 6, rd.get("plate", ""))
            safe_set(ws, r, 7, rd.get("model", ""))
            safe_set(ws, r, 8, rd.get("type", ""))
            safe_set(ws, r, 9, rd.get("pallets", ""))
            safe_set(ws, r, 10, rd.get("capacity", ""))
            safe_set(ws, r, 11, rd.get("serial", ""))
            safe_set(ws, r, 12, rd.get("inspect", ""))
            safe_set(ws, r, 13, rd.get("license", ""))
            safe_set(ws, r, 14, rd.get("drivercard", ""))
            safe_set(ws, r, 15, rd.get("opcard", ""))
            safe_set(ws, r, 16, rd.get("notes", ""))
            safe_set(ws, r, 17, rd.get("phone", ""))

        # Delete unused rows from bottom up to avoid shifting index bugs
        spare_deletions = 8 - len(spare_data) if len(spare_data) < 8 else 0
        main_deletions = 28 - len(main_data) if len(main_data) < 28 else 0

        if spare_deletions > 0:
            ws.delete_rows(42 + len(spare_data), spare_deletions)
        if main_deletions > 0:
            ws.delete_rows(9 + len(main_data), main_deletions)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@schedule_bp.route("/api/generate_washing", methods=["POST"])
@role_required('admin', 'operations', 'maintenance')
def generate_washing():
    try:
        data = request.json
        vehicles = data.get("vehicles", [])

        template_path = os.path.join(app.root_path, "washing_template.xlsx")
        if not os.path.exists(template_path):
            return (
                jsonify({"success": False, "error": "Washing template not found"}),
                500,
            )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Insert Logo
        logo_path = os.path.join(
            app.root_path, "templates", "ApplicationFrameHost_KUIZUuJ46O (1).png"
        )
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.width = 300
            img.height = 90
            ws.add_image(img, "F1")

        # Unmerge ALL merged cells to avoid MergedCell write errors
        merged_ranges = list(ws.merged_cells.ranges)
        for mr in merged_ranges:
            ws.unmerge_cells(str(mr))

        # Copy header style from row 4 for reuse
        from copy import copy as shallow_copy
        header_fills = {}
        header_fonts = {}
        header_aligns = {}
        header_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=4, column=c)
            header_fills[c] = shallow_copy(cell.fill)
            header_fonts[c] = shallow_copy(cell.font)
            header_aligns[c] = shallow_copy(cell.alignment)
            header_borders[c] = shallow_copy(cell.border)

        # Copy a data row style (row 5) for styling data rows
        data_fills = {}
        data_fonts = {}
        data_aligns = {}
        data_borders = {}
        for c in range(1, 19):
            cell = ws.cell(row=5, column=c)
            data_fills[c] = shallow_copy(cell.fill)
            data_fonts[c] = shallow_copy(cell.font)
            data_aligns[c] = shallow_copy(cell.alignment)
            data_borders[c] = shallow_copy(cell.border)

        # Write vehicle data starting at row 5
        for idx, v in enumerate(vehicles):
            r = 5 + idx
            ws.cell(row=r, column=1, value=v.get("id", idx + 1))
            ws.cell(row=r, column=2, value=v.get("plate", ""))
            ws.cell(row=r, column=3, value=v.get("type", ""))
            ws.cell(row=r, column=4, value=v.get("driver", ""))
            months = v.get("m", [])
            total = sum(months)
            for m_idx in range(12):
                val = "استلم" if m_idx < len(months) and months[m_idx] == 1 else None
                ws.cell(row=r, column=5 + m_idx, value=val)
            ws.cell(row=r, column=17, value=total)
            # Apply data styling
            for c in range(1, 19):
                cell = ws.cell(row=r, column=c)
                cell.fill = shallow_copy(data_fills.get(c, data_fills[1]))
                cell.font = shallow_copy(data_fonts.get(c, data_fonts[1]))
                cell.alignment = shallow_copy(data_aligns.get(c, data_aligns[1]))
                cell.border = shallow_copy(data_borders.get(c, data_borders[1]))

        # Summary row: right after last vehicle
        summary_row = 5 + len(vehicles)
        ws.cell(row=summary_row, column=1, value="إجمالي الغسيل الشهري")
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
        for m_idx in range(12):
            col = 5 + m_idx
            start_cell = ws.cell(row=5, column=col).coordinate.replace("5", "")
            formula = '=COUNTIF(%s5:%s%d,"استلم")' % (start_cell, start_cell, summary_row - 1)
            ws.cell(row=summary_row, column=col, value=formula)
        ws.cell(row=summary_row, column=17, value="=SUM(Q5:Q%d)" % (summary_row - 1))

        # Style summary row bold
        from openpyxl.styles import Font, PatternFill, Alignment
        summary_font = Font(name="Cairo", size=11, bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        summary_align = Alignment(horizontal="center", vertical="center")
        for c in range(1, 19):
            cell = ws.cell(row=summary_row, column=c)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = summary_align

        # Footer row
        footer_row = summary_row + 2
        ws.cell(row=footer_row, column=1, value="تم إعداد هذا الجدول بواسطة قسم الحركة - فرع الدمام")
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=18)

        # Re-merge header rows
        ws.merge_cells("A1:R1")
        ws.merge_cells("A2:D2")
        ws.merge_cells("A3:R3")

        # Update total washes in header
        total_washes = sum(sum(v.get("m", [])) for v in vehicles)
        ws.cell(row=2, column=12, value=total_washes)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@schedule_bp.route("/api/export_schedule_exact", methods=["POST"])
@login_required
def export_schedule_exact():
    try:
        data = request.json or {}
        template_path = os.path.join(app.root_path, "weekly_schedule_template.xlsx")
        
        if not os.path.exists(template_path):
            return jsonify({"success": False, "error": "قالب التصدير غير موجود"}), 500

        wb = openpyxl.load_workbook(template_path)
        
        def safe_set(sheet, row, col, val):
            cell = sheet.cell(row=row, column=col)
            try:
                cell.value = val
            except AttributeError:
                pass # Merged cell

        def safe_set_num(sheet, row, col, val):
            try:
                val = int(val)
            except (TypeError, ValueError):
                pass
            safe_set(sheet, row, col, val)

        # --- Active Vehicles (المركبات النشطة) ---
        ws_main = wb["المركبات النشطة"]
        main_data = data.get("main", [])
        for idx, rd in enumerate(main_data):
            r = 5 + idx
            safe_set(ws_main, r, 1, idx + 1)
            safe_set(ws_main, r, 2, rd.get("empid", ""))
            safe_set(ws_main, r, 3, rd.get("name", ""))
            safe_set(ws_main, r, 4, rd.get("iqama", ""))
            safe_set(ws_main, r, 5, rd.get("job", ""))
            safe_set(ws_main, r, 6, rd.get("plate", ""))
            safe_set(ws_main, r, 7, rd.get("model", ""))
            safe_set(ws_main, r, 8, rd.get("vtype", ""))
            safe_set(ws_main, r, 9, rd.get("pallets", ""))
            safe_set(ws_main, r, 10, rd.get("load", ""))
            safe_set(ws_main, r, 11, rd.get("vserial", ""))
            safe_set(ws_main, r, 12, rd.get("inspect", ""))
            # Col 13 is formula (rem_days1)
            safe_set(ws_main, r, 14, rd.get("license", ""))
            # Col 15 is formula (rem_days2)
            safe_set(ws_main, r, 16, rd.get("drivercard", ""))
            # Col 17 is formula (rem_days3)
            safe_set(ws_main, r, 18, rd.get("opcard", ""))
            # Col 19 is formula (rem_days4)
            safe_set(ws_main, r, 20, rd.get("empNotes", ""))
            safe_set(ws_main, r, 21, rd.get("phone", ""))

        # --- Spare and Broken (الأسبير والمعطلة) ---
        if "الأسبير والمعطلة" in wb.sheetnames:
            ws_spare = wb["الأسبير والمعطلة"]
            spare_data = data.get("spare", [])
            for idx, rd in enumerate(spare_data):
                r = 4 + idx
                safe_set(ws_spare, r, 1, idx + 1)
                safe_set(ws_spare, r, 2, rd.get("status", "اسبير"))
                safe_set(ws_spare, r, 3, rd.get("plate", ""))
                safe_set(ws_spare, r, 4, rd.get("model", ""))
                safe_set(ws_spare, r, 5, rd.get("vtype", ""))
                safe_set(ws_spare, r, 6, rd.get("pallets", ""))
                safe_set(ws_spare, r, 7, rd.get("load", ""))
                safe_set(ws_spare, r, 8, rd.get("vserial", ""))
                safe_set(ws_spare, r, 9, rd.get("inspect", ""))
                # Col 10 is the formula for remaining days
                safe_set(ws_spare, r, 11, rd.get("license", ""))
                # Col 12 is the formula for remaining days
                safe_set(ws_spare, r, 13, rd.get("opcard", ""))
                # Col 14 is the formula for remaining days
                safe_set(ws_spare, r, 15, rd.get("empNotes", ""))

        # --- Vacation (السائقون في إجازة) ---
        if "السائقون في إجازة" in wb.sheetnames:
            ws_vac = wb["السائقون في إجازة"]
            vac_data = data.get("vacation", [])
            for idx, rd in enumerate(vac_data):
                r = 4 + idx
                safe_set(ws_vac, r, 1, idx + 1)
                safe_set(ws_vac, r, 2, rd.get("empid", ""))
                safe_set(ws_vac, r, 3, rd.get("name", ""))
                safe_set(ws_vac, r, 4, rd.get("iqama", ""))
                safe_set(ws_vac, r, 5, rd.get("job", ""))
                safe_set(ws_vac, r, 6, rd.get("drivercard", ""))
                # Col 7 is the formula
                safe_set(ws_vac, r, 8, rd.get("phone", ""))
                safe_set(ws_vac, r, 9, rd.get("empNotes", ""))

        # --- Dashboard (لوحة المعلومات) — KPIs, expiry alert list, and the two summary
        # tables, matching whatever's currently shown/edited on the site's dashboard tab.
        # G5 (vacation KPI) is left untouched: it's a live COUNTA() formula in the template.
        if "لوحة المعلومات" in wb.sheetnames:
            ws_dash = wb["لوحة المعلومات"]
            dash = data.get("dashboard", {})

            if data.get("date"):
                safe_set(ws_dash, 2, 2, data.get("date"))

            kpis = dash.get("kpis", {})
            safe_set_num(ws_dash, 5, 1, kpis.get("drivers", ""))
            safe_set_num(ws_dash, 5, 3, kpis.get("delivery", ""))
            safe_set_num(ws_dash, 5, 5, kpis.get("distributors", ""))
            safe_set_num(ws_dash, 5, 9, kpis.get("spare", ""))

            expiring = dash.get("expiring", [])[:7]
            for idx, rd in enumerate(expiring):
                r = 9 + idx
                safe_set(ws_dash, r, 1, idx + 1)
                safe_set(ws_dash, r, 2, rd.get("name", ""))
                safe_set(ws_dash, r, 3, rd.get("plate", ""))
                safe_set(ws_dash, r, 4, rd.get("doc", ""))
                safe_set(ws_dash, r, 5, rd.get("date", ""))
                safe_set_num(ws_dash, r, 6, rd.get("days", ""))

            job_split = dash.get("jobSplit", {})
            for idx, (label, val) in enumerate(list(job_split.items())[:2]):
                r = 19 + idx
                safe_set(ws_dash, r, 1, label)
                safe_set_num(ws_dash, r, 2, val)

            def vt_sort_key(kv):
                try:
                    return -int(kv[1])
                except (TypeError, ValueError):
                    return 0
            vehicle_types = sorted(dash.get("vehicleTypes", {}).items(), key=vt_sort_key)[:5]
            for idx, (label, val) in enumerate(vehicle_types):
                r = 21 + idx
                safe_set(ws_dash, r, 1, label)
                safe_set_num(ws_dash, r, 2, val)

        import io, base64
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"success": True, "file_b64": b64})
    except Exception as e:
        import traceback
        logger.exception("export_schedule_exact error")
        return jsonify({"success": False, "error": str(e)}), 500

def _harvest_driver_registry(sd):
    """Best-effort: fold any non-empty personal fields from the schedule rows into the
    per-branch driver registry, keyed by iqama. Latest non-empty value wins; a blank NEVER
    clears a remembered value — so a swap that clears a row keeps the driver's memory intact.
    Never raises into the caller (the schedule save must always succeed)."""
    if not isinstance(sd, dict):
        return
    with _DRIVER_REG_LOCK:  # atomic get→mutate→set within the single gunicorn worker
        reg = blob_get("driver_registry") or {}
        if not isinstance(reg, dict):
            reg = {}
        changed = False
        for section in ("main", "spare", "vacation"):
            for row in (sd.get(section) or []):
                if not isinstance(row, dict):
                    continue
                key = _norm_iqama(row.get("iqama"))
                if not key:
                    continue
                existing = key in reg
                if not existing and len(reg) >= DRIVER_REG_MAX:
                    continue  # cap reached: stop remembering NEW drivers, still update existing ones
                entry = reg.get(key) if existing else {}
                if not isinstance(entry, dict):
                    entry = {}
                local_changed = False
                for f in _REG_FIELDS:
                    val = str(row.get(f, "") or "").strip()[:_REG_FIELD_MAX.get(f, 80)]
                    if val and entry.get(f) != val:
                        entry[f] = val
                        local_changed = True
                if local_changed and entry:
                    reg[key] = entry
                    changed = True
        if changed:
            blob_set("driver_registry", reg)

def _harvest_vehicle_registry(sd):
    """Best-effort: fold any non-empty technical-spec fields from schedule rows (main/spare —
    vacation rows carry no vehicle) into the per-branch vehicle registry, keyed by normalized
    plate. Latest non-empty value wins; a blank NEVER clears a remembered value. Never raises
    into the caller (the schedule save must always succeed)."""
    if not isinstance(sd, dict):
        return
    with _VEHICLE_REG_LOCK:
        reg = blob_get("vehicle_registry") or {}
        if not isinstance(reg, dict):
            reg = {}
        changed = False
        for section in ("main", "spare"):
            for row in (sd.get(section) or []):
                if not isinstance(row, dict):
                    continue
                key = _norm_plate(row.get("plate"))
                if not key:
                    continue
                existing = key in reg
                if not existing and len(reg) >= VEHICLE_REG_MAX:
                    continue
                entry = reg.get(key) if existing else {}
                if not isinstance(entry, dict):
                    entry = {}
                local_changed = False
                for f in _VEH_REG_FIELDS:
                    val = str(row.get(f, "") or "").strip()[:_VEH_REG_FIELD_MAX.get(f, 80)]
                    if val and entry.get(f) != val:
                        entry[f] = val
                        local_changed = True
                if local_changed and entry:
                    reg[key] = entry
                    changed = True
        if changed:
            blob_set("vehicle_registry", reg)

def _norm_iqama(v):
    return re.sub(r"\D", "", str(v or ""))

def _norm_plate(v):
    """Digits-then-letters, Arabic-Indic digits folded to Latin — matches window.normalizePlate
    in app_ux.js exactly, so the same plate always resolves to the same registry key regardless
    of digit script or letter/digit typing order."""
    s = str(v or "")
    for i, ch in enumerate(_AR_DIGITS):
        s = s.replace(ch, str(i))
    s = re.sub(r"\s+", "", s)
    digits = "".join(re.findall(r"\d+", s))
    letters = "".join(re.findall(r"[^\d]+", s))
    return digits + letters