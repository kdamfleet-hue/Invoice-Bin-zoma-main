import re

with open('app.py', 'r', encoding='utf-8') as f:
    app_py = f.read()

seed_route = """
@app.route("/api/seed_from_template")
def seed_from_template():
    import openpyxl
    template_path = os.path.join(app.root_path, "weekly_schedule_template.xlsx")
    if not os.path.exists(template_path):
        return "Template not found!", 404

    wb = openpyxl.load_workbook(template_path, data_only=True)

    def extract_rows(sheet_name, start_row, keys):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        data = []
        for r in range(start_row, ws.max_row + 1):
            row_data = {}
            has_value = False
            for c, key in enumerate(keys, start=2): # Start from col 2 (skip 'م')
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    has_value = True
                row_data[key] = str(val) if val is not None else ""
            if has_value:
                data.append(row_data)
        return data

    main_keys = ["empid", "name", "iqama", "job", "plate", "model", "vtype", "pallets", "load", "vserial", "inspect", "rem_days1", "license", "rem_days2", "drivercard", "rem_days3", "opcard", "rem_days4", "empNotes", "phone"]
    spare_keys = ["status", "plate", "model", "vtype", "pallets", "load", "vserial", "inspect", "rem_days1", "license", "rem_days2", "opcard", "rem_days3", "empNotes"]
    vac_keys = ["empid", "name", "iqama", "job", "drivercard", "rem_days1", "phone", "empNotes"]

    sd = {
        "main": extract_rows("المركبات النشطة", 5, main_keys),
        "spare": extract_rows("الأسبير والمعطلة", 4, spare_keys),
        "vacation": extract_rows("السائقون في إجازة", 4, vac_keys),
        "summary": {}
    }

    blob_set("schedule_data", sd)
    db.session.commit()
    return "تم سحب البيانات من ملف الإكسل وإدخالها في الموقع بنجاح!"
"""

if "def seed_from_template" not in app_py:
    app_py = app_py + "\n" + seed_route

with open('app.py', 'w', encoding='utf-8') as f:
    f.write(app_py)

print("Added seed_from_template route")
