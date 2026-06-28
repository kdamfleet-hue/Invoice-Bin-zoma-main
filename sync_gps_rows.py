"""
sync_gps_rows.py
----------------
يكتب 77 صف من ملف المصدر مباشرة إلى ملف الكشف مع الحفاظ على التنسيق.
"""
import os
import sys
# Fix Windows console encoding
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ── مسارات الملفات ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_SRC    = os.path.join(BASE_DIR, "GPS_source_data.xlsx")
FILE_TARGET = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026-1-26 - Copy.xlsx")
FILE_OUTPUT = os.path.join(BASE_DIR, "كشف اجهزة GPSللمركبات فرع الدمام 2026_محدث_مطابق.xlsx")

# ── التحقق من وجود الملفات ───────────────────────────────────────────────────
for f in [FILE_SRC, FILE_TARGET]:
    if not os.path.exists(f):
        print(f"[ERROR] الملف غير موجود: {f}")
        sys.exit(1)

# ── خريطة الأعمدة (اسم العمود في المصدر → رقم العمود في الهدف) ──────────────
COL_MAP = {
    'رقم الهيكل':    1,   # A
    'الرقم التسلسلي': 2,  # B
    'سنة الصنع':     3,   # C
    'الطراز':        4,   # D
    'الماركة':       5,   # E
    'نوع التسجيل':   6,   # F
    'الفرع':         7,   # G
    'رقم اللوحة':    9,   # I
}

# ── كتابة آمنة تدعم الخلايا المدمجة ─────────────────────────────────────────
def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
    else:
        cell.value = value

# ── تحميل البيانات ───────────────────────────────────────────────────────────
print("قراءة بيانات المصدر (77 صف)...")
df_src_full = pd.read_excel(FILE_SRC, header=3)
df_src_data = df_src_full.iloc[0:77].copy()
print(f"تم قراءة {len(df_src_data)} صف.")

# التحقق من الأعمدة
missing = [c for c in COL_MAP.keys() if c not in df_src_data.columns]
if missing:
    print(f"[WARNING] أعمدة ناقصة في المصدر: {missing}")

print("تحميل ملف الهدف...")
wb = load_workbook(FILE_TARGET)
ws = wb.active

# ── الكتابة ───────────────────────────────────────────────────────────────────
print("مزامنة الصفوف...")
START_ROW = 6

for i, (_, src_row) in enumerate(df_src_data.iterrows()):
    target_row = START_ROW + i
    for src_col, col_idx in COL_MAP.items():
        if src_col not in df_src_data.columns:
            continue
        val = src_row[src_col]
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            safe_write(ws, target_row, col_idx, val)
    # رقم تسلسلي في العمود J (10)
    safe_write(ws, target_row, 10, i + 1)

print(f"تمت مزامنة {len(df_src_data)} صف بنجاح.")
wb.save(FILE_OUTPUT)
print(f"تم الحفظ: {FILE_OUTPUT}")

