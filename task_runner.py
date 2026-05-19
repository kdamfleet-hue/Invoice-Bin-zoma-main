import os, glob, re, sqlite3
import openpyxl

def main():
    print("--- Fixing Templates ---")
    whatsapp_btn = '<a href="https://wa.me/?text=مرحباً" target="_blank" style="position: fixed; bottom: 20px; right: 20px; z-index: 999; background-color: #25d366; color: white; border-radius: 50px; padding: 15px 25px; box-shadow: 0 4px 10px rgba(0,0,0,0.2); font-weight: bold; text-decoration: none; display: flex; align-items: center; gap: 10px; transition: 0.3s; font-family: Cairo;"><svg width="24" height="24" viewBox="0 0 24 24" fill="currentColor" xmlns="http://www.w3.org/2000/svg"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.888-.788-1.489-1.761-1.662-2.06-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51a12.8 12.8 0 0 0-.57-.01c-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 0 1-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 0 1-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 0 1 2.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0 0 12.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 0 0 5.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 0 0-3.48-8.413z"/></svg>تواصل واتساب / إرسال نماذج</a>'
    logout_btn = '<a href="/logout" style="position: absolute; top: 1.5rem; left: 1.5rem; z-index: 100; background: #e74c3c; color: white; padding: 8px 20px; border-radius: 50px; font-weight: bold; font-size: 0.95rem; text-decoration: none; font-family: Cairo; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">تسجيل الخروج</a>'
    logo_replace = '<img src="{{ url_for(\'static\', filename=\'site_logo.jpg\') }}" alt="Logo" class="site-logo" style="width: 300px; max-width: 90%; height: auto; object-fit: contain; margin: 0 auto 1.5rem auto; display: block; filter: drop-shadow(0 4px 6px rgba(0,0,0,0.1));">'

    for html_file in glob.glob('templates/*.html'):
        with open(html_file, 'r', encoding='utf-8') as f:
            html = f.read()

        # Fix Logo
        # If it has the b64 block:
        html = re.sub(r'\{\%\s*if\s*b64_en\s*\%\}.*?\{\%\s*endif\s*\%\}', logo_replace, html, flags=re.DOTALL)
        # If it has the new_logo_clean.png
        html = re.sub(r'<img[^>]*src="\{\{\s*url_for\(\'static\',\s*filename=\'new_logo_clean\.png\'\)\s*\}\}"[^>]*>', logo_replace, html)

        # Add buttons if not login/password
        if 'login.html' not in html_file and 'password.html' not in html_file:
            if 'تواصل واتساب' not in html:
                html = html.replace('</body>', whatsapp_btn + '\n</body>')
            if 'تسجيل الخروج' not in html:
                html = html.replace('<body>', '<body>\n' + logout_btn)

        # Purchase specifics
        if 'purchase.html' in html_file:
            # 1. Add odometer
            if 'id="odometer"' not in html:
                html = html.replace('<div><label>رقم اللوحة:</label><input type="text" id="poPlate"></div>',
                                    '<div><label>رقم اللوحة:</label><input type="text" id="poPlate"></div>\n            <div><label>عداد السيارة (كم):</label><input type="number" id="odometer" required placeholder="مثال: 150000"></div>')
            # 2. Update summary modal payload (composeModal)
            if 'poPlate:' in html and 'odometer:' not in html:
                html = html.replace("model: document.getElementById('poModel').value,", "model: document.getElementById('poModel').value,\n                odometer: document.getElementById('odometer').value,")
            # 3. Ascending serial number
            html = html.replace('Math.floor(100000 + Math.random() * 900000).toString()', 'Math.floor(Date.now() / 1000).toString()')

        # Schedule print button
        if 'schedule.html' in html_file:
            html = html.replace('onclick="window.print()"', 'onclick="exportExcel()"')

        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html)

    print("--- Fixing app.py ---")
    with open('app.py', 'r', encoding='utf-8') as f:
        app_code = f.read()

    # 1. Fix generate_washing
    washing_fix = """
        def write_unmerged(sheet, r, c, val):
            for merge in sheet.merged_cells.ranges:
                if r >= merge.min_row and r <= merge.max_row and c >= merge.min_col and c <= merge.max_col:
                    sheet.cell(row=merge.min_row, column=merge.min_col).value = val
                    return
            sheet.cell(row=r, column=c).value = val

        for idx, item in enumerate(vehicles_data):
            row = 11 + idx
            write_unmerged(ws, row, 3, str(item.get("plate", "")))
            write_unmerged(ws, row, 4, str(item.get("car", "")))
            write_unmerged(ws, row, 6, str(item.get("driver", "")))
"""
    loop_pattern = r'for idx, item in enumerate\(vehicles_data\):\s+row = 11 \+ idx\s+ws\.cell\(row=row, column=3\)\.value = [^\n]+\s+ws\.cell\(row=row, column=4\)\.value = [^\n]+\s+ws\.cell\(row=row, column=6\)\.value = [^\n]+'
    if 'def write_unmerged' not in app_code:
        app_code = re.sub(loop_pattern, washing_fix.strip(), app_code)

    # 2. Fix generate_po to include odometer
    if 'ws["I8"] = data.get("model", "")' in app_code and 'odometer' not in app_code:
        app_code = app_code.replace('ws["I8"] = data.get("model", "")', 'ws["I8"] = data.get("model", "")\n        ws["D9"] = data.get("odometer", "")')

    # 3. Add /logout route
    if '@app.route("/logout")' not in app_code:
        logout_route = '''
@app.route("/logout")
def logout():
    session.pop("google_user", None)
    return redirect(url_for("login"))
'''
        app_code += logout_route

    with open('app.py', 'w', encoding='utf-8') as f:
        f.write(app_code)

    print("--- Modifying po_template.xlsx ---")
    try:
        wb = openpyxl.load_workbook('po_template.xlsx')
        ws = wb.active
        ws['C9'] = 'عداد السيارة:'
        wb.save('po_template.xlsx')
        print("Updated po_template.xlsx")
    except Exception as e:
        print("Error updating po_template.xlsx:", e)

    print("--- Updating DB with vehicle_lookup_v2.xlsx ---")
    try:
        wb = openpyxl.load_workbook(r'New\vehicle_lookup_v2.xlsx', data_only=True)
        ws = wb['البيانات']
        
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS drivers (
                     id INTEGER PRIMARY KEY AUTOINCREMENT,
                     name TEXT UNIQUE,
                     iqama TEXT,
                     car TEXT,
                     plate TEXT,
                     branch TEXT DEFAULT 'الدمام')''')
        
        for r in range(2, ws.max_row + 1):
            plate = str(ws.cell(row=r, column=1).value or '').strip()
            brand = str(ws.cell(row=r, column=4).value or '').strip()
            model = str(ws.cell(row=r, column=5).value or '').strip()
            year = str(ws.cell(row=r, column=6).value or '').strip()
            iqama = str(ws.cell(row=r, column=14).value or '').strip()
            name = str(ws.cell(row=r, column=15).value or '').strip()
            
            if name and name != 'None' and plate and plate != 'None':
                car = f"{brand} {model} {year}".replace('None', '').strip()
                # Upsert
                try:
                    c.execute("INSERT INTO drivers (name, iqama, car, plate, branch) VALUES (?, ?, ?, ?, 'الدمام')", (name, iqama, car, plate))
                except sqlite3.IntegrityError:
                    c.execute("UPDATE drivers SET iqama=?, car=?, plate=? WHERE name=?", (iqama, car, plate, name))
        
        conn.commit()
        conn.close()
        print("Database updated with employee names!")
    except Exception as e:
        print("Error updating DB:", e)

if __name__ == "__main__":
    main()
