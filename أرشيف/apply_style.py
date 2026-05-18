# -*- coding: utf-8 -*-
"""
Apply the formatting/style from بيان_استهلاك_الزيوت_نهائي_مميز (1).xlsx
to 09e1a2c0-3b83-4591-8159-16994670f3c9.xlsx

Source style summary:
- RTL direction
- Font: Cairo throughout
- Header row: blue (#2E56A5), white text, bold, Cairo 11
- Data rows: alternating white (#FFFFFF) / cream (#F0EDE6)
- Title: Cairo 14, bold, blue text (#2E56A5), merged across all cols
- Rows 1-3: merged for logo/header area with 2 images
- Thin black borders on all cells
- All text centered horizontally and vertically
- Column A (م): numbered, bold
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))

# ===== COLORS (from source file) =====
BLUE_HEADER = "2E56A5"    # Header background, title text, section headers
WHITE = "FFFFFF"           # White fill for odd rows
CREAM = "F0EDE6"          # Cream fill for even rows
ACCENT_GOLD = "C8A45A"    # Gold accent for section headers
TEXT_DARK = "2C3E50"       # Dark text color for data
DARK_BLUE = "1A3A5C"      # Darker blue for row numbers
LIGHT_BG = "F5F3EE"       # Light background for notes
BLACK = "000000"           # Border color

# ===== STYLES =====
thin_side = Side(style='thin', color=BLACK)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# ===== LOAD TARGET FILE =====
target_path = os.path.join(script_dir, '09e1a2c0-3b83-4591-8159-16994670f3c9.xlsx')
wb = load_workbook(target_path)
ws = wb.active

# ===== READ ALL DATA FIRST =====
# Store all cell values before we modify anything
all_data = {}
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            all_data[(row, col)] = cell.value

# Extract images from source file
source_path = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_نهائي_مميز (1).xlsx')
source_wb = load_workbook(source_path)
source_ws = source_wb.active
source_images = []
for img in source_ws._images:
    img_data = img._data()
    anchor = img.anchor
    from_col = anchor._from.col if hasattr(anchor, '_from') else 0
    from_row = anchor._from.row if hasattr(anchor, '_from') else 0
    from_colOff = anchor._from.colOff if hasattr(anchor, '_from') else 0
    from_rowOff = anchor._from.rowOff if hasattr(anchor, '_from') else 0
    source_images.append({
        'data': img_data,
        'width': img.width,
        'height': img.height,
        'format': img.format or 'png',
        'from_col': from_col,
        'from_row': from_row,
        'from_colOff': from_colOff,
        'from_rowOff': from_rowOff,
    })

# ===== CREATE NEW WORKBOOK =====
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "جدول تحديث المركبات"
new_ws.sheet_view.rightToLeft = True

# ===== TARGET FILE STRUCTURE =====
# Rows 1-6: Logo area (merged A1:K6) -> map to rows 1-3 (source style)
# Row 7: Title (merged A7:K7) -> map to row 4 (source style)
# Row 8: Headers -> map to row 5 (source style)
# Rows 9-36: Main data (28 rows) -> map starting row 6
# Row 37: spacer
# Row 38: Sub-header "المركبات الاسبير"
# Rows 39-46: Spare vehicles data (8 rows)
# Row 47: spacer
# Row 48: Summary header
# Row 49: Summary data

# ===== COLUMN WIDTHS (11 columns A-K) =====
# Match source style proportions but for 11 columns
col_widths = {
    'A': 5,      # م
    'B': 15,     # الرقم الوظيفي
    'C': 22,     # اسم السائق
    'D': 18,     # رقم الإقامة
    'E': 12,     # الوظيفة
    'F': 14,     # رقم اللوحة
    'G': 18,     # موديل المركبة
    'H': 16,     # نوع المركبة
    'I': 10,     # عدد الطبالي
    'J': 12,     # حمولة المركبة
    'K': 25,     # الملاحظات
}
for letter, w in col_widths.items():
    new_ws.column_dimensions[letter].width = w

MAX_COL = 11  # K

row_cursor = 1

# ===== ROWS 1-3: LOGO AREA (like source rows 1-3) =====
new_ws.merge_cells(f'A{row_cursor}:K{row_cursor + 2}')
for r in range(row_cursor, row_cursor + 3):
    new_ws.row_dimensions[r].height = 22

# Add images from source file
for i, img_info in enumerate(source_images):
    img_path = os.path.join(script_dir, f'source_img_{i}.{img_info["format"]}')
    if os.path.exists(img_path):
        img = XlImage(img_path)
        img.width = img_info['width']
        img.height = img_info['height']
        if i == 0:  # Right-side image (company logo)
            new_ws.add_image(img, 'H1')
        else:  # Left-side image
            new_ws.add_image(img, 'A1')

row_cursor = 4

# ===== ROW 4: TITLE (like source row 4) =====
new_ws.merge_cells(f'A{row_cursor}:K{row_cursor}')
title_text = all_data.get((7, 1), 'جدول تحديث مركبات وموصلين و موزعين فرع الدمام')
new_ws[f'A{row_cursor}'] = title_text
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=14, bold=True, color=BLUE_HEADER)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
new_ws.row_dimensions[row_cursor].height = 25.5
row_cursor = 5

# ===== ROW 5: HEADERS (like source row 5) =====
target_headers = [
    all_data.get((8, 1), 'No.'),        # A - م
    all_data.get((8, 2), 'الرقم الوظيفي'),  # B
    all_data.get((8, 3), 'اسم السائق'),     # C
    all_data.get((8, 4), 'رقم الإقامة'),     # D
    all_data.get((8, 5), 'الوظيفة'),         # E
    all_data.get((8, 6), 'رقم اللوحة'),      # F
    all_data.get((8, 7), 'موديل المركبة'),    # G
    all_data.get((8, 8), 'نوع المركبة'),      # H
    all_data.get((8, 9), 'عدد الطبالي'),      # I
    all_data.get((8, 10), 'حمولة المركبة(عدد الركاب)'),  # J
    all_data.get((8, 11), 'الملاحظات'),       # K
]

# Rename first header to match source style
target_headers[0] = 'م'

new_ws.row_dimensions[row_cursor].height = 24.75
for i, h in enumerate(target_headers):
    cell = new_ws.cell(row=row_cursor, column=i + 1, value=h)
    cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
    cell.fill = fl(BLUE_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
row_cursor = 6

# ===== DATA ROWS (rows 9-36 in target -> starting at row 6 in new) =====
data_start_row_target = 9
data_end_row_target = 36

for target_row in range(data_start_row_target, data_end_row_target + 1):
    new_ws.row_dimensions[row_cursor].height = 17.25
    r_idx = target_row - data_start_row_target  # 0-based index for alternating colors
    
    for col in range(1, MAX_COL + 1):
        cell = new_ws.cell(row=row_cursor, column=col)
        val = all_data.get((target_row, col), '')
        
        # Apply value
        cell.value = val if val != '' else None
        
        # Apply alternating row colors (odd index = white, even index = cream - like source)
        if r_idx % 2 == 0:
            cell.fill = fl(WHITE)
        else:
            cell.fill = fl(CREAM)
        
        # Apply font
        if col == 1:  # Row number column
            cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        else:
            cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        
        # Apply alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply borders
        cell.border = thin_border
    
    row_cursor += 1

# ===== SPACER ROW =====
new_ws.row_dimensions[row_cursor].height = 10
row_cursor += 1

# ===== SECTION HEADER: المركبات الاسبير (like source notes section - gold) =====
new_ws.merge_cells(f'A{row_cursor}:K{row_cursor}')
spare_title = all_data.get((38, 2), 'المركبات الاسبير')
# Make it look like the source "ملاحظات وتفاصيل الفلاتر" header
new_ws[f'A{row_cursor}'] = f'  {spare_title}'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
new_ws[f'A{row_cursor}'].fill = fl(ACCENT_GOLD)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
new_ws.row_dimensions[row_cursor].height = 38
row_cursor += 1

# Spacer
new_ws.row_dimensions[row_cursor].height = 8
row_cursor += 1

# ===== SUB-HEADERS for spare vehicles (like source filter headers - blue) =====
spare_headers = [
    all_data.get((38, 1), 'م'),
    all_data.get((38, 2), 'المركبات الاسبير'),
    all_data.get((38, 3), 'نوع المركبه و الموديل'),
    all_data.get((38, 4), 'عدد الطبالي'),
    all_data.get((38, 5), 'حمولة المركبة(عدد الركاب)'),
    all_data.get((38, 6), 'الملاحظات'),
]
# Use 6 columns for spare section, but merge rest
spare_col_map = [1, 2, 3, 4, 5, 6]  # A-F for this section, then merge G-K for notes

new_ws.merge_cells(f'A{row_cursor}:K{row_cursor}')
new_ws[f'A{row_cursor}'] = 'بيان المركبات الاسبير والمعطلة'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
new_ws[f'A{row_cursor}'].fill = fl(BLUE_HEADER)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
new_ws.row_dimensions[row_cursor].height = 30
row_cursor += 1

# Spare headers row
spare_headers_display = ['م', 'رقم اللوحة', 'نوع المركبه و الموديل', 'عدد الطبالي', 'حمولة المركبة', 'الملاحظات']
# Merge F-K for الملاحظات
new_ws.merge_cells(f'F{row_cursor}:K{row_cursor}')
new_ws.row_dimensions[row_cursor].height = 30
for i, h in enumerate(spare_headers_display):
    col_idx = i + 1
    cell = new_ws.cell(row=row_cursor, column=col_idx, value=h)
    cell.font = Font(name='Cairo', size=10, bold=True, color=WHITE)
    cell.fill = fl(BLUE_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
# Fill merged cells with style
for c in range(7, MAX_COL + 1):
    cell = new_ws.cell(row=row_cursor, column=c)
    cell.fill = fl(BLUE_HEADER)
    cell.border = thin_border
row_cursor += 1

# ===== SPARE VEHICLES DATA (rows 39-46) =====
spare_start = 39
spare_end = 46
for target_row in range(spare_start, spare_end + 1):
    new_ws.row_dimensions[row_cursor].height = 17.25
    r_idx = target_row - spare_start
    
    # Map: A=م, B=رقم اللوحة, C=نوع, D=عدد الطبالي, E=حمولة, F-K=الملاحظات
    spare_data = [
        all_data.get((target_row, 1), ''),  # م
        all_data.get((target_row, 2), ''),  # رقم اللوحة
        all_data.get((target_row, 3), ''),  # نوع المركبه
        all_data.get((target_row, 4), ''),  # عدد الطبالي
        all_data.get((target_row, 5), ''),  # حمولة المركبة
        all_data.get((target_row, 6), ''),  # الملاحظات
    ]
    
    # Merge F-K for notes
    new_ws.merge_cells(f'F{row_cursor}:K{row_cursor}')
    
    for i, val in enumerate(spare_data):
        col_idx = i + 1
        cell = new_ws.cell(row=row_cursor, column=col_idx)
        cell.value = val if val != '' else None
        
        if r_idx % 2 == 0:
            cell.fill = fl(WHITE)
        else:
            cell.fill = fl(CREAM)
        
        if col_idx == 1:
            cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        else:
            cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Style merged cells
    for c in range(7, MAX_COL + 1):
        cell = new_ws.cell(row=row_cursor, column=c)
        if r_idx % 2 == 0:
            cell.fill = fl(WHITE)
        else:
            cell.fill = fl(CREAM)
        cell.border = thin_border
    
    row_cursor += 1

# ===== SPACER =====
new_ws.row_dimensions[row_cursor].height = 10
row_cursor += 1

# ===== SUMMARY SECTION (like source diesel section - blue header) =====
new_ws.merge_cells(f'A{row_cursor}:K{row_cursor}')
new_ws[f'A{row_cursor}'] = 'ملخص الأعداد'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
new_ws[f'A{row_cursor}'].fill = fl(BLUE_HEADER)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
new_ws.row_dimensions[row_cursor].height = 32
row_cursor += 1

# Summary headers (row 48 in target)
summary_headers = [
    all_data.get((48, 1), 'م'),
    all_data.get((48, 2), 'عدد السائقيين'),
    all_data.get((48, 3), 'عدد الدينات'),
    all_data.get((48, 4), 'عدد اللواري'),
    all_data.get((48, 5), 'عدد الموصلين'),
    all_data.get((48, 6), 'عدد الموزعين'),
]
# Merge G-K for this section
new_ws.merge_cells(f'G{row_cursor}:K{row_cursor}')
new_ws.row_dimensions[row_cursor].height = 30
for i, h in enumerate(summary_headers):
    cell = new_ws.cell(row=row_cursor, column=i + 1, value=h)
    cell.font = Font(name='Cairo', size=10, bold=True, color=WHITE)
    cell.fill = fl(BLUE_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
for c in range(7, MAX_COL + 1):
    cell = new_ws.cell(row=row_cursor, column=c)
    cell.fill = fl(BLUE_HEADER)
    cell.border = thin_border
row_cursor += 1

# Summary data (row 49 in target)
summary_data = [
    all_data.get((49, 1), 1),
    all_data.get((49, 2), ''),
    all_data.get((49, 3), ''),
    all_data.get((49, 4), ''),
    all_data.get((49, 5), ''),
    all_data.get((49, 6), ''),
]
new_ws.merge_cells(f'G{row_cursor}:K{row_cursor}')
new_ws.row_dimensions[row_cursor].height = 25
for i, val in enumerate(summary_data):
    cell = new_ws.cell(row=row_cursor, column=i + 1)
    cell.value = val if val != '' else None
    cell.fill = fl(WHITE)
    cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
for c in range(7, MAX_COL + 1):
    cell = new_ws.cell(row=row_cursor, column=c)
    cell.fill = fl(WHITE)
    cell.border = thin_border
row_cursor += 1

# ===== SPACER =====
new_ws.row_dimensions[row_cursor].height = 12
row_cursor += 1

# ===== CLOSING NOTE (like source) =====
new_ws.merge_cells(f'A{row_cursor}:K{row_cursor}')
new_ws[f'A{row_cursor}'] = 'تم إعداد هذا الجدول بواسطة قسم الحركة'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=12, bold=True, color=BLUE_HEADER)
new_ws[f'A{row_cursor}'].fill = fl(LIGHT_BG)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
new_ws[f'A{row_cursor}'].border = Border(
    top=Side(style='medium', color=ACCENT_GOLD),
    bottom=Side(style='medium', color=ACCENT_GOLD)
)
new_ws.row_dimensions[row_cursor].height = 38
row_cursor += 1

# ===== SPACER =====
new_ws.row_dimensions[row_cursor].height = 30
row_cursor += 1

# ===== SIGNATURES (like source) =====
new_ws.merge_cells(f'A{row_cursor}:E{row_cursor}')
new_ws[f'A{row_cursor}'] = 'قسم الحركة'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=11, bold=True, color=DARK_BLUE)
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells(f'F{row_cursor}:K{row_cursor}')
new_ws[f'F{row_cursor}'] = 'التعميد'
new_ws[f'F{row_cursor}'].font = Font(name='Cairo', size=11, bold=True, color=DARK_BLUE)
new_ws[f'F{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
row_cursor += 1

new_ws.merge_cells(f'A{row_cursor}:E{row_cursor}')
new_ws[f'A{row_cursor}'] = 'خالد الغامدي'
new_ws[f'A{row_cursor}'].font = Font(name='Cairo', size=10, color='6B7B8D')
new_ws[f'A{row_cursor}'].alignment = Alignment(horizontal='center', vertical='center')
row_cursor += 1

# Signature lines
new_ws.row_dimensions[row_cursor].height = 8
row_cursor += 1
for c in range(1, 6):
    new_ws.cell(row=row_cursor, column=c).border = Border(bottom=Side(style='thin', color='CCCCCC'))
for c in range(6, MAX_COL + 1):
    new_ws.cell(row=row_cursor, column=c).border = Border(bottom=Side(style='thin', color='CCCCCC'))

# ===== PRINT SETUP =====
new_ws.page_setup.orientation = 'landscape'
new_ws.page_setup.fitToWidth = 1
new_ws.page_setup.fitToHeight = 0
new_ws.sheet_properties.pageSetUpPr.fitToPage = True
new_ws.page_margins.left = 0.4
new_ws.page_margins.right = 0.4
new_ws.page_margins.top = 0.3
new_ws.page_margins.bottom = 0.3

# ===== SAVE =====
output_path = os.path.join(script_dir, '09e1a2c0-3b83-4591-8159-16994670f3c9_styled.xlsx')
new_wb.save(output_path)
print(f"Styled Excel saved: {output_path}")
print("Done!")
