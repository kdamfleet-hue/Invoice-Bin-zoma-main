import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
doc_path = os.path.join(script_dir, 'نموذج طلب شراء رقم 2222.doc')
excel_path = os.path.join(script_dir, 'نموذج_طلب_شراء_رقم_2222.xlsx')

# Unzip and parse XML
ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}

with zipfile.ZipFile(doc_path, 'r') as z:
    xml_content = z.read('word/document.xml')

root = ET.fromstring(xml_content)

wb = openpyxl.Workbook()
ws = wb.active
ws.sheet_view.rightToLeft = True
ws.title = "نموذج طلب شراء"

row_idx = 1
# Find all paragraphs and tables. 
# Body contains p and tbl.
body = root.find('w:body', ns)
for element in body:
    if element.tag.endswith('p'):
        # Extract text from paragraph
        texts = [t.text for t in element.findall('.//w:t', ns) if t.text]
        if texts:
            text = "".join(texts)
            ws.cell(row=row_idx, column=1, value=text)
            row_idx += 1
    elif element.tag.endswith('tbl'):
        # Extract table
        for tr in element.findall('.//w:tr', ns):
            col_idx = 1
            for tc in tr.findall('.//w:tc', ns):
                # Text in a cell
                texts = [t.text for t in tc.findall('.//w:t', ns) if t.text]
                ws.cell(row=row_idx, column=col_idx, value="".join(texts))
                col_idx += 1
            row_idx += 1
        row_idx += 1 # Add empty row after table

wb.save(excel_path)
print("Saved word contents to", excel_path)
