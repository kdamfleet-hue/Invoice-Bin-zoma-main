import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
output_file = os.path.join(script_dir, 'نموذج_طلب_شراء_رقم_2222_محول.xlsx')
img_0_path = os.path.join(script_dir, 'source_img_0.png')
img_1_path = os.path.join(script_dir, 'source_img_1.png')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "نموذج طلب شراء"
ws.sheet_view.rightToLeft = True

# ===== COLORS & STYLES =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
ODD_ROW = "FFFFFF"
TEXT_DARK = "2C3E50"
LIGHT_BG = "F5F3EE"

thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
accent_side = Side(style='medium', color=ACCENT)

thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'), top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

ws.sheet_view.showGridLines = False

# Base font
for r in range(1, 80):
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name='Cairo', size=11, color=TEXT_DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row = 1

# Header Images
try:
    img_ar = XlImage(img_1_path)
    img_ar.width = 200
    img_ar.height = 70
    ws.add_image(img_ar, 'A1')
    
    img_en = XlImage(img_0_path)
    img_en.width = 200
    img_en.height = 70
    ws.add_image(img_en, 'G1')
except Exception as e:
    pass

ws.row_dimensions[row].height = 40
row += 1
ws.row_dimensions[row].height = 40
row += 1

# Title
ws.merge_cells(f'A{row}:H{row}')
title_cell = ws[f'A{row}']
title_cell.value = "نموذج طلب شراء وإصلاح سيارة"
title_cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
ws.row_dimensions[row].height = 40
for c in range(1, 9):
    ws.cell(row=row, column=c).border = Border(bottom=accent_side)
row += 2

# Section 1: بيانات السيارة والسائق
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "بيانات السيارة والسائق"
ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl(PRIMARY)
ws[f'A{row}'].border = header_border
ws.row_dimensions[row].height = 30
row += 1

ws[f'A{row}'] = "اسم السائق:"; ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws.merge_cells(f'B{row}:C{row}')
ws[f'D{row}'] = "الفرع:"; ws[f'D{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws[f'E{row}'] = "الدمام"; ws[f'E{row}'].font = Font(name='Cairo', size=11)
ws[f'F{row}'] = "الوظيفة:"; ws[f'F{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws.merge_cells(f'G{row}:H{row}')
ws.row_dimensions[row].height = 30
for c in range(1, 9): 
    ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=c).fill = fl(EVEN_ROW)
row += 1

ws[f'A{row}'] = "نوع السيارة:"; ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws.merge_cells(f'B{row}:C{row}')
ws[f'D{row}'] = "رقم اللوحة:"; ws[f'D{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws[f'E{row}'] = " "
ws[f'F{row}'] = "الموديل:"; ws[f'F{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws.merge_cells(f'G{row}:H{row}')
ws.row_dimensions[row].height = 30
for c in range(1, 9): 
    ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=c).fill = fl(ODD_ROW)
row += 2

# Section 2: قطع الغيار
def draw_table(title, cols, num_rows, bg_color):
    global row
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = title
    ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
    ws[f'A{row}'].fill = fl(bg_color)
    ws.row_dimensions[row].height = 30
    for c in range(1, 9): ws.cell(row=row, column=c).border = header_border
    row += 1
    
    col_idx = 1
    for cw, cname in cols:
        cell = ws.cell(row=row, column=col_idx)
        cell.value = cname
        cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        cell.fill = fl(PRIMARY)
        if cw > 1:
            ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx+cw-1)
        for i in range(cw):
            ws.cell(row=row, column=col_idx+i).border = header_border
        col_idx += cw
    ws.row_dimensions[row].height = 30
    row += 1
    
    for i in range(num_rows):
        col_idx = 1
        bg = fl(EVEN_ROW) if i % 2 == 0 else fl(ODD_ROW)
        for cw, cname in cols:
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = bg
            if cname == 'م': cell.value = i+1
            if cw > 1:
                ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx+cw-1)
            for j in range(cw):
                ws.cell(row=row, column=col_idx+j).border = thin_border
                ws.cell(row=row, column=col_idx+j).fill = bg
            col_idx += cw
        ws.row_dimensions[row].height = 25
        row += 1
    row += 1

draw_table("قطع الغيار", [(1, 'م'), (3, 'البيان'), (1, 'الكمية'), (1, 'السعر'), (1, 'القيمة'), (1, 'ملاحظات')], 3, PRIMARY)
draw_table("الإصلاح", [(1, 'م'), (4, 'الوصف'), (1, 'القيمة'), (2, 'ملاحظات')], 3, "34495E")
draw_table("الكفرات", [(1, 'م'), (1, 'تاريخ التغيير'), (1, 'العدد'), (1, 'أمامية'), (1, 'خلفية'), (1, 'قراءة سابقة'), (1, 'قراءة حالية'), (1, 'المسافة')], 2, PRIMARY)
draw_table("البطاريات", [(1, 'م'), (2, 'الوصف'), (1, 'العدد'), (1, 'المقاس'), (1, 'أمبير'), (1, 'السعر للوحدة'), (1, 'التاريخ')], 2, "34495E")

# الإجمالي
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "الإجمالي"
ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl(ACCENT)
for c in range(1, 9): ws.cell(row=row, column=c).border = header_border
ws.row_dimensions[row].height = 30
row += 1

cols = ["قطع الغيار", "الإصلاح", "الكفرات", "البطاريات", "الإجمالي شامل الضريبة", "ملاحظات"]
for i, cname in enumerate(cols, 1):
    cw = 1 if i < 5 else 2
    col_idx = i if i < 5 else (5 if i == 5 else 7)
    cell = ws.cell(row=row, column=col_idx, value=cname)
    cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
    cell.fill = fl(PRIMARY)
    if cw > 1:
        ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx+cw-1)
    for j in range(cw): ws.cell(row=row, column=col_idx+j).border = header_border
ws.row_dimensions[row].height = 30
row += 1

# row 2 of total
for i in range(1, 7):
    cw = 1 if i < 5 else 2
    col_idx = i if i < 5 else (5 if i == 5 else 7)
    if cw > 1:
        ws.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx+cw-1)
    for j in range(cw): 
        ws.cell(row=row, column=col_idx+j).border = thin_border
        ws.cell(row=row, column=col_idx+j).fill = fl(EVEN_ROW)
ws.row_dimensions[row].height = 30
row += 2

# Signatures
sigs = ["السائق", "الميكانيكي", "مسئول الورشة", "قسم الحركة", "مسئول الشراء", "الحسابات", "مدير الفرع", "اعتماد الإدارة"]
for i, sig in enumerate(sigs, 1):
    cell = ws.cell(row=row, column=i, value=sig)
    cell.font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws.row_dimensions[row].height = 30
row += 1

for i in range(1, 9):
    cell = ws.cell(row=row, column=i, value="")
    cell.border = Border(bottom=accent_side)
ws.row_dimensions[row].height = 40
row += 2

# Notes
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "ملاحظات هامة"
ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl(PRIMARY)
for c in range(1, 9): ws.cell(row=row, column=c).border = header_border
ws.row_dimensions[row].height = 30
row += 1

notes = """1- يعبأ الطلب من قبل مسئول الورشة بعد توقيعه وتوقيع السائق والميكانيكي.
2- يوقع بعد ذلك من المحاسبة ويتم ارساله للإدارة عن طريق محاسب الفرع ليتم اعتماده من الإدارة.
3- عند وجود ملاحظة لأي من الأطراف يمكن تدوينها في حقل الملاحظات أعلاه.
4- ترفق صورة للإدارة المالية وصورة لقسم الحركة بالإدارة، ويتم حفظ صورة من الطلب لدى الورشة."""
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = notes
ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=TEXT_DARK)
ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
ws[f'A{row}'].fill = fl(LIGHT_BG)
for c in range(1, 9): ws.cell(row=row, column=c).border = thin_border
ws.row_dimensions[row].height = 90

# Set column widths
widths = [6, 14, 12, 14, 14, 15, 15, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25

wb.save(output_file)


