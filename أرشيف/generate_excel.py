# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
from data import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(script_dir, 'output-onlinepngtools.png')

wb = Workbook()
ws = wb.active
ws.title = "بيان الاستهلاك"
ws.sheet_view.rightToLeft = True

# ===== COLORS =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
TEXT_DARK = "2C3E50"
TEXT_LIGHT = "6B7B8D"
LIGHT_BG = "F5F3EE"
GREEN_TEXT = "27AE60"

# ===== STYLES =====
thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'),
                       top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# ===== COLUMN WIDTHS (8 columns) =====
widths = [5, 16, 14, 12, 12, 16, 10, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1

# ===== LOGO (rows 1-4) =====
ws.merge_cells(f'A{row}:H{row+3}')
for r in range(row, row+4):
    ws.row_dimensions[r].height = 22
try:
    img = XlImage(logo_path)
    img.width = 400
    img.height = 80
    ws.add_image(img, 'D1')
except:
    ws[f'A{row}'] = "شركة بن زومة للتجارة الدولية والإنماء المحدودة"
    ws[f'A{row}'].font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
row = 5

# ===== ACCENT LINE =====
for c in range(1, 9):
    ws.cell(row=row, column=c).border = Border(bottom=Side(style='medium', color=ACCENT))
ws.row_dimensions[row].height = 6
row = 6

# ===== TITLE =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = TITLE
ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=PRIMARY)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[row].height = 45
row = 7

# ===== SPACER =====
ws.row_dimensions[row].height = 10
row = 8

# ===== TABLE HEADERS =====
ws.row_dimensions[row].height = 35
for i, h in enumerate(HEADERS):
    start_c = i + 1
    end_c = i + 1
    if i == 6: # Last header "عدد فلاتر..."
        end_c = 8
        ws.merge_cells(start_row=row, start_column=start_c, end_row=row, end_column=end_c)
    
    cell = ws.cell(row=row, column=start_c, value=h)
    cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
    cell.fill = fl(PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = header_border
    if start_c != end_c:
        ws.cell(row=row, column=end_c).fill = fl(PRIMARY)
        ws.cell(row=row, column=end_c).border = header_border
row = 9

# ===== TABLE DATA (with plate letters) =====
for r_idx, data_row in enumerate(TABLE_DATA):
    ws.row_dimensions[row].height = 27
    for c_idx, val in enumerate(data_row):
        start_c = c_idx + 1
        end_c = c_idx + 1
        if c_idx == 6:
            end_c = 8
            ws.merge_cells(start_row=row, start_column=start_c, end_row=row, end_column=end_c)
            
        cell = ws.cell(row=row, column=start_c)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if start_c != end_c:
            ws.cell(row=row, column=end_c).border = thin_border
            
        cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)
        if r_idx % 2 == 0:
            cell.fill = fl(EVEN_ROW)
            if start_c != end_c: ws.cell(row=row, column=end_c).fill = fl(EVEN_ROW)

        if c_idx == 0:  # Row number
            cell.value = val
            cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
        elif c_idx == 1:  # Plate number - add letters
            plate_num = val
            cell.value = plate_display(plate_num)
            cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        elif c_idx == 6:  # Filter column
            if str(val).strip() == "0":
                cell.value = "لا يوجد"
                cell.font = Font(name='Cairo', size=9, color=TEXT_LIGHT)
            else:
                cell.value = val
                cell.font = Font(name='Cairo', size=10, bold=True, color=GREEN_TEXT)
        else:
            cell.value = val
    row += 1

# ===== SPACER =====
row += 1
ws.row_dimensions[row].height = 15
row += 1

# ===== SECTION: NOTES HEADER =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "📋  ملاحظات وتفاصيل الفلاتر"
ws[f'A{row}'].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl(ACCENT)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 38
row += 1
ws.row_dimensions[row].height = 8
row += 1

# ===== OIL PURCHASE INFO =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = NOTES_TITLE
ws[f'A{row}'].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[row].height = 35
row += 1

ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = NOTES_SUBTITLE.replace('\n', '  |  ')
ws[f'A{row}'].font = Font(name='Cairo', size=10, color=TEXT_LIGHT)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws[f'A{row}'].fill = fl(LIGHT_BG)
ws.row_dimensions[row].height = 32
row += 1
ws.row_dimensions[row].height = 8
row += 1

# ===== FILTERS TITLE =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = FILTERS_TITLE
ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl(PRIMARY)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[row].height = 38
row += 1

# ===== FILTERS TABLE HEADERS =====
# Layout: [م | نوع الفلتر | المستهلك | العدد | م | نوع الفلتر | المستهلك | العدد]
filter_headers = [
    (1, "م"), (2, "نوع الفلتر"), (3, "المستهلك"), (4, "العدد"),
    (5, "م"), (6, "نوع الفلتر"), (7, "المستهلك"), (8, "العدد")
]

ws.row_dimensions[row].height = 30
for col, label in filter_headers:
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = Font(name='Cairo', size=10, bold=True, color=WHITE)
    cell.fill = fl(PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = header_border
row += 1

# ===== FILTERS DATA (2-column layout with المستهلك and العدد) =====
half = (len(FILTERS_LIST) + 1) // 2
for i in range(half):
    ws.row_dimensions[row].height = 25
    bg = fl(EVEN_ROW) if i % 2 == 0 else fl(WHITE)

    # LEFT side
    idx_l = i
    name_l, orig_l, count_l, used_l = FILTERS_LIST[idx_l]
    # Add "- 2" to name if used exists, per user saying "المستهلك في الشهر السابق.. والـ= بالعدد"
    display_l = f"{name_l} - {used_l}" if used_l else name_l

    ws.cell(row=row, column=1, value=idx_l + 1).font = Font(name='Cairo', size=9, bold=True, color=PRIMARY)
    ws.cell(row=row, column=2, value=display_l).font = Font(name='Cairo', size=9, color=TEXT_DARK)
    ws.cell(row=row, column=3, value=orig_l).font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
    ws.cell(row=row, column=4, value=f"= {count_l}").font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
    
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = bg
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).alignment = Alignment(horizontal='center' if c!=2 else 'right', vertical='center')

    # RIGHT side
    idx_r = half + i
    if idx_r < len(FILTERS_LIST):
        name_r, orig_r, count_r, used_r = FILTERS_LIST[idx_r]
        display_r = f"{name_r} - {used_r}" if used_r else name_r

        ws.cell(row=row, column=5, value=idx_r + 1).font = Font(name='Cairo', size=9, bold=True, color=PRIMARY)
        ws.cell(row=row, column=6, value=display_r).font = Font(name='Cairo', size=9, color=TEXT_DARK)
        ws.cell(row=row, column=7, value=orig_r).font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
        ws.cell(row=row, column=8, value=f"= {count_r}").font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
        
        for c in range(5, 9):
            ws.cell(row=row, column=c).fill = bg
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).alignment = Alignment(horizontal='center' if c!=6 else 'right', vertical='center')
    else:
        for c in range(5, 9):
            ws.cell(row=row, column=c).fill = bg
            ws.cell(row=row, column=c).border = thin_border

    row += 1

# ===== SPACER =====
ws.row_dimensions[row].height = 10
row += 1

# ===== DIESEL FILTERS =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = "فلاتر الديزل"
ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
ws[f'A{row}'].fill = fl("34495E")
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 32
row += 1

for d_item in DIESEL_FILTERS:
    ws.merge_cells(f'A{row}:H{row}')
    is_sub = d_item.startswith("تركب")
    ws[f'A{row}'] = ("    ↳ " + d_item) if is_sub else ("  •  " + d_item)
    ws[f'A{row}'].font = Font(name='Cairo', size=10, color=TEXT_LIGHT if is_sub else TEXT_DARK, italic=is_sub)
    ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

# ===== SPACER =====
ws.row_dimensions[row].height = 12
row += 1

# ===== CLOSING NOTE =====
ws.merge_cells(f'A{row}:H{row}')
ws[f'A{row}'] = CLOSING_NOTE.replace('\n', '  —  ')
ws[f'A{row}'].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
ws[f'A{row}'].fill = fl(LIGHT_BG)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'A{row}'].border = Border(top=Side(style='medium', color=ACCENT), bottom=Side(style='medium', color=ACCENT))
ws.row_dimensions[row].height = 38
row += 1

# ===== SPACER =====
ws.row_dimensions[row].height = 30
row += 1

# ===== SIGNATURES =====
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "التعميد"
ws[f'A{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells(f'E{row}:H{row}')
ws[f'E{row}'] = "قسم الحركة"
ws[f'E{row}'].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

ws.merge_cells(f'E{row}:H{row}')
ws[f'E{row}'] = "خالد الغامدي"
ws[f'E{row}'].font = Font(name='Cairo', size=10, color=TEXT_LIGHT)
ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

ws.row_dimensions[row].height = 8
row += 1
for c in range(1, 5):
    ws.cell(row=row, column=c).border = Border(bottom=Side(style='thin', color='CCCCCC'))
for c in range(5, 9):
    ws.cell(row=row, column=c).border = Border(bottom=Side(style='thin', color='CCCCCC'))

# ===== PRINT SETUP =====
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3

excel_path = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_محسن.xlsx')
wb.save(excel_path)
print(f"Excel saved: {excel_path}")
print("Done!")
