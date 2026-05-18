import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_بدون_فراغات.xlsx')
output_file = os.path.join(script_dir, 'بيان_استهلاك_الزيوت_النهائي.xlsx')

wb = openpyxl.load_workbook(input_file)
ws = wb.active
ws.sheet_view.rightToLeft = True

# ===== COLORS & STYLES =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
ODD_ROW = "FFFFFF"
TEXT_DARK = "2C3E50"
TEXT_LIGHT = "6B7B8D"
GREEN_TEXT = "27AE60"
RED_TEXT = "C0392B"

thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'), top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# Apply base font
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.font = Font(name='Cairo', size=10, color=TEXT_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Find sections dynamically
main_header_row = None
filter_header_row = None

for row in ws.iter_rows(min_row=1, max_row=100):
    val = str(row[0].value).strip() if row[0].value else ""
    
    if "بيان استهلاك الزيوت" in val:
        ws.row_dimensions[row[0].row].height = 45
        row[0].font = Font(name='Cairo', size=14, bold=True, color=PRIMARY)
        if row[0].row > 1:
            for c in range(1, 11):
                ws.cell(row=row[0].row - 1, column=c).border = Border(bottom=Side(style='medium', color=ACCENT))
            
    elif val == 'م' and not main_header_row:
        main_header_row = row[0].row
        ws.row_dimensions[main_header_row].height = 35
        for cell in row:
            if cell.value:
                cell.font = Font(name='Cairo', size=11, bold=True, color=WHITE)
                cell.fill = fl(PRIMARY)
                cell.border = header_border
                
    elif "ملاحظات وتفاصيل الفلاتر" in val:
        ws.row_dimensions[row[0].row].height = 38
        row[0].font = Font(name='Cairo', size=13, bold=True, color=WHITE)
        row[0].fill = fl(ACCENT)
        
    elif "شراء عدد" in val:
        ws.row_dimensions[row[0].row].height = 35
        row[0].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
        
    elif "الكمية=" in val:
        ws.row_dimensions[row[0].row].height = 32
        row[0].font = Font(name='Cairo', size=10, color=TEXT_LIGHT)
        
    elif "شراء فلاتر زيت وديزل" in val:
        ws.row_dimensions[row[0].row].height = 38
        row[0].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        row[0].fill = fl(PRIMARY)
        
    elif val == 'م' and main_header_row and row[0].row > main_header_row:
        filter_header_row = row[0].row
        ws.row_dimensions[filter_header_row].height = 30
        for cell in row:
            if cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=WHITE)
                cell.fill = fl(PRIMARY)
                cell.border = header_border
                
    elif "فلاتر الديزل" in val:
        ws.row_dimensions[row[0].row].height = 32
        row[0].font = Font(name='Cairo', size=11, bold=True, color=WHITE)
        row[0].fill = fl("34495E")
        
    elif "وبناءا عليه تم طلب" in val:
        ws.row_dimensions[row[0].row].height = 38
        row[0].font = Font(name='Cairo', size=12, bold=True, color=PRIMARY)
        row[0].border = Border(top=Side(style='medium', color=ACCENT), bottom=Side(style='medium', color=ACCENT))
        
    elif "التعميد" in val:
        row[0].font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
        if ws.cell(row=row[0].row, column=5).value:
            ws.cell(row=row[0].row, column=5).font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)

# Style Main Data
if main_header_row:
    for r in range(main_header_row + 1, 100):
        if ws.cell(row=r, column=1).value == "📋  ملاحظات وتفاصيل الفلاتر" or str(ws.cell(row=r, column=1).value).strip() == "":
            break
        ws.row_dimensions[r].height = 27
        bg = fl(EVEN_ROW) if r % 2 == 0 else fl(ODD_ROW)
        
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.fill = bg
            if c == 1 and cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)
            elif c == 2 and cell.value:
                cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
            elif c in (9, 10) and cell.value:
                if "لا يوجد" in str(cell.value):
                    cell.font = Font(name='Cairo', size=9, color=TEXT_LIGHT)
                else:
                    cell.font = Font(name='Cairo', size=10, bold=True, color=GREEN_TEXT)

# Style Filter Data
if filter_header_row:
    for r in range(filter_header_row + 1, 100):
        if ws.cell(row=r, column=1).value == "فلاتر الديزل" or str(ws.cell(row=r, column=1).value).strip() == "":
            break
        ws.row_dimensions[r].height = 25
        bg = fl(EVEN_ROW) if r % 2 == 0 else fl(ODD_ROW)
        
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.fill = bg
            cell.border = thin_border
            
            val = str(cell.value) if cell.value is not None else ""
            if c in (1, 6) and val: # م
                cell.font = Font(name='Cairo', size=9, bold=True, color=PRIMARY)
            elif c in (2, 7) and val: # نوع الفلتر
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.font = Font(name='Cairo', size=9, color=TEXT_DARK)
            elif c in (3, 8) and val: # السابق
                cell.font = Font(name='Cairo', size=10, bold=True, color=TEXT_DARK)
            elif c in (4, 9) and val: # المستهلك
                cell.font = Font(name='Cairo', size=10, bold=True, color=RED_TEXT)
            elif c in (5, 10) and val: # المتبقي
                cell.font = Font(name='Cairo', size=10, bold=True, color=PRIMARY)

# Adjust widths based on 10 columns:
widths = [5, 15, 8, 12, 11, 5, 15, 8, 12, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(output_file)



