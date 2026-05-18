import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(script_dir, 'جدول غسيل 2026-4-25.xlsx')
output_file = os.path.join(script_dir, 'جدول_غسيل_احترافي_2026.xlsx')
img_0_path = os.path.join(script_dir, 'source_img_0.png')
img_1_path = os.path.join(script_dir, 'source_img_1.png')

wb = openpyxl.load_workbook(input_file)
ws = wb.active
ws.sheet_view.rightToLeft = True

# Hide gridlines for clarity
ws.sheet_view.showGridLines = False

# ===== COLORS & STYLES =====
PRIMARY = "1A3A5C"
ACCENT = "C8A45A"
WHITE = "FFFFFF"
EVEN_ROW = "F0EDE6"
ODD_ROW = "FFFFFF"
TEXT_DARK = "2C3E50"
GREEN_TEXT = "27AE60"
RED_TEXT = "C0392B"

thin_side = Side(style='thin', color='C0BDB5')
med_side = Side(style='medium', color=PRIMARY)
accent_side = Side(style='medium', color=ACCENT)

thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_border = Border(left=Side(style='thin', color='3A5A7C'), right=Side(style='thin', color='3A5A7C'), top=med_side, bottom=med_side)

def fl(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

# Insert rows at the top for logos if not enough space
ws.insert_rows(1, 4)

# Try to add logos
try:
    img_ar = XlImage(img_1_path)
    img_ar.width = 200
    img_ar.height = 70
    ws.add_image(img_ar, 'A1')
    
    img_en = XlImage(img_0_path)
    img_en.width = 200
    img_en.height = 70
    ws.add_image(img_en, 'N1') # Far right for landscape
except Exception as e:
    pass

ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 40

# Locate header row after inserting rows
header_row_idx = None
for r in range(1, 15):
    if str(ws.cell(row=r, column=1).value).strip() == 'م':
        header_row_idx = r
        # Ensure month 2 is not missing
        if not ws.cell(row=r, column=6).value or "شهر" not in str(ws.cell(row=r, column=6).value):
            ws.cell(row=r, column=6).value = "شهر 2"
        break

# Title
title_row = header_row_idx - 2 if header_row_idx and header_row_idx > 2 else 5
ws.merge_cells(f'A{title_row}:P{title_row}')
title_cell = ws.cell(row=title_row, column=1)
title_cell.value = "جدول غسيل السيارات لفرع الدمام لعام 2026"
title_cell.font = Font(name='Cairo', size=16, bold=True, color=PRIMARY)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
for c in range(1, 17):
    ws.cell(row=title_row, column=c).border = Border(bottom=accent_side)
ws.row_dimensions[title_row].height = 40

# Format Header
if header_row_idx:
    ws.row_dimensions[header_row_idx].height = 35
    for c in range(1, 17):
        cell = ws.cell(row=header_row_idx, column=c)
        cell.font = Font(name='Cairo', size=12, bold=True, color=WHITE)
        cell.fill = fl(PRIMARY)
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Format Data Rows
    last_row = header_row_idx
    for r in range(header_row_idx + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None and ws.cell(row=r, column=2).value is None:
            continue
        last_row = r
            
        ws.row_dimensions[r].height = 25
        bg = fl(EVEN_ROW) if r % 2 == 0 else fl(ODD_ROW)
        
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            cell.fill = bg
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            val = str(cell.value).strip() if cell.value else ""
            if c == 1 and val:
                cell.font = Font(name='Cairo', size=11, bold=True, color=PRIMARY)
            elif c in (2, 3, 4) and val:
                cell.font = Font(name='Cairo', size=11, bold=True, color=TEXT_DARK)
            elif c >= 5 and val:
                if "استلم" in val:
                    cell.font = Font(name='Cairo', size=11, bold=True, color=GREEN_TEXT)
                elif "لا" in val or "لم" in val:
                    cell.font = Font(name='Cairo', size=11, bold=True, color=RED_TEXT)
                else:
                    cell.font = Font(name='Cairo', size=11, color=TEXT_DARK)

# Set Column Widths
widths = [5, 20, 20, 15] + [10]*12
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25

wb.save(output_file)


